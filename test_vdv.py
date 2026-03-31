# test_vdv.py — Compare old vs new model on real VDV bookings
# Old model: trained on Kaggle only (Portuguese hotels 2015-2017)
# New model: trained on Kaggle + Van der Valk Mechelen 2025

import pickle
import re
import numpy as np
import pandas as pd
import openpyxl
from datetime import datetime, timedelta

BASE_DIR   = r'C:\Users\jpdou\Desktop\Occupado'
ANON_DIR   = rf'{BASE_DIR}\VDV-MEC\anonymized'
OLD_MODEL  = rf'{BASE_DIR}\occupado_model_vdv.pkl'   # backup of old pkl saved earlier
NEW_MODEL  = rf'{BASE_DIR}\occupado_model.pkl'

FEATURES = [
    'lead_time', 'arrival_date_week_number',
    'stays_in_weekend_nights', 'stays_in_week_nights',
    'adults', 'is_repeated_guest',
    'previous_cancellations', 'previous_bookings_not_canceled',
    'booking_changes', 'days_in_waiting_list',
    'adr', 'total_of_special_requests',
]

# ---------------------------------------------------------------------------
# Helpers (same as train_vdv.py)
# ---------------------------------------------------------------------------

def parse_date(val):
    if val is None: return None
    s = str(val).strip()[:16]
    for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%Y'):
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

def parse_rate(val):
    if val is None: return None
    s = re.sub(r'[^\d,\.]', '', str(val))
    s = s.replace('.','').replace(',','.') if (',' in s and '.' in s) else s.replace(',','.')
    try: return float(s)
    except: return None

def parse_adults(val):
    if val is None: return None
    try: return int(str(val).split('/')[0])
    except: return None

def is_hash(val):
    return bool(val and re.match(r'^[0-9a-f]{60,}', str(val)))

def count_nights(arr, n):
    we = wd = 0
    for i in range(max(0, int(n))):
        d = arr + timedelta(days=i)
        (we if d.weekday() >= 5 else wd).__class__  # dummy
        if d.weekday() >= 5: we += 1
        else: wd += 1
    return we, wd

# ---------------------------------------------------------------------------
# Load 30 cancelled + 30 completed stays from VDV data
# ---------------------------------------------------------------------------

def sample_cancelled(path, n=30):
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    records = []
    i = 0
    while i < len(rows) and len(records) < n:
        row = rows[i]
        if is_hash(row[0]):
            for j in range(i+1, min(i+5, len(rows))):
                r = rows[j]
                if r[0] is None and r[9] is not None and '/' in str(r[9]):
                    arr  = parse_date(row[8])
                    dep  = parse_date(r[8])
                    con  = parse_date(row[14])
                    adlt = parse_adults(r[9])
                    rate = parse_rate(r[11])
                    n_   = row[9]
                    if arr:
                        try: nights = int(n_)
                        except: nights = 0
                        if nights == 0 and dep: nights = max(0,(dep-arr).days)
                        we, wd = count_nights(arr, nights)
                        lead = (arr-con).days if con else 68
                        records.append({
                            'lead_time': lead,
                            'arrival_date_week_number': int(arr.isocalendar()[1]),
                            'stays_in_weekend_nights': we,
                            'stays_in_week_nights': wd,
                            'adults': adlt or 1,
                            'is_repeated_guest': 0,
                            'previous_cancellations': 0,
                            'previous_bookings_not_canceled': 0,
                            'booking_changes': 0,
                            'days_in_waiting_list': 0,
                            'adr': rate or 150,
                            'total_of_special_requests': 0,
                            'actual': 'CANCELLED',
                        })
                    break
        i += 1
    return records


def sample_completed(path, n=30):
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    STATUSES = {'CO', 'NS', 'CX', 'IH', 'DI', 'DO'}
    records = []
    pending1 = pending3 = None

    for row in rows:
        if len(records) >= n: break
        col1 = row[1] if len(row) > 1 else None
        col7 = row[7] if len(row) > 7 else None
        if is_hash(col1) and row[4] is not None:
            pending1 = row; pending3 = None
        elif row[0] == 'DNM' and pending1 is not None:
            pending3 = row
        elif (col1 and str(col1).startswith('MEC-') and
              col7 and str(col7).strip() in STATUSES and pending1 is not None):
            if str(col7).strip() == 'CO':
                arr  = parse_date(pending1[4])
                dep  = parse_date(row[4])
                adlt = parse_adults(pending1[8])
                rate = parse_rate(row[12])
                n_   = pending1[7]
                if arr:
                    try: nights = int(n_)
                    except: nights = 0
                    if nights == 0 and dep: nights = max(0,(dep-arr).days)
                    we, wd = count_nights(arr, nights)
                    records.append({
                        'lead_time': 68,
                        'arrival_date_week_number': int(arr.isocalendar()[1]),
                        'stays_in_weekend_nights': we,
                        'stays_in_week_nights': wd,
                        'adults': adlt or 1,
                        'is_repeated_guest': 0,
                        'previous_cancellations': 0,
                        'previous_bookings_not_canceled': 0,
                        'booking_changes': 0,
                        'days_in_waiting_list': 0,
                        'adr': rate or 150,
                        'total_of_special_requests': 0,
                        'actual': 'COMPLETED',
                    })
            pending1 = pending3 = None

    return records

# ---------------------------------------------------------------------------
# Run comparison
# ---------------------------------------------------------------------------

print('=' * 60)
print('  Occupado — Old vs New Model Test on VDV Data')
print('=' * 60)

print('\nLoading models...')
with open(NEW_MODEL, 'rb') as f:
    new_model = pickle.load(f)

# Try to load old model backup (occupado_model_vdv.pkl was saved before overwrite)
try:
    with open(OLD_MODEL, 'rb') as f:
        raw = pickle.load(f)
        old_model = raw['model'] if isinstance(raw, dict) else raw
    has_old = True
    print('  Old model (VDV-only 10-feat): loaded')
except:
    has_old = False
    print('  Old model: not available — showing new model only')
print('  New model (Kaggle+VDV 12-feat): loaded')

print('\nSampling VDV bookings...')
cancelled  = sample_cancelled(rf'{ANON_DIR}\RES_036_CancelledReservations (1).xlsx', 30)
completed  = sample_completed(rf'{ANON_DIR}\RES_001_ArrivalDetailed.xlsx', 30)
all_bookings = cancelled + completed
print(f'  {len(cancelled)} cancelled + {len(completed)} completed stays')

df = pd.DataFrame(all_bookings)
X  = df[FEATURES]

new_scores = new_model.predict_proba(X)[:, 1] * 100
new_preds  = ['CANCELLED' if s >= 50 else 'COMPLETED' for s in new_scores]

if has_old:
    old_scores = old_model.predict_proba(X[old_model.feature_names_in_])[:, 1] * 100
    old_preds  = ['CANCELLED' if s >= 50 else 'COMPLETED' for s in old_scores]

# Summary stats
print('\n' + '=' * 60)
print('  RESULTS SUMMARY')
print('=' * 60)

new_correct = sum(p == a for p, a in zip(new_preds, df['actual']))
print(f'\n  New model accuracy on VDV sample: {new_correct}/{len(df)} ({100*new_correct//len(df)}%)')

if has_old:
    old_correct = sum(p == a for p, a in zip(old_preds, df['actual']))
    print(f'  Old model accuracy on VDV sample: {old_correct}/{len(df)} ({100*old_correct//len(df)}%)')

# Cancelled bookings — risk scores
print('\n--- CANCELLED bookings (should score HIGH) ---')
print(f'  {"Actual":<12} {"New Score":>10}' + (f' {"Old Score":>10}' if has_old else ''))
cxl_new = [new_scores[i] for i in range(len(cancelled))]
for i in range(min(10, len(cancelled))):
    line = f'  CANCELLED    {new_scores[i]:>8.1f}%'
    if has_old: line += f'  {old_scores[i]:>8.1f}%'
    print(line)
print(f'  Avg risk score: {np.mean(cxl_new):.1f}%' + (f'  (old: {np.mean(old_scores[:len(cancelled)]):.1f}%)' if has_old else ''))

# Completed stays — risk scores
print('\n--- COMPLETED stays (should score LOW) ---')
co_new = [new_scores[i] for i in range(len(cancelled), len(df))]
for i in range(len(cancelled), min(len(cancelled)+10, len(df))):
    line = f'  COMPLETED    {new_scores[i]:>8.1f}%'
    if has_old: line += f'  {old_scores[i]:>8.1f}%'
    print(line)
print(f'  Avg risk score: {np.mean(co_new):.1f}%' + (f'  (old: {np.mean(old_scores[len(cancelled):]):.1f}%)' if has_old else ''))

print('\n' + '=' * 60)
print('  A good model shows HIGH scores for cancelled,')
print('  LOW scores for completed stays.')
print('=' * 60)
