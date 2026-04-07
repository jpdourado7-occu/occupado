"""
Train a VdV-specific cancellation risk model using only features
reliably available from Shiji PMS exports.

Features used (8):
  arrival_date_week_number, arrival_month, arrival_day_of_week,
  stays_in_weekend_nights, stays_in_week_nights,
  is_repeated_guest, lead_time, channel_encoded

Data sources:
  Completed stays : RES_004 (5/6/7/8) — past arrivals with booking creation dates
  Cancellations   : RES_036 (all) + RES_037 (all)
  Repeat guests   : RES_042 (all) — marks known repeat guests in training data

channel_encoded: 0=OTA/Web, 1=Direct, 2=Corporate, 3=Group/Package, NaN=Unknown
lead_time: days between booking creation and arrival (real value for all records)

Output: occupado_model_vdv.pkl
"""
import os, re, pickle, warnings
import numpy as np
import pandas as pd
from datetime import datetime

warnings.filterwarnings('ignore')

VDV_DIR = 'VDV-MEC'
FEATURES = [
    'lead_time',
    'arrival_date_week_number',
    'arrival_month',
    'arrival_day_of_week',
    'stays_in_weekend_nights', 'stays_in_week_nights',
    'is_repeated_guest',
    'channel_encoded',
    'channel_cancel_rate',
    'seasonal_cancel_rate',
    'avg_days_to_cancel_for_channel',
    'is_last_minute',
    'is_early_bird',
    'is_business_pattern',
    'deposit_risk',
]

# GTD codes → deposit risk score (mirrors app.py _VDV_GTD_RISK)
_VDV_GTD_RISK = {
    'PRE':    0.05,
    'ADV':    0.10,
    'CREDIT': 0.20,
    'CRP':    0.20,
    'CRPCL':  0.25,
    'VCC':    0.35,
    'HOLD18': 0.75,
    'NONE':   0.90,
    'None':   0.90,
    '':       0.90,
}

def gtd_to_deposit_risk(gtd_code):
    return _VDV_GTD_RISK.get(str(gtd_code).strip(), 0.90)

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def parse_date(s):
    s = str(s).strip()[:10]
    for fmt in ('%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d'):
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

def weekend_split(arrival, nights):
    if arrival is None or nights <= 0: return 0, max(nights, 0)
    wknd = sum(1 for d in range(int(nights))
               if (arrival + pd.Timedelta(days=d)).weekday() in (4, 5))
    return wknd, int(nights) - wknd

def safe_float(s):
    try: return float(str(s).replace('.','').replace(',','.').strip())
    except: return None

# RES_036 sub-row col[3] market segment → channel
_SEG_MAP = {
    'BARWEB': 0, 'BAROTAGROSS': 0, 'DEALSOTA': 0, 'DISCOTAGROSS': 0,
    'DISCWEB': 0, 'DISCOTA': 0, 'BAROTA': 0, 'DEALSOTAWEB': 0,
    'DISCDIR': 1, 'BARDIR': 1, 'DISCBBDIR': 1,
    'CORPFIX': 2, 'CORPDYN': 2,
    'MTGBNS': 3, 'BNSGRP': 3, 'PACK': 3,
}

# RES_004 col[25] channel/source → channel
_CH25_MAP = {
    'IBE': 0,        # Internet Booking Engine (OTA/Web)
    'OTH': 0,        # Other online
    'GDS': 0,        # Global Distribution System
    'DIRECT': 1,     # Direct
    'PHONE': 1,      # Phone/direct
    'EMAIL': 1,      # Email/direct
    'CORP': 2,       # Corporate
    'GROUP': 3,      # Group
}

def seg_to_channel(seg):
    return float(_SEG_MAP.get(str(seg).upper().strip(), float('nan')))

def ch25_to_channel(val):
    """Map RES_004 col[25] channel code to channel int."""
    if not val: return float('nan')
    v = str(val).upper().strip()
    if v in _CH25_MAP: return float(_CH25_MAP[v])
    if 'OTA' in v or 'WEB' in v or 'IBE' in v: return 0.0
    if 'DIRECT' in v or 'PHONE' in v: return 1.0
    if 'CORP' in v: return 2.0
    if 'GROUP' in v or 'GRP' in v or 'MTG' in v: return 3.0
    return float('nan')

def rate_plan_to_channel(code):
    """Map rate plan code (col[13] in RES_004) to channel int as fallback."""
    if not code: return float('nan')
    c = str(code).upper()
    if 'OTA' in c or ('WEB' in c and 'DIR' not in c): return 0.0
    if 'DIR' in c: return 1.0
    if 'CORP' in c or c.startswith('CRPL'): return 2.0
    if any(x in c for x in ('MTG', 'BNS', 'PACK', 'GRP')): return 3.0
    return float('nan')


# ─────────────────────────────────────────────────────────────────────────────
# PARSE CANCELLATIONS (RES_036 x3 + RES_037 x2)
# ─────────────────────────────────────────────────────────────────────────────
def parse_cancellations():
    """
    RES_036 row structure (0-indexed):
      Guest row:  col[8]=arrival, col[9]=nights, col[13]=rate_amt, col[14]=created
      Sub-rows:   col[3]=market_segment, col[9]=adults/children
    Returns (DataFrame, set of (arrival_date_str, nights) keys for cross-ref).
    """
    import openpyxl
    records = []
    seen_keys = set()  # (arrival_str, nights_or_dep, created_str) — dedup across files
    files = sorted([f for f in os.listdir(VDV_DIR)
                    if f.startswith(('RES_036', 'RES_037')) and f.endswith('.xlsx')])
    print(f'Cancellation files: {files}')

    for fn in files:
        path = os.path.join(VDV_DIR, fn)
        wb = openpyxl.load_workbook(path, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        file_count = 0

        if fn.startswith('RES_037'):
            # No-show layout: col[1]=name, col[2]=conf#, col[9]=market_seg,
            #   col[11]=arrival, col[13]=departure, col[16]=created, col[24]=rate
            for i, row in enumerate(rows):
                if i < 7: continue
                if not row or len(row) < 14: continue
                c1  = str(row[1]).strip() if row[1]  is not None else ''
                c2  = str(row[2]).strip() if row[2]  is not None else ''
                c11 = str(row[11]).strip() if row[11] is not None else ''
                c13 = str(row[13]).strip() if row[13] is not None else ''
                c16 = str(row[16]).strip() if row[16] is not None else ''
                # Guest row: name has comma, conf# starts with MEC- (not folio)
                if ',' not in c1: continue
                if not c2.startswith('MEC-') or c2.startswith('MEC-F'): continue
                if not re.match(r'\d{2}/\d{2}/\d{4}$', c11): continue
                dedup_key = (c2, c11)  # conf# + arrival date
                if dedup_key in seen_keys: continue
                seen_keys.add(dedup_key)
                arr  = parse_date(c11)
                dep  = parse_date(c13)
                if arr is None: continue
                nights = max(0, (dep - arr).days) if dep else 1
                created = parse_date(c16) if re.match(r'\d{2}/\d{2}/\d{4}', c16) else None
                rate = safe_float(str(row[24]).strip() if len(row) > 24 and row[24] else '')
                seg  = str(row[9]).strip() if row[9] is not None else ''
                channel = seg_to_channel(seg)
                lead = max(0, (arr - created).days) if created else 25
                adr  = rate / max(nights, 1) if rate and nights > 0 else rate
                wknd, wkday = weekend_split(arr, nights)
                records.append({
                    'arrival': arr, 'nights': nights, 'adults': 1,
                    'lead_time': lead,
                    'arrival_date_week_number': int(arr.isocalendar()[1]),
                    'arrival_month': arr.month,
                    'arrival_day_of_week': arr.weekday(),
                    'stays_in_weekend_nights': wknd,
                    'stays_in_week_nights': wkday,
                    'adr': adr,
                    'is_repeated_guest': 0,
                    'channel_encoded': channel,
                    'deposit_risk': 0.90,  # GTD not available in RES_037 — default no guarantee
                    'is_canceled': 1,
                    'source': fn
                })
                file_count += 1
        else:
            # RES_036 layout: col[8]=arrival, col[9]=nights, col[13]=rate, col[14]=created
            for i, row in enumerate(rows):
                if i < 7: continue
                if not row or len(row) < 15: continue
                c8  = str(row[8]).strip()  if row[8]  is not None else ''
                c9  = str(row[9]).strip()  if row[9]  is not None else ''
                c13 = str(row[13]).strip() if row[13] is not None else ''
                c14 = str(row[14]).strip() if row[14] is not None else ''

                if not re.match(r'\d{2}/\d{2}/\d{4}$', c8): continue
                if not c9.isdigit(): continue

                # col[20] = MEC-CXL-XXXXXX (unique cancellation reference)
                cxl_ref = str(row[20]).strip() if len(row) > 20 and row[20] is not None else ''
                dedup_key = cxl_ref if cxl_ref.startswith('MEC-CXL') else (c8, c9, c14)
                if dedup_key in seen_keys: continue
                seen_keys.add(dedup_key)

                arr     = parse_date(c8)
                nights  = int(c9)
                rate    = safe_float(c13)
                created = parse_date(c14)
                if arr is None: continue

                adults = 1
                channel = float('nan')
                for j in range(i + 1, min(i + 6, len(rows))):
                    sub = rows[j]
                    if not sub or len(sub) < 10: continue
                    v = str(sub[9]).strip() if sub[9] is not None else ''
                    if '/' in v and adults == 1:
                        parts = v.split('/')
                        if len(parts) == 2 and all(p.strip().isdigit() for p in parts):
                            a = int(parts[0].strip())
                            if 1 <= a <= 10: adults = a
                    if sub[3] is not None and (isinstance(channel, float) and channel != channel):
                        ch = seg_to_channel(str(sub[3]))
                        if not (isinstance(ch, float) and ch != ch):
                            channel = ch

                lead = max(0, (arr - created).days) if created else 25
                adr  = rate / max(nights, 1) if rate and nights > 0 else rate
                wknd, wkday = weekend_split(arr, nights)

                records.append({
                    'arrival': arr, 'nights': nights, 'adults': adults,
                    'lead_time': lead,
                    'arrival_date_week_number': int(arr.isocalendar()[1]),
                    'arrival_month': arr.month,
                    'arrival_day_of_week': arr.weekday(),
                    'stays_in_weekend_nights': wknd,
                    'stays_in_week_nights': wkday,
                    'adr': adr,
                    'is_repeated_guest': 0,
                    'channel_encoded': channel,
                    'deposit_risk': 0.90,  # GTD not available in RES_036 — default no guarantee
                    'is_canceled': 1,
                    'source': fn
                })
                file_count += 1

        print(f'  {fn}: {file_count:,} records')

    df = pd.DataFrame(records) if records else pd.DataFrame()
    if len(df): df = df[df['lead_time'] < 730]
    ch_known = df['channel_encoded'].notna().sum() if len(df) else 0
    print(f'  Cancellations parsed: {len(df):,}  (channel known: {ch_known:,})')

    # Build cross-ref set: (arr_str, nights) for mislabel prevention
    cxl_keys = set()
    for _, r in df.iterrows():
        cxl_keys.add((r['arrival'].strftime('%Y-%m-%d'), int(r['nights'])))
    return df, cxl_keys


# ─────────────────────────────────────────────────────────────────────────────
# PARSE COMPLETED STAYS — RES_004 historical (past arrival dates)
# ─────────────────────────────────────────────────────────────────────────────
def parse_res004_stays(cxl_keys):
    """
    Parse RES_004 files for past arrivals. These records have real lead_time
    (booking creation date in col[28]) and channel (col[25]).
    Exclude any record whose (arrival_date, nights) key appears in cxl_keys
    (known cancellations from RES_036/037) to avoid mislabeling.

    RES_004 column map (verified):
      col[0]  = guest name (Lastname, Firstname)
      col[3]  = MEC-XXXXXX confirmation number
      col[8]  = arrival datetime (DD/MM/YYYY HH:MM)
      col[9]  = nights
      col[12] = pax ("2 / 0")
      col[13] = rate plan code
      col[16] = total rate amount
      col[25] = channel/source code
      col[28] = booking creation datetime
    """
    import openpyxl
    today = datetime.now().date()
    records = []
    seen_keys = set()  # (arr_str, nights, created_str) for dedup across files

    files = sorted([f for f in os.listdir(VDV_DIR)
                    if f.startswith('RES_004') and f.endswith('.xlsx')])
    print(f'RES_004 historical files: {files}')

    for fn in files:
        path = os.path.join(VDV_DIR, fn)
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            print(f'  Skip {fn}: {e}'); continue

        # Build an index of sub-rows for fast GTD lookup: row_index → gtd_code
        subrow_gtd = {}
        for idx, row in enumerate(rows):
            if (row and len(row) > 12
                    and len(row) > 9
                    and str(row[9]).strip() == 'RS'
                    and row[12] is not None):
                subrow_gtd[idx] = str(row[12]).strip()

        file_records = 0
        for i, r in enumerate(rows):
            if not r or len(r) < 29: continue
            c0 = str(r[0]).strip() if r[0] else ''
            c3 = str(r[3]).strip() if len(r) > 3 and r[3] else ''
            if not (',' in c0 and c3.startswith('MEC-') and not c3.startswith('MEC-F')):
                continue
            if not r[8]: continue

            arr_str = str(r[8])[:10]
            try: arr = datetime.strptime(arr_str, '%d/%m/%Y').date()
            except: continue
            if arr >= today: continue  # only past arrivals

            nights_raw = str(r[9]).strip() if r[9] else ''
            if not nights_raw.isdigit(): continue
            nights = int(nights_raw)

            created_str = str(r[28])[:10] if r[28] else ''
            if not created_str: continue
            try: created = datetime.strptime(created_str, '%d/%m/%Y').date()
            except: continue

            lead = max(0, (arr - created).days)
            if lead > 730: continue

            # Skip if this is a known cancellation
            arr_key = arr.strftime('%Y-%m-%d')
            if (arr_key, nights) in cxl_keys: continue

            # Dedup across files
            dedup_key = (arr_key, nights, created_str)
            if dedup_key in seen_keys: continue
            seen_keys.add(dedup_key)

            # Pax
            adults = 1
            pax = str(r[12]).strip() if len(r) > 12 and r[12] else ''
            if '/' in pax:
                parts = pax.split('/')
                if all(p.strip().isdigit() for p in parts):
                    a = int(parts[0].strip())
                    if 1 <= a <= 10: adults = a

            # Channel: try col[25] first, fall back to rate plan col[13]
            ch25 = str(r[25]).strip() if len(r) > 25 and r[25] else ''
            channel = ch25_to_channel(ch25)
            if isinstance(channel, float) and channel != channel:  # still NaN
                rp = str(r[13]).strip() if len(r) > 13 and r[13] else ''
                channel = rate_plan_to_channel(rp)

            # Rate
            rate_raw = str(r[16]).strip() if len(r) > 16 and r[16] else ''
            rate = safe_float(rate_raw)
            adr = rate / max(nights, 1) if rate and rate > 0 else None

            arr_dt = datetime.combine(arr, datetime.min.time())
            wknd, wkday = weekend_split(arr_dt, nights)

            # GTD from sub-row (col9 == 'RS', col12 == GTD code)
            gtd_code = 'NONE'
            for j in range(i + 1, min(i + 8, len(rows))):
                if j in subrow_gtd:
                    gtd_code = subrow_gtd[j]
                    break

            records.append({
                'arrival': arr_dt, 'nights': nights, 'adults': adults,
                'lead_time': lead,
                'arrival_date_week_number': int(arr_dt.isocalendar()[1]),
                'arrival_month': arr_dt.month,
                'arrival_day_of_week': arr_dt.weekday(),
                'stays_in_weekend_nights': wknd,
                'stays_in_week_nights': wkday,
                'adr': adr,
                'is_repeated_guest': 0,
                'channel_encoded': channel,
                'deposit_risk': gtd_to_deposit_risk(gtd_code),
                'is_canceled': 0,
                'source': fn
            })
            file_records += 1

        print(f'  {fn}: {file_records:,} completed stays')

    df = pd.DataFrame(records) if records else pd.DataFrame()
    ch_known = df['channel_encoded'].notna().sum() if len(df) else 0
    print(f'  RES_004 completed stays total: {len(df):,}  (channel known: {ch_known:,})')
    return df


# ─────────────────────────────────────────────────────────────────────────────
# ENRICH WITH REPEAT GUESTS (RES_042)
# ─────────────────────────────────────────────────────────────────────────────
def enrich_repeat_guests(df):
    import openpyxl
    repeat_dates = set()
    files = sorted([f for f in os.listdir(VDV_DIR)
                    if re.match(r'RES_042_RepeatReservationsReport.*\.xlsx$', f)])
    if not files:
        return df
    print(f'  RES_042 files: {files}')
    for fn in files:
        path = os.path.join(VDV_DIR, fn)
        wb = openpyxl.load_workbook(path, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        for row in rows:
            if not row or row[0] is None: continue
            col4 = row[4] if len(row) > 4 else None
            if col4 and '/' in str(col4):
                d = parse_date(str(col4))
                if d: repeat_dates.add(d.strftime('%Y-%m-%d'))
    df['is_repeated_guest'] = df['arrival'].apply(
        lambda x: 1 if x.strftime('%Y-%m-%d') in repeat_dates else 0
    )
    print(f'  Repeat guests marked: {df["is_repeated_guest"].sum():,} '
          f'(from {len(files)} RES_042 file(s), {len(repeat_dates):,} repeat dates)')
    return df


# ─────────────────────────────────────────────────────────────────────────────
# BUILD DATASET
# ─────────────────────────────────────────────────────────────────────────────
print('='*60)
print('BUILDING VdV TRAINING DATASET')
print('='*60)

cancels, cxl_keys = parse_cancellations()
stays = parse_res004_stays(set())  # no cross-ref dedup — key (arr,nights) is not unique per guest

df = pd.concat([cancels, stays], ignore_index=True)
df = enrich_repeat_guests(df)

median_adr = df['adr'].median() if df['adr'].notna().any() else 160.0
df['adr'] = df['adr'].fillna(median_adr)
df = df.dropna(subset=['lead_time'])

ch_coverage = df['channel_encoded'].notna().mean()
print(f'\nFinal dataset: {len(df):,} records')
print(f'  Cancellations: {df["is_canceled"].sum():,} ({df["is_canceled"].mean():.1%})')
print(f'  Completed stays: {(df["is_canceled"]==0).sum():,}')
print(f'  Channel coverage: {ch_coverage:.1%}')
print(f'  Lead time: median {df["lead_time"].median():.0f}d, mean {df["lead_time"].mean():.0f}d, max {df["lead_time"].max():.0f}d')

# ── Compute channel-level features from training data ────────────────────────
# channel_cancel_rate: historical cancel rate per channel in training data
_ch_rate = df.groupby('channel_encoded')['is_canceled'].mean()
overall_rate = float(df['is_canceled'].mean())
df['channel_cancel_rate'] = df['channel_encoded'].map(_ch_rate).fillna(overall_rate)

# seasonal_cancel_rate: cancel rate per (channel_encoded, arrival_month)
_sea_rate = df.groupby(['channel_encoded', 'arrival_month'])['is_canceled'].mean()
_sea_rate_dict = _sea_rate.to_dict()
df['seasonal_cancel_rate'] = df.apply(
    lambda r: _sea_rate_dict.get((r['channel_encoded'], r['arrival_month']),
                                  r['channel_cancel_rate']), axis=1
)

# avg_days_to_cancel_for_channel: mean lead_time of cancelled records per channel
# (lead_time used as proxy; actual CXL date not separately extracted in training data)
_avg_dtc = df[df['is_canceled'] == 1].groupby('channel_encoded')['lead_time'].mean()
df['avg_days_to_cancel_for_channel'] = df['channel_encoded'].map(_avg_dtc).fillna(30.0)

print(f'  Channel cancel rates: { {int(k): round(v,2) for k,v in _ch_rate.items()} }')

df['is_last_minute'] = (df['lead_time'] <= 3).astype(int)
df['is_early_bird']  = (df['lead_time'] >= 60).astype(int)
df['is_business_pattern'] = (
    (df['stays_in_week_nights'] >= 3) &
    (df['stays_in_weekend_nights'] == 0)
).astype(int)

print("[TRAIN] Micro-segment distribution:")
print(f"  Last minute: {df['is_last_minute'].sum()}")
print(f"  Early bird:  {df['is_early_bird'].sum()}")
print(f"  Business:    {df['is_business_pattern'].sum()}")

print("[TRAIN] GTD distribution (deposit_risk):")
print(df['deposit_risk'].value_counts().sort_index())

print("[TRAIN] Feature null check:")
print(df[FEATURES].isnull().sum())
print("[TRAIN] Feature sample:")
print(df[FEATURES].head(3))

# Keep channel_encoded as float with NaN — XGBoost handles natively
X = df[FEATURES].apply(pd.to_numeric, errors='coerce').astype(float)
y = df['is_canceled'].astype(int)

# ─────────────────────────────────────────────────────────────────────────────
# TRAIN — XGBoost with class weight balancing
# ─────────────────────────────────────────────────────────────────────────────
from xgboost import XGBClassifier
from sklearn.model_selection import StratifiedKFold, cross_validate
from sklearn.metrics import (accuracy_score, precision_score, recall_score,
                             roc_auc_score)

neg, pos = (y==0).sum(), (y==1).sum()
scale_pos = neg / pos
print(f'\nClass ratio: {neg:,} stays / {pos:,} cancellations -> scale_pos_weight={scale_pos:.2f}')

model = XGBClassifier(
    n_estimators=300,
    max_depth=5,
    learning_rate=0.05,
    subsample=0.8,
    colsample_bytree=0.8,
    scale_pos_weight=scale_pos,
    eval_metric='auc',
    random_state=42,
    verbosity=0
)

print('\nRunning 5-fold cross-validation...')
cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
cv_results = cross_validate(model, X, y, cv=cv,
    scoring=['accuracy','precision','recall','roc_auc'],
    return_train_score=False)

print('\n=== CROSS-VALIDATION RESULTS (5-fold) ===')
print(f'  AUC-ROC:   {cv_results["test_roc_auc"].mean():.3f} ± {cv_results["test_roc_auc"].std():.3f}')
print(f'  Accuracy:  {cv_results["test_accuracy"].mean():.1%} ± {cv_results["test_accuracy"].std():.1%}')
print(f'  Precision: {cv_results["test_precision"].mean():.1%} ± {cv_results["test_precision"].std():.1%}')
print(f'  Recall:    {cv_results["test_recall"].mean():.1%} ± {cv_results["test_recall"].std():.1%}')

print('\nTraining final model on full dataset...')
model.fit(X, y)

probs = model.predict_proba(X)[:, 1]
preds = (probs >= 0.5).astype(int)

print('\n=== FINAL MODEL — TRAINING SET METRICS ===')
print(f'  AUC-ROC:   {roc_auc_score(y, probs):.3f}')
print(f'  Accuracy:  {accuracy_score(y, preds):.1%}')
print(f'  Precision: {precision_score(y, preds):.1%}')
print(f'  Recall:    {recall_score(y, preds):.1%}')

df2 = df.copy()
df2['score'] = probs
df2['tier'] = pd.cut(probs, bins=[0, .4, .7, 1.0], labels=['Low', 'Medium', 'High'])
tier = df2.groupby('tier', observed=True)['is_canceled'].agg(['mean', 'count'])
print('\n=== ACTUAL CANCEL RATE BY RISK TIER (VdV data) ===')
for t, row in tier.iterrows():
    print(f'  {t:6s}: {row["mean"]:.0%}  ({int(row["count"]):,} bookings)')

ch_labels = {0: 'OTA/Web', 1: 'Direct', 2: 'Corporate', 3: 'Group/Pkg'}
print('\n=== CANCEL RATE BY CHANNEL (where known) ===')
for code, label in ch_labels.items():
    sub = df2[df2['channel_encoded'] == code]
    if len(sub) > 0:
        print(f'  {label:<12s}: {sub["is_canceled"].mean():.0%}  ({len(sub):,} records)')

fi = pd.Series(model.feature_importances_, index=FEATURES).sort_values(ascending=False)
print('\n=== FEATURE IMPORTANCES ===')
for feat, imp in fi.items():
    print(f'  {feat:<35s} {imp:.3f}')

out_path = 'occupado_model_vdv.pkl'
with open(out_path, 'wb') as f:
    pickle.dump(model, f)
print(f'\nModel saved -> {out_path}')
print(f'Features: {FEATURES}')
print('\nDone!')
