"""
Adds to Occupado/app.py:
1. _parse_forecast_data()  — FOR_042 occupancy + revenue trend
2. _parse_group_pipeline() — GRP_017 upcoming group bookings
3. Wires them into build_vdv_dashboard() with new chart + group section
"""
import re

with open('app.py', encoding='utf-8') as f:
    src = f.read()

# ─────────────────────────────────────────────────────────────────────────────
# 1. NEW PARSE FUNCTIONS — inject after _parse_mice_data()
# ─────────────────────────────────────────────────────────────────────────────
NEW_FUNCTIONS = '''

def _parse_forecast_data():
    """Parse FOR_042 HistoryAndForecast — returns last 60d history + next 60d forecast."""
    path = os.path.join(_VDV_DIR, "FOR_042_HistoryAndForecast.xlsx")
    if not os.path.exists(path):
        return {'history': [], 'forecast': []}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        history, forecast, mode = [], [], None
        for row in rows:
            if not row or row[0] is None: continue
            v0 = str(row[0]).strip()
            if v0 == 'HISTORY':   mode = 'h'; continue
            if v0.startswith('FORECAST'): mode = 'f'; continue
            if v0 in ('Totals','', 'Date', 'Sold'): continue
            # Data row: date string like "Wed  01/01/2025"
            date_part = v0.split()[-1] if v0 else ''
            if not re.match(r'\d{2}/\d{2}/\d{4}', date_part): continue
            try:
                dt = datetime.strptime(date_part, '%d/%m/%Y')
                rooms_sold = int(str(row[1]).replace(',','').split('.')[0]) if row[1] else 0
                arrivals   = int(str(row[2]).replace(',','').split('.')[0]) if row[2] else 0
                departures = int(str(row[3]).replace(',','').split('.')[0]) if row[3] else 0
                occ_pct    = float(str(row[4]).replace(',','.')) if row[4] else 0.0
                rev_raw    = str(row[10]).replace('.','').replace(',','.').strip() if row[10] else '0'
                revenue    = float(rev_raw) if rev_raw else 0.0
                entry = {'date': dt.strftime('%d %b'), 'dt': dt.strftime('%Y-%m-%d'),
                         'rooms': rooms_sold, 'arrivals': arrivals, 'departures': departures,
                         'occ': round(occ_pct, 1), 'rev': round(revenue)}
                if mode == 'h':   history.append(entry)
                elif mode == 'f': forecast.append(entry)
            except: continue
        # Return last 60 days of history + next 60 days of forecast
        return {'history': history[-60:], 'forecast': forecast[:60]}
    except Exception as e:
        print(f"[VDV] Forecast parse error: {e}")
        return {'history': [], 'forecast': []}


def _parse_group_pipeline():
    """Parse GRP_017 for upcoming confirmed/deposit/tentative group bookings."""
    path = os.path.join(_VDV_DIR, "GRP_017_GroupAndMEBookingListWithRevenueCSVOnly.xlsx")
    if not os.path.exists(path):
        return []
    STATUS_LABELS = {'DEF': 'Definite', 'DEP': 'Deposit', 'TEN': 'Tentative',
                     'LOS': 'Lost', 'CXL': 'Cancelled', 'PRO': 'Prospect'}
    MARKET_LABELS = {'MTGBNS': 'Meeting', 'CORPFIX': 'Corporate', 'CORPDYN': 'Corporate Dyn',
                     'BNSGRP': 'Business Group', 'LEISURE': 'Leisure'}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        groups = []
        today = datetime.now()
        for i, row in enumerate(rows):
            if i < 5 or not row or row[0] is None: continue
            cols = [str(c).strip() if c is not None else '' for c in row]
            # Col pattern: Profile | M&E Code | Group Name | Arrival | Dep | Status | Market | Owner ...
            if len(cols) < 7: continue
            code   = cols[1]
            name   = cols[2]
            arr_s  = cols[3]
            dep_s  = cols[4]
            status = cols[5]
            market = cols[6]
            # Skip lost/cancelled and past
            if status in ('LOS', 'CXL', ''): continue
            if not re.match(r'\d{2}/\d{2}/\d{4}', arr_s): continue
            try:
                arr_dt = datetime.strptime(arr_s, '%d/%m/%Y')
                dep_dt = datetime.strptime(dep_s, '%d/%m/%Y') if re.match(r'\d{2}/\d{2}/\d{4}', dep_s) else arr_dt
                if arr_dt.date() < today.date(): continue  # past
                nights = (dep_dt - arr_dt).days
                groups.append({
                    'code': code, 'name': name if name else code,
                    'arrival': arr_dt.strftime('%d %b %Y'),
                    'departure': dep_dt.strftime('%d %b %Y'),
                    'arr_dt': arr_dt.strftime('%Y-%m-%d'),
                    'nights': nights,
                    'status': status,
                    'status_label': STATUS_LABELS.get(status, status),
                    'market': MARKET_LABELS.get(market, market),
                })
            except: continue
        # Sort by arrival, cap at 20
        return sorted(groups, key=lambda x: x['arr_dt'])[:20]
    except Exception as e:
        print(f"[VDV] Group pipeline parse error: {e}")
        return []

'''

# Inject after _parse_mice_data function (find its end)
insert_after = "    return result\n\n\n# Load VdV data once at startup"
if insert_after in src:
    src = src.replace(insert_after, "    return result\n" + NEW_FUNCTIONS + "\n# Load VdV data once at startup")
    print("Parse functions injected: OK")
else:
    print("WARNING: injection point not found")

# ─────────────────────────────────────────────────────────────────────────────
# 2. GLOBAL VARIABLES + STARTUP LOADING
# ─────────────────────────────────────────────────────────────────────────────
src = src.replace(
    "VDV_MICE_DATA       = {}",
    "VDV_MICE_DATA       = {}\nVDV_FORECAST_DATA   = {'history': [], 'forecast': []}\nVDV_GROUP_PIPELINE  = []"
)
src = src.replace(
    "    VDV_MICE_DATA       = _parse_mice_data()",
    "    VDV_MICE_DATA       = _parse_mice_data()\n    VDV_FORECAST_DATA   = _parse_forecast_data()\n    VDV_GROUP_PIPELINE  = _parse_group_pipeline()"
)
src = src.replace(
    'print(f"[VDV] Loaded {len(VDV_GUESTS_RAW)} repeat guests, "',
    'print(f"[VDV] Loaded {len(VDV_GUESTS_RAW)} repeat guests, "\n          f"{len(VDV_FORECAST_DATA[\'history\'])}h/{len(VDV_FORECAST_DATA[\'forecast\'])}f forecast days, "\n          f"{len(VDV_GROUP_PIPELINE)} upcoming groups, ",'
)
print("Global vars + startup: OK")

# ─────────────────────────────────────────────────────────────────────────────
# 3. WIRE DATA INTO build_vdv_dashboard()
# ─────────────────────────────────────────────────────────────────────────────
# Add after mice_est_rev line (already in the function preamble)
src = src.replace(
    "    # Risk score per corporate client (rule-based)",
    """    # ── Forecast + group pipeline ────────────────────────────────────────────
    fc_data     = VDV_FORECAST_DATA
    fc_hist     = fc_data.get('history', [])
    fc_fore     = fc_data.get('forecast', [])
    grp_pipe    = VDV_GROUP_PIPELINE
    # Build JS arrays for the occupancy chart
    fc_labels   = json.dumps([r['date'] for r in fc_hist] + [r['date'] for r in fc_fore])
    fc_occ_hist = json.dumps([r['occ'] for r in fc_hist] + [None] * len(fc_fore))
    fc_occ_fore = json.dumps([None] * len(fc_hist) + [r['occ'] for r in fc_fore])
    fc_rev_hist = json.dumps([r['rev'] for r in fc_hist] + [None] * len(fc_fore))
    fc_rev_fore = json.dumps([None] * len(fc_hist) + [r['rev'] for r in fc_fore])
    fc_arr_hist = json.dumps([r['arrivals'] for r in fc_hist] + [None] * len(fc_fore))
    fc_arr_fore = json.dumps([None] * len(fc_hist) + [r['arrivals'] for r in fc_fore])
    fc_boundary = len(fc_hist)  # index where forecast starts
    # Group pipeline status badges
    STATUS_COLORS = {{'Definite': '#00d165', 'Deposit': '#f59e0b', 'Tentative': '#9ca3af', 'Prospect': '#e5e7eb'}}

    # Risk score per corporate client (rule-based)"""
)
print("Dashboard data wiring: OK")

# ─────────────────────────────────────────────────────────────────────────────
# 4. NEW HTML SECTION — Occupancy & Revenue Trend (inject after today-strip)
# ─────────────────────────────────────────────────────────────────────────────
OCCUPANCY_SECTION = """
<!-- OCCUPANCY & REVENUE TREND ──────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Occupancy & Revenue Trend</span><span class="sh-line"></span><span class="sh-sub">{len(fc_hist)} days history · {len(fc_fore)} days forecast · FOR_042</span></div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:0;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;margin-bottom:32px;">
  <div style="padding:24px;border-right:1px solid #e5e7eb;">
    <div style="font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;margin-bottom:16px;">Occupancy %<span style="margin-left:8px;color:#d1d5db;">——</span> History &nbsp;<span style="color:#93c5fd;">- - -</span> Forecast</div>
    <div style="position:relative;height:160px;"><canvas id="occChart"></canvas></div>
  </div>
  <div style="padding:24px;">
    <div style="font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;margin-bottom:16px;">Room Revenue (EUR)</div>
    <div style="position:relative;height:160px;"><canvas id="revChart"></canvas></div>
  </div>
</div>

<!-- GROUP PIPELINE ──────────────────────────────────────────────────────── -->
{'<div class="sh"><span class="sh-title">Group Pipeline</span><span class="sh-line"></span><span class="sh-sub">Upcoming confirmed & tentative groups · GRP_017</span></div>' + _grp_pipe_html(grp_pipe, STATUS_COLORS) if grp_pipe else ''}

"""

# Helper function for group pipeline HTML — inject into the function body
GROUP_PIPE_HELPER = '''
def _grp_pipe_html(groups, status_colors):
    if not groups: return ''
    rows = ''
    for g in groups:
        sc = status_colors.get(g['status_label'], '#e5e7eb')
        rows += f\'\'\'<tr>
          <td style="padding:10px 14px;font-size:13px;font-weight:500;color:#111827;">{g['name']}</td>
          <td style="padding:10px 14px;font-size:12px;color:#6b7280;">{g['arrival']}</td>
          <td style="padding:10px 14px;font-size:12px;color:#6b7280;">{g['nights']}n</td>
          <td style="padding:10px 14px;font-size:11px;color:#6b7280;">{g['market']}</td>
          <td style="padding:10px 14px;"><span style="background:{sc}22;color:{sc};border:1px solid {sc}66;border-radius:99px;padding:3px 10px;font-size:11px;font-weight:500;">{g['status_label']}</span></td>
        </tr>\'\'\'
    return (
        \'<div style="border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;margin-bottom:32px;">\' +
        \'<table style="width:100%;border-collapse:collapse;">\' +
        \'<thead><tr>\' +
        \'<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Group</th>\' +
        \'<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Arrival</th>\' +
        \'<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Nights</th>\' +
        \'<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Segment</th>\' +
        \'<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Status</th>\' +
        \'</tr></thead><tbody>\' + rows + \'</tbody></table></div>\'
    )

'''

# Inject helper before build_vdv_dashboard
src = src.replace(
    "def build_vdv_dashboard(",
    GROUP_PIPE_HELPER + "def build_vdv_dashboard("
)
print("Group pipeline helper: OK")

# ─────────────────────────────────────────────────────────────────────────────
# 5. INJECT SECTION HTML — after the today-strip section
# ─────────────────────────────────────────────────────────────────────────────
# Find the today strip end and inject after it
insert_after_today = '<!-- REPEAT GUESTS TABLE'
if insert_after_today in src:
    src = src.replace(insert_after_today, OCCUPANCY_SECTION + '<!-- REPEAT GUESTS TABLE')
    print("Section HTML injected: OK")
else:
    print("WARNING: today strip insertion point not found")

# ─────────────────────────────────────────────────────────────────────────────
# 6. INJECT CHART JS — before closing </script> in VdV dashboard
# ─────────────────────────────────────────────────────────────────────────────
CHART_JS = """
// ── Occupancy & Revenue charts ────────────────────────────────────────────
(function() {{
  var labels   = {fc_labels};
  var occHist  = {fc_occ_hist};
  var occFore  = {fc_occ_fore};
  var revHist  = {fc_rev_hist};
  var revFore  = {fc_rev_fore};
  var boundary = {fc_boundary};

  var chartDefaults = {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }}, tooltip: {{ mode: 'index', intersect: false }} }},
    scales: {{
      x: {{ grid: {{ color: '#f3f4f6' }}, ticks: {{ font: {{ size: 9, family: 'JetBrains Mono' }}, color: '#9ca3af',
             callback: function(v, i) {{ return i % 7 === 0 ? labels[i] : ''; }} }} }},
      y: {{ grid: {{ color: '#f3f4f6' }}, ticks: {{ font: {{ size: 9, family: 'JetBrains Mono' }}, color: '#9ca3af' }} }}
    }}
  }};

  // Occupancy chart
  var oc = document.getElementById('occChart');
  if (oc) new Chart(oc, {{ type: 'line', data: {{
    labels: labels,
    datasets: [
      {{ label: 'History', data: occHist, borderColor: '#111827', borderWidth: 1.5,
         pointRadius: 0, tension: 0.3, spanGaps: false }},
      {{ label: 'Forecast', data: occFore, borderColor: '#93c5fd', borderWidth: 1.5,
         borderDash: [4,3], pointRadius: 0, tension: 0.3, spanGaps: false }}
    ]
  }}, options: Object.assign({{}}, chartDefaults, {{
    plugins: Object.assign({{}}, chartDefaults.plugins, {{
      annotation: {{ annotations: {{ line1: {{ type: 'line', x: boundary-0.5,
        borderColor: '#e5e7eb', borderWidth: 1, borderDash: [4,3] }} }} }}
    }}),
    scales: Object.assign({{}}, chartDefaults.scales, {{
      y: Object.assign({{}}, chartDefaults.scales.y, {{ min: 0, max: 100,
         ticks: Object.assign({{}}, chartDefaults.scales.y.ticks, {{ callback: function(v) {{ return v + '%'; }} }}) }})
    }})
  }}) }});

  // Revenue chart
  var rc = document.getElementById('revChart');
  if (rc) new Chart(rc, {{ type: 'bar', data: {{
    labels: labels,
    datasets: [
      {{ label: 'History', data: revHist, backgroundColor: '#f3f4f6', borderColor: '#e5e7eb', borderWidth: 1, borderRadius: 2 }},
      {{ label: 'Forecast', data: revFore, backgroundColor: '#dbeafe', borderColor: '#bfdbfe', borderWidth: 1, borderRadius: 2 }}
    ]
  }}, options: Object.assign({{}}, chartDefaults, {{
    scales: Object.assign({{}}, chartDefaults.scales, {{
      y: Object.assign({{}}, chartDefaults.scales.y, {{
         ticks: Object.assign({{}}, chartDefaults.scales.y.ticks, {{
           callback: function(v) {{ return '€' + (v >= 1000 ? Math.round(v/1000) + 'k' : v); }} }}) }})
    }})
  }}) }});
}})();
"""

# Find the VdV dashboard's DYNAMICS_JS section and inject chart JS before it
vdv_script_end = '// ── DYNAMICS JS ──────────────────────────────────────────────'
if vdv_script_end in src:
    # Find first occurrence (in VdV dashboard)
    idx = src.find(vdv_script_end)
    src = src[:idx] + CHART_JS + '\n' + src[idx:]
    print("Chart JS injected: OK")
else:
    print("WARNING: chart JS injection point not found")

# ─────────────────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────────────────
with open('app.py', 'w', encoding='utf-8') as f:
    f.write(src)
print("\nDone!")
