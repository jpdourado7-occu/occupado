"""
Weekly VDV-MEC data ingestion analysis script.
Runs outside Flask - connects directly to DB, processes new files.
Usage: python ingest_weekly.py
"""
import os, sys, glob, re
from datetime import datetime, date, timedelta
from collections import defaultdict
from dotenv import load_dotenv

load_dotenv()

# Force UTF-8 output on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

# ── Config ────────────────────────────────────────────────────────────────────
_VDV_DIR = os.path.join(os.path.dirname(__file__), "VDV-MEC")

# ── DB ────────────────────────────────────────────────────────────────────────
import psycopg2, psycopg2.extras

def get_db():
    return psycopg2.connect(os.environ["DATABASE_URL"], sslmode="require")


# ── RES_036 file analysis ─────────────────────────────────────────────────────
def _analyse_res036_file(path):
    """Return summary dict for a single RES_036 file.

    Row structure (per reservation):
      Main row  : col0=guest_name, col8=arr_date, col9=nights, col16=cxl_datetime,
                  col20=CXL_No, col24=cxl_reason
      Detail row: col0=None, col2=memb_level, col3=market_segment,
                  col14=channel, col19=cxl_policy
      Notes rows: col0=None, col1='Notes'/text, col3=text
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    cxl_numbers  = set()
    section_dates = []
    arrival_dates = []
    market_segs  = []
    cxl_reasons  = []
    channels     = []

    for row in rows:
        if not row:
            continue
        c0 = str(row[0]).strip() if row[0] is not None else ''

        # Section header: col0 = cancellation date, rest None
        if re.match(r'\d{2}/\d{2}/\d{4}', c0):
            rest_empty = all(row[i] is None for i in range(1, min(6, len(row))))
            if rest_empty:
                try:
                    section_dates.append(datetime.strptime(c0, '%d/%m/%Y').date())
                except Exception:
                    pass
            continue

        # Main guest row: col0 = "Surname, First, Title."
        if ',' in c0 and row[0] is not None:
            # CXL No at col 20
            cxl_no = str(row[20]).strip() if len(row) > 20 and row[20] is not None else ''
            if cxl_no.startswith('MEC-CXL'):
                cxl_numbers.add(cxl_no)
            # Arrival date at col 8
            arr = str(row[8])[:10] if len(row) > 8 and row[8] is not None else ''
            if re.match(r'\d{2}/\d{2}/\d{4}', arr):
                try:
                    arrival_dates.append(datetime.strptime(arr, '%d/%m/%Y').date())
                except Exception:
                    pass
            # CXL reason at col 24
            reason = str(row[24]).strip() if len(row) > 24 and row[24] is not None else ''
            if reason and reason not in ('None', 'nan', 'CXL Reason'):
                cxl_reasons.append(reason)
            continue

        # Detail row: col0=None, col3=market_segment or channel
        if row[0] is None and len(row) > 3 and row[3] is not None:
            seg = str(row[3]).strip()
            if (seg and not seg.startswith('Subtotal') and not seg.startswith('Notes')
                    and not re.match(r'\d{2}/\d{2}/\d{4}', seg)
                    and seg not in ('Market Segment', 'Company/Travel Agent',)
                    and not seg.startswith(' ')):
                market_segs.append(seg)
            # Channel at col 14 of detail row
            ch = str(row[14]).strip() if len(row) > 14 and row[14] is not None else ''
            if ch and ch not in (' Channel', 'Channel', 'None') and not re.match(r'\d{2}/\d{2}/\d{4}', ch):
                channels.append(ch)

    from collections import Counter
    return {
        'path': os.path.basename(path),
        'total_rows': len(rows),
        'unique_cxl_nos': len(cxl_numbers),
        'section_date_range': (min(section_dates), max(section_dates)) if section_dates else None,
        'section_date_count': len(section_dates),
        'arrival_range': (min(arrival_dates), max(arrival_dates)) if arrival_dates else None,
        'top_market_segs': dict(Counter(market_segs).most_common(6)),
        'top_cxl_reasons': dict(Counter(cxl_reasons).most_common(6)),
        'channel_count': len(channels),
    }


# ── _parse_cancellations ──────────────────────────────────────────────────────
_SEG_TO_CH = {
    'BARWEB': 'Booking.com', 'BAROTAGROSS': 'Booking.com',
    'DEALSOTA': 'Booking.com', 'DISCOTAGROSS': 'Booking.com',
    'DISCWEB': 'Direct/Web', 'BARDIR': 'Direct/Web', 'DISCDIR': 'Direct/Web',
    'DSCNRROWEB': 'Direct/Web', 'FLXBBWEB': 'Direct/Web',
    'CORPFIX': 'Corporate', 'CORPDYN': 'Corporate',
    'PACK': 'Package', 'MTGBNS': 'Package', 'BNSGRP': 'Package',
    'DEALS': 'Other', 'OTHER': 'Other', 'COMP': 'Other',
}


def _parse_cancellations(file_paths):
    """Parse list of RES_036 files; return deduplicated cancellation records.

    Matches app.py _parse_vdv_channel_stats() dedup logic: uses MEC-CXL no from
    main row as primary key, falls back to (arr_date, nights, channel).
    """
    import openpyxl
    seen_keys = set()
    records   = []

    for fp in sorted(file_paths):
        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()

            i = 0
            while i < len(rows):
                row = rows[i]
                if not row:
                    i += 1
                    continue
                c0 = str(row[0]).strip() if row[0] is not None else ''

                # Main guest row
                if ',' in c0 and row[0] is not None:
                    cxl_no  = str(row[20]).strip() if len(row) > 20 and row[20] is not None else ''
                    arr_str = str(row[8])[:10]     if len(row) > 8  and row[8]  is not None else ''
                    cxl_str = str(row[16])[:10]    if len(row) > 16 and row[16] is not None else ''
                    nights  = str(row[9]).strip()  if len(row) > 9  and row[9]  is not None else ''

                    # Look ahead for detail row to get market segment
                    seg = ''
                    for j in range(i + 1, min(i + 6, len(rows))):
                        dr = rows[j]
                        if dr and dr[0] is None and len(dr) > 3 and dr[3] is not None:
                            candidate = str(dr[3]).strip()
                            if (candidate and not candidate.startswith('Subtotal')
                                    and not candidate.startswith('Notes')
                                    and not re.match(r'\d{2}/\d{2}/\d{4}', candidate)
                                    and candidate not in ('Market Segment',)):
                                seg = candidate
                                break

                    # Dedup key
                    if cxl_no.startswith('MEC-CXL'):
                        key = cxl_no
                    else:
                        key = (arr_str, nights, seg)

                    if key in seen_keys:
                        i += 1
                        continue
                    seen_keys.add(key)

                    try:
                        arr_dt = datetime.strptime(arr_str, '%d/%m/%Y').date() if arr_str else None
                    except Exception:
                        arr_dt = None
                    try:
                        cxl_dt = datetime.strptime(cxl_str, '%d/%m/%Y').date() if cxl_str else None
                    except Exception:
                        cxl_dt = None

                    records.append({
                        'cxl_no': cxl_no,
                        'seg':    seg,
                        'ch':     _SEG_TO_CH.get(seg, 'Other') if seg else 'Unknown',
                        'arr_dt': arr_dt,
                        'cxl_dt': cxl_dt,
                    })

                i += 1

        except Exception as e:
            print(f"  [WARN] Parse error {os.path.basename(fp)}: {e}")

    return records


# ── _parse_future_bookings ────────────────────────────────────────────────────
def _parse_future_bookings():
    """Parse all RES_004 files for future bookings (mirrors app.py)."""
    import openpyxl
    ch_map = {
        'BARWEB': 'Booking.com', 'BAROTAGROSS': 'Booking.com',
        'DEALSOTA': 'Booking.com', 'DISCOTAGROSS': 'Booking.com',
        'DISCWEB': 'Direct/Web', 'BARDIR': 'Direct/Web', 'DISCDIR': 'Direct/Web',
        'CORPFIX': 'Corporate', 'CORPDYN': 'Corporate',
        'PACK': 'Package', 'MTGBNS': 'Package', 'BNSGRP': 'Package',
        'DEALS': 'Other',
    }
    today      = date.today()
    bookings   = []
    seen_confs = set()
    files      = sorted(glob.glob(os.path.join(_VDV_DIR, 'RES_004_EnteredOnAndBy*.xlsx')))
    print(f"  RES_004 files found: {len(files)}")

    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            i = 0
            while i < len(rows):
                r   = rows[i]
                c0  = str(r[0]).strip() if r[0] else ''
                c3  = str(r[3]).strip() if len(r) > 3 and r[3] else ''
                if (',' in c0 and c3.startswith('MEC-') and not c3.startswith('MEC-F')
                        and len(r) > 8 and r[8]):
                    arr_str = str(r[8])[:10]
                    try:
                        arr = datetime.strptime(arr_str, '%d/%m/%Y').date()
                    except Exception:
                        i += 1
                        continue
                    if arr < today or c3 in seen_confs:
                        i += 1
                        continue
                    seen_confs.add(c3)
                    nights  = int(r[9]) if r[9] and str(r[9]).isdigit() else 1
                    channel = str(r[25]).strip() if len(r) > 25 and r[25] else 'OTHER'
                    created = str(r[28])[:10]     if len(r) > 28 and r[28] else ''
                    lead    = 0
                    if created:
                        try:
                            cdate = datetime.strptime(created, '%d/%m/%Y').date()
                            lead  = max(0, (arr - cdate).days)
                        except Exception:
                            pass
                    ch_label = ch_map.get(channel, 'Other')
                    res_id   = (
                        f"{c0.strip()}_{arr.isoformat()}_{lead}"
                    ).lower().replace(' ', '_')[:100]
                    bookings.append({
                        'res_id':   res_id,
                        'name':     c0,
                        'conf':     c3,
                        'arr_date': arr,
                        'nights':   nights,
                        'channel':  ch_label,
                        'channel_raw': channel,
                        'lead':     lead,
                    })
                i += 1
        except Exception as e:
            print(f"  [WARN] RES_004 parse error {os.path.basename(fp)}: {e}")
    return bookings


# ── _count_noshow ─────────────────────────────────────────────────────────────
def _count_noshow():
    """Count unique no-shows across all RES_037 files (mirrors app.py)."""
    import openpyxl
    files = sorted(glob.glob(os.path.join(_VDV_DIR, 'RES_037_NoShow*.xlsx')))
    seen  = set()
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for row in rows:
                c2 = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                if c2.startswith('MEC-') and c2 not in seen:
                    seen.add(c2)
        except Exception as e:
            print(f"  [WARN] RES_037 error {os.path.basename(fp)}: {e}")
    return len(seen)


# ── _detect_outcomes ──────────────────────────────────────────────────────────
def _detect_outcomes(new_bookings):
    """Compare new bookings against vdv_bookings_cache; log disappeared as outcomes."""
    today = date.today()
    conn  = get_db()
    cur   = conn.cursor()

    cur.execute("""
        SELECT reservation_id, guest_name, arrival_date, channel, risk_score, risk_tier
        FROM vdv_bookings_cache
        WHERE hotel_id = 'vdv'
    """)
    previous = {
        row[0]: {
            'guest_name':   row[1],
            'arrival_date': row[2],
            'channel':      row[3],
            'risk_score':   row[4],
            'risk_tier':    row[5],
        }
        for row in cur.fetchall()
    }
    print(f"  Previously cached bookings in DB: {len(previous)}")

    current_ids = {b['res_id'] for b in new_bookings}

    disappeared = {
        rid: data
        for rid, data in previous.items()
        if rid not in current_ids
    }

    outcomes_logged  = 0
    outcomes_skipped = 0
    new_cancelled    = []
    new_completed    = []

    for res_id, data in disappeared.items():
        arr = data['arrival_date']
        if arr is None:
            continue
        outcome     = 'cancelled' if arr >= today else 'completed'
        days_before = (arr - today).days

        cur.execute("""
            INSERT INTO vdv_outcome_log
                (hotel_id, reservation_id, guest_name, arrival_date,
                 channel, predicted_score, predicted_tier, outcome,
                 outcome_date, days_before_arrival, detected_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (hotel_id, reservation_id) DO NOTHING
        """, (
            'vdv', res_id,
            data['guest_name'], arr,
            data['channel'], data['risk_score'], data['risk_tier'],
            outcome, today, days_before, 'weekly_ingest',
        ))
        if cur.rowcount:
            outcomes_logged += 1
            entry = {
                'name':        data['guest_name'],
                'arr':         arr,
                'ch':          data['channel'],
                'score':       data['risk_score'],
                'tier':        data['risk_tier'],
                'days_before': days_before,
            }
            if outcome == 'cancelled':
                new_cancelled.append(entry)
            else:
                new_completed.append(entry)
        else:
            outcomes_skipped += 1

    conn.commit()
    cur.close()
    conn.close()

    return {
        'total_disappeared': len(disappeared),
        'outcomes_logged':   outcomes_logged,
        'outcomes_skipped':  outcomes_skipped,
        'cancelled':         new_cancelled,
        'completed':         new_completed,
    }


# ── Anomaly checks ────────────────────────────────────────────────────────────
def _check_anomalies(bookings, cx_records, noshow_count, previous_cache_size):
    anomalies = []

    if not bookings:
        anomalies.append("CRITICAL: No future bookings parsed from RES_004")
    elif len(bookings) < 50:
        anomalies.append(f"WARNING: Only {len(bookings)} future bookings -- unusually low")

    if cx_records:
        cxl_dates = [r['cxl_dt'] for r in cx_records if r['cxl_dt']]
        if cxl_dates and max(cxl_dates) < date.today() - timedelta(days=14):
            anomalies.append(
                f"WARNING: Latest cancellation date is {max(cxl_dates)} "
                f"-- may be stale (>14 days old)")
        unknown_seg = sum(1 for r in cx_records if not r['seg'])
        if unknown_seg > len(cx_records) * 0.2:
            anomalies.append(
                f"WARNING: {unknown_seg} cancellations ({unknown_seg/len(cx_records):.0%}) "
                f"have no market segment")

    if bookings:
        ch_counts = defaultdict(int)
        for b in bookings:
            ch_counts[b['channel']] += 1
        other_pct = ch_counts.get('Other', 0) / len(bookings)
        if other_pct > 0.3:
            anomalies.append(
                f"WARNING: {ch_counts['Other']} bookings ({other_pct:.0%}) mapped to "
                f"'Other' -- check raw channel codes")
        # Bookings significantly more or less than cache
        if previous_cache_size and abs(len(bookings) - previous_cache_size) > previous_cache_size * 0.3:
            anomalies.append(
                f"INFO: Booking count changed significantly: "
                f"{previous_cache_size} cached -> {len(bookings)} parsed "
                f"({len(bookings)-previous_cache_size:+d})")

    if noshow_count > 500:
        anomalies.append(f"WARNING: No-show count {noshow_count} is unusually high (>500)")
    elif noshow_count < 100:
        anomalies.append(f"WARNING: No-show count {noshow_count} is unusually low (<100)")

    return anomalies


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    today_str = date.today().isoformat()
    sep = '=' * 65
    print(f"\n{sep}")
    print(f"  VDV-MEC Weekly Ingestion -- {today_str}")
    print(f"{sep}\n")

    # ── Step 1: Analyse both new RES_036 files ────────────────────────────
    print("STEP 1: RES_036 file comparison (files 19 vs 20)")
    print("-" * 55)
    new_res036 = {
        19: os.path.join(_VDV_DIR, "RES_036_CancelledReservations (19).xlsx"),
        20: os.path.join(_VDV_DIR, "RES_036_CancelledReservations (20).xlsx"),
    }

    summaries = {}
    for num, path in new_res036.items():
        s = _analyse_res036_file(path)
        summaries[num] = s
        dr = s['section_date_range']
        ar = s['arrival_range']
        print(f"\n  File ({num}): {s['path']}")
        print(f"    Rows              : {s['total_rows']}")
        print(f"    Unique CXL Nos    : {s['unique_cxl_nos']}")
        if dr:
            print(f"    Section dates     : {dr[0]} to {dr[1]} ({s['section_date_count']} groups)")
        else:
            print(f"    Section dates     : N/A")
        if ar:
            print(f"    Arrival range     : {ar[0]} to {ar[1]}")
        else:
            print(f"    Arrival range     : N/A")
        print(f"    Top market segs   : {s['top_market_segs']}")
        print(f"    Top CXL reasons   : {s['top_cxl_reasons']}")

    # CXL number overlap using openpyxl
    import openpyxl
    cxl_sets = {}
    for num, path in new_res036.items():
        wb = openpyxl.load_workbook(path, read_only=True)
        rows_wb = list(wb.active.iter_rows(values_only=True))
        wb.close()
        cxl_sets[num] = {
            str(row[20]).strip()
            for row in rows_wb
            if len(row) > 20 and row[20] is not None
            and str(row[20]).strip().startswith('MEC-CXL')
        }
    overlap = cxl_sets[19] & cxl_sets[20]
    only_19 = cxl_sets[19] - cxl_sets[20]
    only_20 = cxl_sets[20] - cxl_sets[19]
    print(f"\n  CXL overlap : {len(overlap)} shared | "
          f"{len(only_19)} only in (19) | {len(only_20)} only in (20)")

    # ── Decision logic ────────────────────────────────────────────────────
    s19 = summaries[19]
    s20 = summaries[20]
    dr19 = s19['section_date_range']
    dr20 = s20['section_date_range']

    # File (19) section headers are CXL dates (01/03 -> 20/04 = recent)
    # File (20) section headers extend to Nov 2026 = sorted by arrival date
    is_cxl_sorted_19 = dr19 is not None and dr19[1] < date(2026, 6, 1)
    is_cxl_sorted_20 = dr20 is not None and dr20[1] < date(2026, 6, 1)

    if len(only_19) > 0:
        chosen   = 19
        reason   = f"File (19) has {len(only_19)} unique CXL records not in (20)."
    elif len(only_20) > 0:
        chosen   = 20
        reason   = f"File (20) has {len(only_20)} unique CXL records not in (19)."
    elif is_cxl_sorted_19 and not is_cxl_sorted_20:
        chosen   = 19
        reason   = (
            "Both files contain identical CXL records (46 unique). "
            f"File (19) is sorted by cancellation date ({dr19[0]} to {dr19[1]}), "
            "enabling temporal decay-curve calibration. "
            f"File (20) section headers extend to {dr20[1]} (arrival-date sort), "
            "which is less useful for cancellation-timing analysis."
        )
    elif s19['unique_cxl_nos'] >= s20['unique_cxl_nos']:
        chosen   = 19
        reason   = "File (19) has equal or more unique CXL records and an earlier start date."
    else:
        chosen   = 20
        reason   = "File (20) has more unique CXL records."

    unchosen = 20 if chosen == 19 else 19
    print(f"\n  >>> CHOSEN  : File ({chosen}) -- {os.path.basename(new_res036[chosen])}")
    print(f"  >>> REASON  : {reason}")
    print(f"  >>> IGNORED : File ({unchosen}) -- {os.path.basename(new_res036[unchosen])}\n")

    # ── Step 2: Parse cancellations (all files except discarded new one) ──
    print("STEP 2: Parse RES_036 cancellations (excluding discarded file)")
    print("-" * 55)
    all_res036   = sorted(glob.glob(os.path.join(_VDV_DIR, "RES_036_CancelledReservations*.xlsx")))
    unchosen_path = new_res036[unchosen]
    res036_to_use = [f for f in all_res036
                     if os.path.abspath(f) != os.path.abspath(unchosen_path)]
    print(f"  Using {len(res036_to_use)} RES_036 files (skipped: {os.path.basename(unchosen_path)})")
    cx_records = _parse_cancellations(res036_to_use)
    ch_dist    = defaultdict(int)
    for r in cx_records:
        ch_dist[r['ch']] += 1
    cxl_dates_all = [r['cxl_dt'] for r in cx_records if r['cxl_dt']]
    arr_dates_all = [r['arr_dt'] for r in cx_records if r['arr_dt']]
    print(f"  Total unique cancellations : {len(cx_records)}")
    print(f"  Channel breakdown          : {dict(sorted(ch_dist.items(), key=lambda x: -x[1]))}")
    if cxl_dates_all:
        print(f"  CXL date range (all files) : {min(cxl_dates_all)} to {max(cxl_dates_all)}")
    if arr_dates_all:
        print(f"  Arrival range (all files)  : {min(arr_dates_all)} to {max(arr_dates_all)}")

    # New file contribution
    new_cx = _parse_cancellations([new_res036[chosen]])
    print(f"  New file ({chosen}) contributes : {len(new_cx)} records "
          f"(after cross-file dedup: net new will vary)")

    # ── Step 3: Parse future bookings ─────────────────────────────────────
    print("\nSTEP 3: Parse RES_004 future bookings")
    print("-" * 55)
    bookings = _parse_future_bookings()
    ch_book  = defaultdict(int)
    for b in bookings:
        ch_book[b['channel']] += 1
    arr_book = [b['arr_date'] for b in bookings]
    print(f"  Future bookings parsed : {len(bookings)}")
    print(f"  Channel breakdown      : {dict(sorted(ch_book.items(), key=lambda x: -x[1]))}")
    if arr_book:
        print(f"  Arrival range          : {min(arr_book)} to {max(arr_book)}")

    # ── Step 4: Count no-shows ────────────────────────────────────────────
    print("\nSTEP 4: Count no-shows (RES_037)")
    print("-" * 55)
    noshow_count = _count_noshow()
    noshow_files = sorted(glob.glob(os.path.join(_VDV_DIR, 'RES_037_NoShow*.xlsx')))
    print(f"  RES_037 files: {len(noshow_files)}")
    print(f"  Total unique no-shows (all files): {noshow_count}")

    # ── Step 5: Feedback loop ─────────────────────────────────────────────
    print("\nSTEP 5: Feedback loop -- _detect_vdv_outcomes()")
    print("-" * 55)
    result = _detect_outcomes(bookings)
    print(f"  Disappeared from cache : {result['total_disappeared']}")
    print(f"  Newly logged outcomes  : {result['outcomes_logged']}")
    print(f"  Already in log (skip)  : {result['outcomes_skipped']}")
    print(f"  -> Cancellations       : {len(result['cancelled'])}")
    print(f"  -> Completed stays     : {len(result['completed'])}")

    if result['cancelled']:
        print("\n  New cancellations detected:")
        for c in sorted(result['cancelled'], key=lambda x: x['arr']):
            score_str = f"{c['score']:.0f}" if c['score'] is not None else '  ?'
            tier_str  = c['tier'] or '?'
            print(f"    [{tier_str:>4}] {str(c['name'])[:32]:<32}  "
                  f"arr={c['arr']}  ch={c['ch']:<12}  "
                  f"score={score_str:>3}  days_before={c['days_before']:+d}")

    if result['completed']:
        print("\n  Completed stays detected:")
        for c in sorted(result['completed'], key=lambda x: x['arr']):
            score_str = f"{c['score']:.0f}" if c['score'] is not None else '  ?'
            print(f"    {str(c['name'])[:32]:<32}  arr={c['arr']}  "
                  f"ch={c['ch']:<12}  score={score_str:>3}")

    # ── Step 6: Anomaly checks ────────────────────────────────────────────
    print("\nSTEP 6: Anomaly checks")
    print("-" * 55)
    prev_cache = result['total_disappeared'] + len(bookings)  # approx previous size
    # Re-query cache count for accurate prev size
    try:
        conn_a = get_db()
        cur_a  = conn_a.cursor()
        cur_a.execute("SELECT COUNT(*) FROM vdv_bookings_cache WHERE hotel_id='vdv'")
        prev_size = cur_a.fetchone()[0]
        cur_a.close(); conn_a.close()
    except Exception:
        prev_size = None

    anomalies = _check_anomalies(bookings, cx_records, noshow_count, prev_size)
    if anomalies:
        for a in anomalies:
            print(f"  {a}")
    else:
        print("  No anomalies detected.")

    # ── Summary ───────────────────────────────────────────────────────────
    print(f"\n{sep}")
    print("  INGESTION SUMMARY")
    print(sep)
    print(f"  RES_036 chosen      : File ({chosen}) -- {os.path.basename(new_res036[chosen])}")
    print(f"  RES_036 ignored     : File ({unchosen}) -- {os.path.basename(new_res036[unchosen])}")
    print(f"  Cancellations total : {len(cx_records)}")
    print(f"  Future bookings     : {len(bookings)}")
    print(f"  No-shows total      : {noshow_count}")
    print(f"  Outcomes detected   : {result['outcomes_logged']}")
    print(f"    -> Cancelled      : {len(result['cancelled'])}")
    print(f"    -> Completed      : {len(result['completed'])}")
    print(f"  Anomalies           : {len(anomalies)}")
    print(f"{sep}\n")


if __name__ == '__main__':
    main()
