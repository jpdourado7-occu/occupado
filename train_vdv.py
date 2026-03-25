# train_vdv.py — Retrain Occupado model combining Kaggle + Van der Valk Mechelen data
# All 12 original features preserved. Saves as occupado_model.pkl (replaces current model).

import re
import pickle
import numpy as np
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from xgboost import XGBClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, classification_report

BASE_DIR = r'C:\Users\jpdou\Desktop\Occupado'
ANON_DIR = rf'{BASE_DIR}\VDV-Data\anonymized'

KAGGLE_CSV     = rf'{BASE_DIR}\hotel_bookings.csv'
ARRIVAL_FILE   = rf'{ANON_DIR}\RES_001_ArrivalDetailed.xlsx'
CANCELLED_FILE = rf'{ANON_DIR}\RES_036_CancelledReservations (1).xlsx'
NOSHOW_FILE    = rf'{ANON_DIR}\RES_037_NoShow (1).xlsx'

MODEL_OUT = rf'{BASE_DIR}\occupado_model.pkl'

FEATURES = [
    'lead_time',
    'arrival_date_week_number',
    'stays_in_weekend_nights',
    'stays_in_week_nights',
    'adults',
    'is_repeated_guest',
    'previous_cancellations',
    'previous_bookings_not_canceled',
    'booking_changes',
    'days_in_waiting_list',
    'adr',
    'total_of_special_requests',
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_date(val):
    if val is None:
        return None
    s = str(val).strip()[:16]
    for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    return None

def parse_rate(val):
    if val is None:
        return None
    s = re.sub(r'[^\d,\.]', '', str(val))
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.')
    try:
        return float(s)
    except:
        return None

def parse_adults(val):
    if val is None:
        return None
    try:
        return int(str(val).split('/')[0])
    except:
        return None

def is_hash(val):
    return bool(val and re.match(r'^[0-9a-f]{60,}', str(val)))

def count_nights(arr_date, n):
    weekend = weekday = 0
    for i in range(max(0, int(n))):
        d = arr_date + timedelta(days=i)
        if d.weekday() >= 5:
            weekend += 1
        else:
            weekday += 1
    return weekend, weekday

# ---------------------------------------------------------------------------
# VDV Parsers
# ---------------------------------------------------------------------------

def parse_arrivals_co(path):
    """RES_001 — completed stays (status CO). Returns flat DataFrame."""
    print('  Parsing Arrival Detailed (CO)...')
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    STATUSES = {'CO', 'NS', 'CX', 'IH', 'DI', 'DO'}
    records = []
    pending1 = pending3 = None

    for row in rows:
        col1 = row[1] if len(row) > 1 else None
        col7 = row[7] if len(row) > 7 else None

        if is_hash(col1) and row[4] is not None:
            pending1 = row
            pending3 = None
        elif row[0] == 'DNM' and pending1 is not None:
            pending3 = row
        elif (col1 and str(col1).startswith('MEC-') and
              col7 and str(col7).strip() in STATUSES and pending1 is not None):

            if str(col7).strip() == 'CO':
                arr  = parse_date(pending1[4])
                dep  = parse_date(row[4])
                n    = pending1[7]
                adlt = parse_adults(pending1[8])
                rate = parse_rate(row[12])

                if arr:
                    try:    nights = int(n)
                    except: nights = 0
                    if nights == 0 and dep:
                        nights = max(0, (dep - arr).days)
                    wkend, wkday = count_nights(arr, nights)
                    records.append({
                        'guest_hash':                    pending1[1],
                        'arr_date':                      arr,
                        'lead_time':                     np.nan,   # imputed later
                        'arrival_date_week_number':      int(arr.isocalendar()[1]),
                        'stays_in_weekend_nights':       wkend,
                        'stays_in_week_nights':          wkday,
                        'adults':                        adlt if adlt is not None else np.nan,
                        'is_repeated_guest':             0,
                        'previous_cancellations':        0,
                        'previous_bookings_not_canceled':0,
                        'booking_changes':               0,
                        'days_in_waiting_list':          0,
                        'adr':                           rate if rate is not None else np.nan,
                        'total_of_special_requests':     0,
                        'is_canceled':                   0,
                    })
            pending1 = pending3 = None

    print(f'    -> {len(records)} completed stays')
    return pd.DataFrame(records)


def parse_cancelled(path):
    """RES_036 — cancelled bookings."""
    print('  Parsing Cancelled Reservations...')
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    records = []
    i = 0
    while i < len(rows):
        row = rows[i]
        if is_hash(row[0]):
            row_b = None
            for j in range(i + 1, min(i + 5, len(rows))):
                r = rows[j]
                if r[0] is None and r[9] is not None and '/' in str(r[9]):
                    row_b = r; break
            if row_b is not None:
                arr        = parse_date(row[8])
                created_on = parse_date(row[14])
                dep        = parse_date(row_b[8])
                adlt       = parse_adults(row_b[9])
                rate       = parse_rate(row_b[11])
                n          = row[9]
                if arr:
                    try:    nights = int(n)
                    except: nights = 0
                    if nights == 0 and dep:
                        nights = max(0, (dep - arr).days)
                    wkend, wkday = count_nights(arr, nights)
                    lead = (arr - created_on).days if created_on else np.nan
                    records.append({
                        'guest_hash':                    row[0],
                        'arr_date':                      arr,
                        'lead_time':                     lead,
                        'arrival_date_week_number':      int(arr.isocalendar()[1]),
                        'stays_in_weekend_nights':       wkend,
                        'stays_in_week_nights':          wkday,
                        'adults':                        adlt if adlt is not None else np.nan,
                        'is_repeated_guest':             0,
                        'previous_cancellations':        0,
                        'previous_bookings_not_canceled':0,
                        'booking_changes':               0,
                        'days_in_waiting_list':          0,
                        'adr':                           rate if rate is not None else np.nan,
                        'total_of_special_requests':     0,
                        'is_canceled':                   1,
                    })
        i += 1
    print(f'    -> {len(records)} cancelled bookings')
    return pd.DataFrame(records)


def parse_noshow(path):
    """RES_037 — no-show bookings."""
    print('  Parsing No-Show Reservations...')
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    records = []
    i = 0
    while i < len(rows):
        row = rows[i]
        if is_hash(row[1] if len(row) > 1 else None):
            row_b = None
            for j in range(i + 1, min(i + 4, len(rows))):
                r = rows[j]
                if r[2] and str(r[2]).startswith('MEC-F') and r[16] is not None:
                    row_b = r; break
            if row_b is not None:
                arr  = parse_date(row[11])
                dep  = parse_date(row[13])
                n    = row[18]
                adlt = parse_adults(row[20])
                rate = parse_rate(row_b[24])
                created_on = parse_date(row_b[16])
                if arr:
                    try:    nights = int(n)
                    except: nights = 0
                    if nights == 0 and dep:
                        nights = max(0, (dep - arr).days)
                    wkend, wkday = count_nights(arr, nights)
                    lead = (arr - created_on).days if created_on else np.nan
                    records.append({
                        'guest_hash':                    row[1],
                        'arr_date':                      arr,
                        'lead_time':                     lead,
                        'arrival_date_week_number':      int(arr.isocalendar()[1]),
                        'stays_in_weekend_nights':       wkend,
                        'stays_in_week_nights':          wkday,
                        'adults':                        adlt if adlt is not None else np.nan,
                        'is_repeated_guest':             0,
                        'previous_cancellations':        0,
                        'previous_bookings_not_canceled':0,
                        'booking_changes':               0,
                        'days_in_waiting_list':          0,
                        'adr':                           rate if rate is not None else np.nan,
                        'total_of_special_requests':     0,
                        'is_canceled':                   1,
                    })
        i += 1
    print(f'    -> {len(records)} no-show bookings')
    return pd.DataFrame(records)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

print('=' * 55)
print('  Occupado - Model Retraining (Kaggle + VDV)')
print('=' * 55)

print('\n[1] Loading Kaggle base dataset...')
df_kaggle = pd.read_csv(KAGGLE_CSV)
df_kaggle = df_kaggle[FEATURES + ['is_canceled']].dropna()
print(f'    -> {len(df_kaggle):,} Kaggle bookings')

print('\n[2] Parsing VDV files...')
df_co  = parse_arrivals_co(ARRIVAL_FILE)
df_cxl = parse_cancelled(CANCELLED_FILE)
df_ns  = parse_noshow(NOSHOW_FILE)

df_vdv = pd.concat([df_co, df_cxl, df_ns], ignore_index=True)
print(f'    -> {len(df_vdv):,} VDV bookings total')

print('\n[3] Imputing lead_time for VDV completed stays...')
# lead_time not available in historical Shiji arrival exports.
# Impute with median of known lead_times (from Kaggle non-cancelled + VDV cancelled).
known_lead = pd.concat([
    df_kaggle['lead_time'],
    df_cxl['lead_time'].dropna(),
    df_ns['lead_time'].dropna(),
])
median_lead = float(known_lead.median())
print(f'    Median lead_time from known data: {median_lead:.0f} days')
df_vdv['lead_time'] = df_vdv['lead_time'].fillna(median_lead)

print('\n[4] Imputing remaining NaNs in VDV (adults, adr)...')
for col in ['adults', 'adr']:
    if df_vdv[col].isna().any():
        med = float(pd.concat([df_kaggle[col], df_vdv[col].dropna()]).median())
        df_vdv[col] = df_vdv[col].fillna(med)
        print(f'    {col}: imputed with median {med:.2f}')

print('\n[5] Combining datasets...')
df_vdv_train = df_vdv[FEATURES + ['is_canceled']]
df_combined  = pd.concat([df_kaggle, df_vdv_train], ignore_index=True)
print(f'    Kaggle:  {len(df_kaggle):,} bookings')
print(f'    VDV:     {len(df_vdv_train):,} bookings')
print(f'    Total:   {len(df_combined):,} bookings')
print(f'    Cancelled/No-show: {df_combined["is_canceled"].sum():,}')
print(f'    Completed stays:   {(df_combined["is_canceled"]==0).sum():,}')

print('\n[6] Training XGBoost on combined dataset...')
X = df_combined[FEATURES]
y = df_combined['is_canceled']

X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=42, stratify=y
)
print(f'    Train: {len(X_train):,} | Test: {len(X_test):,}')

model = XGBClassifier(
    n_estimators=200,
    max_depth=6,
    learning_rate=0.05,
    random_state=42,
    eval_metric='logloss',
)
model.fit(X_train, y_train)

print('\n[7] Evaluating...')
y_pred   = model.predict(X_test)
accuracy = accuracy_score(y_test, y_pred) * 100
print(f'\n    Accuracy: {accuracy:.1f}%\n')
print(classification_report(y_test, y_pred, target_names=['Completed', 'Cancelled']))

print('[8] Feature importances:')
fi = sorted(zip(FEATURES, model.feature_importances_), key=lambda x: -x[1])
for feat, imp in fi:
    bar = '#' * int(imp * 40)
    print(f'    {feat:<38} {bar} {imp:.3f}')

print(f'\n[9] Saving model -> {MODEL_OUT}')
with open(MODEL_OUT, 'wb') as f:
    pickle.dump(model, f)
print('    Done! occupado_model.pkl updated.')
print('    All 12 features preserved. Deploy as normal.')
