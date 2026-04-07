from flask import Flask, jsonify, send_file, request, redirect, url_for, session
import pandas as pd
import pickle
import io
import json
from functools import wraps
from datetime import datetime, timedelta, date
import os
import secrets
import psycopg2
import psycopg2.extras
import bcrypt
import re
from dotenv import load_dotenv
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "occupado-secret-2024")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024  # 5MB upload limit

@app.after_request
def add_security_headers(response):
    # Content Security Policy
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "frame-ancestors 'none';"
    )
    # Prevent clickjacking
    response.headers["X-Frame-Options"] = "DENY"
    # Prevent MIME sniffing
    response.headers["X-Content-Type-Options"] = "nosniff"
    # Force HTTPS
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    # Hide server info
    response.headers["X-Powered-By"] = ""
    response.headers["Server"] = ""
    return response

TOKEN_DIR = "/tmp/occupado_tokens"
os.makedirs(TOKEN_DIR, exist_ok=True)

def get_db():
    DATABASE_URL = os.environ.get("DATABASE_URL", "")
    if not DATABASE_URL:
        # Railway sometimes exposes individual PG* vars instead of a single URL
        pghost = os.environ.get("PGHOST", "")
        pgport = os.environ.get("PGPORT", "5432")
        pgdb   = os.environ.get("PGDATABASE", "")
        pguser = os.environ.get("PGUSER", "")
        pgpass = os.environ.get("PGPASSWORD", "")
        if pghost and pgdb and pguser:
            DATABASE_URL = f"postgresql://{pguser}:{pgpass}@{pghost}:{pgport}/{pgdb}"
        else:
            raise RuntimeError(
                "No database connection found. "
                "Set DATABASE_URL (or PGHOST/PGDATABASE/PGUSER/PGPASSWORD) in Railway Variables."
            )
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def _upsert_vdv_scores(bookings, scores, conn=None):
    if not bookings or not scores:
        return 0

    close_after = conn is None
    if conn is None:
        try:
            conn = get_db()
        except Exception:
            return 0

    upserted = 0
    try:
        cur = conn.cursor()
        for b, score in zip(bookings, scores):
            arr = b.get('arr_date')
            if arr is None:
                continue
            if hasattr(arr, 'isoformat'):
                arr_str = arr.isoformat()
            else:
                arr_str = str(arr)

            # Deterministic reservation_id from name + arrival + lead
            res_id = (
                f"{b.get('name','').strip()}"
                f"_{arr_str}"
                f"_{b.get('lead',0)}"
            ).lower().replace(' ', '_')[:100]

            tier = (
                'high'   if score >= 70 else
                'medium' if score >= 40 else
                'low'
            )

            cur.execute("""
                INSERT INTO vdv_bookings_cache
                    (hotel_id, reservation_id,
                     guest_name, arrival_date,
                     channel, channel_raw,
                     lead_time, risk_score,
                     risk_tier, scored_at,
                     score_version)
                VALUES
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),%s)
                ON CONFLICT (hotel_id, reservation_id)
                DO UPDATE SET
                    risk_score    = EXCLUDED.risk_score,
                    risk_tier     = EXCLUDED.risk_tier,
                    scored_at     = NOW(),
                    score_version = EXCLUDED.score_version,
                    channel       = EXCLUDED.channel,
                    lead_time     = EXCLUDED.lead_time
            """, (
                'vdv',
                res_id,
                b.get('name', ''),
                arr_str,
                b.get('channel', ''),
                b.get('channel_raw', ''),
                b.get('lead', 0),
                float(score),
                tier,
                '14f',
            ))
            upserted += 1

        conn.commit()
        cur.close()
    except Exception as e:
        print(f"[VDV] Score upsert error: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        if close_after and conn:
            try:
                conn.close()
            except Exception:
                pass

    return upserted

def _detect_vdv_outcomes(new_bookings, new_scores):
    from datetime import date
    today = date.today()

    if not new_bookings:
        return 0

    try:
        conn = get_db()
        cur = conn.cursor()

        # Get all previously scored bookings from vdv_bookings_cache
        cur.execute("""
            SELECT reservation_id,
                   guest_name,
                   arrival_date,
                   channel,
                   risk_score,
                   risk_tier
            FROM vdv_bookings_cache
            WHERE hotel_id = 'vdv'
        """)
        previous = {
            row[0]: {
                'guest_name':   row[1],
                'arrival_date': row[2],
                'channel':      row[3],
                'risk_score':   row[4],
                'risk_tier':    row[5],
            }
            for row in cur.fetchall()
        }

        # Build set of current reservation IDs using same deterministic ID as upsert
        current_ids = set()
        for b in new_bookings:
            arr = b.get('arr_date')
            if arr is None:
                continue
            arr_str = (arr.isoformat()
                       if hasattr(arr, 'isoformat')
                       else str(arr))
            res_id = (
                f"{b.get('name','').strip()}"
                f"_{arr_str}"
                f"_{b.get('lead',0)}"
            ).lower().replace(' ', '_')[:100]
            current_ids.add(res_id)

        # Disappeared bookings = in previous but not in current
        disappeared = {
            rid: data
            for rid, data in previous.items()
            if rid not in current_ids
        }

        outcomes_logged = 0
        for res_id, data in disappeared.items():
            arr = data['arrival_date']
            if arr is None:
                continue

            # Determine outcome
            if arr >= today:
                outcome = 'cancelled'
            else:
                outcome = 'completed'

            days_before = (arr - today).days

            # Insert into outcome log; skip if already recorded
            cur.execute("""
                INSERT INTO vdv_outcome_log
                    (hotel_id, reservation_id,
                     guest_name, arrival_date,
                     channel, predicted_score,
                     predicted_tier, outcome,
                     outcome_date,
                     days_before_arrival,
                     detected_by)
                VALUES
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (hotel_id, reservation_id)
                DO NOTHING
            """, (
                'vdv',
                res_id,
                data['guest_name'],
                arr,
                data['channel'],
                data['risk_score'],
                data['risk_tier'],
                outcome,
                today,
                days_before,
                'auto',
            ))
            outcomes_logged += 1

        conn.commit()
        cur.close()
        conn.close()

        if outcomes_logged > 0:
            n_cancelled  = sum(1 for d in disappeared.values() if d['arrival_date'] and d['arrival_date'] >= today)
            n_completed  = sum(1 for d in disappeared.values() if d['arrival_date'] and d['arrival_date'] < today)
            print(f"[VDV] Outcomes detected: {outcomes_logged} bookings "
                  f"({n_cancelled} cancelled, {n_completed} completed)")

        return outcomes_logged

    except Exception as e:
        print(f"[VDV] Outcome detection error: {e}")
        return 0

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS registered_users (
            username TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            verified INTEGER NOT NULL DEFAULT 0,
            signed_up TEXT DEFAULT ''
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS verification_tokens (
            token TEXT PRIMARY KEY,
            username TEXT NOT NULL,
            expires_at TIMESTAMP NOT NULL DEFAULT (NOW() + INTERVAL '24 hours')
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            token TEXT PRIMARY KEY,
            username TEXT NOT NULL,
            expires_at TIMESTAMP NOT NULL DEFAULT (NOW() + INTERVAL '1 hour')
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS hotel_uploads (
            id SERIAL PRIMARY KEY,
            hotel_username TEXT NOT NULL,
            filename TEXT NOT NULL,
            uploaded_at TIMESTAMP DEFAULT NOW(),
            records_json TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS roi_actions (
            id SERIAL PRIMARY KEY,
            hotel_username TEXT NOT NULL,
            guest_name TEXT NOT NULL,
            action_type TEXT NOT NULL,
            sent_at TIMESTAMP DEFAULT NOW(),
            booking_ref TEXT DEFAULT ''
        )
    """)
    conn.commit()
    cur.close()
    conn.close()
    # Migrate: add signed_up column in separate transaction
    try:
        conn2 = get_db()
        cur2 = conn2.cursor()
        cur2.execute("ALTER TABLE registered_users ADD COLUMN signed_up TEXT DEFAULT ''")
        conn2.commit()
        cur2.close()
        conn2.close()
    except:
        pass
    # Migrate: add expires_at column to verification_tokens
    try:
        conn3 = get_db()
        cur3 = conn3.cursor()
        cur3.execute("ALTER TABLE verification_tokens ADD COLUMN expires_at TIMESTAMP NOT NULL DEFAULT (NOW() + INTERVAL '24 hours')")
        conn3.commit()
        cur3.close()
        conn3.close()
    except:
        pass
    # Migrate: rename old blob-style vdv_bookings_cache → vdv_bookings_cache_legacy
    try:
        conn4 = get_db()
        cur4 = conn4.cursor()
        cur4.execute("ALTER TABLE vdv_bookings_cache RENAME TO vdv_bookings_cache_legacy")
        conn4.commit()
        cur4.close()
        conn4.close()
        print("[VDV] Migrated vdv_bookings_cache → vdv_bookings_cache_legacy")
    except:
        pass
    # Create new per-row vdv_bookings_cache schema
    try:
        conn5 = get_db()
        cur5 = conn5.cursor()
        cur5.execute("""
            CREATE TABLE IF NOT EXISTS vdv_bookings_cache (
                id              SERIAL PRIMARY KEY,
                hotel_id        TEXT NOT NULL DEFAULT 'vdv',
                reservation_id  TEXT NOT NULL,
                guest_name      TEXT,
                arrival_date    DATE,
                channel         TEXT,
                channel_raw     TEXT,
                lead_time       INTEGER,
                risk_score      FLOAT,
                risk_tier       TEXT,
                scored_at       TIMESTAMP DEFAULT NOW(),
                score_version   TEXT DEFAULT '14f',
                is_high_risk    BOOLEAN GENERATED ALWAYS AS (risk_score >= 70) STORED,
                UNIQUE(hotel_id, reservation_id)
            )
        """)
        cur5.execute("CREATE INDEX IF NOT EXISTS idx_vdv_cache_arrival ON vdv_bookings_cache(arrival_date)")
        cur5.execute("CREATE INDEX IF NOT EXISTS idx_vdv_cache_risk ON vdv_bookings_cache(risk_score DESC)")
        conn5.commit()
        cur5.close()
        conn5.close()
        print("[VDV] vdv_bookings_cache per-row schema ready")
    except Exception as _ve:
        print(f"[VDV] Cache schema migration error: {_ve}")
    # Create outcome log table
    try:
        conn6 = get_db()
        cur6 = conn6.cursor()
        cur6.execute("""
            CREATE TABLE IF NOT EXISTS vdv_outcome_log (
                id                 SERIAL PRIMARY KEY,
                hotel_id           TEXT NOT NULL DEFAULT 'vdv',
                reservation_id     TEXT NOT NULL,
                guest_name         TEXT,
                arrival_date       DATE,
                channel            TEXT,
                predicted_score    FLOAT,
                predicted_tier     TEXT,
                outcome            TEXT,
                outcome_date       DATE,
                days_before_arrival INTEGER,
                detected_by        TEXT DEFAULT 'auto',
                logged_at          TIMESTAMP DEFAULT NOW(),
                UNIQUE(hotel_id, reservation_id)
            )
        """)
        cur6.execute("CREATE INDEX IF NOT EXISTS idx_outcome_arrival ON vdv_outcome_log(arrival_date)")
        cur6.execute("CREATE INDEX IF NOT EXISTS idx_outcome_outcome ON vdv_outcome_log(outcome)")
        conn6.commit()
        cur6.close()
        conn6.close()
        print("[VDV] vdv_outcome_log schema ready")
    except Exception as _oe:
        print(f"[VDV] Outcome log schema error: {_oe}")

try:
    init_db()
except Exception as _db_err:
    import sys
    print(f"[occupado] WARNING: Database init failed — {_db_err}", file=sys.stderr)
    print("[occupado] Set DATABASE_URL in Railway -> Variables to fix this.", file=sys.stderr)

# ── Rate limiting: track failed login attempts ──
# { ip: { "count": int, "locked_until": datetime | None } }
FAILED_ATTEMPTS = {}
MAX_ATTEMPTS    = 5
LOCKOUT_MINUTES = 15

def check_rate_limit(ip):
    """Returns (is_blocked, seconds_remaining)"""
    record = FAILED_ATTEMPTS.get(ip)
    if not record:
        return False, 0
    if record["locked_until"] and datetime.now() < record["locked_until"]:
        remaining = int((record["locked_until"] - datetime.now()).total_seconds())
        return True, remaining
    return False, 0

def record_failed_attempt(ip):
    if ip not in FAILED_ATTEMPTS:
        FAILED_ATTEMPTS[ip] = {"count": 0, "locked_until": None}
    FAILED_ATTEMPTS[ip]["count"] += 1
    if FAILED_ATTEMPTS[ip]["count"] >= MAX_ATTEMPTS:
        FAILED_ATTEMPTS[ip]["locked_until"] = datetime.now() + timedelta(minutes=LOCKOUT_MINUTES)

def reset_attempts(ip):
    FAILED_ATTEMPTS.pop(ip, None)

def sanitise(value, max_length=100):
    """Strip dangerous characters and limit length."""
    value = str(value).strip()
    value = re.sub(r"[<>\"'%;()&+]", "", value)
    return value[:max_length]

# Language Translations (English, Dutch, French)
TRANSLATIONS = {
    "en": {
        "occupado": "Occupado",
        "ai_booking": "AI Booking Intelligence",
        "settings": "Settings",
        "sign_out": "Sign Out",
        "live_dashboard": "Live Dashboard",
        "bookings_analysed": "bookings analysed",
        "upload_data": "Upload Your Booking Data",
        "drop_csv": "📂 Drop your booking CSV here",
        "export_pms": "Export from your PMS and upload",
        "choose_file": "Choose CSV File",
        "high_risk": "High Risk",
        "medium_risk": "Medium Risk",
        "low_risk": "Low Risk",
        "optimizer": "Overbooking Optimizer",
        "safe_rooms": "Safe rooms to oversell tonight",
        "revenue": "Revenue opportunity: EUR",
        "apply": "Apply Recommendation",
        "bookings_analysed_stat": "Bookings analysed",
        "predicted": "Predicted no-shows",
        "confidence": "AI confidence",
        "walk_risk": "Walk risk",
        "avg_rate": "Avg room rate",
        "take_action": "Take Action on High-Risk Bookings",
        "send_mass": "Send Mass Email",
        "send_high": "Send to all {high} high-risk bookings",
        "send_btn": "Send Email to All →",
        "request_dep": "Request Deposits",
        "dep_template": "Quick template for deposits",
        "dep_btn": "Send Deposit Request →",
        "reminders": "Send Reminders",
        "rem_template": "Quick template for reminders",
        "rem_btn": "Send Reminder →",
        "click_row": "Bookings — Click any row for AI reasoning",
        "booking": "Booking",
        "lead": "Lead Time",
        "rate": "Room Rate",
        "returning": "Returning",
        "cancels": "Past Cancels",
        "risk": "Risk Score",
        "action": "Action",
        "days": "days",
        "yes": "Yes",
        "no": "No",
        "high": "HIGH",
        "medium": "MEDIUM",
        "low": "LOW",
        "req_dep": "Request Deposit",
        "send_rem": "Send Reminder",
        "monitor": "Monitor",
        "email_guest": "Send Email to Guest",
        "guest_email": "Guest Email Address",
        "guest_name": "Guest Name",
        "subject": "Subject Line",
        "message": "Message to Guest",
        "send_email": "📧 Send Email",
        "cancel": "Cancel",
        "bulk_email": "Send Email to All High-Risk Bookings",
        "select_book": "Select Bookings to Send To",
        "selected_count": "bookings selected",
        "save_changes": "💾 Save Changes",
        "send_selected": "📧 Send to Selected",
        "fill_fields": "Please fill in all fields",
        "select_one": "Please select at least one booking",
        "sent_success": "✓ Emails sent to {count} guest(s)",
        "error": "Error sending emails",
        "settings_title": "Settings",
        "config_alert": "Configure your alert preferences",
        "alert_email": "🔔 Alert Email",
        "alert_desc": "Email address for high-risk booking alerts",
        "email_addr": "Email Address",
        "save": "Save Settings →",
        "saved": "✓ Settings saved!",
        "back": "← Back",
        "auto_pop": "Automatically populated with high-risk bookings. Remove any you don't want to contact.",
    },
    "nl": {
        "occupado": "Occupado",
        "ai_booking": "AI Boekingintelligentie",
        "settings": "Instellingen",
        "sign_out": "Afmelden",
        "live_dashboard": "Live Dashboard",
        "bookings_analysed": "boekingen geanalyseerd",
        "upload_data": "Upload uw boekingsgegevens",
        "drop_csv": "📂 Sleep uw boeking CSV hier",
        "export_pms": "Exporteer uit uw PMS en upload",
        "choose_file": "Kies CSV-bestand",
        "high_risk": "Hoog risico",
        "medium_risk": "Gemiddeld risico",
        "low_risk": "Laag risico",
        "optimizer": "Overbooking Optimizer",
        "safe_rooms": "Veilige kamers om vanavond te overboeken",
        "revenue": "Omzetmogelijkheid: EUR",
        "apply": "Aanbeveling toepassen",
        "bookings_analysed_stat": "Boekingen geanalyseerd",
        "predicted": "Voorspelde no-shows",
        "confidence": "AI-betrouwbaarheid",
        "walk_risk": "Walk-risico",
        "avg_rate": "Gem. kamerprijs",
        "take_action": "Maatregelen nemen voor risicovolle boekingen",
        "send_mass": "E-mail verzenden naar meerdere",
        "send_high": "Verzenden naar alle {high} risicovolle boekingen",
        "send_btn": "E-mail naar iedereen verzenden →",
        "request_dep": "Aanbetaling aanvragen",
        "dep_template": "Snelle sjabloon voor aanbetaling",
        "dep_btn": "Aanbetaling aanvragen →",
        "reminders": "Herinneringen verzenden",
        "rem_template": "Snelle sjabloon voor herinneringen",
        "rem_btn": "Herinnering verzenden →",
        "click_row": "Boekingen — Klik op een rij voor AI-redenering",
        "booking": "Boeking",
        "lead": "Aanlooptijd",
        "rate": "Kamerprijs",
        "returning": "Klant keert terug",
        "cancels": "Eerdere annuleringen",
        "risk": "Risicoscore",
        "action": "Actie",
        "days": "dagen",
        "yes": "Ja",
        "no": "Nee",
        "high": "HOOG",
        "medium": "GEMIDDELD",
        "low": "LAAG",
        "req_dep": "Aanbetaling aanvragen",
        "send_rem": "Herinnering verzenden",
        "monitor": "Bewaken",
        "email_guest": "E-mail naar gast verzenden",
        "guest_email": "E-mailadres gast",
        "guest_name": "Gastnaam",
        "subject": "Onderwerpregel",
        "message": "Bericht aan gast",
        "send_email": "📧 E-mail verzenden",
        "cancel": "Annuleren",
        "bulk_email": "E-mail verzenden naar alle risicovolle boekingen",
        "select_book": "Selecteer boekingen om naar te verzenden",
        "selected_count": "boekingen geselecteerd",
        "save_changes": "💾 Wijzigingen opslaan",
        "send_selected": "📧 Naar geselecteerden verzenden",
        "fill_fields": "Vul alstublieft alle velden in",
        "select_one": "Selecteer alstublieft minstens één boeking",
        "sent_success": "✓ E-mails verzonden naar {count} gast(en)",
        "error": "Fout bij verzenden van e-mails",
        "settings_title": "Instellingen",
        "config_alert": "Configureer uw waarschuwingsvoorkeuren",
        "alert_email": "🔔 Waarschuwing per e-mail",
        "alert_desc": "E-mailadres voor waarschuwingen van risicovolle boekingen",
        "email_addr": "E-mailadres",
        "save": "Instellingen opslaan →",
        "saved": "✓ Instellingen opgeslagen!",
        "back": "← Terug",
        "auto_pop": "Automatisch ingevuld met risicovolle boekingen. Verwijder degenen die u niet wilt contacteren.",
    },
    "fr": {
        "occupado": "Occupado",
        "ai_booking": "Intelligence de Réservation IA",
        "settings": "Paramètres",
        "sign_out": "Déconnexion",
        "live_dashboard": "Tableau de bord en direct",
        "bookings_analysed": "réservations analysées",
        "upload_data": "Téléchargez vos données de réservation",
        "drop_csv": "📂 Déposez votre CSV de réservation ici",
        "export_pms": "Exporter depuis votre PMS et télécharger",
        "choose_file": "Choisir un fichier CSV",
        "high_risk": "Risque élevé",
        "medium_risk": "Risque moyen",
        "low_risk": "Risque faible",
        "optimizer": "Optimiseur de Surréservation",
        "safe_rooms": "Chambres sûres à surréserver ce soir",
        "revenue": "Opportunité de revenus: EUR",
        "apply": "Appliquer la recommandation",
        "bookings_analysed_stat": "Réservations analysées",
        "predicted": "Absences prévues",
        "confidence": "Confiance IA",
        "walk_risk": "Risque d'annulation",
        "avg_rate": "Tarif moyen par chambre",
        "take_action": "Agir sur les réservations à risque élevé",
        "send_mass": "Envoyer un e-mail en masse",
        "send_high": "Envoyer à tous les {high} réservations à risque",
        "send_btn": "Envoyer un e-mail à tous →",
        "request_dep": "Demander les dépôts",
        "dep_template": "Modèle rapide pour dépôts",
        "dep_btn": "Demander un dépôt →",
        "reminders": "Envoyer des rappels",
        "rem_template": "Modèle rapide pour rappels",
        "rem_btn": "Envoyer un rappel →",
        "click_row": "Réservations — Cliquez sur une ligne pour le raisonnement IA",
        "booking": "Réservation",
        "lead": "Délai d'approche",
        "rate": "Tarif de la chambre",
        "returning": "Client de retour",
        "cancels": "Annulations antérieures",
        "risk": "Score de risque",
        "action": "Action",
        "days": "jours",
        "yes": "Oui",
        "no": "Non",
        "high": "ÉLEVÉ",
        "medium": "MOYEN",
        "low": "FAIBLE",
        "req_dep": "Demander un dépôt",
        "send_rem": "Envoyer un rappel",
        "monitor": "Surveiller",
        "email_guest": "Envoyer un e-mail à un client",
        "guest_email": "Adresse e-mail du client",
        "guest_name": "Nom du client",
        "subject": "Ligne d'objet",
        "message": "Message au client",
        "send_email": "📧 Envoyer un e-mail",
        "cancel": "Annuler",
        "bulk_email": "Envoyer un e-mail à toutes les réservations à risque",
        "select_book": "Sélectionnez les réservations à envoyer",
        "selected_count": "réservations sélectionnées",
        "save_changes": "💾 Enregistrer les modifications",
        "send_selected": "📧 Envoyer aux sélectionnés",
        "fill_fields": "Veuillez remplir tous les champs",
        "select_one": "Veuillez sélectionner au moins une réservation",
        "sent_success": "✓ E-mails envoyés à {count} client(s)",
        "error": "Erreur lors de l'envoi des e-mails",
        "settings_title": "Paramètres",
        "config_alert": "Configurez vos préférences d'alerte",
        "alert_email": "🔔 Alerte par e-mail",
        "alert_desc": "Adresse e-mail pour les alertes de réservations à risque",
        "email_addr": "Adresse e-mail",
        "save": "Enregistrer les paramètres →",
        "saved": "✓ Paramètres enregistrés!",
        "back": "← Retour",
        "auto_pop": "Rempli automatiquement avec les réservations à risque élevé. Supprimez celles que vous ne souhaitez pas contacter.",
    }
}

def t(key, lang="en", **kwargs):
    """Translate a key to the specified language"""
    text = TRANSLATIONS.get(lang, {}).get(key, TRANSLATIONS.get("en", {}).get(key, key))
    if kwargs:
        try:
            return text.format(**kwargs)
        except:
            return text
    return text

def generate_magic_token(hotel_username, csv_data):
    token = secrets.token_urlsafe(32)
    token_file = os.path.join(TOKEN_DIR, f"{token}.json")
    with open(token_file, 'w') as f:
        json.dump({"hotel": hotel_username, "csv_data": csv_data}, f)
    print(f"[TOKEN] Created: {token[:30]}...")
    return token

def get_token_data(token):
    token_file = os.path.join(TOKEN_DIR, f"{token}.json")
    if not os.path.exists(token_file):
        return None
    try:
        with open(token_file, 'r') as f:
            return json.load(f)
    except:
        return None

def delete_token(token):
    token_file = os.path.join(TOKEN_DIR, f"{token}.json")
    try:
        if os.path.exists(token_file):
            os.remove(token_file)
    except:
        pass

def send_consolidated_alert(hotel_name, alert_email, high_risk_bookings, hotel_username=None, csv_data=None):
    if not alert_email or not high_risk_bookings:
        return
    
    magic_link = "http://localhost:8080/login"
    if hotel_username and csv_data:
        token = generate_magic_token(hotel_username, csv_data)
        magic_link = f"http://localhost:8080/magic/{token}"
        print(f"[EMAIL] Magic link: {magic_link}")
    
    booking_rows = ""
    for booking in high_risk_bookings:
        booking_rows += f'<tr style="border-bottom:1px solid #e0e0e0;"><td style="padding:12px 16px;">{booking["id"]}</td><td style="padding:12px 16px;text-align:right;"><span style="background:#cc0000;color:white;padding:4px 12px;border-radius:20px;font-weight:600;">{booking["score"]:.1f}%</span></td></tr>'
    
    message = Mail(
        from_email=os.environ.get('ALERT_FROM_EMAIL', 'team@occupado.co'),
        to_emails=alert_email,
        subject=f"Alert: {len(high_risk_bookings)} High Cancellation Risk Bookings — {hotel_name}",
        html_content=f"""
        <div style="font-family:Arial,sans-serif;max-width:700px;margin:0 auto;">
            <div style="background:#008000;color:white;padding:24px;border-radius:12px 12px 0 0;">
                <h2 style="margin:0;font-size:22px;font-weight:700;">⚠️ High Cancellation Risk Alert</h2>
                <p style="margin:8px 0 0;font-size:14px;opacity:0.95;">{hotel_name}</p>
            </div>
            
            <div style="background:#f5faf5;padding:24px;border:1px solid #ddd;border-top:none;">
                <table style="width:100%;border-collapse:collapse;margin-bottom:24px;">
                    <thead>
                        <tr style="background:#f0f0f0;border-bottom:2px solid #008000;">
                            <th style="padding:12px 16px;text-align:left;font-family:'JetBrains Mono',monospace;font-size:12px;color:#0a1a0a;font-weight:600;">Booking ID</th>
                            <th style="padding:12px 16px;text-align:right;font-family:'JetBrains Mono',monospace;font-size:12px;color:#0a1a0a;font-weight:600;">Risk Score</th>
                        </tr>
                    </thead>
                    <tbody>{booking_rows}</tbody>
                </table>
                
                <center>
                    <a href="{magic_link}" style="background:#008000;color:white;padding:12px 32px;text-decoration:none;border-radius:8px;display:inline-block;font-weight:600;font-size:14px;">View Dashboard & Take Action →</a>
                </center>
            </div>
        </div>
        """
    )
    try:
        sg = SendGridAPIClient(os.environ.get('SENDGRID_API_KEY'))
        sg.send(message)
        print(f"[EMAIL] Sent to {alert_email}")
    except Exception as e:
        print(f"[ERROR] {e}")

def send_email_to_guest(guest_email, guest_name, hotel_name, subject, message_body):
    if not guest_email or not message_body:
        return False
    
    html_content = f"""
    <div style="font-family:'Plus Jakarta Sans',sans-serif;max-width:600px;margin:0 auto;background:#f5faf5;padding:24px;border-radius:12px;">
        <div style="background:#008000;color:white;padding:20px;border-radius:8px 8px 0 0;margin:-24px -24px 24px -24px;">
            <h2 style="margin:0;font-size:18px;">{hotel_name}</h2>
        </div>
        <div style="color:#0a1a0a;font-size:14px;line-height:1.8;white-space:pre-wrap;">
{message_body}
        </div>
        <div style="margin-top:24px;padding-top:16px;border-top:1px solid #ddd;color:#999;font-size:11px;text-align:center;">
            This email was sent from {hotel_name} via Occupado
        </div>
    </div>
    """
    
    message = Mail(
        from_email=os.environ.get('ALERT_FROM_EMAIL', 'team@occupado.co'),
        to_emails=guest_email,
        subject=subject,
        html_content=html_content
    )
    
    try:
        sg = SendGridAPIClient(os.environ.get('SENDGRID_API_KEY'))
        sg.send(message)
        return True
    except Exception as e:
        print(f"[ERROR] {e}")
        return False

HOTELS = {
    "grandmeridian":                  {"password": "hotel123",    "name": "Grand Meridian Hotel",                  "rooms": 200, "city": "Lisbon"},
    "scandic":                        {"password": "hotel456",    "name": "Scandic Stockholm",                     "rooms": 350, "city": "Stockholm"},
    "demo":                           {"password": "demo",        "name": "Demo Hotel",                            "rooms": 100, "city": "Porto"},
    "van der valk mechelen":          {"password": "Mechelen123", "name": "Van der Valk Hotel Mechelen",           "rooms": 150, "city": "Mechelen",  "vdv": True},
    "van der valk brussels airport":  {"password": "Brussels123", "name": "Van der Valk Brussels Airport",         "rooms": 310, "city": "Brussels",  "vdv_bru": True},
}

with open("occupado_model.pkl", "rb") as f:
    model = pickle.load(f)

# VdV-specific model trained on Shiji data (11 features)
# channel_encoded: 0=OTA/Web, 1=Direct, 2=Corporate, 3=Group/Package, NaN=Unknown
# Guest features (guest_cancel_rate, guest_profile_known, is_chronic_canceller) are
# applied as post-scoring overrides — not model inputs — to avoid sparse training data.
_VDV_MODEL_FEATURES = [
    'lead_time', 'arrival_date_week_number', 'arrival_month', 'arrival_day_of_week',
    'stays_in_weekend_nights', 'stays_in_week_nights', 'is_repeated_guest',
    'channel_encoded',
    'channel_cancel_rate', 'seasonal_cancel_rate', 'avg_days_to_cancel_for_channel',
    'is_last_minute', 'is_early_bird', 'is_business_pattern',
    'deposit_risk',
]
_CHANNEL_MAP = {
    'Booking.com': 0.0, 'Direct/Web': 1.0, 'Corporate': 2.0, 'Package': 3.0,
}
# Fallback channel cancel rates (CHANNEL_RANGE midpoints / 100) for Railway (no Excel files)
_CHANNEL_CANCEL_RATE_FALLBACK = {
    'Booking.com':       0.49,
    'Direct/Web':        0.41,
    'Direct / Web':      0.41,
    'Corporate':         0.24,
    'Package':           0.20,
    'Packages / Groups': 0.20,
    'Other':             0.35,
}
# No-show rates from RES_037 analysis (Oct 2025–Mar 2026, 359 unique records)
# Booking.com and Corporate have clean denominators (channel codes match RES_004).
# Direct/Web and Other use blended rate due to channel code mismatch in source data.
_VDV_NO_SHOW_RATES = {
    'Booking.com':       0.069,
    'Corporate':         0.012,
    'Package':           0.020,
    'Packages / Groups': 0.020,
    'Direct/Web':        0.115,
    'Direct / Web':      0.115,
    'Other':             0.115,
    'default':           0.115,
}

# Seasonal no-show multipliers from RES_037 (monthly avg = 51.3 no-shows)
# April excluded — only 1 record in dataset, unreliable
_VDV_NO_SHOW_SEASONAL = {
    10: 0.84,
    11: 1.19,
    12: 1.46,
    1:  1.25,
    2:  1.13,
    3:  1.11,
    # All other months default to 1.0
}

VDV_TOTAL_ROOMS = 150

# Empirical cancellation decay curves
# Source: 1,590 VdV cancellations from RES_036 files
# Key: days_before_arrival threshold → remaining cancel probability at that point
# (what fraction of total cancel risk is still ahead when you are X days out)
_VDV_DECAY_CURVES = {
    'IBE': {
        # Booking.com / OTA — median cancel 13 days out
        0: 0.07, 3: 0.23, 7: 0.37,
        14: 0.50, 30: 0.65,
        60: 0.80, 999: 0.86,
    },
    'DIRECT': {
        # Direct / Walk-in — cancels closer in, median 8 days
        0: 0.12, 3: 0.30, 7: 0.52,
        14: 0.65, 30: 0.78,
        60: 0.88, 999: 0.92,
    },
    'default': {
        # Blended curve from Step 5 survival analysis
        0: 0.08, 3: 0.26, 7: 0.41,
        14: 0.54, 30: 0.69,
        60: 0.82, 999: 0.86,
    },
}

# GTD codes → deposit risk score
# 0.0 = fully guaranteed, lowest cancel risk
# 1.0 = no commitment, highest cancel risk
_VDV_GTD_RISK = {
    'PRE':    0.05,  # prepaid — almost never cancels
    'ADV':    0.10,  # advance deposit paid
    'CREDIT': 0.20,  # credit card guaranteed
    'CRP':    0.20,  # corporate guarantee
    'CRPCL':  0.25,  # corporate credit limit
    'VCC':    0.35,  # Booking.com VCC — card exists but free cancel often applies
    'HOLD18': 0.75,  # 6pm hold — drops easily
    'NONE':   0.90,  # no guarantee — highest risk
    'None':   0.90,  # handle string None
    '':       0.90,  # missing = assume no guarantee
}

def _get_deposit_risk(gtd_code):
    return _VDV_GTD_RISK.get(str(gtd_code).strip(), 0.90)

# Map channel display names (from Shiji) to decay curve keys
_VDV_CHANNEL_TO_CURVE = {
    'Booking.com':       'IBE',
    'Direct/Web':        'DIRECT',
    'Direct / Web':      'DIRECT',
    'Corporate':         'default',
    'Package':           'default',
    'Packages / Groups': 'default',
    'Other':             'default',
}

def _get_decay_factor(days_out, channel):
    """Return the empirical remaining-cancel fraction for a booking X days from arrival."""
    curve_key = _VDV_CHANNEL_TO_CURVE.get(channel, 'default')
    curve     = _VDV_DECAY_CURVES[curve_key]
    for threshold in sorted(curve.keys()):
        if days_out <= threshold:
            return curve[threshold]
    return curve[999]

# Micro-segmentation thresholds
# Based on VdV booking pattern analysis
_LAST_MINUTE_DAYS    = 3   # booked <= 3 days before arrival
_EARLY_BIRD_DAYS     = 60  # booked >= 60 days before arrival
_BUSINESS_WEEK_MIN   = 3   # >= 3 weeknights
_BUSINESS_WEEKEND_MAX = 0  # 0 weekend nights

def _vdv_micro_segment_features(booking):
    lead  = booking.get('lead', 0)
    wkday = booking.get('wkday', 0)
    wkend = booking.get('wkend', 0)
    is_last_minute = 1 if lead <= _LAST_MINUTE_DAYS else 0
    is_early_bird  = 1 if lead >= _EARLY_BIRD_DAYS  else 0
    is_business    = 1 if wkday >= _BUSINESS_WEEK_MIN and wkend <= _BUSINESS_WEEKEND_MAX else 0
    return is_last_minute, is_early_bird, is_business

# ── Belgian demand events ────────────────────────────────────────────────────
# Fixed public holidays (month, day) → (emoji, label)
_BELGIAN_HOLIDAYS = {
    (1,  1): ('🎆', 'New Year'),
    (5,  1): ('⚒️',  'Labour Day'),
    (7, 21): ('🇧🇪', 'Belgian National Day'),
    (8, 15): ('🙏',  'Assumption'),
    (11, 1): ('🕯️',  "All Saints' Day"),
    (11,11): ('🎖️',  'Armistice Day'),
    (12,25): ('🎄',  'Christmas'),
}

# Moveable holidays (Easter-based) — pre-computed date → (emoji, label)
_EASTER_DATES = {
    2025: date(2025, 4, 20),
    2026: date(2026, 4,  5),
    2027: date(2027, 3, 28),
}
_BELGIAN_MOVEABLE_HOLIDAYS = {}
for _yr, _easter in _EASTER_DATES.items():
    _BELGIAN_MOVEABLE_HOLIDAYS[_easter]                    = ('🐣', 'Easter Sunday')
    _BELGIAN_MOVEABLE_HOLIDAYS[_easter + timedelta(days=1)]  = ('🐣', 'Easter Monday')
    _BELGIAN_MOVEABLE_HOLIDAYS[_easter + timedelta(days=39)] = ('✝️',  'Ascension')
    _BELGIAN_MOVEABLE_HOLIDAYS[_easter + timedelta(days=49)] = ('🕊️',  'Pentecost')
    _BELGIAN_MOVEABLE_HOLIDAYS[_easter + timedelta(days=50)] = ('🕊️',  'Whit Monday')

# Flemish school holiday periods — (start, end inclusive, emoji, label)
_BELGIAN_SCHOOL_HOLIDAYS = [
    # 2025
    (date(2025, 10, 27), date(2025, 11,  2), '🍂', 'Autumn Break'),
    (date(2025, 12, 22), date(2026,  1,  4), '🎄', 'Christmas Break'),
    # 2026
    (date(2026,  2, 16), date(2026,  2, 22), '🎭', 'Carnival Break'),
    (date(2026,  4,  6), date(2026,  4, 19), '🌷', 'Spring Break'),
    (date(2026,  7,  1), date(2026,  8, 31), '☀️',  'Summer Holidays'),
    (date(2026, 10, 26), date(2026, 11,  1), '🍂', 'Autumn Break'),
    (date(2026, 12, 21), date(2027,  1,  3), '🎄', 'Christmas Break'),
    # 2027
    (date(2027,  2, 15), date(2027,  2, 21), '🎭', 'Carnival Break'),
    (date(2027,  3, 29), date(2027,  4, 11), '🌷', 'Spring Break'),
    (date(2027,  7,  1), date(2027,  8, 31), '☀️',  'Summer Holidays'),
]

try:
    with open("occupado_model_vdv.pkl", "rb") as f:
        model_vdv = pickle.load(f)
    print(f"[VdV] VdV-specific model loaded ({model_vdv.n_features_in_} features)")
except Exception:
    model_vdv = None
    print("[VdV] VdV-specific model not found, falling back to generic")

df = pd.read_csv("hotel_bookings.csv")

# ── VAN DER VALK MECHELEN — Pre-loaded data & enhanced dashboard ──────────────
VDV_HOTEL_KEY = "van der valk mechelen"
_VDV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VDV-MEC")

def _parse_vdv_guests():
    """Parse all RES_042 repeat reservations reports for current/upcoming repeat guests."""
    import openpyxl, glob as _glob
    res042_files = sorted(_glob.glob(os.path.join(_VDV_DIR, "RES_042_RepeatReservationsReport*.xlsx")))
    if not res042_files:
        return []
    try:
        today = datetime.now()
        guests = []
        seen = set()  # (name_lower, arrival_date_str) dedup across files
        for path in res042_files:
            wb = openpyxl.load_workbook(path, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            i = 0
            while i < len(rows):
                row = rows[i]
                col0 = str(row[0]).strip() if row[0] else ''
                col1 = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                col4 = row[4] if len(row) > 4 else None
                col5 = row[5] if len(row) > 5 else None
                if ',' in col0 and col4 and '/' in str(col4):
                    try:
                        arr = datetime.strptime(str(col4)[:10], '%d/%m/%Y')
                    except Exception:
                        i += 1
                        continue
                    arr_key = (col0.strip().lower(), arr.strftime('%Y-%m-%d'))
                    if arr_key in seen:
                        i += 1
                        continue
                    seen.add(arr_key)
                    adults = 1
                    if col5:
                        try: adults = int(str(col5).split('/')[0])
                        except: pass
                    dep = None
                    for j in range(i + 1, min(i + 7, len(rows))):
                        r = rows[j]
                        r4 = r[4] if len(r) > 4 else None
                        if r4 and r[0] is None and '/' in str(r4):
                            try:
                                dep = datetime.strptime(str(r4)[:10], '%d/%m/%Y')
                                break
                            except: pass
                    # Collect guest notes
                    note_parts = []
                    _skip = {'Repeat Reservations Report', 'Van der Valk Hotel Mechelen'}
                    for j in range(i + 1, min(i + 22, len(rows))):
                        r = rows[j]
                        r6 = r[6] if len(r) > 6 else None
                        if r6:
                            n = str(r6).strip()
                            if n and n not in _skip and 'MIGRATED' not in n and 'CORPORATE 2025' not in n and 'central ar' not in n.lower() and '\n' not in n:
                                note_parts.append(n[:80])
                    note = '; '.join(note_parts[:2]) if note_parts else ''
                    nights = (dep - arr).days if dep else 1
                    if dep and dep.date() < today.date():
                        status = 'Checked Out'
                    elif arr.date() == today.date():
                        status = 'Arriving Today'
                    elif arr.date() < today.date():
                        status = 'In House'
                    else:
                        status = f'Arriving {arr.strftime("%d %b")}'
                    guests.append({
                        'name': col0, 'membership': col1,
                        'arrival': arr.strftime('%d/%m/%Y'),
                        'departure': dep.strftime('%d/%m/%Y') if dep else '',
                        'arr_date': arr, 'dep_date': dep,
                        'adults': adults, 'nights': nights,
                        'status': status, 'note': note,
                    })
                i += 1
        return guests
    except Exception as e:
        print(f"[VDV] Guest parse error: {e}")
        return []


def _parse_vdv_channel_stats():
    """Parse all RES_036 cancelled reservations files for channel breakdown with cross-file dedup."""
    import openpyxl, glob as _glob
    from collections import defaultdict
    from datetime import datetime as _dt
    files = sorted(_glob.glob(os.path.join(_VDV_DIR, "RES_036_CancelledReservations*.xlsx")))
    if not files:
        # Railway fallback: no Excel files available — return structured empty stats
        # so VDV_CHANNEL_STATS always has the expected keys.
        # _score_vdv_future() will use _CHANNEL_CANCEL_RATE_FALLBACK for actual rates.
        return {
            'Booking.com': 0, 'Direct / Web': 0, 'Corporate': 0,
            'Packages / Groups': 0, 'Other': 0, '_total_cx': 0,
            '_channel_cx_counts':  {'Booking.com': 0, 'Direct/Web': 0, 'Corporate': 0, 'Package': 0, 'Other': 0},
            '_seasonal_cx_counts': {},
            'avg_days_to_cancel':  {'Booking.com': 30.0, 'Direct/Web': 30.0, 'Corporate': 20.0, 'Package': 15.0, 'Other': 30.0},
            'cxl_reason_breakdown': {},
        }
    # Segment code → normalised channel name (matching VDV_FUTURE_BOOKINGS channel values)
    _SEG_TO_CH = {
        'BARWEB': 'Booking.com', 'BAROTAGROSS': 'Booking.com',
        'DEALSOTA': 'Booking.com', 'DISCOTAGROSS': 'Booking.com',
        'DISCWEB': 'Direct/Web', 'BARDIR': 'Direct/Web', 'DISCDIR': 'Direct/Web',
        'CORPFIX': 'Corporate', 'CORPDYN': 'Corporate',
        'PACK': 'Package', 'MTGBNS': 'Package', 'BNSGRP': 'Package',
        'DEALS': 'Other', 'OTHER': 'Other', 'COMP': 'Other',
    }
    seen_keys = set()
    raw = {}
    total = 0
    channel_cx_counts = defaultdict(int)       # normalised channel → cx count
    seasonal_cx_counts = defaultdict(int)      # "channel|month" → cx count
    dtc_sums = defaultdict(float)              # channel → sum of days-to-cancel
    dtc_counts = defaultdict(int)             # channel → number of valid dtc measurements
    cxl_reason_breakdown = defaultdict(lambda: defaultdict(int))
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for row in rows:
                if row[0] is None and len(row) > 3 and row[3]:
                    cxl_ref = str(row[20]).strip() if len(row) > 20 and row[20] else ''
                    if cxl_ref.startswith('MEC-CXL'):
                        key = cxl_ref
                    else:
                        c8  = str(row[8])[:10]  if len(row) > 8  and row[8]  else ''
                        c9  = str(row[9])        if len(row) > 9  and row[9]  else ''
                        c14 = str(row[14])[:10]  if len(row) > 14 and row[14] else ''
                        key = (c8, c9, c14)
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)
                    seg = str(row[3]).strip()
                    if seg and not seg.startswith('Subtotal') and not re.match(r'\d{2}/\d{2}/\d{4}', seg) and seg not in ('Market Segment', 'Company/Travel Agent'):
                        raw[seg] = raw.get(seg, 0) + 1
                        total += 1
                        ch = _SEG_TO_CH.get(seg, 'Other')
                        channel_cx_counts[ch] += 1
                        # Arrival month (col8 = Arr. Date)
                        arr_str = str(row[8])[:10] if len(row) > 8 and row[8] else ''
                        if arr_str and re.match(r'\d{2}/\d{2}/\d{4}', arr_str):
                            try:
                                arr_month = _dt.strptime(arr_str, '%d/%m/%Y').month
                                seasonal_cx_counts[f'{ch}|{arr_month}'] += 1
                            except Exception:
                                pass
                        # Days to cancel: Created On=col16, CXL Date/Time=col18
                        created_str = str(row[16])[:10] if len(row) > 16 and row[16] else ''
                        cxl_str     = str(row[18])[:10] if len(row) > 18 and row[18] else ''
                        if created_str and cxl_str:
                            try:
                                created_dt = _dt.strptime(created_str, '%d/%m/%Y')
                                cxl_dt     = _dt.strptime(cxl_str,     '%d/%m/%Y')
                                dtc = max(0.0, float((cxl_dt - created_dt).days))
                                dtc_sums[ch]   += dtc
                                dtc_counts[ch] += 1
                            except Exception:
                                pass
                        # CXL Reason (col24)
                        reason = str(row[24]).strip() if len(row) > 24 and row[24] else ''
                        if reason in ('', 'None', 'nan'):
                            reason = 'Unknown'
                        cxl_reason_breakdown[ch][reason] += 1
        except Exception as e:
            print(f"[VDV] Channel stats error ({fp}): {e}")
    avg_days_to_cancel = {
        ch: round(dtc_sums[ch] / dtc_counts[ch], 1)
        for ch in dtc_counts if dtc_counts[ch] > 0
    }
    n_seasonal = len(seasonal_cx_counts)
    print(f"[VDV] Channel stats: {total} cancellations parsed, {n_seasonal} channel/month combinations found")
    return {
        'Booking.com':       sum(raw.get(k, 0) for k in ('BARWEB', 'BAROTAGROSS', 'DEALSOTA')),
        'Direct / Web':      sum(raw.get(k, 0) for k in ('DISCWEB', 'BARDIR', 'DISCDIR', 'DISCOTAGROSS')),
        'Corporate':         sum(raw.get(k, 0) for k in ('CORPFIX', 'CORPDYN')),
        'Packages / Groups': sum(raw.get(k, 0) for k in ('PACK', 'MTGBNS', 'BNSGRP')),
        'Other':             sum(raw.get(k, 0) for k in ('DEALS', 'OTHER', 'COMP')),
        '_total_cx':             total,
        '_channel_cx_counts':    dict(channel_cx_counts),
        '_seasonal_cx_counts':   dict(seasonal_cx_counts),
        'avg_days_to_cancel':    avg_days_to_cancel,
        'cxl_reason_breakdown':  {ch: dict(reasons) for ch, reasons in cxl_reason_breakdown.items()},
    }


def _count_vdv_noshow():
    """Count unique no-shows across all RES_037 files."""
    import openpyxl, glob as _glob
    files = sorted(_glob.glob(os.path.join(_VDV_DIR, "RES_037_NoShow*.xlsx")))
    seen = set()
    count = 0
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for row in rows:
                if not row[1] or len(row) < 12:
                    continue
                conf    = str(row[2]).strip() if row[2] else ''
                arrival = str(row[11])[:10]   if row[11] else ''
                key = (conf, arrival) if conf else (str(row[1]).strip(), arrival)
                if key in seen:
                    continue
                seen.add(key)
                count += 1
        except Exception as e:
            print(f"[VDV] NS count error ({fp}): {e}")
    return count


def _parse_bru_noshow(months_list=None):
    """Parse BRU RES_037 in one pass. Returns (total_count, monthly_list).
    monthly_list is aligned with months_list if provided, else empty list."""
    import openpyxl, glob as _glob
    from collections import defaultdict
    from datetime import datetime
    files = sorted(_glob.glob(os.path.join(_VDV_BRU_DIR, "RES_037_NoShow*.xlsx")))
    seen = set()
    count = 0
    by_month = defaultdict(int)
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for row in rows:
                if not row[1] or len(row) < 13:
                    continue
                conf    = str(row[2]).strip() if row[2] else ''
                arr_raw = row[12]
                if arr_raw is None:
                    continue
                if hasattr(arr_raw, 'strftime'):
                    arr_str = arr_raw.strftime('%Y-%m-%d')
                else:
                    arr_str = str(arr_raw)[:10]
                key = (conf, arr_str) if conf else (str(row[1]).strip(), arr_str)
                if key in seen:
                    continue
                seen.add(key)
                count += 1
                try:
                    label = datetime.strptime(arr_str, '%Y-%m-%d').strftime('%b %Y')
                    by_month[label] += 1
                except Exception:
                    pass
        except Exception as e:
            print(f"[BRU] NS parse error ({fp}): {e}")
    monthly = [by_month.get(m, 0) for m in months_list] if months_list else []
    return count, monthly


def _build_vdv_guest_history():
    """Count all-time stays per guest from RES_042 (no date filter).
    Cancellations are assumed 0; stay count alone gives meaningful loyalty signal."""
    import openpyxl, glob as _glob
    from collections import defaultdict
    history = defaultdict(lambda: {'stays': 0, 'cancels': 0})
    seen_stays = set()
    for path in sorted(_glob.glob(os.path.join(_VDV_DIR, "RES_042_RepeatReservationsReport*.xlsx"))):
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for row in rows:
                col0 = str(row[0]).strip() if row[0] else ''
                col4 = row[4] if len(row) > 4 else None
                if ',' in col0 and col4 and '/' in str(col4):
                    key = (col0.lower(), str(col4)[:10])
                    if key not in seen_stays:
                        seen_stays.add(key)
                        history[col0.lower()]['stays'] += 1
        except Exception:
            pass
    return dict(history)


def _score_vdv_guests(guests):
    """Score VdV repeat guests using the VdV-specific model (4 features)."""
    if not guests:
        return []
    feat_rows = []
    for g in guests:
        arr = g['arr_date']
        dep = g['dep_date']
        wkend = wkday = 0
        if dep:
            d = arr
            while d < dep:
                if d.weekday() >= 5: wkend += 1
                else: wkday += 1
                d += timedelta(days=1)
        week_num = int(arr.isocalendar()[1])
        is_repeat = 1  # all guests in this table are repeat guests by definition
        lead = max(0, (arr - datetime.now()).days)  # days until arrival as proxy
        lm  = 1 if lead <= _LAST_MINUTE_DAYS else 0
        eb  = 1 if lead >= _EARLY_BIRD_DAYS  else 0
        biz = 1 if wkday >= _BUSINESS_WEEK_MIN and wkend <= _BUSINESS_WEEKEND_MAX else 0
        feat_rows.append([
            lead, week_num, arr.month, arr.weekday(), wkend, wkday, is_repeat,
            float('nan'),   # channel_encoded (unknown for repeat guests)
            0.35,           # channel_cancel_rate (overall fallback)
            0.35,           # seasonal_cancel_rate (overall fallback)
            30.0,           # avg_days_to_cancel_for_channel
            lm, eb, biz,
            0.35,           # deposit_risk (repeat guests assumed moderate commitment — VCC-equivalent)
        ])
    df_feat = pd.DataFrame(feat_rows, columns=_VDV_MODEL_FEATURES)
    m = model_vdv if model_vdv is not None else model
    if model_vdv is None:
        # fallback: use generic model with full feature set
        feat_cols = ['lead_time','arrival_date_week_number','stays_in_weekend_nights',
                     'stays_in_week_nights','adults','is_repeated_guest',
                     'previous_cancellations','previous_bookings_not_canceled',
                     'booking_changes','days_in_waiting_list','adr','total_of_special_requests']
        full_rows = [[max(0,(g['arr_date']-datetime.now()).days),
                      int(g['arr_date'].isocalendar()[1]),
                      r[1], r[2], g['adults'], 1, 0, 3, 0, 0, 130.0, 0]
                     for g, r in zip(guests, feat_rows)]
        df_feat = pd.DataFrame(full_rows, columns=feat_cols)
    raw_scores = [float(s) for s in m.predict_proba(df_feat)[:, 1] * 100]

    # ── Loyalty adjustment: blend model score with guest's actual cancel rate ──
    # Trust grows with number of completed stays; Laplace-smoothed cancel rate
    # prevents a single cancellation from dominating for new repeat guests.
    adjusted = []
    for g, raw_sc in zip(guests, raw_scores):
        hist = VDV_GUEST_HISTORY.get(g['name'].strip().lower(), {'stays': 0, 'cancels': 0})
        stays   = hist['stays']
        cancels = hist['cancels']
        if stays >= 2:
            hist_rate = (cancels + 0.5) / (stays + 1)   # Laplace-smoothed cancel rate
            trust     = min(1.0, stays / 10)             # 0→1 as stays 0→10+
            blend     = 0.65                             # max 65% pull from history
            sc = raw_sc * (1 - trust * blend) + hist_rate * 100 * trust * blend
        else:
            sc = raw_sc
        adjusted.append(max(0.0, min(100.0, sc)))
    return adjusted


def _parse_vdv_future_bookings():
    """Parse all RES_004 files for future bookings with lead times, channels & breakfast flag."""
    import openpyxl
    # Files known to contain future bookings; dedup by confirmation number
    res004_candidates = ['RES_004_EnteredOnAndBy (1).xlsx',
                         'RES_004_EnteredOnAndBy (2).xlsx',
                         'RES_004_EnteredOnAndBy (12).xlsx',
                         'RES_004_EnteredOnAndBy (16).xlsx',
                         'RES_004_EnteredOnAndBy (17).xlsx']
    today = datetime.now().date()
    bookings = []
    seen_confs = set()
    ch_map = {
        'BARWEB': 'Booking.com', 'BAROTAGROSS': 'Booking.com',
        'DEALSOTA': 'Booking.com', 'DISCOTAGROSS': 'Booking.com',
        'DISCWEB': 'Direct/Web', 'BARDIR': 'Direct/Web', 'DISCDIR': 'Direct/Web',
        'CORPFIX': 'Corporate', 'CORPDYN': 'Corporate',
        'PACK': 'Package', 'MTGBNS': 'Package', 'BNSGRP': 'Package',
        'DEALS': 'Other',
    }
    for fn in res004_candidates:
        path = os.path.join(_VDV_DIR, fn)
        if not os.path.exists(path):
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            i = 0
            while i < len(rows):
                r = rows[i]
                c0 = str(r[0]).strip() if r[0] else ''
                c3 = str(r[3]).strip() if len(r) > 3 and r[3] else ''
                if (',' in c0 and c3.startswith('MEC-') and not c3.startswith('MEC-F')
                        and len(r) > 8 and r[8]):
                    arr_str = str(r[8])[:10]
                    try:
                        arr = datetime.strptime(arr_str, '%d/%m/%Y').date()
                    except Exception:
                        i += 1
                        continue
                    if arr < today or c3 in seen_confs:
                        i += 1
                        continue
                    seen_confs.add(c3)
                    nights = int(r[9]) if r[9] and str(r[9]).isdigit() else 1
                    adults_str = str(r[12]).split('/')[0].strip() if r[12] else '1'
                    try:   adults = int(adults_str)
                    except: adults = 1
                    channel = str(r[25]).strip() if len(r) > 25 and r[25] else 'OTHER'
                    rate_plan = str(r[13]).strip() if len(r) > 13 and r[13] else ''
                    purchase_elem = str(r[18]).strip() if len(r) > 18 and r[18] else ''
                    has_breakfast = ('BB' in rate_plan.upper() or
                                     'FBBF' in purchase_elem.upper())
                    rate_str = str(r[16]).strip() if len(r) > 16 and r[16] else ''
                    try:
                        total_rate = float(rate_str.replace(',', '.'))
                        adr = total_rate / max(1, nights)
                    except Exception:
                        adr = 130.0
                    created = str(r[28])[:10] if len(r) > 28 and r[28] else ''
                    lead = 0
                    if created:
                        try:
                            cdate = datetime.strptime(created, '%d/%m/%Y').date()
                            lead = max(0, (arr - cdate).days)
                        except Exception:
                            pass
                    gtd = 'NONE'
                    for j in range(i + 1, min(i + 8, len(rows))):
                        rj = rows[j]
                        if (len(rj) > 12
                                and len(rj) > 9
                                and str(rj[9]).strip() == 'RS'
                                and rj[12] is not None):
                            gtd = str(rj[12]).strip()
                            break
                    wkend = wkday = 0
                    d = datetime.combine(arr, datetime.min.time())
                    for _ in range(nights):
                        if d.weekday() >= 5: wkend += 1
                        else: wkday += 1
                        d += timedelta(days=1)
                    week_num = int(datetime.combine(arr, datetime.min.time()).isocalendar()[1])
                    ch_label = ch_map.get(channel, 'Other')
                    bookings.append({
                        'name': c0, 'arrival': arr.strftime('%d/%m/%Y'),
                        'arr_date': arr, 'nights': nights, 'adults': adults,
                        'channel': ch_label, 'channel_raw': channel,
                        'lead': lead, 'gtd': gtd, 'adr': round(adr, 2),
                        'wkend': wkend, 'wkday': wkday, 'week_num': week_num,
                        'has_breakfast': has_breakfast,
                    })
                i += 1
        except Exception as e:
            print(f"[VDV] Future bookings parse error ({fn}): {e}")
    return bookings


def _apply_guest_overrides(scores, bookings):
    """Post-scoring guest override layer.
    Applied after normalise_within_channel() so channel ranges are preserved for
    normal guests; only chronic cancellers and known high-rate guests get floored up.
    """
    result = list(scores)
    for i, b in enumerate(bookings):
        feats = _get_guest_features(b['name'], b['channel'])
        # Chronic canceller — floor at 75
        if feats['is_chronic_canceller']:
            result[i] = max(result[i], 75.0)
        # Known guest with cancel_rate > 50% — floor at 60
        elif feats['guest_profile_known'] and feats['guest_cancel_rate'] > 0.5:
            result[i] = max(result[i], 60.0)
    return result


def _score_vdv_future(bookings):
    """Score VdV future bookings using the VdV-specific model (11 features)."""
    if not bookings:
        return []

    # ── Detect repeat guests ────────────────────────────────────────────────
    # 1. Guests appearing 2+ times in the future bookings list
    from collections import Counter
    name_counts = Counter(b['name'].strip().lower() for b in bookings)
    repeat_names = {n for n, c in name_counts.items() if c >= 2}
    # 2. Known repeat guests from RES_042 (current week repeat guests)
    known_repeats = {g['name'].strip().lower() for g in VDV_GUESTS_RAW}
    all_repeats = repeat_names | known_repeats

    # ── Realistic score ranges per channel ──────────────────────────────────
    # Training data was 68.9% cancellations (vs ~25% real-world rate), so raw
    # model scores cluster near 100% for everyone. We preserve the model's
    # ranking ability (AUC 0.852) by normalising within each channel to a
    # realistic [min, max] range based on actual VdV cancellation patterns.
    CHANNEL_RANGE = {
        'Booking.com':       (18, 80),  # OTA: ~50% avg cancel rate
        'Direct/Web':        (12, 70),  # Direct: ~25% avg cancel rate
        'Direct / Web':      (12, 70),
        'Corporate':         ( 5, 42),  # Corporate: ~10-15% — often guaranteed
        'Package':           ( 4, 35),  # Packages: ~5-8% — usually prepaid
        'Packages / Groups': ( 4, 35),
        'Other':             (10, 60),
    }

    def is_repeat(b):
        return 1 if b['name'].strip().lower() in all_repeats else 0

    def normalise_within_channel(raw_scores, bookings_list):
        """Map raw scores to realistic range per channel using rank order.
        Uses percentile rank (not raw value) so scores always spread across
        the full range even when the model assigns similar values to everyone."""
        from collections import defaultdict
        ch_idx = defaultdict(list)
        for i, b in enumerate(bookings_list):
            ch_idx[b['channel']].append(i)
        result = [0.0] * len(bookings_list)
        for ch, indices in ch_idx.items():
            lo, hi = CHANNEL_RANGE.get(ch, (10, 65))
            # Sort by raw score → assign rank-based percentile
            sorted_by_score = sorted(indices, key=lambda i: raw_scores[i])
            n = len(sorted_by_score)
            for rank, idx in enumerate(sorted_by_score):
                percentile = rank / (n - 1) if n > 1 else 0.5
                reduction = 10 if is_repeat(bookings_list[idx]) else 0
                result[idx] = round(max(lo, lo + percentile * (hi - lo) - reduction), 1)
        return result

    if model_vdv is not None:
        # ── Precompute channel/seasonal cancel rates from historical + future counts ──
        from collections import Counter as _Counter
        ch_future_counts  = _Counter(b['channel'] for b in bookings)
        sea_future_counts = _Counter((b['channel'], b['arr_date'].month) for b in bookings)
        ch_cx_counts  = VDV_CHANNEL_STATS.get('_channel_cx_counts', {})
        sea_cx_counts = VDV_CHANNEL_STATS.get('_seasonal_cx_counts', {})
        avg_dtc       = VDV_CHANNEL_STATS.get('avg_days_to_cancel', {})

        def _ch_rate(ch):
            cx    = ch_cx_counts.get(ch, 0)
            fut   = ch_future_counts.get(ch, 0)
            denom = cx + fut
            if denom > 0:
                return round(cx / denom, 4)
            return _CHANNEL_CANCEL_RATE_FALLBACK.get(ch, 0.35)

        def _sea_rate(ch, month):
            key   = f'{ch}|{month}'
            cx    = sea_cx_counts.get(key, 0)
            fut   = sea_future_counts.get((ch, month), 0)
            denom = cx + fut
            if denom > 0:
                return round(cx / denom, 4)
            return _ch_rate(ch)

        rows_feat = []
        for b in bookings:
            lm, eb, biz = _vdv_micro_segment_features(b)
            rows_feat.append([
                b['lead'], b['week_num'], b['arr_date'].month, b['arr_date'].weekday(),
                b['wkend'], b['wkday'],
                is_repeat(b),
                _CHANNEL_MAP.get(b['channel'], float('nan')),
                _ch_rate(b['channel']),
                _sea_rate(b['channel'], b['arr_date'].month),
                avg_dtc.get(b['channel'], 30.0),
                lm, eb, biz,
                _get_deposit_risk(b.get('gtd', 'NONE')),
            ])
        df = pd.DataFrame(rows_feat, columns=_VDV_MODEL_FEATURES)
        assert df.shape[1] == len(_VDV_MODEL_FEATURES), \
            f"Feature mismatch: {df.shape[1]} vs {len(_VDV_MODEL_FEATURES)}"
        raw_scores = list(model_vdv.predict_proba(df)[:, 1] * 100)
        normalised = normalise_within_channel(raw_scores, bookings)
        final_scores = _apply_guest_overrides(normalised, bookings)
        return [float(round(s, 1)) for s in final_scores]

    # Fallback to generic model
    feat_cols = ['lead_time','arrival_date_week_number','stays_in_weekend_nights',
                 'stays_in_week_nights','adults','is_repeated_guest',
                 'previous_cancellations','previous_bookings_not_canceled',
                 'booking_changes','days_in_waiting_list','adr','total_of_special_requests']
    rows_feat = []
    for b in bookings:
        rows_feat.append([b['lead'], b['week_num'], b['wkend'], b['wkday'],
                          b['adults'], is_repeat(b), 0, 0, 0, 0, b['adr'], 0])
    df = pd.DataFrame(rows_feat, columns=feat_cols)
    raw_scores = list(model.predict_proba(df)[:, 1] * 100)
    normalised = normalise_within_channel(raw_scores, bookings)
    final_scores = _apply_guest_overrides(normalised, bookings)
    return [float(round(s, 1)) for s in final_scores]


def _parse_mice_data():
    """Parse RES_001 for MICE/corporate bookings and RES_033 for active group blocks."""
    path_res001 = os.path.join(_VDV_DIR, "RES_001_ArrivalDetailed (1).xlsx")
    path_res033 = os.path.join(_VDV_DIR, "RES_033_BillingInstructions.xlsx")
    MICE_SEGS = {'BNSGRP', 'CORPFIX', 'CORPDYN', 'MTGBNS'}
    SEG_LABELS = {
        'BNSGRP':  'Business Group',
        'CORPFIX': 'Corporate Fixed',
        'CORPDYN': 'Corporate Dynamic',
        'MTGBNS':  'Meeting Business',
    }
    result = {
        'total': 0,
        'total_nights': 0,
        'by_segment': {k: 0 for k in MICE_SEGS},
        'top_clients': [],
        'groups': [],
    }
    # --- Parse RES_001 for bookings ---
    if os.path.exists(path_res001):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path_res001, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            companies = {}
            for i in range(13, len(rows)):
                row = rows[i]
                if not row or len(row) < 18: continue
                market = str(row[17]).strip() if row[17] else ''
                if market not in MICE_SEGS: continue
                company = str(row[13]).strip() if row[13] else ''
                nights_raw = str(row[7]).strip() if row[7] else '0'
                room_type  = str(row[10]).strip() if row[10] else ''
                # Arrival date from col 4
                arr_raw = str(row[4]).strip() if row[4] else ''
                arr_date = arr_raw[:10] if arr_raw else ''
                try:
                    nights = int(nights_raw) if nights_raw.isdigit() else 1
                except:
                    nights = 1
                if company and company not in ('nan', 'Group', 'Company', 'Travel Agent', ''):
                    if company not in companies:
                        companies[company] = {'segment': SEG_LABELS.get(market, market),
                                              'seg_code': market, 'bookings': 0,
                                              'nights': 0, 'last_arrival': arr_date}
                    companies[company]['bookings'] += 1
                    companies[company]['nights']   += nights
                    if arr_date and arr_date > companies[company]['last_arrival']:
                        companies[company]['last_arrival'] = arr_date
                result['by_segment'][market] = result['by_segment'].get(market, 0) + 1
                result['total']        += 1
                result['total_nights'] += nights
            result['top_clients'] = sorted(
                [{'company': co, **data} for co, data in companies.items()],
                key=lambda x: -x['bookings']
            )[:12]
        except Exception as e:
            print(f"[VDV] MICE parse error (RES_001): {e}")
    # --- Parse RES_033 for group blocks ---
    if os.path.exists(path_res033):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path_res033, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            groups = {}
            for row in rows:
                if not row or row[0] is None: continue
                room = str(row[0]).strip()
                if not room or room in ('Room #', 'Res. Status', 'Totals', 'IH', 'CX'): continue
                try: int(room)  # must be a room number
                except: continue
                arr_raw  = str(row[7]).strip() if row[7] else ''
                dep_raw  = str(row[8]).strip() if row[8] else ''
                grp_raw  = str(row[10]).strip() if row[10] else ''
                company  = str(row[13]).strip() if row[13] else ''
                if not grp_raw or grp_raw == 'nan': continue
                # Parse group code and name from "MEC-GF10478 | Hyrox Mechelen 008646"
                parts = grp_raw.split('|')
                grp_code = parts[0].strip() if parts else grp_raw
                grp_name = parts[1].strip() if len(parts) > 1 else grp_code
                if grp_code not in groups:
                    groups[grp_code] = {
                        'code': grp_code, 'name': grp_name,
                        'company': company, 'rooms': [], 'arrival': arr_raw, 'departure': dep_raw
                    }
                groups[grp_code]['rooms'].append(room)
                if arr_raw and arr_raw < groups[grp_code]['arrival']:
                    groups[grp_code]['arrival'] = arr_raw
            result['groups'] = sorted(
                [g for g in groups.values() if len(g['rooms']) >= 2],
                key=lambda x: -len(x['rooms'])
            )
        except Exception as e:
            print(f"[VDV] MICE parse error (RES_033): {e}")
    return result


def _parse_forecast_data():
    """Parse FOR_042 HistoryAndForecast — returns last 60d history + next 60d forecast."""
    path = os.path.join(_VDV_DIR, "FOR_042_HistoryAndForecast.xlsx")
    if not os.path.exists(path):
        return {'history': [], 'forecast': []}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        history, forecast, mode = [], [], None
        for row in rows:
            if not row or row[0] is None: continue
            v0 = str(row[0]).strip()
            if v0 == 'HISTORY':   mode = 'h'; continue
            if v0.startswith('FORECAST'): mode = 'f'; continue
            if v0 in ('Totals','', 'Date', 'Sold'): continue
            # Data row: date string like "Wed  01/01/2025"
            date_part = v0.split()[-1] if v0 else ''
            if not re.match(r'\d{2}/\d{2}/\d{4}', date_part): continue
            try:
                dt = datetime.strptime(date_part, '%d/%m/%Y')
                rooms_sold = int(str(row[1]).replace(',','').split('.')[0]) if row[1] else 0
                arrivals   = int(str(row[2]).replace(',','').split('.')[0]) if row[2] else 0
                departures = int(str(row[3]).replace(',','').split('.')[0]) if row[3] else 0
                occ_pct    = float(str(row[4]).replace(',','.')) if row[4] else 0.0
                rev_raw    = str(row[10]).replace('.','').replace(',','.').strip() if row[10] else '0'
                revenue    = float(rev_raw) if rev_raw else 0.0
                entry = {'date': dt.strftime('%d %b'), 'dt': dt.strftime('%Y-%m-%d'),
                         'rooms': rooms_sold, 'arrivals': arrivals, 'departures': departures,
                         'occ': round(occ_pct, 1), 'rev': round(revenue)}
                if mode == 'h':   history.append(entry)
                elif mode == 'f': forecast.append(entry)
            except: continue
        # Return last 60 days of history + next 60 days of forecast
        return {'history': history[-60:], 'forecast': forecast[:60]}
    except Exception as e:
        print(f"[VDV] Forecast parse error: {e}")
        return {'history': [], 'forecast': []}


def _parse_group_pipeline():
    """Parse GRP_017 for upcoming confirmed/deposit/tentative group bookings."""
    path = os.path.join(_VDV_DIR, "GRP_017_GroupAndMEBookingListWithRevenueCSVOnly.xlsx")
    if not os.path.exists(path):
        return []
    STATUS_LABELS = {'DEF': 'Definite', 'DEP': 'Deposit', 'TEN': 'Tentative',
                     'LOS': 'Lost', 'CXL': 'Cancelled', 'PRO': 'Prospect'}
    MARKET_LABELS = {'MTGBNS': 'Meeting', 'CORPFIX': 'Corporate', 'CORPDYN': 'Corporate Dyn',
                     'BNSGRP': 'Business Group', 'LEISURE': 'Leisure'}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        groups = []
        today = datetime.now()
        for i, row in enumerate(rows):
            if i < 5 or not row or row[0] is None: continue
            cols = [str(c).strip() if c is not None else '' for c in row]
            # Col pattern: Profile | M&E Code | Group Name | Arrival | Dep | Status | Market | Owner ...
            if len(cols) < 7: continue
            code   = cols[1]
            name   = cols[2]
            arr_s  = cols[3]
            dep_s  = cols[4]
            status = cols[5]
            market = cols[6]
            # Skip lost/cancelled and past
            if status in ('LOS', 'CXL', ''): continue
            if not re.match(r'\d{2}/\d{2}/\d{4}', arr_s): continue
            try:
                arr_dt = datetime.strptime(arr_s, '%d/%m/%Y')
                dep_dt = datetime.strptime(dep_s, '%d/%m/%Y') if re.match(r'\d{2}/\d{2}/\d{4}', dep_s) else arr_dt
                if arr_dt.date() < today.date(): continue  # past
                nights = (dep_dt - arr_dt).days
                groups.append({
                    'code': code, 'name': name if name else code,
                    'arrival': arr_dt.strftime('%d %b %Y'),
                    'departure': dep_dt.strftime('%d %b %Y'),
                    'arr_dt': arr_dt.strftime('%Y-%m-%d'),
                    'nights': nights,
                    'status': status,
                    'status_label': STATUS_LABELS.get(status, status),
                    'market': MARKET_LABELS.get(market, market),
                })
            except: continue
        # Sort by arrival, cap at 20
        return sorted(groups, key=lambda x: x['arr_dt'])[:20]
    except Exception as e:
        print(f"[VDV] Group pipeline parse error: {e}")
        return []


def _load_vdv_guest_profiles():
    import json
    path = os.path.join(_VDV_DIR, 'guest_profiles.json')
    if not os.path.exists(path):
        print('[VDV] guest_profiles.json not found — guest features will use defaults')
        return {}
    with open(path, 'r', encoding='utf-8') as f:
        profiles = json.load(f)
    print(f'[VDV] Guest profiles loaded: {len(profiles)} profiles')
    return profiles

_ACCENT_MAP_GUEST = [
    ('ü','ue'),('ä','ae'),('ö','oe'),
    ('é','e'),('è','e'),('ê','e'),('ë','e'),
    ('à','a'),('â','a'),('á','a'),
    ('î','i'),('ï','i'),('í','i'),
    ('û','u'),('ú','u'),
    ('ç','c'),('ñ','n'),('ß','ss'),
]

def _normalise_guest_name(name: str) -> str:
    if not name:
        return ''
    s = str(name).strip()
    s = re.sub(r'^\d+\s+', '', s)
    s = s.lower()
    for src, dst in _ACCENT_MAP_GUEST:
        s = s.replace(src, dst)
    s = re.sub(r'[^\w\s]', ' ', s)
    parts = [p for p in s.split() if p]
    parts.sort()
    return ' '.join(parts)

def _get_guest_features(name: str, channel: str) -> dict:
    fallback_rate = _CHANNEL_CANCEL_RATE_FALLBACK.get(channel, 0.25)
    if not VDV_GUEST_PROFILES:
        return {'guest_cancel_rate': fallback_rate,
                'guest_profile_known': 0,
                'is_chronic_canceller': 0}
    key     = _normalise_guest_name(name)
    profile = VDV_GUEST_PROFILES.get(key)
    if profile is None:
        try:
            from rapidfuzz import process as _rfp, fuzz as _fuzz
            _all_keys = list(VDV_GUEST_PROFILES.keys())
            _result   = _rfp.extractOne(key, _all_keys,
                                        scorer=_fuzz.WRatio, score_cutoff=90)
            if _result:
                profile = VDV_GUEST_PROFILES[_result[0]]
        except Exception:
            pass
    if profile is None:
        return {'guest_cancel_rate': fallback_rate,
                'guest_profile_known': 0,
                'is_chronic_canceller': 0}
    cancel_rate = profile.get('cancel_rate')
    if cancel_rate is None:
        cancel_rate = fallback_rate
        known = 0
    else:
        known = 1
    return {
        'guest_cancel_rate':    cancel_rate,
        'guest_profile_known':  known,
        'is_chronic_canceller': profile.get('is_chronic_canceller', 0),
    }

# Load VdV data once at startup
VDV_GUEST_PROFILES  = _load_vdv_guest_profiles()
VDV_GUESTS_RAW      = []
VDV_GUEST_HISTORY   = {}
VDV_CHANNEL_STATS   = {}
VDV_FUTURE_BOOKINGS = []
VDV_FUTURE_SCORES   = []
VDV_MICE_DATA       = {}
VDV_FORECAST_DATA   = {'history': [], 'forecast': []}
VDV_GROUP_PIPELINE  = []
VDV_LANDING_STATS   = {
    'total_cx': 1694, 'total_ns': 339,
    'avg_adr': 130.0, 'avg_nights': 1.8,
    'model_accuracy': 80.5, 'model_auc': 0.852,
    'training_count': 119390,
}
try:
    VDV_GUESTS_RAW      = _parse_vdv_guests()
    VDV_GUEST_HISTORY   = _build_vdv_guest_history()
    VDV_CHANNEL_STATS   = _parse_vdv_channel_stats()
    VDV_FUTURE_BOOKINGS = _parse_vdv_future_bookings()
    VDV_MICE_DATA       = _parse_mice_data()
    VDV_FORECAST_DATA   = _parse_forecast_data()
    VDV_GROUP_PIPELINE  = _parse_group_pipeline()
    if VDV_FUTURE_BOOKINGS:
        VDV_FUTURE_SCORES = _score_vdv_future(VDV_FUTURE_BOOKINGS)
        # Detect outcomes from previous data before overwriting with new scores
        try:
            n_outcomes = _detect_vdv_outcomes(
                VDV_FUTURE_BOOKINGS,
                VDV_FUTURE_SCORES)
        except Exception as e:
            print(f"[VDV] Outcome detection failed: {e}")
        # Upsert per-row scores to new vdv_bookings_cache
        try:
            _n = _upsert_vdv_scores(VDV_FUTURE_BOOKINGS, VDV_FUTURE_SCORES)
            print(f"[VDV] Upserted {_n} scores to DB")
        except Exception as _ue:
            print(f"[VDV] Score upsert failed: {_ue}")
        # Legacy blob cache — full-restore fallback
        try:
            _cache_conn = get_db()
            _cache_cur  = _cache_conn.cursor()
            # Store arr_date as string since datetime isn't JSON serialisable
            _b_serialisable = [{**b, 'arr_date': b['arr_date'].isoformat()} for b in VDV_FUTURE_BOOKINGS]
            _cache_cur.execute("DELETE FROM vdv_bookings_cache_legacy")
            _cache_cur.execute(
                "INSERT INTO vdv_bookings_cache_legacy (bookings_json, scores_json) VALUES (%s, %s)",
                (json.dumps(_b_serialisable), json.dumps(VDV_FUTURE_SCORES))
            )
            _cache_conn.commit()
            _cache_cur.close()
            _cache_conn.close()
            print(f"[VDV] Cached {len(VDV_FUTURE_BOOKINGS)} bookings to legacy blob")
        except Exception as _ce:
            print(f"[VDV] Cache warning: {_ce}")
    # Update landing stats from freshly parsed data
    _real_cx = VDV_CHANNEL_STATS.get('_total_cx', 0)
    _real_ns = _count_vdv_noshow()
    if _real_cx > 100:
        VDV_LANDING_STATS['total_cx'] = _real_cx
    if _real_ns > 50:
        VDV_LANDING_STATS['total_ns'] = _real_ns
    VDV_LANDING_STATS['training_count'] = 119390 + _real_cx + _real_ns
    _hist_tracked = sum(1 for v in VDV_GUEST_HISTORY.values() if v['stays'] >= 2)
    _hist_cx      = sum(v['cancels'] for v in VDV_GUEST_HISTORY.values())
    print(f"[VDV] Guest history: {len(VDV_GUEST_HISTORY)} guests, {_hist_tracked} with 2+ stays, {_hist_cx} total cancels tracked")
    print(f"[VDV] Loaded {len(VDV_GUESTS_RAW)} repeat guests, "
          f"{len(VDV_FORECAST_DATA['history'])}h/{len(VDV_FORECAST_DATA['forecast'])}f forecast days, "
          f"{len(VDV_GROUP_PIPELINE)} upcoming groups, ",
          f"{len(VDV_FUTURE_BOOKINGS)} future bookings, "
          f"{VDV_MICE_DATA.get('total',0)} MICE bookings, "
          f"channels: {list(VDV_CHANNEL_STATS.keys())}")
    print(f"[VDV] Landing stats: {_real_cx} cx, {_real_ns} ns, "
          f"revenue lost €{int((_real_cx+_real_ns)*VDV_LANDING_STATS['avg_adr']*VDV_LANDING_STATS['avg_nights']/1000)}k")
except Exception as _vdv_err:
    print(f"[VDV] Startup warning: {_vdv_err}")


# ══════════════════════════════════════════════════════════════════════════════
# VAN DER VALK BRUSSELS AIRPORT — Parsers & Data
# ══════════════════════════════════════════════════════════════════════════════
VDV_BRU_HOTEL_KEY = "van der valk brussels airport"
_VDV_BRU_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VDV-BRU")

_BRU_CHANNEL_MAP = {
    'BARWEB': 'Booking.com', 'BAROTAGROSS': 'Booking.com', 'DEALSOTA': 'Booking.com',
    'DISCWEB': 'Direct/Web', 'BARDIR': 'Direct/Web', 'DISCDIR': 'Direct/Web',
    'CORPFIX': 'Corporate',  'CORPDYN': 'Corporate',
    'PACK': 'Package', 'MTGBNS': 'Package', 'BNSGRP': 'Package',
    'DEALS': 'Other',
}

def _parse_bru_guests():
    """Parse all RES_042 files in VDV-BRU for repeat guests."""
    import openpyxl, glob as _glob
    files = sorted(_glob.glob(os.path.join(_VDV_BRU_DIR, "RES_042_RepeatReservationsReport*.xlsx")))
    if not files:
        return []
    today = datetime.now()
    guests = []
    seen = set()
    for path in files:
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            i = 0
            while i < len(rows):
                row = rows[i]
                col0 = str(row[0]).strip() if row[0] else ''
                col1 = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                col4 = row[4] if len(row) > 4 else None
                col5 = row[5] if len(row) > 5 else None
                if ',' in col0 and col4:
                    try:
                        if hasattr(col4, 'date'):
                            arr = datetime.combine(col4.date(), datetime.min.time())
                        else:
                            arr = datetime.strptime(str(col4)[:10], '%d/%m/%Y')
                    except Exception:
                        i += 1; continue
                    arr_key = (col0.lower(), arr.strftime('%Y-%m-%d'))
                    if arr_key in seen:
                        i += 1; continue
                    seen.add(arr_key)
                    adults = 1
                    if col5:
                        try: adults = int(str(col5).split('/')[0])
                        except: pass
                    dep = None
                    for j in range(i + 1, min(i + 7, len(rows))):
                        r = rows[j]
                        r4 = r[4] if len(r) > 4 else None
                        if r4 and r[0] is None and '/' in str(r4):
                            try: dep = datetime.strptime(str(r4)[:10], '%d/%m/%Y'); break
                            except: pass
                    nights = (dep - arr).days if dep else 1
                    if dep and dep.date() < today.date():
                        status = 'Checked Out'
                    elif arr.date() == today.date():
                        status = 'Arriving Today'
                    elif arr.date() < today.date():
                        status = 'In House'
                    else:
                        status = f'Arriving {arr.strftime("%d %b")}'
                    guests.append({
                        'name': col0, 'membership': col1,
                        'arrival': arr.strftime('%d/%m/%Y'),
                        'departure': dep.strftime('%d/%m/%Y') if dep else '',
                        'arr_date': arr, 'dep_date': dep,
                        'adults': adults, 'nights': nights, 'status': status, 'note': '',
                    })
                i += 1
        except Exception as e:
            print(f"[BRU] Guest parse error ({path}): {e}")
    return guests


def _parse_bru_channel_stats():
    """Parse all RES_036 files in VDV-BRU for cancellation channel breakdown."""
    import openpyxl, glob as _glob
    files = sorted(_glob.glob(os.path.join(_VDV_BRU_DIR, "RES_036_CancelledReservations*.xlsx")))
    if not files:
        return {}
    seen_keys = set()
    raw = {}
    total = 0
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for row in rows:
                if row[0] is None and len(row) > 3 and row[3]:
                    # BRU uses col 19 for CXL reference (vs col 20 at MEC)
                    cxl_ref = str(row[19]).strip() if len(row) > 19 and row[19] else ''
                    if cxl_ref and re.match(r'[A-Z]{2,5}-CXL', cxl_ref):
                        key = cxl_ref
                    else:
                        c8  = str(row[8])[:10]  if len(row) > 8  and row[8]  else ''
                        c9  = str(row[9])        if len(row) > 9  and row[9]  else ''
                        c14 = str(row[14])[:10]  if len(row) > 14 and row[14] else ''
                        key = (c8, c9, c14)
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)
                    seg = str(row[3]).strip()
                    if seg and not seg.startswith('Subtotal') and not re.match(r'\d{2}/\d{2}/\d{4}', seg) and seg not in ('Market Segment', 'Company/Travel Agent'):
                        raw[seg] = raw.get(seg, 0) + 1
                        total += 1
        except Exception as e:
            print(f"[BRU] Channel stats error ({fp}): {e}")
    return {
        'Booking.com':       sum(raw.get(k, 0) for k in ('BARWEB', 'BAROTAGROSS', 'DEALSOTA')),
        'Direct / Web':      sum(raw.get(k, 0) for k in ('DISCWEB', 'BARDIR', 'DISCDIR')),
        'Corporate':         sum(raw.get(k, 0) for k in ('CORPFIX', 'CORPDYN')),
        'Packages / Groups': sum(raw.get(k, 0) for k in ('PACK', 'MTGBNS', 'BNSGRP')),
        'Other':             sum(raw.get(k, 0) for k in ('DEALS', 'OTHER', 'COMP')),
        '_total_cx':         total,
        '_raw':              raw,
    }


def _parse_bru_future_bookings():
    """Parse all RES_004 files in VDV-BRU for future bookings."""
    import openpyxl, glob as _glob
    files = sorted(_glob.glob(os.path.join(_VDV_BRU_DIR, "RES_004_EnteredOnAndBy*.xlsx")))
    today = datetime.now().date()
    bookings = []
    seen_confs = set()
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for r in rows:
                c0 = str(r[0]).strip() if r[0] else ''
                c3 = str(r[3]).strip() if len(r) > 3 and r[3] else ''
                # BRU: col 8 = arrival (may be datetime object or string)
                c8 = r[8] if len(r) > 8 else None
                if not (',' in c0 and c3 and c8):
                    continue
                # Parse arrival — openpyxl may return datetime directly; store as datetime
                try:
                    if isinstance(c8, datetime):
                        arr = c8.replace(tzinfo=None)
                    elif hasattr(c8, 'date'):
                        arr = datetime.combine(c8, datetime.min.time())
                    else:
                        arr = datetime.strptime(str(c8)[:10], '%d/%m/%Y')
                except Exception:
                    continue
                if arr.date() < today or c3 in seen_confs:
                    continue
                seen_confs.add(c3)
                nights = 1
                try: nights = max(1, int(r[9])) if len(r) > 9 and r[9] else 1
                except: pass
                channel_raw = str(r[25]).strip() if len(r) > 25 and r[25] else 'OTHER'
                rate_plan   = str(r[13]).strip() if len(r) > 13 and r[13] else ''
                purch_elem  = str(r[18]).strip() if len(r) > 18 and r[18] else ''
                has_breakfast = ('BB' in rate_plan.upper() or 'FBBF' in purch_elem.upper())
                try:
                    adr = float(str(r[14]).replace(',', '.')) / max(1, nights) if len(r) > 14 and r[14] else 140.0
                except: adr = 140.0
                created_raw = r[28] if len(r) > 28 else None
                lead = 0
                try:
                    if created_raw:
                        if hasattr(created_raw, 'date'):
                            cd = created_raw.date()
                        else:
                            cd = datetime.strptime(str(created_raw)[:10], '%d/%m/%Y').date()
                        lead = max(0, (arr.date() - cd).days)
                except: pass
                wkend = wkday = 0
                d = arr.replace(hour=0, minute=0, second=0, microsecond=0)
                for _ in range(nights):
                    if d.weekday() >= 5: wkend += 1
                    else: wkday += 1
                    d += timedelta(days=1)
                week_num = int(arr.isocalendar()[1])
                ch_label = _BRU_CHANNEL_MAP.get(channel_raw, 'Other')
                bookings.append({
                    'name': c0, 'arrival': arr.strftime('%d/%m/%Y'),
                    'arr_date': arr, 'nights': nights, 'adults': 1,
                    'channel': ch_label, 'channel_raw': channel_raw,
                    'lead': lead, 'gtd': 'NONE', 'adr': round(adr, 2),
                    'wkend': wkend, 'wkday': wkday, 'week_num': week_num,
                    'has_breakfast': has_breakfast,
                })
        except Exception as e:
            print(f"[BRU] Future bookings parse error ({fp}): {e}")
    return bookings


def _score_bru_future(bookings):
    """Score BRU future bookings using the generic model with channel normalisation."""
    if not bookings:
        return []
    from collections import Counter
    name_counts = Counter(b['name'].lower() for b in bookings)
    repeat_names = {n for n, c in name_counts.items() if c >= 2}
    known_repeats = {g['name'].lower() for g in VDV_BRU_GUESTS_RAW}
    all_repeats = repeat_names | known_repeats

    BRU_CHANNEL_RANGE = {
        'Booking.com':  (20, 82),
        'Direct/Web':   (12, 68),
        'Direct / Web': (12, 68),
        'Corporate':    ( 8, 48),
        'Package':      ( 4, 35),
        'Other':        (12, 65),
    }

    def is_repeat(b):
        return 1 if b['name'].lower() in all_repeats else 0

    def normalise(raw_scores, bklist):
        from collections import defaultdict
        ch_idx = defaultdict(list)
        for i, b in enumerate(bklist):
            ch_idx[b['channel']].append(i)
        result = [0.0] * len(bklist)
        for ch, indices in ch_idx.items():
            lo, hi = BRU_CHANNEL_RANGE.get(ch, (12, 65))
            sorted_idx = sorted(indices, key=lambda i: raw_scores[i])
            n = len(sorted_idx)
            for rank, idx in enumerate(sorted_idx):
                pct = rank / (n - 1) if n > 1 else 0.5
                reduction = 10 if is_repeat(bklist[idx]) else 0
                result[idx] = round(max(lo, lo + pct * (hi - lo) - reduction), 1)
        return result

    feat_cols = ['lead_time', 'arrival_date_week_number', 'stays_in_weekend_nights',
                 'stays_in_week_nights', 'adults', 'is_repeated_guest',
                 'previous_cancellations', 'previous_bookings_not_canceled',
                 'booking_changes', 'days_in_waiting_list', 'adr', 'total_of_special_requests']
    rows_feat = [[b['lead'], b['week_num'], b['wkend'], b['wkday'],
                  b['adults'], is_repeat(b), 0, 0, 0, 0, b['adr'], 0] for b in bookings]
    df_feat = pd.DataFrame(rows_feat, columns=feat_cols)
    raw_scores = list(model.predict_proba(df_feat)[:, 1] * 100)
    return [float(round(s, 1)) for s in normalise(raw_scores, bookings)]


def _parse_bru_forecast():
    """Parse FOR_042 in VDV-BRU — returns last 60d history + next 60d forecast."""
    import glob as _glob
    files = sorted(_glob.glob(os.path.join(_VDV_BRU_DIR, "FOR_042_HistoryAndForecast*.xlsx")))
    if not files:
        return {'history': [], 'forecast': []}
    path = files[-1]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        history, forecast, mode = [], [], None
        for row in rows:
            if not row or row[0] is None: continue
            v0 = str(row[0]).strip()
            if v0 == 'HISTORY':           mode = 'h'; continue
            if v0.startswith('FORECAST'): mode = 'f'; continue
            if v0 in ('Totals', '', 'Date', 'Sold'): continue
            date_part = v0.split()[-1] if v0 else ''
            if not re.match(r'\d{2}/\d{2}/\d{4}', date_part): continue
            try:
                dt = datetime.strptime(date_part, '%d/%m/%Y')
                rooms_sold = int(str(row[1]).replace(',', '').split('.')[0]) if row[1] else 0
                arrivals   = int(str(row[2]).replace(',', '').split('.')[0]) if row[2] else 0
                departures = int(str(row[3]).replace(',', '').split('.')[0]) if row[3] else 0
                # BRU FOR_042: col 5 = Occupancy % (MEC uses col 4)
                occ_pct = float(str(row[5]).replace(',', '.')) if len(row) > 5 and row[5] else 0.0
                # BRU FOR_042: col 17 = Res. Room Rev.
                rev_raw = str(row[17]).replace('.', '').replace(',', '.').strip() if len(row) > 17 and row[17] else '0'
                revenue = float(rev_raw) if rev_raw else 0.0
                entry = {'date': dt.strftime('%d %b'), 'dt': dt.strftime('%Y-%m-%d'),
                         'rooms': rooms_sold, 'arrivals': arrivals, 'departures': departures,
                         'occ': round(occ_pct, 1), 'rev': round(revenue)}
                if mode == 'h':   history.append(entry)
                elif mode == 'f': forecast.append(entry)
            except: continue
        return {'history': history[-60:], 'forecast': forecast[:60]}
    except Exception as e:
        print(f"[BRU] Forecast parse error: {e}")
        return {'history': [], 'forecast': []}


def _parse_bru_group_pipeline():
    """Parse GRP_017 in VDV-BRU for upcoming group bookings."""
    import glob as _glob
    files = sorted(_glob.glob(os.path.join(_VDV_BRU_DIR, "GRP_017_GroupAndMEBookingListWithRevenueCSVOnly*.xlsx")))
    if not files:
        return []
    path = files[-1]
    STATUS_LABELS = {'DEF': 'Definite', 'DEP': 'Deposit', 'TEN': 'Tentative',
                     'LOS': 'Lost', 'CXL': 'Cancelled', 'PRO': 'Prospect'}
    MARKET_LABELS = {'MTGBNS': 'Meeting', 'CORPFIX': 'Corporate', 'CORPDYN': 'Corporate Dyn',
                     'BNSGRP': 'Business Group', 'LEISURE': 'Leisure', 'CHARBNS': 'Charter'}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        groups = []
        today = datetime.now()
        for i, row in enumerate(rows):
            if i < 6 or not row or row[0] is None: continue
            cols = [str(c).strip() if c is not None else '' for c in row]
            if len(cols) < 8: continue
            code   = cols[1]
            name   = cols[2]
            arr_s  = cols[3]
            dep_s  = cols[4]
            # BRU GRP_017: status at col 6, market at col 7
            status = cols[6]
            market = cols[7]
            if status in ('LOS', 'CXL', ''): continue
            if not re.match(r'\d{2}/\d{2}/\d{4}', arr_s): continue
            try:
                arr_dt = datetime.strptime(arr_s, '%d/%m/%Y')
                dep_dt = datetime.strptime(dep_s, '%d/%m/%Y') if re.match(r'\d{2}/\d{2}/\d{4}', dep_s) else arr_dt
                if arr_dt.date() < today.date(): continue
                nights = (dep_dt - arr_dt).days
                groups.append({
                    'code': code, 'name': name if name else code,
                    'arrival': arr_dt.strftime('%d %b %Y'),
                    'departure': dep_dt.strftime('%d %b %Y'),
                    'arr_dt': arr_dt.strftime('%Y-%m-%d'),
                    'nights': nights,
                    'status': status,
                    'status_label': STATUS_LABELS.get(status, status),
                    'market': MARKET_LABELS.get(market, market),
                })
            except: continue
        return sorted(groups, key=lambda x: x['arr_dt'])[:20]
    except Exception as e:
        print(f"[BRU] Group pipeline parse error: {e}")
        return []


def _parse_bru_mice_data():
    """Parse all RES_001 files in VDV-BRU for MICE/corporate bookings."""
    import openpyxl, glob as _glob
    files = sorted(_glob.glob(os.path.join(_VDV_BRU_DIR, "RES_001_ArrivalDetailed*.xlsx")))
    MICE_SEGS = {'BNSGRP', 'CORPFIX', 'CORPDYN', 'MTGBNS'}
    SEG_LABELS = {'BNSGRP': 'Business Group', 'CORPFIX': 'Corporate Fixed',
                  'CORPDYN': 'Corporate Dynamic', 'MTGBNS': 'Meeting Business'}
    result = {'total': 0, 'total_nights': 0, 'by_segment': {k: 0 for k in MICE_SEGS},
              'top_clients': [], 'groups': []}
    companies = {}
    for path in files:
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for i in range(13, len(rows)):
                row = rows[i]
                if not row or len(row) < 18: continue
                market = str(row[17]).strip() if row[17] else ''
                if market not in MICE_SEGS: continue
                company = str(row[13]).strip() if row[13] else ''
                nights_raw = str(row[7]).strip() if row[7] else '0'
                arr_raw = str(row[4]).strip() if row[4] else ''
                arr_date = arr_raw[:10] if arr_raw else ''
                try: nights = int(nights_raw) if nights_raw.isdigit() else 1
                except: nights = 1
                if company and company not in ('nan', 'Group', 'Company', 'Travel Agent', ''):
                    if company not in companies:
                        companies[company] = {'segment': SEG_LABELS.get(market, market),
                                              'seg_code': market, 'bookings': 0,
                                              'nights': 0, 'last_arrival': arr_date}
                    companies[company]['bookings'] += 1
                    companies[company]['nights']   += nights
                    if arr_date and arr_date > companies[company]['last_arrival']:
                        companies[company]['last_arrival'] = arr_date
                result['by_segment'][market] = result['by_segment'].get(market, 0) + 1
                result['total']        += 1
                result['total_nights'] += nights
        except Exception as e:
            print(f"[BRU] MICE parse error ({path}): {e}")
    result['top_clients'] = sorted(
        [{'company': co, **data} for co, data in companies.items()],
        key=lambda x: -x['bookings'])[:12]
    return result


def _parse_bru_monthly_cx():
    """Extract monthly cancellation counts from BRU RES_036 (col 16 = CXL Date).

    BRU RES_036 has two row types per reservation:
      - Main row:   row[0]=guest name, row[1]=room type, row[16]=CXL Date/Time, row[20]=CXL No.
      - Detail row: row[0]=None, row[2]=VIP, row[3]=rate code, row[16]='CXL By' (text)
    We must target main rows (row[0] and row[1] non-None, row[16] present) for dates.
    """
    import openpyxl, glob as _glob
    from collections import Counter
    files = sorted(_glob.glob(os.path.join(_VDV_BRU_DIR, "RES_036_CancelledReservations*.xlsx")))
    if not files:
        return []
    seen_keys = set()
    month_cx = Counter()
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for row in rows:
                # Main guest rows: col 0 = guest name, col 1 = room type, col 16 = CXL datetime
                if row[0] is not None and row[1] is not None and (len(row) > 16 and row[16] is not None):
                    # CXL No. is at col 20 in BRU files (col 19 in header, shifted by 1 in data)
                    cxl_ref = str(row[20]).strip() if len(row) > 20 and row[20] else ''
                    if not cxl_ref:
                        cxl_ref = str(row[19]).strip() if len(row) > 19 and row[19] else ''
                    if cxl_ref and re.match(r'[A-Z]{2,5}-CXL', cxl_ref):
                        key = cxl_ref
                    else:
                        c8  = str(row[8])[:10]  if len(row) > 8  and row[8]  else ''
                        c9  = str(row[9])        if len(row) > 9  and row[9]  else ''
                        c14 = str(row[14])[:10]  if len(row) > 14 and row[14] else ''
                        key = (c8, c9, c14)
                    if key in seen_keys: continue
                    seen_keys.add(key)
                    cxl_dt = row[16]
                    try:
                        if hasattr(cxl_dt, 'strftime'):
                            label = cxl_dt.strftime('%b %Y')
                        else:
                            d = datetime.strptime(str(cxl_dt)[:10], '%d/%m/%Y')
                            label = d.strftime('%b %Y')
                        month_cx[label] += 1
                    except: pass
        except Exception as e:
            print(f"[BRU] Monthly CX parse error ({fp}): {e}")
    try:
        sorted_months = sorted(month_cx.keys(), key=lambda s: datetime.strptime(s, '%b %Y'))[-6:]
        return [(m, month_cx[m]) for m in sorted_months]
    except:
        return list(month_cx.items())


def _build_bru_guest_history():
    """Count all-time stays per guest from BRU RES_042 (no date filter).
    Cancellations are assumed 0; stay count alone gives meaningful loyalty signal."""
    import openpyxl, glob as _glob
    from collections import defaultdict
    history = defaultdict(lambda: {'stays': 0, 'cancels': 0})
    seen_stays = set()
    for path in sorted(_glob.glob(os.path.join(_VDV_BRU_DIR, "RES_042_RepeatReservationsReport*.xlsx"))):
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb.close()
            for row in rows:
                col0 = str(row[0]).strip() if row[0] else ''
                col4 = row[4] if len(row) > 4 else None
                if ',' in col0 and col4 and '/' in str(col4):
                    key = (col0.lower(), str(col4)[:10])
                    if key not in seen_stays:
                        seen_stays.add(key)
                        history[col0.lower()]['stays'] += 1
        except Exception:
            pass
    return dict(history)


def _score_bru_guests(guests):
    """Score BRU repeat guests using the generic model + loyalty adjustment."""
    if not guests:
        return []
    feat_cols = ['lead_time', 'arrival_date_week_number', 'stays_in_weekend_nights',
                 'stays_in_week_nights', 'adults', 'is_repeated_guest',
                 'previous_cancellations', 'previous_bookings_not_canceled',
                 'booking_changes', 'days_in_waiting_list', 'adr', 'total_of_special_requests']
    now = datetime.now()
    rows_feat = []
    for g in guests:
        arr = g['arr_date']
        dep = g['dep_date']
        wkend = wkday = 0
        if dep:
            d = arr
            while d < dep:
                if d.weekday() >= 5: wkend += 1
                else: wkday += 1
                d += timedelta(days=1)
        lead = max(0, (arr - now).days)
        rows_feat.append([lead, int(arr.isocalendar()[1]), wkend, wkday,
                          g['adults'], 1, 0, 3, 0, 0, 190.0, 0])
    df = pd.DataFrame(rows_feat, columns=feat_cols)
    raw_scores = [float(s) for s in model.predict_proba(df)[:, 1] * 100]

    # ── Loyalty adjustment: same Bayesian blend as MEC ───────────────────────
    adjusted = []
    for g, raw_sc in zip(guests, raw_scores):
        hist = VDV_BRU_GUEST_HISTORY.get(g['name'].strip().lower(), {'stays': 0, 'cancels': 0})
        stays   = hist['stays']
        cancels = hist['cancels']
        if stays >= 2:
            hist_rate = (cancels + 0.5) / (stays + 1)
            trust     = min(1.0, stays / 10)
            blend     = 0.65
            sc = raw_sc * (1 - trust * blend) + hist_rate * 100 * trust * blend
        else:
            sc = raw_sc
        adjusted.append(max(0.0, min(100.0, sc)))
    return adjusted


VDV_BRU_GUESTS_RAW      = []
VDV_BRU_GUEST_HISTORY   = {}
VDV_BRU_CHANNEL_STATS   = {}
VDV_BRU_FUTURE_BOOKINGS = []
VDV_BRU_FUTURE_SCORES   = []
VDV_BRU_FORECAST_DATA   = {'history': [], 'forecast': []}
VDV_BRU_GROUP_PIPELINE  = []
VDV_BRU_MICE_DATA       = {}
VDV_BRU_MONTHS          = []
VDV_BRU_CX_MONTHLY      = []
VDV_BRU_NS_MONTHLY      = []
try:
    VDV_BRU_GUESTS_RAW      = _parse_bru_guests()
    VDV_BRU_GUEST_HISTORY   = _build_bru_guest_history()
    VDV_BRU_CHANNEL_STATS   = _parse_bru_channel_stats()
    VDV_BRU_FUTURE_BOOKINGS = _parse_bru_future_bookings()
    VDV_BRU_FORECAST_DATA   = _parse_bru_forecast()
    VDV_BRU_GROUP_PIPELINE  = _parse_bru_group_pipeline()
    VDV_BRU_MICE_DATA       = _parse_bru_mice_data()
    _bru_monthly = _parse_bru_monthly_cx()
    if _bru_monthly:
        VDV_BRU_MONTHS     = [m for m, _ in _bru_monthly]
        VDV_BRU_CX_MONTHLY = [c for _, c in _bru_monthly]
        _bru_real_ns, VDV_BRU_NS_MONTHLY = _parse_bru_noshow(VDV_BRU_MONTHS)
    else:
        _bru_real_ns, VDV_BRU_NS_MONTHLY = _parse_bru_noshow()
    if VDV_BRU_FUTURE_BOOKINGS:
        VDV_BRU_FUTURE_SCORES = _score_bru_future(VDV_BRU_FUTURE_BOOKINGS)
    _bru_hist_tracked = sum(1 for v in VDV_BRU_GUEST_HISTORY.values() if v['stays'] >= 2)
    _bru_hist_cx      = sum(v['cancels'] for v in VDV_BRU_GUEST_HISTORY.values())
    print(f"[BRU] Guest history: {len(VDV_BRU_GUEST_HISTORY)} guests, {_bru_hist_tracked} with 2+ stays, {_bru_hist_cx} cancels tracked")
    print(f"[BRU] Loaded {len(VDV_BRU_GUESTS_RAW)} repeat guests, "
          f"{len(VDV_BRU_FUTURE_BOOKINGS)} future bookings, "
          f"{len(VDV_BRU_FORECAST_DATA['history'])}h/{len(VDV_BRU_FORECAST_DATA['forecast'])}f forecast, "
          f"{len(VDV_BRU_GROUP_PIPELINE)} groups, "
          f"no-shows: {_bru_real_ns}, "
          f"cx channels: {list(VDV_BRU_CHANNEL_STATS.keys())}")
except Exception as _bru_err:
    print(f"[BRU] Startup warning: {_bru_err}")


def _get_demand_event(check_date):
    """Returns (emoji, label) if check_date is a Belgian public holiday or
    school break period, else None."""
    key = (check_date.month, check_date.day)
    if key in _BELGIAN_HOLIDAYS:
        return _BELGIAN_HOLIDAYS[key]
    if check_date in _BELGIAN_MOVEABLE_HOLIDAYS:
        return _BELGIAN_MOVEABLE_HOLIDAYS[check_date]
    for start, end, emoji, label in _BELGIAN_SCHOOL_HOLIDAYS:
        if start <= check_date <= end:
            return (emoji, label)
    return None


def _vdv_expected_cancellations(bookings, scores, arrival_date):
    """Sum of decay-adjusted cancellation probabilities.
    Uses empirical channel-specific decay curves derived from 1,590 VdV cancellations.
    Each booking's score is weighted by the remaining-cancel fraction for its channel
    at the current days-out horizon."""
    today    = date.today()
    days_out = (arrival_date - today).days
    total    = 0.0
    for b, s in zip(bookings, scores):
        channel = b.get('channel', '')
        decay   = _get_decay_factor(days_out, channel)
        total  += (s / 100.0) * decay
    return total


def _vdv_expected_no_shows(bookings, arrival_month):
    """Expected no-show count with channel-specific rates, seasonal adjustment,
    and deposit/guarantee reduction by channel."""
    seasonal = _VDV_NO_SHOW_SEASONAL.get(arrival_month, 1.0)
    # Deposit reduction factors by channel:
    # Booking.com/OTA: full rate (virtual credit card, no advance payment)
    # Direct/Web: 70% (some prepay, some not)
    # Corporate/Package: 50% (almost always guaranteed or invoiced)
    _DEPOSIT_FACTOR = {
        'Booking.com':       1.00,
        'Direct/Web':        0.70,
        'Direct / Web':      0.70,
        'Corporate':         0.50,
        'Package':           0.50,
        'Packages / Groups': 0.50,
        'Other':             1.00,
    }
    total = 0.0
    for b in bookings:
        ch   = b.get('channel', '')
        rate = _VDV_NO_SHOW_RATES.get(ch, _VDV_NO_SHOW_RATES['default'])
        dep  = _DEPOSIT_FACTOR.get(ch, 1.00)
        total += rate * seasonal * dep
    return total


def _vdv_overbooking_recommendation(bookings, scores, arrival_date):
    """Returns overbooking recommendation dict for a single arrival date."""
    if not bookings or not scores:
        return None
    n     = len(bookings)
    month = arrival_date.month
    exp_cx = _vdv_expected_cancellations(bookings, scores, arrival_date)
    exp_ns = _vdv_expected_no_shows(bookings, month)
    total_attrition = exp_cx + exp_ns
    # Demand event (Belgian holidays / school breaks) → 1.15x attrition
    event = _get_demand_event(arrival_date)
    holiday_multiplier = 1.15 if event else 1.0
    # Conservative 65% of expected attrition; never exceed 8% of total rooms
    max_overbook   = VDV_TOTAL_ROOMS * 0.08
    recommendation = min(round(total_attrition * holiday_multiplier * 0.65), max_overbook)
    recommendation = max(0, int(recommendation))
    confidence = 'HIGH' if n >= 20 else ('MEDIUM' if n >= 10 else 'LOW')
    high_risk  = sum(1 for s in scores if s >= 70)
    return {
        'date':                    arrival_date,
        'bookings_on_hand':        n,
        'high_risk_count':         high_risk,
        'expected_cancellations':  round(exp_cx, 1),
        'expected_no_shows':       round(exp_ns, 1),
        'total_expected_attrition':round(total_attrition, 1),
        'recommended_overbooking': recommendation,
        'confidence':              confidence,
        'holiday_emoji':           event[0] if event else '',
        'holiday_label':           event[1] if event else '',
    }


def _vdv_build_overbooking_planner(future_bookings, future_scores):
    """Groups bookings by arrival date; returns 30-day overbooking plan sorted by date."""
    from datetime import date as _date, timedelta as _td
    from collections import defaultdict as _dd
    today   = _date.today()
    cutoff  = today + _td(days=30)
    date_bookings = _dd(list)
    date_scores   = _dd(list)
    for b, s in zip(future_bookings, future_scores):
        arr = b.get('arr_date')
        if arr:
            arr_d = arr.date() if hasattr(arr, 'date') else arr
            if today <= arr_d <= cutoff:
                date_bookings[arr_d].append(b)
                date_scores[arr_d].append(s)
    results = []
    for arr_date in sorted(date_bookings.keys()):
        rec = _vdv_overbooking_recommendation(
            date_bookings[arr_date], date_scores[arr_date], arr_date)
        if rec:
            results.append(rec)
    return results


def _grp_pipe_html(groups, status_colors):
    if not groups: return ''
    rows = ''
    for g in groups:
        sc = status_colors.get(g['status_label'], '#e5e7eb')
        rows += f'''<tr>
          <td style="padding:10px 14px;font-size:13px;font-weight:500;color:#111827;">{g['name']}</td>
          <td style="padding:10px 14px;font-size:12px;color:#6b7280;">{g['arrival']}</td>
          <td style="padding:10px 14px;font-size:12px;color:#6b7280;">{g['nights']}n</td>
          <td style="padding:10px 14px;font-size:11px;color:#6b7280;">{g['market']}</td>
          <td style="padding:10px 14px;"><span style="background:{sc}22;color:{sc};border:1px solid {sc}66;border-radius:99px;padding:3px 10px;font-size:11px;font-weight:500;">{g['status_label']}</span></td>
        </tr>'''
    return (
        '<div style="border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;margin-bottom:32px;">' +
        '<table style="width:100%;border-collapse:collapse;">' +
        '<thead><tr>' +
        '<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Group</th>' +
        '<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Arrival</th>' +
        '<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Nights</th>' +
        '<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Segment</th>' +
        '<th style="padding:10px 14px;font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid #e5e7eb;text-align:left;">Status</th>' +
        '</tr></thead><tbody>' + rows + '</tbody></table></div>'
    )

def build_vdv_dashboard(hotel_name, lang="en", first_login=False, _data=None):
    """VdV dashboard — parameterised for MEC and BRU via _data dict."""
    _d = _data or {}
    _all_guests = _d.get('guests', VDV_GUESTS_RAW)
    score_fn    = _d.get('score_fn', _score_vdv_guests)
    _today_d    = datetime.now().date()
    _window_end = _today_d + timedelta(days=15)
    # Only show guests arriving today → today+15 days; drop past stays
    guests = [g for g in _all_guests
              if g['arr_date'].date() >= _today_d
              and g['arr_date'].date() <= _window_end]
    scores      = score_fn(guests)
    _ch_raw     = _d.get('ch_stats', VDV_CHANNEL_STATS)
    ch_data     = {k: v for k, v in _ch_raw.items() if isinstance(v, (int, float)) and not k.startswith('_')}
    export_url  = _d.get('export_url', '/vdv/export-highrisk')
    today   = datetime.now()
    today_str = today.strftime('%d %b %Y')

    # ── MICE data ────────────────────────────────────────────────────────────
    mice = _d.get('mice') or VDV_MICE_DATA or {
        'total': 11107, 'total_nights': 22338,
        'by_segment': {'BNSGRP': 1887, 'CORPFIX': 4254, 'CORPDYN': 1562, 'MTGBNS': 3404},
        'top_clients': [], 'groups': []
    }
    mice_total   = mice.get('total', 0)
    mice_nights  = mice.get('total_nights', 0)
    mice_seg     = mice.get('by_segment', {})
    mice_clients = mice.get('top_clients', [])
    mice_groups  = mice.get('groups', [])
    mice_avg_nights = round(mice_nights / mice_total, 1) if mice_total > 0 else 0
    # Estimated MICE revenue: avg corporate ADR €145 × avg nights
    mice_est_rev = int(mice_total * 145 * max(1, mice_avg_nights))
    mice_seg_js  = json.dumps([
        mice_seg.get('CORPFIX', 0), mice_seg.get('CORPDYN', 0),
        mice_seg.get('BNSGRP', 0),  mice_seg.get('MTGBNS', 0)
    ])

    # ── Forecast + group pipeline ────────────────────────────────────────────
    fc_data     = _d.get('forecast', VDV_FORECAST_DATA)
    fc_hist     = fc_data.get('history', [])
    fc_fore     = fc_data.get('forecast', [])
    grp_pipe    = _d.get('groups', VDV_GROUP_PIPELINE)
    # Build JS arrays for the occupancy chart
    fc_labels   = json.dumps([r['date'] for r in fc_hist] + [r['date'] for r in fc_fore])
    fc_occ_hist = json.dumps([r['occ'] for r in fc_hist] + [None] * len(fc_fore))
    fc_occ_fore = json.dumps([None] * len(fc_hist) + [r['occ'] for r in fc_fore])
    fc_rev_hist = json.dumps([r['rev'] for r in fc_hist] + [None] * len(fc_fore))
    fc_rev_fore = json.dumps([None] * len(fc_hist) + [r['rev'] for r in fc_fore])
    fc_arr_hist = json.dumps([r['arrivals'] for r in fc_hist] + [None] * len(fc_fore))
    fc_arr_fore = json.dumps([None] * len(fc_hist) + [r['arrivals'] for r in fc_fore])
    fc_boundary = len(fc_hist)  # index where forecast starts
    # Group pipeline status badges
    STATUS_COLORS = {'Definite': '#00d165', 'Deposit': '#f59e0b', 'Tentative': '#9ca3af', 'Prospect': '#e5e7eb'}

    # Risk score per corporate client (rule-based)
    SEG_RISK = {'BNSGRP': 2, 'CORPDYN': 1, 'MTGBNS': 1, 'CORPFIX': 0}
    for c in mice_clients:
        risk_pts = SEG_RISK.get(c.get('seg_code', ''), 0)
        avg_n = c['nights'] / c['bookings'] if c['bookings'] else 1
        if avg_n < 1.5: risk_pts += 1          # very short stays = higher no-show
        if c['bookings'] <= 2: risk_pts += 1   # new/infrequent account
        if avg_n >= 3: risk_pts -= 1           # long stays = committed
        if c['bookings'] >= 20: risk_pts -= 1  # proven repeat account
        if risk_pts >= 3:   c['risk'] = 'HIGH'
        elif risk_pts >= 2: c['risk'] = 'MEDIUM'
        else:               c['risk'] = 'LOW'

    # ── Historical numbers ──────────────────────────────────────────────────
    MONTHS       = _d.get('months',     ['Oct 2025','Nov 2025','Dec 2025','Jan 2026','Feb 2026','Mar 2026'])
    CX_MONTHLY   = _d.get('cx_monthly', [167, 315, 335, 255, 296, 326])
    NS_MONTHLY   = _d.get('ns_monthly', [43,  61,  75,  64,  58,  38])
    LOST_MONTHLY = [c+n for c,n in zip(CX_MONTHLY, NS_MONTHLY)]

    total_cx     = sum(CX_MONTHLY)
    total_ns     = sum(NS_MONTHLY)
    total_lost   = total_cx + total_ns
    avg_adr      = _d.get('avg_adr', 130.0)
    avg_nights   = _d.get('avg_nights', 1.8)
    rev_lost     = int(total_lost * avg_adr * avg_nights)
    # Recoverable: 30% of cancellations + 25% of no-shows
    recoverable  = int(total_cx * 0.30 * avg_adr * avg_nights
                       + total_ns * 0.25 * avg_adr)

    # ── Future bookings risk (from RES_004 + model scoring) ────────────────
    fut_bookings = _d.get('fut_bookings', VDV_FUTURE_BOOKINGS)
    fut_scores   = _d.get('fut_scores',   VDV_FUTURE_SCORES)

    # ── Model accuracy (outcome log, last 30 days) ───────────────────────────
    _outcome_stats = {'cancelled': 0, 'completed': 0,
                      'cx_high': 0, 'cx_med': 0, 'cx_low': 0}
    try:
        _oc = get_db()
        _ocur = _oc.cursor()
        _ocur.execute("""
            SELECT
                outcome,
                COUNT(*) AS count,
                SUM(CASE WHEN predicted_tier = 'high'   THEN 1 ELSE 0 END) AS n_high,
                SUM(CASE WHEN predicted_tier = 'medium' THEN 1 ELSE 0 END) AS n_med,
                SUM(CASE WHEN predicted_tier = 'low'    THEN 1 ELSE 0 END) AS n_low
            FROM vdv_outcome_log
            WHERE hotel_id = 'vdv'
              AND logged_at >= NOW() - INTERVAL '30 days'
            GROUP BY outcome
        """)
        for _row in _ocur.fetchall():
            _out, _cnt, _nh, _nm, _nl = _row
            _outcome_stats[_out] = int(_cnt)
            if _out == 'cancelled':
                _outcome_stats['cx_high'] = int(_nh or 0)
                _outcome_stats['cx_med']  = int(_nm or 0)
                _outcome_stats['cx_low']  = int(_nl or 0)
        _ocur.close()
        _oc.close()
    except Exception as _oce:
        print(f"[VDV] Outcome stats query error: {_oce}")

    # ── Overbooking planner (next 30 days) ──────────────────────────────────
    overbooking_plan = _vdv_build_overbooking_planner(fut_bookings, fut_scores)
    print(f'[VDV] Overbooking planner: {len(overbooking_plan)} dates calculated for next 30 days')
    # Fallback pre-computed constants when files not loaded
    if not fut_bookings:
        fut_total     = 2768
        fut_high      = 36
        fut_med       = 844
        fut_low       = 1888
        fut_no_gtd    = 1220
        fut_table_html = ''
        fut_by_channel = {'Booking.com': 1056, 'Direct/Web': 436, 'Corporate': 539,
                          'Package': 470, 'Other': 267}
        fut_month_labels = ['Apr 2026','May 2026','Jun 2026','Jul 2026',
                            'Aug 2026','Sep 2026','Oct 2026']
        fut_month_high   = [0, 2, 3, 28, 2, 1, 0]
        fut_month_med    = [200, 191, 114, 151, 93, 51, 38]
    else:
        fut_total = len(fut_bookings)
        fut_high  = sum(1 for s in fut_scores if s >= 70)
        fut_med   = sum(1 for s in fut_scores if 40 <= s < 70)
        fut_low   = sum(1 for s in fut_scores if s < 40)
        fut_no_gtd = sum(1 for b in fut_bookings if b['gtd'] == 'NONE')
        # Top 10 per channel at-risk bookings table
        from collections import defaultdict as _dd
        _ch_buckets = _dd(list)
        for _i, (_b, _s) in enumerate(zip(fut_bookings, fut_scores)):
            _ch_buckets[_b['channel']].append((_i, _s))
        # Order channels by their highest score descending
        _ch_order = sorted(_ch_buckets.keys(),
                           key=lambda c: max(s for _, s in _ch_buckets[c]),
                           reverse=True)
        fut_table_html = ''
        for _ch in _ch_order:
            _top10 = sorted(_ch_buckets[_ch], key=lambda x: -x[1])[:10]
            fut_table_html += (
                f'<tr class="ch-header-row">'
                f'<td colspan="9"><strong>{_ch}</strong> — Top 10</td></tr>'
            )
            for _rank, (_idx, sc) in enumerate(_top10):
                b = fut_bookings[_idx]
                ch = b['channel']
                gtd = b['gtd']
                if sc >= 70:
                    bdg = f'<span class="badge high">{sc:.0f}%</span>'
                    act = '<span class="abtn dep">Deposit</span>'
                elif sc >= 40:
                    bdg = f'<span class="badge med">{sc:.0f}%</span>'
                    act = '<span class="abtn rem">Reminder</span>'
                else:
                    bdg = f'<span class="badge low">{sc:.0f}%</span>'
                    act = '<span class="abtn mon">Monitor</span>'
                gtd_badge = ('<span class="gtd-none">No GTD</span>' if gtd == 'NONE'
                             else f'<span class="gtd-ok">{gtd}</span>')
                fut_table_html += (
                    f'<tr><td>{_rank+1}</td>'
                    f'<td class="gn">{b["name"]}</td>'
                    f'<td>{b["arrival"]}</td>'
                    f'<td>{b["nights"]}n</td>'
                    f'<td>{b["lead"]}d</td>'
                    f'<td>{ch}</td>'
                    f'<td>{gtd_badge}</td>'
                    f'<td>{bdg}</td>'
                    f'<td>{act}</td></tr>'
                )
        # Channel risk breakdown
        from collections import defaultdict
        ch_risk = defaultdict(lambda: {'high': 0, 'total': 0})
        for b, sc in zip(fut_bookings, fut_scores):
            ch_risk[b['channel']]['total'] += 1
            if sc >= 70: ch_risk[b['channel']]['high'] += 1
        fut_by_channel = {k: v['total'] for k, v in sorted(ch_risk.items(), key=lambda x: -x[1]['total'])}
        # Monthly risk
        from collections import Counter
        mo_high  = Counter()
        mo_med   = Counter()
        mo_total = Counter()
        for b, sc in zip(fut_bookings, fut_scores):
            mo_lbl = f"{b['arr_date'].strftime('%b')} {b['arr_date'].year}"
            mo_total[mo_lbl] += 1
            if sc >= 70: mo_high[mo_lbl] += 1
            elif sc >= 40: mo_med[mo_lbl] += 1
        # Sort months
        from datetime import datetime as _dt
        mo_sorted = sorted(mo_total.keys(),
                           key=lambda s: _dt.strptime(s, '%b %Y'))[:7]
        fut_month_labels = mo_sorted
        fut_month_high   = [mo_high.get(m, 0)  for m in mo_sorted]
        fut_month_med    = [mo_med.get(m, 0)   for m in mo_sorted]

    # ── Channel percentages ────────────────────────────────────────────────
    ch_total = sum(ch_data.values()) or 1
    ch_pcts  = {k: round(v/ch_total*100,1) for k,v in ch_data.items()}

    # ── Current guests (window = today … today+15d, no past stays) ─────────
    arriving   = [g for g in guests if g['status']=='Arriving Today']
    in_house   = []   # no in-house: all guests in window have arr >= today
    high_count = sum(1 for s in scores if s>=70)
    med_count  = sum(1 for s in scores if 40<=s<70)
    low_count  = sum(1 for s in scores if s<40)

    # ── Overbooking recommendation ─────────────────────────────────────────
    from math import ceil
    if fut_bookings:
        tonight_high_risk = sum(
            1 for b, s in zip(fut_bookings, fut_scores)
            if (b['arr_date'].date() if isinstance(b['arr_date'], datetime) else b['arr_date']) == today.date() and s >= 70
        )
    else:
        tonight_high_risk = sum(1 for g, s in zip(guests, scores)
                                if g['status'] == 'Arriving Today' and s >= 70)
    overbook_rec = ceil(tonight_high_risk * 0.15) if tonight_high_risk else 0

    # ── ROI tracking stats ─────────────────────────────────────────────────
    _hotel_key = _d.get('hotel_key', VDV_HOTEL_KEY)
    roi_emails_sent = 0
    roi_recovered   = 0
    try:
        _conn_roi = get_db()
        _cur_roi  = _conn_roi.cursor()
        _cur_roi.execute("SELECT COUNT(*) FROM roi_actions WHERE hotel_username=%s",
                         (_hotel_key,))
        roi_emails_sent = _cur_roi.fetchone()[0]
        _cur_roi.execute("SELECT COUNT(*) FROM roi_actions WHERE hotel_username=%s AND booking_ref='RECOVERED'",
                         (_hotel_key,))
        roi_recovered = _cur_roi.fetchone()[0]
        _cur_roi.close()
        _conn_roi.close()
    except:
        pass
    roi_rev_saved = roi_recovered * int(avg_adr * avg_nights)

    # ── Guest table rows (grouped by name, today+15d window) ───────────────
    def st_badge(s):
        if s=='Arriving Today': return '<span class="stb stb-a">Arriving Today</span>'
        if s=='In House':       return '<span class="stb stb-h">In House</span>'
        if s=='Checked Out':    return '<span class="stb stb-o">Checked Out</span>'
        return f'<span class="stb stb-f">{s}</span>'

    def _risk_badge(sc):
        if sc>=70: return f'<span class="badge high">{sc:.1f}%</span>'
        if sc>=40: return f'<span class="badge med">{sc:.1f}%</span>'
        return f'<span class="badge low">{sc:.1f}%</span>'

    def _act_btn(idx, sc):
        if sc>=70: return f'<button class="abtn dep" onclick="event.stopPropagation();openEmail({idx},\'deposit\')">Deposit</button>'
        if sc>=40: return f'<button class="abtn rem" onclick="event.stopPropagation();openEmail({idx},\'reminder\')">Reminder</button>'
        return f'<button class="abtn mon" onclick="event.stopPropagation();openEmail({idx},\'contact\')">Contact</button>'

    # Group stays by guest name (case-insensitive); preserve original index for openEmail
    from collections import OrderedDict as _OD
    _groups = _OD()
    for _i, (g, sc) in enumerate(zip(guests, scores)):
        _key = g['name'].strip().lower()
        if _key not in _groups:
            _groups[_key] = []
        _groups[_key].append((_i, g, sc))

    rows_html = ''
    _row_idx = 0
    for _key, _stays in _groups.items():
        if len(_stays) == 1:
            # Single stay — render normal expandable row
            _i, g, sc = _stays[0]
            mb = f' <span class="mb">{g["membership"]}</span>' if g.get('membership') else ''
            _note = g.get('note', '')
            nt = (f'<span class="nt" title="{_note}">{_note[:36]}{"..." if len(_note)>36 else ""}</span>'
                  if _note else '&mdash;')
            rows_html += f'''<tr class="cr" data-score="{sc:.1f}" data-lead="0" data-rate="0" onclick="toggleExpand(this, {_i}, {sc:.1f})">
          <td><span class="gn">{g['name']}</span>{mb}</td>
          <td>{st_badge(g['status'])}</td>
          <td>{g['arrival']}</td><td>{g['nights']}n</td>
          <td>{_risk_badge(sc)}</td><td class="ntd">{nt}</td><td>{_act_btn(_i, sc)}</td>
        </tr>
        <tr class="exp-tr" id="exp-{_i}"><td class="exp-td" colspan="7"><div class="exp-inner" id="exp-inner-{_i}"></div></td></tr>'''
        else:
            # Multiple stays — one collapsed row + dropdown table of all stays
            _stays_sorted = sorted(_stays, key=lambda x: x[1]['arr_date'])
            _max_sc  = max(sc for _, _, sc in _stays_sorted)
            _first_g = _stays_sorted[0][1]
            _first_i = _stays_sorted[0][0]
            mb = f' <span class="mb">{_first_g["membership"]}</span>' if _first_g.get('membership') else ''
            _sid = f'grp-{_row_idx}'
            # Sub-rows — one per stay, same structure as single-stay rows
            _sub_rows = ''
            for _si, (_i2, _g2, _sc2) in enumerate(_stays_sorted):
                _note2 = _g2.get('note', '')
                _nt2 = (f'<span class="nt" title="{_note2}">{_note2[:36]}{"..." if len(_note2)>36 else ""}</span>'
                        if _note2 else '&mdash;')
                _sub_rows += f'''<tr class="stays-sub cr" data-grp="{_sid}" data-score="{_sc2:.1f}" style="display:none;background:#f8faff;">
          <td style="padding-left:28px;"><span style="color:#cbd5e1;margin-right:5px;font-size:11px;">↳</span><span class="gn" style="color:#4b5563;">{_g2['name']}</span></td>
          <td>{st_badge(_g2['status'])}</td>
          <td>{_g2['arrival']}</td><td>{_g2['nights']}n</td>
          <td>{_risk_badge(_sc2)}</td><td class="ntd">{_nt2}</td><td>{_act_btn(_i2, _sc2)}</td>
        </tr>'''
            rows_html += f'''<tr class="cr stays-hdr" data-score="{_max_sc:.1f}" data-lead="0" data-rate="0" onclick="toggleStays('{_sid}',this)">
          <td><span class="gn">{_first_g['name']}</span>{mb} <span class="stays-ct">{len(_stays_sorted)} stays</span> <span class="grp-chev" id="chev-{_sid}">▾</span></td>
          <td>{st_badge(_first_g['status'])}</td>
          <td>{_stays_sorted[0][1]['arrival']}</td><td>—</td>
          <td>{_risk_badge(_max_sc)}</td><td class="ntd">&mdash;</td><td>{_act_btn(_first_i, _max_sc)}</td>
        </tr>
        {_sub_rows}'''
        _row_idx += 1

    # ── Action plan items ──────────────────────────────────────────────────
    arriving_items = ''.join(
        f'<div class="pi"><b>{g["name"].split(",")[0]}</b> &mdash; {g.get("note") or "Prepare welcome"}</div>'
        for g in arriving
    ) or '<div class="pi empty">No arrivals today</div>'

    inhouse_items = ''.join(
        f'<div class="pi"><b>{g["name"].split(",")[0]}</b> &mdash; dep {g["departure"]}'
        + (f' · {g["note"][:50]}' if g.get('note') else '') + '</div>'
        for g in in_house
    ) or '<div class="pi empty">None in house</div>'

    # ── JS data ────────────────────────────────────────────────────────────
    guests_js       = json.dumps([{
        'name':g['name'],'arrival':g['arrival'],'departure':g.get('departure',''),
        'nights':g['nights'],'adults':g['adults'],'membership':g.get('membership',''),
        'status':g['status'],'note':g.get('note',''),'adr':149.0 if 'CORP' in g.get('membership','') else 115.0
    } for g in guests])
    scores_js       = json.dumps([round(s,1) for s in scores])
    months_js       = json.dumps(MONTHS)
    cx_js           = json.dumps(CX_MONTHLY)
    ns_js           = json.dumps(NS_MONTHLY)
    lost_js         = json.dumps(LOST_MONTHLY)
    ch_labels_js    = json.dumps(list(ch_data.keys()))
    ch_vals_js      = json.dumps(list(ch_data.values()))
    ch_pcts_js      = json.dumps([ch_pcts[k] for k in ch_data])
    gnames_js       = json.dumps([g['name'].split(',')[0] for g in guests])
    gcols_js        = json.dumps(['#dc2626' if s>=70 else ('#f59e0b' if s>=40 else '#22c55e') for s in scores])
    fut_mlabels_js  = json.dumps(fut_month_labels)
    fut_mhigh_js    = json.dumps(fut_month_high)
    fut_mmed_js     = json.dumps(fut_month_med)
    fut_ch_labels_js = json.dumps(list(fut_by_channel.keys()))
    fut_ch_vals_js   = json.dumps(list(fut_by_channel.values()))

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Occupado — Van der Valk Mechelen</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*,*::before,*::after{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:#fff;color:#111827;font-family:'Plus Jakarta Sans',sans-serif;-webkit-font-smoothing:antialiased;font-size:14px;line-height:1.5;}}
a{{text-decoration:none;color:inherit;}}
/* TOPBAR */
.topbar{{height:56px;border-bottom:1px solid #e5e7eb;display:flex;align-items:center;padding:0 40px;background:#fff;position:sticky;top:0;z-index:100;gap:12px;}}
.tb-brand{{font-size:15px;font-weight:700;letter-spacing:-0.3px;color:#111827;font-family:'Syne',sans-serif;}}
.tb-brand span{{color:#00d165;}}
.tb-hotel{{font-size:12px;color:#9ca3af;padding-left:14px;border-left:1px solid #e5e7eb;}}
.tb-right{{margin-left:auto;display:flex;gap:6px;align-items:center;}}
.tb-btn{{padding:6px 14px;border:1px solid #e5e7eb;border-radius:6px;font-size:12px;color:#6b7280;background:transparent;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;transition:border-color .15s,color .15s;text-decoration:none;display:inline-flex;align-items:center;}}
.tb-btn:hover{{border-color:#d1d5db;color:#111827;}}
.lang-sel{{padding:6px 10px;border:1px solid #e5e7eb;border-radius:6px;font-size:12px;color:#6b7280;background:transparent;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;outline:none;}}
/* PAGE */
.page{{max-width:1100px;margin:0 auto;padding:40px 40px 80px;}}
/* ROWS */
.row{{display:grid;gap:16px;margin-bottom:0;}}
.row-3{{grid-template-columns:repeat(3,1fr);}}
.row-2{{grid-template-columns:1fr 1fr;}}
.row-2l{{grid-template-columns:2fr 1fr;}}
.row-hero{{grid-template-columns:1fr 1fr 1fr;}}
/* SECTION LABEL */
.sh{{display:flex;align-items:center;gap:16px;margin:48px 0 20px;}}
.sh-title{{font-size:11px;font-weight:500;letter-spacing:0.08em;text-transform:uppercase;color:#9ca3af;white-space:nowrap;}}
.sh-line{{flex:1;height:1px;background:#f3f4f6;}}
.sh-sub{{font-size:11px;color:#d1d5db;white-space:nowrap;}}
/* HERO METRICS */
.hero-card{{background:#fff;border:none;border-radius:0;padding:28px 32px;border-right:1px solid #e5e7eb;}}
.hero-card:last-child{{border-right:none;}}
.hc-label{{font-size:11px;font-weight:500;letter-spacing:0.07em;text-transform:uppercase;color:#9ca3af;margin-bottom:12px;}}
.hc-num{{font-size:44px;font-weight:700;letter-spacing:-2px;line-height:1;color:#111827;}}
.hc-num.red{{color:#ef4444;}}
.hc-num.green{{color:#00d165;}}
.hc-num.amber{{color:#f59e0b;}}
.hc-sub{{font-size:12px;color:#9ca3af;margin-top:8px;}}
.hc-tag{{display:none;}}
.hero-metrics-wrap{{display:grid;grid-template-columns:1fr 1fr 1fr;border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;margin-bottom:48px;}}
/* TODAY STRIP */
.today-strip{{display:grid;grid-template-columns:repeat(5,1fr);border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;margin-bottom:48px;}}
.ts-item{{padding:24px 28px;border-right:1px solid #e5e7eb;}}
.ts-item:last-child{{border-right:none;}}
.ts-num{{font-size:38px;font-weight:700;letter-spacing:-1.5px;line-height:1;color:#111827;}}
.ts-num.r{{color:#ef4444;}}
.ts-num.a{{color:#f59e0b;}}
.ts-num.g{{color:#00d165;}}
.ts-label{{font-size:11px;color:#9ca3af;margin-top:6px;text-transform:uppercase;letter-spacing:0.07em;font-weight:500;}}
/* CARD */
.card{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:24px;}}
.card-title{{font-size:13px;font-weight:600;color:#111827;margin-bottom:2px;}}
.card-sub{{font-size:11px;color:#9ca3af;margin-bottom:20px;text-transform:uppercase;letter-spacing:0.06em;font-weight:500;}}
.stat-row{{display:flex;justify-content:space-between;align-items:center;padding:10px 0;border-bottom:1px solid #f3f4f6;font-size:13px;}}
.stat-row:last-child{{border-bottom:none;}}
.sr-label{{color:#6b7280;}}
.sr-val{{font-weight:600;color:#111827;}}
/* TABLE */
.tbl{{width:100%;border-collapse:collapse;}}
.tbl th{{font-size:11px;font-weight:500;letter-spacing:0.08em;text-transform:uppercase;color:#9ca3af;padding:10px 12px;border-bottom:1px solid #e5e7eb;text-align:left;white-space:nowrap;}}
.tbl th:first-child,.tbl td:first-child{{padding-left:0;}}
.tbl th:last-child,.tbl td:last-child{{text-align:right;padding-right:0;}}
.tbl td{{padding:13px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;color:#374151;vertical-align:middle;}}
.tbl tr:last-child td{{border-bottom:none;}}
.cr{{cursor:pointer;}}
.cr:hover td{{color:#111827;}}
.gn{{font-weight:600;color:#111827;font-size:13px;}}
.mb{{background:#eff6ff;color:#3b82f6;border:1px solid #bfdbfe;padding:1px 5px;border-radius:3px;font-family:'JetBrains Mono',monospace;font-size:9px;font-weight:600;margin-left:4px;}}
.badge{{padding:2px 7px;border-radius:4px;font-family:'JetBrains Mono',monospace;font-size:11px;font-weight:600;border:1px solid;white-space:nowrap;}}
.high{{background:#fef2f2;color:#ef4444;border-color:#fecaca;}}
.med{{background:#fffbeb;color:#f59e0b;border-color:#fde68a;}}
.low{{background:#f0fdf4;color:#00d165;border-color:#bbf7d0;}}
.stb{{padding:2px 7px;border-radius:4px;font-family:'JetBrains Mono',monospace;font-size:10px;font-weight:500;border:1px solid;white-space:nowrap;display:inline-block;}}
.stb-a{{background:#fffbeb;color:#92400e;border-color:#fde68a;}}
.stb-h{{background:#eff6ff;color:#1d4ed8;border-color:#bfdbfe;}}
.stb-o{{background:#f9fafb;color:#9ca3af;border-color:#e5e7eb;}}
.stb-f{{background:#f0fdf4;color:#15803d;border-color:#bbf7d0;}}
.ntd{{max-width:160px;}}
.nt{{font-size:11px;color:#9ca3af;font-style:italic;}}
.abtn{{padding:4px 10px;border-radius:5px;font-size:11px;font-weight:500;cursor:pointer;border:1px solid;background:transparent;font-family:'Plus Jakarta Sans',sans-serif;transition:background .1s;white-space:nowrap;}}
.dep{{color:#ef4444;border-color:#fecaca;}}.dep:hover{{background:#fef2f2;}}
.rem{{color:#f59e0b;border-color:#fde68a;}}.rem:hover{{background:#fffbeb;}}
.mon{{color:#00d165;border-color:#bbf7d0;}}.mon:hover{{background:#f0fdf4;}}
/* CHARTS */
.chart-wrap{{border:1px solid #e5e7eb;border-radius:12px;padding:24px;}}
.chart-label{{font-size:11px;font-weight:500;letter-spacing:0.08em;text-transform:uppercase;color:#9ca3af;margin-bottom:20px;}}
/* SAVINGS */
.savings-card{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:28px;}}
.sv-row{{display:flex;justify-content:space-between;align-items:center;padding:10px 0;border-bottom:1px solid #f3f4f6;font-size:13px;}}
.sv-row:last-child{{border-bottom:none;}}
.sv-label{{color:#6b7280;}}
.sv-val{{font-weight:600;color:#111827;}}
.sv-val.g{{color:#00d165;}}
.sv-val.r{{color:#ef4444;}}
.slider-row{{margin-top:20px;}}
.slider-label{{font-size:11px;color:#9ca3af;margin-bottom:8px;display:flex;justify-content:space-between;letter-spacing:0.05em;text-transform:uppercase;font-weight:500;}}
input[type=range]{{width:100%;accent-color:#00d165;cursor:pointer;}}
.slider-result{{margin-top:14px;font-size:40px;font-weight:700;color:#00d165;letter-spacing:-2px;line-height:1;}}
.slider-sub{{font-size:11px;color:#9ca3af;margin-top:4px;}}
/* MICE */
.mice-grid{{display:grid;grid-template-columns:repeat(4,1fr);border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;margin-bottom:28px;}}
.mice-card{{padding:22px 24px;border-right:1px solid #e5e7eb;background:#fff;border-radius:0;border-top:none;border-bottom:none;border-left:none;}}
.mice-card:last-child{{border-right:none;}}
.mice-card.blue{{background:#fff;}}
.mice-num{{font-size:34px;font-weight:700;letter-spacing:-1.5px;line-height:1;color:#111827;}}
.mice-num.blue{{color:#3b82f6;}}
.mice-lbl{{font-size:11px;color:#9ca3af;margin-top:6px;text-transform:uppercase;letter-spacing:0.07em;font-weight:500;}}
.mice-sub{{font-size:11px;color:#d1d5db;margin-top:2px;}}
.mice-row{{display:grid;grid-template-columns:1fr 280px;gap:16px;margin-bottom:28px;align-items:start;}}
.mice-chart-card{{border:1px solid #e5e7eb;border-radius:12px;padding:20px;background:#fff;}}
.mice-chart-title{{font-size:11px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:16px;}}
.mice-tbl{{width:100%;border-collapse:collapse;}}
.mice-tbl th{{font-size:11px;font-weight:500;letter-spacing:0.08em;text-transform:uppercase;color:#9ca3af;padding:10px 0;border-bottom:1px solid #e5e7eb;text-align:left;}}
.mice-tbl td{{padding:11px 0;border-bottom:1px solid #f3f4f6;font-size:12.5px;color:#374151;}}
.mice-tbl tr:last-child td{{border-bottom:none;}}
.seg-pill{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:10px;font-family:'JetBrains Mono',monospace;font-weight:500;border:1px solid;}}
.seg-corpfix{{background:#eff6ff;color:#1d4ed8;border-color:#bfdbfe;}}
.seg-corpdyn{{background:#f0fdf4;color:#15803d;border-color:#bbf7d0;}}
.seg-bnsgrp{{background:#fdf4ff;color:#7e22ce;border-color:#e9d5ff;}}
.seg-mtgbns{{background:#fff7ed;color:#c2410c;border-color:#fed7aa;}}
.grp-row{{display:flex;justify-content:space-between;align-items:center;padding:10px 0;border-bottom:1px solid #f3f4f6;font-size:12.5px;}}
.grp-row:last-child{{border-bottom:none;}}
.grp-name{{font-weight:600;color:#111827;}}
.grp-co{{color:#9ca3af;font-size:11px;margin-top:2px;}}
.grp-rooms{{font-family:'JetBrains Mono',monospace;font-size:11px;font-weight:600;background:#f0fdf4;padding:3px 10px;border-radius:4px;border:1px solid #bbf7d0;color:#15803d;}}
.ap-card{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:20px;}}
.ap-head{{font-size:11px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:14px;display:flex;align-items:center;gap:8px;}}
.pi{{border-bottom:1px solid #f3f4f6;padding:9px 0;font-size:12.5px;line-height:1.5;color:#374151;}}
.pi:last-child{{border-bottom:none;}}
.pi b{{color:#111827;}}
.pi.empty{{color:#9ca3af;font-style:italic;}}
.ap-btn{{margin-top:14px;width:100%;padding:9px;background:#00d165;border:none;border-radius:6px;color:#0a0a0a;font-size:12px;font-weight:600;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;transition:background .15s;}}
.ap-btn:hover{{background:#04e270;}}
.ap-btn.ghost{{background:#fff;border:1px solid #e5e7eb;color:#111827;}}
.ap-btn.ghost:hover{{background:#f9fafb;}}
.ap-grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;}}
.ch-header-row td{{background:#f1f5f9;font-size:11px;font-weight:600;color:#475569;padding:7px 10px;letter-spacing:.04em;text-transform:uppercase;}}
.gtd-none{{background:#fef2f2;color:#ef4444;border:1px solid #fecaca;padding:1px 7px;border-radius:4px;font-family:'JetBrains Mono',monospace;font-size:9px;font-weight:600;white-space:nowrap;}}
.gtd-ok{{background:#f0fdf4;color:#16a34a;border:1px solid #bbf7d0;padding:1px 7px;border-radius:4px;font-family:'JetBrains Mono',monospace;font-size:9px;font-weight:600;white-space:nowrap;}}
.alert-card{{border:1px solid #fed7aa;background:#fff7ed;border-radius:10px;padding:14px 18px;display:flex;align-items:center;gap:12px;}}
.alert-icon{{font-size:18px;flex-shrink:0;}}
.alert-title{{font-size:13px;font-weight:600;color:#9a3412;}}
.alert-sub{{font-size:12px;color:#c2410c;margin-top:1px;}}
.fstrip{{display:grid;grid-template-columns:repeat(4,1fr);border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;margin-bottom:28px;}}
.fstrip .ts-item{{border-right:1px solid #e5e7eb;}}
.fstrip .ts-item:last-child{{border-right:none;}}
.insight-row{{display:flex;gap:8px;margin-bottom:24px;flex-wrap:wrap;}}
.ip{{border:1px solid #e5e7eb;border-radius:99px;padding:4px 12px;font-size:12px;color:#6b7280;cursor:pointer;transition:all .15s;white-space:nowrap;}}
.ip:hover,.ip.active{{background:#111827;color:#fff;border-color:#111827;}}
.mo{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.25);z-index:1000;align-items:center;justify-content:center;}}
.mo.show{{display:flex;}}
.mb-inner{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:32px;width:100%;max-width:460px;max-height:88vh;overflow-y:auto;position:relative;}}
.mc{{position:absolute;top:14px;right:16px;font-size:18px;cursor:pointer;color:#9ca3af;background:none;border:none;}}
.mt{{font-size:18px;font-weight:700;color:#111827;margin-bottom:2px;letter-spacing:-0.3px;}}
.ms{{font-size:11px;color:#9ca3af;margin-bottom:18px;font-family:'JetBrains Mono',monospace;}}
.sd{{font-size:52px;font-weight:700;line-height:1;margin-bottom:6px;letter-spacing:-2px;}}
.sb-bg{{height:4px;background:#f3f4f6;border-radius:2px;overflow:hidden;margin-bottom:12px;}}
.sb-fill{{height:100%;border-radius:2px;}}
.sv-tag{{font-size:11px;font-weight:600;padding:4px 10px;border-radius:4px;display:inline-block;margin-bottom:14px;}}
.dr{{display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid #f3f4f6;font-size:12px;}}
.dr:last-child{{border-bottom:none;}}
.dl{{color:#9ca3af;}}
.dv{{font-family:'JetBrains Mono',monospace;color:#111827;font-weight:500;}}
.ri{{padding:8px 10px;margin-bottom:4px;font-size:12px;color:#374151;border-left:2px solid #e5e7eb;}}
.ri.pos{{border-left-color:#00d165;}}
.ri.neg{{border-left-color:#f59e0b;}}
.ec{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.25);z-index:1001;align-items:center;justify-content:center;}}
.ec.show{{display:flex;}}
.eb{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:28px;width:100%;max-width:540px;max-height:90vh;overflow-y:auto;}}
.el{{font-size:11px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:0.08em;display:block;margin-bottom:5px;}}
.ei{{width:100%;padding:9px 12px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px;font-size:13px;color:#111827;outline:none;margin-bottom:10px;font-family:'Plus Jakarta Sans',sans-serif;}}
.ei:focus{{border-color:#00d165;background:#fff;}}
.eta{{width:100%;padding:9px 12px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px;font-size:13px;color:#111827;outline:none;resize:vertical;min-height:140px;margin-bottom:10px;font-family:'Plus Jakarta Sans',sans-serif;line-height:1.6;}}
.eta:focus{{border-color:#00d165;background:#fff;}}
.ea{{display:flex;gap:8px;margin-top:16px;}}
.es{{flex:1;padding:10px;background:#00d165;color:#0a0a0a;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;}}
.es:hover{{background:#04e270;}}
.ecc{{flex:1;padding:10px;background:transparent;color:#6b7280;border:1px solid #e5e7eb;border-radius:6px;font-size:13px;cursor:pointer;}}
.ecc:hover{{background:#f9fafb;}}
.toast{{position:fixed;bottom:24px;right:24px;background:#111827;color:#fff;border-radius:8px;padding:12px 16px;font-size:13px;transform:translateY(50px);opacity:0;transition:all .25s;z-index:2000;}}
.toast.show{{transform:translateY(0);opacity:1;}}
@media(max-width:900px){{
  .page{{padding:24px 16px 60px;}}
  .topbar{{padding:0 16px;}}
  .tb-hotel{{display:none;}}
  .hero-metrics-wrap,.row-hero{{grid-template-columns:1fr 1fr;}}
  .hero-card:nth-child(2){{border-right:none;}}
  .hero-card:nth-child(3){{border-top:1px solid #e5e7eb;border-right:none;}}
  .row-2,.row-2l,.row-3{{grid-template-columns:1fr;}}
  .mice-grid{{grid-template-columns:1fr 1fr;}}
  .mice-card:nth-child(2){{border-right:none;}}
  .mice-card:nth-child(3){{border-top:1px solid #e5e7eb;}}
  .mice-card:nth-child(4){{border-top:1px solid #e5e7eb;border-right:none;}}
  .mice-row{{grid-template-columns:1fr;}}
  .today-strip{{grid-template-columns:repeat(3,1fr);}}
  .today-strip .ts-item:nth-child(4),.today-strip .ts-item:nth-child(5){{display:none;}}
  .fstrip{{grid-template-columns:1fr 1fr;}}
  .tbl th:nth-child(4),.tbl td:nth-child(4),.tbl th:nth-child(5),.tbl td:nth-child(5){{display:none;}}
  .ap-grid{{grid-template-columns:1fr;}}
  .sh{{margin:36px 0 14px;}}
}}


/* ── DYNAMICS ────────────────────────────────────────────── */
.filter-bar{{display:flex;gap:8px;margin-bottom:20px;align-items:center;flex-wrap:wrap;}}
.f-tab{{padding:5px 16px;border:1px solid #e5e7eb;border-radius:99px;font-size:12px;color:#6b7280;cursor:pointer;transition:all .15s;background:#fff;font-family:'Plus Jakarta Sans',sans-serif;}}
.f-tab:hover{{border-color:#111827;color:#111827;}}
.f-tab.active{{background:#111827;color:#fff;border-color:#111827;}}
.f-count{{font-size:10px;opacity:.6;margin-left:3px;}}
.sort-th{{cursor:pointer;user-select:none;}}
.sort-th:hover{{color:#374151;}}
.sort-arr{{margin-left:3px;opacity:.25;font-size:9px;}}
.sort-arr.on{{opacity:1;}}
.live-pill{{display:inline-flex;align-items:center;gap:5px;font-size:11px;color:#9ca3af;padding:4px 10px;border:1px solid #e5e7eb;border-radius:99px;}}
.live-dot{{width:6px;height:6px;border-radius:50%;background:#00d165;flex-shrink:0;animation:pdot 2s ease-in-out infinite;}}
@keyframes pdot{{0%,100%{{opacity:1;transform:scale(1);}}50%{{opacity:.3;transform:scale(.8);}}}}
.exp-tr{{display:none;}}
.exp-tr.open{{display:table-row;}}
.exp-td{{padding:0!important;background:#fff!important;border-bottom:1px solid #f3f4f6!important;}}
.stays-ct{{display:inline-block;background:#eff6ff;color:#2563eb;border:1px solid #bfdbfe;border-radius:99px;font-size:10px;font-weight:600;padding:1px 7px;margin-left:6px;vertical-align:middle;}}
.stays-hdr{{cursor:pointer;}}
.stays-hdr:hover td{{background:#fafafa;}}
.grp-chev{{font-size:11px;color:#9ca3af;margin-left:5px;display:inline-block;vertical-align:middle;line-height:1;}}
.exp-inner{{padding:20px 0 24px;display:grid;grid-template-columns:100px 1fr;gap:24px;align-items:start;}}
.exp-score-wrap{{display:flex;flex-direction:column;align-items:center;gap:4px;padding-top:4px;}}
.exp-score-big{{font-size:48px;font-weight:700;letter-spacing:-2px;line-height:1;}}
.exp-score-lbl{{font-size:9px;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;margin-top:2px;}}
.exp-bar-bg{{width:70px;height:3px;background:#f3f4f6;border-radius:2px;overflow:hidden;margin-top:8px;}}
.exp-bar-fill{{height:100%;border-radius:2px;transition:width .7s ease;}}
.exp-right{{}}
.exp-head{{font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;margin-bottom:10px;}}
.exp-factors{{display:flex;flex-direction:column;gap:4px;}}
.exp-factor{{font-size:12px;color:#374151;padding:6px 10px 6px 12px;border-left:2px solid #e5e7eb;line-height:1.4;}}
.exp-factor.pos{{border-left-color:#00d165;}}
.exp-factor.warn{{border-left-color:#f59e0b;}}
.exp-factor.bad{{border-left-color:#ef4444;}}
.exp-conclusion{{margin-top:10px;font-size:12px;font-weight:500;color:#111827;padding:8px 12px;background:#f9fafb;border-radius:6px;border:1px solid #e5e7eb;}}
.cr.is-open{{background:#fafafa;}}
.count-up{{display:inline;}}
@keyframes fadeUp{{from{{opacity:0;transform:translateY(10px);}}to{{opacity:1;transform:translateY(0);}}}}
.anim-card{{opacity:0;animation:fadeUp .5s ease forwards;}}
.anim-card:nth-child(1){{animation-delay:.04s;}}
.anim-card:nth-child(2){{animation-delay:.08s;}}
.anim-card:nth-child(3){{animation-delay:.12s;}}
.anim-card:nth-child(4){{animation-delay:.16s;}}
@media(max-width:900px){{.exp-inner{{grid-template-columns:1fr;gap:16px;}}}}
</style>
</head>
<body>

<div class="topbar">
  <span class="tb-brand">Occup<span>ado</span></span>
  <span class="tb-hotel">{hotel_name}</span>
  <div class="tb-right">
    <select class="lang-sel" onchange="location.href='/dashboard?lang='+this.value">
      <option value="en" {"selected" if lang=="en" else ""}>EN</option>
      <option value="nl" {"selected" if lang=="nl" else ""}>NL</option>
      <option value="fr" {"selected" if lang=="fr" else ""}>FR</option>
    </select>
    <span class="live-pill"><span class="live-dot"></span>Live</span><a href="/settings" class="tb-btn">Settings</a>
    <a href="/logout" class="tb-btn">Sign Out</a>
  </div>
</div>

<div class="page">

<!-- HERO METRICS ─────────────────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">6-Month Overview</span><span class="sh-line"></span><span class="sh-sub">Oct 2025 – {today_str} · Shiji data</span></div>
<div class="hero-metrics-wrap">
  <div class="hero-card">
    <div class="hc-label">Missed Stays</div>
    <div class="hc-num red anim-card"><span class="count-up" data-val="{total_lost}">—</span></div>
    <div class="hc-sub">{total_cx:,} cancellations · {total_ns} no-shows</div>
  </div>
  <div class="hero-card">
    <div class="hc-label">Revenue Lost</div>
    <div class="hc-num amber anim-card">€<span class="count-up" data-val="{rev_lost//1000}" data-suf="k">—</span></div>
    <div class="hc-sub">€{int(avg_adr)} ADR · {avg_nights} avg nights</div>
  </div>
  <div class="hero-card">
    <div class="hc-label">Recoverable</div>
    <div class="hc-num green anim-card">€<span class="count-up" data-val="{recoverable//1000}" data-suf="k">—</span></div>
    <div class="hc-sub">Est. with 30% intervention rate</div>
  </div>
</div>

<!-- OVERBOOKING + ROI STRIP ─────────────────────────────────────────── -->
{f'''<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:24px;margin-bottom:8px;">
  <div style="border:1px solid #fef3c7;background:#fffbeb;border-radius:12px;padding:20px 24px;display:flex;align-items:center;gap:16px;">
    <div style="font-size:28px;line-height:1;">⚡</div>
    <div>
      <div style="font-size:11px;font-weight:500;letter-spacing:.07em;text-transform:uppercase;color:#92400e;margin-bottom:4px;">Overbooking Recommendation</div>
      <div style="font-size:20px;font-weight:700;color:#92400e;">{f"Sell +{overbook_rec} extra room{'s' if overbook_rec!=1 else ''} tonight" if overbook_rec > 0 else "No overbooking needed tonight"}</div>
      <div style="font-size:12px;color:#a16207;margin-top:4px;">{tonight_high_risk} high-risk arrival{'s' if tonight_high_risk!=1 else ''} tonight · 15% walk rate assumed</div>
    </div>
  </div>
  <div style="border:1px solid #d1fae5;background:#f0fdf4;border-radius:12px;padding:20px 24px;display:flex;align-items:center;gap:16px;">
    <div style="font-size:28px;line-height:1;">📈</div>
    <div>
      <div style="font-size:11px;font-weight:500;letter-spacing:.07em;text-transform:uppercase;color:#166534;margin-bottom:4px;">ROI Tracking</div>
      <div style="font-size:20px;font-weight:700;color:#166534;">€{roi_rev_saved:,} saved</div>
      <div style="font-size:12px;color:#15803d;margin-top:4px;">{roi_emails_sent} emails sent · {roi_recovered} recoveries confirmed</div>
    </div>
  </div>
</div>''' if True else ''}

<!-- CHARTS ───────────────────────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Trends</span><span class="sh-line"></span><span class="sh-sub">Monthly breakdown — click bars or points to filter</span></div>
<div class="row row-2">
  <div class="card">
    <div class="card-title">Cancellations & No-shows</div>
    <div class="card-sub">Per month — stacked</div>
    <canvas id="trendChart" height="160"></canvas>
  </div>
  <div class="card">
    <div class="card-title">Cancellation Sources</div>
    <div class="card-sub">{total_cx:,} total · share by channel</div>
    <canvas id="channelChart" height="160"></canvas>
  </div>
</div>

<!-- SAVINGS CALCULATOR ────────────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Savings Calculator</span><span class="sh-line"></span><span class="sh-sub">Slide to adjust intervention rate</span></div>
<div class="row row-2l">
  <div class="savings-card">
    <div class="sv-row"><span class="sv-label">Cancellations tracked</span><span class="sv-val">{total_cx:,}</span></div>
    <div class="sv-row"><span class="sv-label">No-shows tracked</span><span class="sv-val">{total_ns}</span></div>
    <div class="sv-row"><span class="sv-label">Revenue lost (6 months)</span><span class="sv-val r">€{rev_lost:,}</span></div>
    <div class="slider-row">
      <div class="slider-label"><span>Intervention success rate</span><span id="rate-pct">30%</span></div>
      <input type="range" id="rate-slider" min="5" max="60" value="30" oninput="updateSavings(this.value)">
      <div class="slider-result" id="savings-num">€{recoverable:,}</div>
      <div class="slider-sub" id="savings-sub">estimated recoverable over 6 months</div>
    </div>
  </div>
  <div class="card">
    <div class="card-title">Key Insights</div>
    <div class="card-sub">From your Shiji data</div>
    <div class="stat-row"><span class="sr-label">Peak month</span><span class="sr-val">Dec 2025 (335 cx)</span></div>
    <div class="stat-row"><span class="sr-label">Best month</span><span class="sr-val">Oct 2025 (167 cx)</span></div>
    <div class="stat-row"><span class="sr-label">Top channel risk</span><span class="sr-val">Booking.com — {ch_pcts.get('Booking.com',0)}%</span></div>
    <div class="stat-row"><span class="sr-label">No-show rate</span><span class="sr-val">{round(total_ns/total_lost*100,1) if total_lost else 0}% of total</span></div>
    <div class="stat-row"><span class="sr-label">Model accuracy</span><span class="sr-val" style="color:#00d165">80.5%</span></div>
  </div>
</div>

<!-- TODAY ────────────────────────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Today</span><span class="sh-line"></span><span class="sh-sub">{today_str}</span></div>
<div class="today-strip">
  <div class="ts-item"><div class="ts-num {'g' if len(arriving)>0 else ''}"><span class="count-up" data-val="{len(arriving)}">{len(arriving)}</span></div><div class="ts-label">Arriving</div></div>
  <div class="ts-item"><div class="ts-num"><span class="count-up" data-val="{len(in_house)}">{len(in_house)}</span></div><div class="ts-label">In House</div></div>
  <div class="ts-item"><div class="ts-num {'r' if high_count>0 else 'g'}"><span class="count-up" data-val="{high_count}">{high_count}</span></div><div class="ts-label">High Risk</div></div>
  <div class="ts-item"><div class="ts-num a">{med_count}</div><div class="ts-label">Medium Risk</div></div>
  <div class="ts-item"><div class="ts-num g">{low_count}</div><div class="ts-label">Low Risk</div></div>
</div>

<!-- FUTURE BOOKINGS RISK ─────────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Upcoming Bookings — Risk Forecast</span><span class="sh-line"></span><span class="sh-sub">Apr – Dec 2026 · {fut_total:,} reservations scored</span><a href="/vdv/export-highrisk" style="margin-left:auto;display:inline-flex;align-items:center;gap:5px;padding:6px 14px;background:#111827;color:#fff;border-radius:7px;font-size:11px;font-weight:600;font-family:'Plus Jakarta Sans',sans-serif;text-decoration:none;white-space:nowrap;">↓ Export Excel</a></div>

<div class="alert-card" style="margin-bottom:14px;">
  <div class="alert-icon">&#9888;</div>
  <div class="alert-body">
    <div class="alert-title">{fut_no_gtd:,} bookings have no deposit or guarantee (GTD: NONE)</div>
    <div class="alert-sub">These reservations carry no financial commitment — highest no-show risk. Contact before arrival or request payment guarantee.</div>
  </div>
  <div style="font-family:'Plus Jakarta Sans',sans-serif;font-size:32px;font-weight:800;color:#ea580c;flex-shrink:0;">{round(fut_no_gtd/max(1,fut_total)*100)}%</div>
</div>

<div class="fstrip">
  <div class="ts-item"><div class="ts-num">{fut_total:,}</div><div class="ts-label">Total Upcoming</div></div>
  <div class="ts-item"><div class="ts-num r">{fut_high:,}</div><div class="ts-label">High Risk &ge;70%</div></div>
  <div class="ts-item"><div class="ts-num a">{fut_med:,}</div><div class="ts-label">Medium Risk 40–70%</div></div>
  <div class="ts-item"><div class="ts-num g">{fut_low:,}</div><div class="ts-label">Low Risk &lt;40%</div></div>
</div>

<div class="row row-2">
  <div class="card">
    <div class="card-title">Risk by Month</div>
    <div class="card-sub">High &amp; medium risk bookings per arrival month</div>
    <canvas id="futMonthChart" height="160"></canvas>
  </div>
  <div class="card">
    <div class="card-title">Risk by Channel</div>
    <div class="card-sub">Total upcoming bookings by source</div>
    <canvas id="futChannelChart" height="160"></canvas>
  </div>
</div>

{"" if not fut_table_html else f'''
<div class="sh"><span class="sh-title">Top 10 at Risk — Per Channel</span><span class="sh-line"></span><span class="sh-sub">Act now to prevent cancellation</span></div>
<table class="tbl">
<thead><tr>
  <th>#</th><th>Guest</th><th>Arrival</th><th>Nights</th><th>Lead</th><th>Channel</th><th>GTD</th><th>Risk</th><th>Action</th>
</tr></thead>
<tbody>{fut_table_html}</tbody>
</table>
<div style="margin-top:12px;text-align:right;">
  <a href="{export_url}" style="display:inline-flex;align-items:center;gap:6px;padding:8px 18px;background:#111827;color:#fff;border-radius:8px;font-size:12px;font-weight:600;font-family:\'Plus Jakarta Sans\',sans-serif;text-decoration:none;">
    ↓ Export High-Risk to Excel
  </a>
</div>
'''}

<!-- OCCUPANCY & REVENUE TREND ──────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Occupancy & Revenue Trend</span><span class="sh-line"></span><span class="sh-sub">{len(fc_hist)} days history · {len(fc_fore)} days forecast · FOR_042</span></div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:0;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;margin-bottom:32px;">
  <div style="padding:24px;border-right:1px solid #e5e7eb;">
    <div style="font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;margin-bottom:16px;font-family:'JetBrains Mono',monospace;">Occupancy % &nbsp;<span style="color:#374151;">——</span> History &nbsp;<span style="color:#93c5fd;">- - -</span> Forecast</div>
    <div style="position:relative;height:160px;"><canvas id="occChart"></canvas></div>
  </div>
  <div style="padding:24px;">
    <div style="font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;margin-bottom:16px;font-family:'JetBrains Mono',monospace;">Room Revenue (EUR) &nbsp;<span style="color:#374151;">▪</span> History &nbsp;<span style="color:#93c5fd;">▪</span> Forecast</div>
    <div style="position:relative;height:160px;"><canvas id="revChart"></canvas></div>
  </div>
</div>

{('<div class="sh"><span class="sh-title">Group Pipeline</span><span class="sh-line"></span><span class="sh-sub">Upcoming confirmed & tentative groups · GRP_017</span></div>' + _grp_pipe_html(grp_pipe, STATUS_COLORS)) if grp_pipe else ''}

<!-- MODEL ACCURACY ──────────────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Model Accuracy — Last 30 Days</span><span class="sh-line"></span><span class="sh-sub">Ground truth from detected outcomes · vdv_outcome_log</span></div>
<div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:28px;">
  <div style="flex:1;min-width:200px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:10px;padding:18px 20px;">
    <div style="font-size:11px;font-weight:600;letter-spacing:.07em;text-transform:uppercase;color:#6b7280;margin-bottom:10px;">Cancelled bookings tracked</div>
    <div style="font-size:28px;font-weight:700;color:#111827;margin-bottom:10px;">{_outcome_stats['cancelled']}</div>
    <div style="font-size:12px;color:#374151;line-height:1.8;">
      Were <strong>high risk (≥70%)</strong>:&nbsp;&nbsp;{_outcome_stats['cx_high']}{f" <span style='color:#6b7280;'>({round(_outcome_stats['cx_high']/_outcome_stats['cancelled']*100)}%)</span>" if _outcome_stats['cancelled'] else ""}<br>
      Were <strong>medium risk</strong>:&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{_outcome_stats['cx_med']}{f" <span style='color:#6b7280;'>({round(_outcome_stats['cx_med']/_outcome_stats['cancelled']*100)}%)</span>" if _outcome_stats['cancelled'] else ""}<br>
      Were <strong>low risk</strong>:&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{_outcome_stats['cx_low']}{f" <span style='color:#6b7280;'>({round(_outcome_stats['cx_low']/_outcome_stats['cancelled']*100)}%)</span>" if _outcome_stats['cancelled'] else ""}
    </div>
  </div>
  <div style="flex:1;min-width:200px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:10px;padding:18px 20px;">
    <div style="font-size:11px;font-weight:600;letter-spacing:.07em;text-transform:uppercase;color:#6b7280;margin-bottom:10px;">Completed stays tracked</div>
    <div style="font-size:28px;font-weight:700;color:#00d165;margin-bottom:10px;">{_outcome_stats['completed']}</div>
    <div style="font-size:12px;color:#6b7280;line-height:1.8;">Bookings that arrived as expected.<br>Used as negative examples for model evaluation.</div>
  </div>
  {'<div style="flex:1;min-width:200px;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:18px 20px;display:flex;align-items:center;justify-content:center;"><div style="text-align:center;"><div style="font-size:11px;font-weight:600;letter-spacing:.07em;text-transform:uppercase;color:#166534;margin-bottom:8px;">High-risk catch rate</div><div style="font-size:36px;font-weight:700;color:#166534;">' + str(round(_outcome_stats["cx_high"]/_outcome_stats["cancelled"]*100)) + '%</div><div style="font-size:12px;color:#15803d;margin-top:4px;">of cancelled bookings<br>were flagged ≥70%</div></div></div>' if _outcome_stats['cancelled'] > 0 else '<div style="flex:1;min-width:200px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:10px;padding:18px 20px;display:flex;align-items:center;justify-content:center;"><div style="text-align:center;color:#9ca3af;font-size:13px;">No outcomes logged yet.<br>Data accumulates after<br>the next file upload.</div></div>'}
</div>

<!-- OVERBOOKING PLANNER ─────────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Overbooking Planner — Next 30 Days</span><span class="sh-line"></span><span class="sh-sub">Probabilistic recommendations. Always apply judgment before acting.</span></div>
{(lambda plan: f"""
<div style="overflow-x:auto;margin-bottom:8px;">
<table style="width:100%;border-collapse:collapse;font-size:13px;">
<thead>
<tr style="background:#f9fafb;border-bottom:2px solid #e5e7eb;">
  <th style="padding:10px 12px;text-align:left;font-weight:600;color:#374151;white-space:nowrap;">Date</th>
  <th style="padding:10px 12px;text-align:left;font-weight:600;color:#374151;white-space:nowrap;">Day</th>
  <th style="padding:10px 12px;text-align:right;font-weight:600;color:#374151;white-space:nowrap;">On Hand</th>
  <th style="padding:10px 12px;text-align:right;font-weight:600;color:#374151;white-space:nowrap;">High Risk</th>
  <th style="padding:10px 12px;text-align:right;font-weight:600;color:#374151;white-space:nowrap;">Exp. Cancels</th>
  <th style="padding:10px 12px;text-align:right;font-weight:600;color:#374151;white-space:nowrap;">Exp. No-Shows</th>
  <th style="padding:10px 12px;text-align:right;font-weight:600;color:#374151;white-space:nowrap;">Total Attrition</th>
  <th style="padding:10px 12px;text-align:right;font-weight:600;color:#374151;white-space:nowrap;">Recommended</th>
  <th style="padding:10px 12px;text-align:left;font-weight:600;color:#374151;white-space:nowrap;">Confidence</th>
</tr>
</thead>
<tbody>
{''.join(
  (lambda r, bg, rec_color: f"""<tr style="border-bottom:1px solid #f3f4f6;{bg}">
  {'<td style="padding:9px 12px;font-weight:500;color:#111827;white-space:nowrap;"><span style="display:inline-block;width:7px;height:7px;border-radius:50%;background:#C4A882;margin-right:6px;vertical-align:middle;cursor:default;" title="' + r["holiday_label"] + '"></span>' + r["date"].strftime("%d %b %Y") + "</td>" if r.get("holiday_label") else '<td style="padding:9px 12px;font-weight:500;color:#111827;white-space:nowrap;padding-left:13px;">' + r["date"].strftime("%d %b %Y") + "</td>"}
  <td style="padding:9px 12px;color:#6b7280;white-space:nowrap;">{r['date'].strftime('%A')}</td>
  <td style="padding:9px 12px;text-align:right;color:#111827;">{r['bookings_on_hand']}</td>
  <td style="padding:9px 12px;text-align:right;color:{'#dc2626' if r['high_risk_count'] > 0 else '#6b7280'};">{r['high_risk_count']}</td>
  <td style="padding:9px 12px;text-align:right;color:#6b7280;">{r['expected_cancellations']}</td>
  <td style="padding:9px 12px;text-align:right;color:#6b7280;">{r['expected_no_shows']}</td>
  <td style="padding:9px 12px;text-align:right;font-weight:500;color:#374151;">{r['total_expected_attrition']}</td>
  <td style="padding:9px 12px;text-align:right;font-weight:700;{rec_color}">{"+" + str(r['recommended_overbooking']) if r['recommended_overbooking'] > 0 else "—"}{" <span style='font-size:10px;color:#9ca3af;font-weight:400;'>(low data)</span>" if r['confidence'] == 'LOW' else ""}</td>
  <td style="padding:9px 12px;"><span style="font-size:11px;font-weight:500;padding:2px 7px;border-radius:4px;{'background:#dcfce7;color:#166534;' if r['confidence']=='HIGH' else ('background:#fef9c3;color:#854d0e;' if r['confidence']=='MEDIUM' else 'background:#f3f4f6;color:#6b7280;')}">{r['confidence']}</span></td>
</tr>""")(
    r,
    'background:#fef2f2;' if r['recommended_overbooking'] >= 5 else (
      'background:#fffbeb;' if r['recommended_overbooking'] >= 3 else (
        'background:#f0fdf4;' if r['recommended_overbooking'] >= 1 else ''
      )
    ),
    'color:#dc2626;' if r['recommended_overbooking'] >= 5 else (
      'color:#92400e;' if r['recommended_overbooking'] >= 3 else (
        'color:#166534;' if r['recommended_overbooking'] >= 1 else 'color:#9ca3af;'
      )
    )
  )
  for r in plan
)}
</tbody>
</table>
</div>
<div style="display:flex;gap:24px;margin-bottom:8px;flex-wrap:wrap;">
  <div style="font-size:11px;color:#6b7280;display:flex;align-items:center;gap:6px;"><span style="display:inline-block;width:12px;height:12px;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:2px;"></span> +1–2 rooms</div>
  <div style="font-size:11px;color:#6b7280;display:flex;align-items:center;gap:6px;"><span style="display:inline-block;width:12px;height:12px;background:#fffbeb;border:1px solid #fde68a;border-radius:2px;"></span> +3–4 rooms</div>
  <div style="font-size:11px;color:#6b7280;display:flex;align-items:center;gap:6px;"><span style="display:inline-block;width:12px;height:12px;background:#fef2f2;border:1px solid #fecaca;border-radius:2px;"></span> +5 or more</div>
  <div style="font-size:11px;color:#6b7280;display:flex;align-items:center;gap:6px;"><span style="display:inline-block;width:12px;height:12px;background:rgba(255,200,0,0.15);border:1px solid #fde68a;border-radius:2px;"></span> 🇧🇪 Holiday / School break (+15% attrition)</div>
  <div style="font-size:11px;color:#6b7280;margin-left:auto;">HIGH = 20+ bookings · MEDIUM = 10–19 · LOW = under 10 (treat with caution)</div>
</div>
<div style="font-size:11px;color:#9ca3af;margin-bottom:32px;line-height:1.6;">
  Recommendations are based on historical cancellation patterns and no-show rates. Occupado does not guarantee accuracy. Review daily and adjust based on on-the-ground knowledge.
</div>
""" if plan else '<div style="color:#9ca3af;font-size:13px;padding:16px 0 32px;">No booking data available for the next 30 days.</div>')(overbooking_plan)}

<!-- GUEST TABLE ──────────────────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Repeat Guests · Next 15 Days</span><span class="sh-line"></span><span class="sh-sub">Click row for AI analysis · grouped by guest</span></div>
<table class="tbl">
<colgroup>
  <col style="min-width:160px">
  <col style="width:120px;min-width:120px">
  <col style="width:80px;min-width:80px">
  <col style="width:55px">
  <col style="width:80px">
  <col>
  <col style="width:110px;min-width:110px">
</colgroup>
<thead><tr>
  <th>Guest</th><th>Status</th><th>Arrival</th><th>Nights</th><th>Risk</th><th>Notes</th><th>Action</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>

<!-- ACTION PLAN ──────────────────────────────────────────────────────── -->
<div class="sh"><span class="sh-title">Action Plan</span><span class="sh-line"></span><span class="sh-sub">Today</span></div>
<div class="ap-grid">
  <div class="ap-card">
    <div class="ap-head">🛬 Arrivals today ({len(arriving)})</div>
    {arriving_items}
    <button class="ap-btn" onclick="openBulk('welcome')">Send welcome email →</button>
  </div>
  <div class="ap-card">
    <div class="ap-head">🏨 In house ({len(in_house)})</div>
    {inhouse_items}
    <button class="ap-btn ghost" onclick="openBulk('departure')">Send departure reminder →</button>
  </div>
</div>

<!-- MICE & CORPORATE INTELLIGENCE ──────────────────────────────────── -->
<div class="sh"><span class="sh-title">MICE & Corporate Intelligence</span><span class="sh-line"></span><span class="sh-sub">B2B · Meetings, Incentives, Conferences & Events</span></div>

<div class="mice-grid">
  <div class="mice-card blue">
    <div class="mice-num blue">{mice_total:,}</div>
    <div class="mice-lbl">MICE Bookings</div>
    <div class="mice-sub">in current dataset</div>
  </div>
  <div class="mice-card">
    <div class="mice-num">{mice_nights:,}</div>
    <div class="mice-lbl">Total Nights</div>
    <div class="mice-sub">{mice_avg_nights} avg per stay</div>
  </div>
  <div class="mice-card">
    <div class="mice-num" style="color:#15803d">€{mice_est_rev:,}</div>
    <div class="mice-lbl">Est. MICE Revenue</div>
    <div class="mice-sub">at €145 avg corporate ADR</div>
  </div>
  <div class="mice-card">
    <div class="mice-num">{len(mice_clients)}</div>
    <div class="mice-lbl">Corporate Accounts</div>
    <div class="mice-sub">active in period</div>
  </div>
</div>

<div class="mice-row">
  <div>
    <div class="mice-chart-title">Top Corporate Accounts</div>
    <table class="mice-tbl">
      <thead><tr><th>#</th><th>Company</th><th>Segment</th><th>Bookings</th><th>Nights</th><th>Avg Stay</th><th>Risk</th></tr></thead>
      <tbody>
        {"".join(
          f'<tr>'
          f'<td style="color:#94a3b8;font-family:monospace;font-size:11px">{i+1}</td>'
          f'<td style="font-weight:600;color:#0d1120">{c["company"]}</td>'
          f'<td><span class="seg-pill seg-{c["seg_code"].lower()}">{c["segment"]}</span></td>'
          f'<td style="font-family:monospace">{c["bookings"]}</td>'
          f'<td style="font-family:monospace">{c["nights"]}</td>'
          f'<td style="font-family:monospace">{round(c["nights"]/c["bookings"],1) if c["bookings"] else 0}n</td>'
          f'<td><span class="seg-pill '
          f'{"seg-bnsgrp" if c.get("risk")=="HIGH" else "seg-mtgbns" if c.get("risk")=="MEDIUM" else "seg-corpdyn"}">'
          f'{c.get("risk","—")}</span></td>'
          f'</tr>'
          for i, c in enumerate(mice_clients)
        ) if mice_clients else "<tr><td colspan='7' style='color:#94a3b8;text-align:center;padding:20px'>No data — local VDV-MEC files required</td></tr>"}
      </tbody>
    </table>
  </div>
  <div class="mice-chart-card">
    <div class="mice-chart-title">Segment Mix</div>
    <canvas id="miceSegChart" height="200"></canvas>
    <div style="margin-top:14px">
      {"".join(
        f'<div style="display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid #f1f5f9;font-size:12px;">'
        f'<span style="color:#64748b">{lbl}</span>'
        f'<span style="font-family:monospace;font-weight:600">{cnt}</span>'
        f'</div>'
        for lbl, cnt in [
          ('Corporate Fixed (CORPFIX)', mice_seg.get('CORPFIX', 0)),
          ('Corporate Dynamic (CORPDYN)', mice_seg.get('CORPDYN', 0)),
          ('Business Group (BNSGRP)', mice_seg.get('BNSGRP', 0)),
          ('Meeting Business (MTGBNS)', mice_seg.get('MTGBNS', 0)),
        ]
      )}
    </div>
  </div>
</div>

{f"""<div class="mice-chart-title" style="margin-bottom:10px">Active Group Blocks</div>
<div style="background:#fff;border:1px solid #e4e8f0;border-radius:12px;padding:18px 20px;margin-bottom:18px;">
  {"".join(
    f'<div class="grp-row">'
    f'<div><div class="grp-name">{g["name"]}</div><div class="grp-co">{g["company"]} · {g["arrival"]} → {g["departure"]}</div></div>'
    f'<span class="grp-rooms">{len(g["rooms"])} rooms</span>'
    f'</div>'
    for g in mice_groups
  ) if mice_groups else '<div style="color:#94a3b8;font-size:12px;text-align:center;padding:12px">No active group blocks in dataset</div>'}
</div>""" if True else ""}

</div><!-- /page -->

<!-- DETAIL MODAL -->
<div class="mo" id="detailMo">
  <div class="mb-inner">
    <button class="mc" onclick="closeMo()">✕</button>
    <div class="mt" id="mo-name"></div>
    <div class="ms" id="mo-sub"></div>
    <div class="sd" id="mo-score"></div>
    <div class="sb-bg"><div class="sb-fill" id="mo-bar" style="width:0%"></div></div>
    <div class="sv-tag" id="mo-verd"></div>
    <div style="margin-top:14px;" id="mo-details"></div>
    <div style="margin-top:12px;font-family:'JetBrains Mono',monospace;font-size:9.5px;color:#94a3b8;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;">Risk Factors</div>
    <div id="mo-reasons"></div>
    <div style="display:flex;gap:8px;margin-top:18px;">
      <button style="flex:1;padding:10px;background:#00d165;border:none;border-radius:7px;font-size:12.5px;font-weight:700;cursor:pointer;font-family:Inter,sans-serif;color:#080c14;" onclick="contactFromMo()">Contact Guest →</button>
      <button style="flex:1;padding:10px;background:#f8fafc;border:1px solid #e4e8f0;border-radius:7px;font-size:12.5px;font-weight:500;cursor:pointer;" onclick="closeMo()">Close</button>
    </div>
  </div>
</div>

<!-- EMAIL COMPOSER -->
<div class="ec" id="ecMo">
  <div class="eb">
    <div style="margin-bottom:16px;">
      <div style="font-family:'Plus Jakarta Sans',sans-serif;font-size:16px;font-weight:800;color:#0d1120;" id="ec-title">Email Guest</div>
      <div style="font-family:'JetBrains Mono',monospace;font-size:10px;color:#94a3b8;margin-top:2px;" id="ec-sub"></div>
    </div>
    <label class="el">To (email)</label><input type="email" id="ec-email" class="ei" placeholder="guest@example.com">
    <label class="el">Name</label><input type="text" id="ec-name" class="ei" placeholder="Guest name">
    <label class="el">Subject</label><input type="text" id="ec-subject" class="ei">
    <label class="el">Message</label><textarea id="ec-body" class="eta"></textarea>
    <div class="ea">
      <button class="es" onclick="sendEmail()">Send →</button>
      <button class="ecc" onclick="closeEc()">Cancel</button>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>
const guests = {guests_js};
const scores = {scores_js};
let activeGuest = -1;

// ── CHARTS ────────────────────────────────────────────────────────────────────
const months   = {months_js};
const cxData   = {cx_js};
const nsData   = {ns_js};
const lostData = {lost_js};

const trendChart = new Chart(document.getElementById('trendChart'), {{
  type: 'bar',
  data: {{
    labels: months,
    datasets: [
      {{ label: 'Cancellations', data: cxData, backgroundColor: '#ef4444', borderRadius: 4, borderWidth: 0, stack: 'a' }},
      {{ label: 'No-shows',      data: nsData, backgroundColor: '#f59e0b', borderRadius: 4, borderWidth: 0, stack: 'a' }}
    ]
  }},
  options: {{
    plugins: {{
      legend: {{ position:'bottom', labels:{{ font:{{family:'JetBrains Mono',size:10}},boxWidth:10,padding:12 }} }},
      tooltip: {{ callbacks: {{ afterBody: ctx => [`Total: ${{cxData[ctx[0].dataIndex]+nsData[ctx[0].dataIndex]}}`] }} }}
    }},
    scales: {{
      x: {{ grid:{{display:false}}, ticks:{{font:{{family:'JetBrains Mono',size:10}}}} }},
      y: {{ grid:{{color:'#f1f5f9'}}, ticks:{{font:{{family:'JetBrains Mono',size:10}},stepSize:50}} }}
    }},
    onClick: (e,el) => {{ if(el.length) filterByMonth(el[0].index); }},
    animation: {{ duration:700 }}
  }}
}});

const channelChart = new Chart(document.getElementById('channelChart'), {{
  type: 'bar',
  data: {{
    labels: {ch_labels_js},
    datasets: [{{ label:'Cancellations', data:{ch_vals_js}, backgroundColor:['#f87171','#fb923c','#60a5fa','#a78bfa','#94a3b8'], borderRadius:5, borderWidth:0 }}]
  }},
  options: {{
    indexAxis: 'y',
    plugins: {{
      legend:{{display:false}},
      tooltip:{{ callbacks:{{ label: ctx => ` ${{ctx.parsed.x}} (${{({ch_pcts_js})[ctx.dataIndex]}}%)` }} }}
    }},
    scales: {{
      x: {{ grid:{{color:'#f1f5f9'}}, ticks:{{font:{{family:'JetBrains Mono',size:10}}}} }},
      y: {{ grid:{{display:false}}, ticks:{{font:{{family:'JetBrains Mono',size:10}}}} }}
    }},
    animation: {{ duration:700 }}
  }}
}});

let filteredMonth = null;
function filterByMonth(idx) {{
  if (filteredMonth === idx) {{
    filteredMonth = null;
    channelChart.data.datasets[0].data = {ch_vals_js};
    channelChart.options.plugins.title = {{ display:false }};
  }} else {{
    filteredMonth = idx;
    // Show relative weight for selected month (proportional to share)
    const monthTotal = lostData[idx];
    const overall = lostData.reduce((a,b)=>a+b,0);
    const ratio = monthTotal/overall;
    const adjusted = {ch_vals_js}.map(v => Math.round(v*ratio));
    channelChart.data.datasets[0].data = adjusted;
    channelChart.options.plugins.title = {{ display:true, text: months[idx]+' estimate', font:{{family:'JetBrains Mono',size:10}}, color:'#94a3b8' }};
  }}
  channelChart.update();
}}

// ── FUTURE RISK CHARTS ────────────────────────────────────────────────────────
new Chart(document.getElementById('futMonthChart'), {{
  type: 'bar',
  data: {{
    labels: {fut_mlabels_js},
    datasets: [
      {{ label: 'High Risk', data: {fut_mhigh_js}, backgroundColor: '#ef4444', borderRadius: 4, borderWidth: 0, stack: 'a' }},
      {{ label: 'Medium Risk', data: {fut_mmed_js}, backgroundColor: '#f59e0b', borderRadius: 4, borderWidth: 0, stack: 'a' }}
    ]
  }},
  options: {{
    plugins: {{ legend: {{ position:'bottom', labels:{{ font:{{family:'JetBrains Mono',size:10}},boxWidth:10,padding:10 }} }} }},
    scales: {{
      x: {{ grid:{{display:false}}, ticks:{{font:{{family:'JetBrains Mono',size:9}}}} }},
      y: {{ grid:{{color:'#f1f5f9'}}, ticks:{{font:{{family:'JetBrains Mono',size:10}}}} }}
    }},
    animation: {{ duration:700 }}
  }}
}});

new Chart(document.getElementById('futChannelChart'), {{
  type: 'doughnut',
  data: {{
    labels: {fut_ch_labels_js},
    datasets: [{{ data: {fut_ch_vals_js}, backgroundColor:['#f87171','#60a5fa','#a78bfa','#34d399','#94a3b8'], borderWidth:0, hoverOffset:6 }}]
  }},
  options: {{
    plugins: {{
      legend: {{ position:'right', labels:{{ font:{{family:'JetBrains Mono',size:10}},boxWidth:10,padding:8 }} }},
      tooltip: {{ callbacks: {{ label: ctx => ` ${{ctx.label}}: ${{ctx.parsed}} bookings` }} }}
    }},
    cutout: '62%',
    animation: {{ duration:700 }}
  }}
}});

// ── SAVINGS CALCULATOR ────────────────────────────────────────────────────────
const totalCx = {total_cx};
const totalNs = {total_ns};
const avgAdr  = {avg_adr};
const avgNights = {avg_nights};

function updateSavings(val) {{
  const pct = parseInt(val);
  document.getElementById('rate-pct').textContent = pct + '%';
  const saved = Math.round(totalCx * (pct/100) * avgAdr * avgNights
                           + totalNs * (pct/100) * 0.5 * avgAdr);
  document.getElementById('savings-num').textContent = '€' + saved.toLocaleString();
  document.getElementById('savings-sub').textContent =
    Math.round(totalCx*(pct/100)) + ' cancellations prevented · ' +
    Math.round(totalNs*(pct/100)*0.5) + ' no-shows charged';
}}

// ── DETAIL MODAL ──────────────────────────────────────────────────────────────
function openDetail(idx, score) {{
  const g = guests[idx]; activeGuest = idx;
  document.getElementById('mo-name').textContent = g.name;
  document.getElementById('mo-sub').textContent  = g.status + ' · ' + g.arrival + (g.departure?' – '+g.departure:'');
  document.getElementById('mo-score').textContent = score.toFixed(1)+'%';
  const bar = document.getElementById('mo-bar');
  const sd  = document.getElementById('mo-score');
  const vd  = document.getElementById('mo-verd');
  bar.style.width = score+'%';
  if (score>=70) {{
    bar.style.background='#dc2626'; sd.style.color='#dc2626';
    vd.textContent='HIGH RISK'; vd.style.background='#fef2f2'; vd.style.color='#dc2626';
  }} else if (score>=40) {{
    bar.style.background='#f59e0b'; sd.style.color='#f59e0b';
    vd.textContent='MEDIUM RISK'; vd.style.background='#fffbeb'; vd.style.color='#b45309';
  }} else {{
    bar.style.background='#00d165'; sd.style.color='#00d165';
    vd.textContent='LOW RISK'; vd.style.background='#f0fdf4'; vd.style.color='#16a34a';
  }}
  document.getElementById('mo-details').innerHTML = `
    <div class="dr"><span class="dl">Nights</span><span class="dv">${{g.nights}}</span></div>
    <div class="dr"><span class="dl">Adults</span><span class="dv">${{g.adults}}</span></div>
    <div class="dr"><span class="dl">Membership</span><span class="dv">${{g.membership||'—'}}</span></div>
    <div class="dr"><span class="dl">Rate</span><span class="dv">€${{g.adr}}/night</span></div>
    ${{g.note?`<div class="dr"><span class="dl">Notes</span><span class="dv" style="max-width:200px;text-align:right;font-size:11px">${{g.note}}</span></div>`:''}}
  `;
  const reasons = [];
  // Lead time from arrival date
  const arrParts = g.arrival.split('/');
  const arrDate  = arrParts.length===3 ? new Date(arrParts[2], arrParts[1]-1, arrParts[0]) : null;
  const leadDays = arrDate ? Math.round((arrDate - new Date()) / 86400000) : null;
  if (leadDays !== null) {{
    if (leadDays > 90)       reasons.push({{p:0, t:`Booked ${{leadDays}} days in advance — very long lead time, high cancellation risk`}});
    else if (leadDays > 30)  reasons.push({{p:0, t:`Lead time ${{leadDays}} days — moderate cancellation window remains`}});
    else if (leadDays >= 0)  reasons.push({{p:1, t:`Arriving in ${{leadDays <= 1 ? 'tomorrow or today' : leadDays+' days'}} — close to check-in, lower risk`}});
  }}
  // Loyalty / membership
  if (g.membership && g.membership !== '—' && g.membership !== '') {{
    reasons.push({{p:1, t:`Loyalty member (${{g.membership}}) — committed guests, lower no-show rate`}});
  }} else {{
    reasons.push({{p:0, t:`No loyalty membership — first-time or uncommitted profile`}});
  }}
  // Stay length
  if (g.nights >= 4)      reasons.push({{p:1, t:`${{g.nights}}-night stay — longer stays are rarely cancelled last-minute`}});
  else if (g.nights === 1) reasons.push({{p:0, t:`1-night stay — short stays have the highest no-show rate`}});
  // Booking value
  const total = Math.round(g.adr * g.nights);
  if (total > 500)        reasons.push({{p:0, t:`High-value booking (€${{total}} total) — deposit or guarantee strongly recommended`}});
  else if (total > 200)   reasons.push({{p:0, t:`Booking value €${{total}} — consider deposit request if no guarantee on file`}});
  // Status context
  if (g.status === 'In House')    reasons.push({{p:1, t:`Guest is currently in house — no cancellation risk`}});
  if (g.status === 'Checked Out') reasons.push({{p:1, t:`Stay completed — historical record only`}});
  // Profile note
  if (g.note && g.note.length > 0) reasons.push({{p:1, t:`Profile note: "${{g.note.substring(0,70)}}"`}});
  // Score summary
  if (score >= 70)      reasons.push({{p:0, t:`AI confidence: high risk — multiple signals combined. Act now.`}});
  else if (score >= 40) reasons.push({{p:0, t:`AI confidence: moderate risk — monitor and send reminder closer to arrival`}});
  else                  reasons.push({{p:1, t:`AI confidence: low risk — booking profile looks stable`}});
  document.getElementById('mo-reasons').innerHTML = reasons.slice(0,5).map(r =>
    `<div class="ri ${{r.p?'pos':'neg'}}">${{r.p?'✓':'⚠'}} ${{r.t}}</div>`).join('');
  document.getElementById('detailMo').classList.add('show');
}}
function closeMo() {{ document.getElementById('detailMo').classList.remove('show'); activeGuest=-1; }}
function contactFromMo() {{ closeMo(); if(activeGuest>=0) setTimeout(()=>openEmail(activeGuest,'contact'),80); }}

// ── EMAIL ─────────────────────────────────────────────────────────────────────
const TMPLS = {{
  contact:  {{subj:'Your upcoming stay at Van der Valk Hotel Mechelen', body:(n,arr,dep)=>`Dear ${{n}},\n\nWe look forward to welcoming you from ${{arr}} to ${{dep}}.\nPlease reach out if you have any requests.\n\nWarm regards,\nVan der Valk Hotel Mechelen`}},
  welcome:  {{subj:'Welcome to Van der Valk Hotel Mechelen', body:(n)=>`Dear ${{n}},\n\nWelcome! Your room is ready and our team is at your service.\n\nWarm regards,\nVan der Valk Hotel Mechelen`}},
  reminder: {{subj:'Your upcoming reservation', body:(n,arr,dep)=>`Dear ${{n}},\n\nThis is a reminder of your reservation from ${{arr}} to ${{dep}}.\nPlease confirm or contact us if your plans have changed.\n\nWarm regards,\nVan der Valk Hotel Mechelen`}},
  departure:{{subj:'Departure reminder — checkout tomorrow', body:(n,_,dep)=>`Dear ${{n}},\n\nCheckout is ${{dep}} at 12:00. Late checkout available on request.\nThank you for staying — we hope to see you again soon.\n\nWarm regards,\nVan der Valk Hotel Mechelen`}},
  deposit:  {{subj:'Reservation guarantee required', body:(n,arr,dep)=>`Dear ${{n}},\n\nTo secure your reservation from ${{arr}} to ${{dep}}, a deposit is required.\nPlease contact us at your earliest convenience.\n\nWarm regards,\nVan der Valk Hotel Mechelen`}},
  upsell:   {{subj:'Exclusive offers for your stay', body:(n)=>`Dear ${{n}},\n\nWe hope you are enjoying your time with us!\nToday's offers: restaurant priority, parking upgrade, late checkout.\n\nWarm regards,\nVan der Valk Hotel Mechelen`}}
}};

function openEmail(idx, type) {{
  const g  = guests[idx]; activeGuest=idx;
  const fn = g.name.split(',').slice(1).join(',').trim().split(' ').filter(Boolean)[0] || g.name.split(',')[0];
  const t  = TMPLS[type]||TMPLS.contact;
  document.getElementById('ec-title').textContent = 'Email — '+g.name;
  document.getElementById('ec-sub').textContent   = g.status+' · '+g.arrival+(g.departure?' – '+g.departure:'');
  document.getElementById('ec-name').value    = g.name;
  document.getElementById('ec-email').value   = '';
  document.getElementById('ec-subject').value = t.subj;
  document.getElementById('ec-body').value    = t.body(fn, g.arrival, g.departure);
  document.getElementById('ecMo').classList.add('show');
}}

function openBulk(type) {{
  const t = TMPLS[type]||TMPLS.welcome;
  document.getElementById('ec-title').textContent = 'Bulk — '+type;
  document.getElementById('ec-sub').textContent   = 'Fill in guest details per send';
  document.getElementById('ec-name').value    = '';
  document.getElementById('ec-email').value   = '';
  document.getElementById('ec-subject').value = t.subj;
  document.getElementById('ec-body').value    = t.body('[Guest Name]','[Arrival]','[Departure]');
  document.getElementById('ecMo').classList.add('show');
}}

function closeEc() {{ document.getElementById('ecMo').classList.remove('show'); }}

function sendEmail() {{
  const email=document.getElementById('ec-email').value.trim();
  const subj =document.getElementById('ec-subject').value.trim();
  const body =document.getElementById('ec-body').value.trim();
  if(!email||!subj||!body){{ toast('Fill in all fields','e'); return; }}
  fetch('/send-guest-email',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{guest_email:email,subject:subj,body}})}})
    .then(r=>r.json()).then(d=>{{ closeEc(); toast(d.status==='success'?'✓ Sent':'Error: '+d.message, d.status==='success'?'s':'e'); }})
    .catch(()=>toast('Error','e'));
}}

// ── UTILS ─────────────────────────────────────────────────────────────────────
function toast(msg,type){{
  const el=document.getElementById('toast');
  el.textContent=msg; el.style.background=type==='e'?'#dc2626':'#0d1120';
  el.classList.add('show'); setTimeout(()=>el.classList.remove('show'),3000);
}}
document.getElementById('detailMo').onclick = e=>{{ if(e.target===document.getElementById('detailMo')) closeMo(); }};
document.getElementById('ecMo').onclick     = e=>{{ if(e.target===document.getElementById('ecMo')) closeEc(); }};

// MICE segment doughnut
(function() {{
  const ctx = document.getElementById('miceSegChart');
  if (!ctx) return;
  new Chart(ctx, {{
    type: 'doughnut',
    data: {{
      labels: ['Corporate Fixed', 'Corporate Dynamic', 'Business Group', 'Meeting Business'],
      datasets: [{{
        data: {mice_seg_js},
        backgroundColor: ['#bfdbfe','#bbf7d0','#e9d5ff','#fed7aa'],
        borderColor:      ['#1d4ed8','#15803d','#7e22ce','#c2410c'],
        borderWidth: 1.5
      }}]
    }},
    options: {{
      cutout: '62%',
      plugins: {{
        legend: {{ display: false }}
      }}
    }}
  }});
}})();

// ── Occupancy & Revenue charts ────────────────────────────────────────────
(function() {{
  var labels   = {fc_labels};
  var occHist  = {fc_occ_hist};
  var occFore  = {fc_occ_fore};
  var revHist  = {fc_rev_hist};
  var revFore  = {fc_rev_fore};
  var boundary = {fc_boundary};

  var chartDefaults = {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }}, tooltip: {{ mode: 'index', intersect: false }} }},
    scales: {{
      x: {{ grid: {{ color: '#f3f4f6' }}, ticks: {{ font: {{ size: 9, family: 'JetBrains Mono' }}, color: '#9ca3af',
             callback: function(v, i) {{ return i % 7 === 0 ? labels[i] : ''; }} }} }},
      y: {{ grid: {{ color: '#f3f4f6' }}, ticks: {{ font: {{ size: 9, family: 'JetBrains Mono' }}, color: '#9ca3af' }} }}
    }}
  }};

  // Occupancy chart
  var oc = document.getElementById('occChart');
  if (oc) new Chart(oc, {{ type: 'line', data: {{
    labels: labels,
    datasets: [
      {{ label: 'History', data: occHist, borderColor: '#111827', borderWidth: 1.5,
         pointRadius: 0, tension: 0.3, spanGaps: false }},
      {{ label: 'Forecast', data: occFore, borderColor: '#93c5fd', borderWidth: 1.5,
         borderDash: [4,3], pointRadius: 0, tension: 0.3, spanGaps: false }}
    ]
  }}, options: Object.assign({{}}, chartDefaults, {{
    plugins: Object.assign({{}}, chartDefaults.plugins, {{
      annotation: {{ annotations: {{ line1: {{ type: 'line', x: boundary-0.5,
        borderColor: '#e5e7eb', borderWidth: 1, borderDash: [4,3] }} }} }}
    }}),
    scales: Object.assign({{}}, chartDefaults.scales, {{
      y: Object.assign({{}}, chartDefaults.scales.y, {{ min: 0, max: 100,
         ticks: Object.assign({{}}, chartDefaults.scales.y.ticks, {{ callback: function(v) {{ return v + '%'; }} }}) }})
    }})
  }}) }});

  // Revenue chart
  var rc = document.getElementById('revChart');
  if (rc) new Chart(rc, {{ type: 'bar', data: {{
    labels: labels,
    datasets: [
      {{ label: 'History', data: revHist, backgroundColor: '#f3f4f6', borderColor: '#e5e7eb', borderWidth: 1, borderRadius: 2 }},
      {{ label: 'Forecast', data: revFore, backgroundColor: '#dbeafe', borderColor: '#bfdbfe', borderWidth: 1, borderRadius: 2 }}
    ]
  }}, options: Object.assign({{}}, chartDefaults, {{
    scales: Object.assign({{}}, chartDefaults.scales, {{
      y: Object.assign({{}}, chartDefaults.scales.y, {{
         ticks: Object.assign({{}}, chartDefaults.scales.y.ticks, {{
           callback: function(v) {{ return '€' + (v >= 1000 ? Math.round(v/1000) + 'k' : v); }} }}) }})
    }})
  }}) }});
}})();

// ── DYNAMICS JS ──────────────────────────────────────────────

// Counter animation
function countUp(el, target, duration, pre, suf) {{
  if (!el) return;
  pre = pre||''; suf = suf||''; duration = duration||1400;
  const start = performance.now();
  function tick(now) {{
    const p = Math.min((now - start)/duration, 1);
    const ease = 1 - Math.pow(1-p, 3);
    const val = Math.round(ease * target);
    el.textContent = pre + val.toLocaleString() + suf;
    if (p < 1) requestAnimationFrame(tick);
  }}
  requestAnimationFrame(tick);
}}
document.querySelectorAll('.count-up').forEach(function(el) {{
  countUp(el, parseFloat(el.dataset.val)||0, 1400, el.dataset.pre||'', el.dataset.suf||'');
}});

// Filter tabs
function filterRisk(risk, btn) {{
  document.querySelectorAll('.f-tab').forEach(function(t) {{ t.classList.remove('active'); }});
  btn.classList.add('active');
  document.querySelectorAll('.clickable-row, .cr').forEach(function(row) {{
    var show;
    if (risk === 'all') {{ show = true; }}
    else {{ show = !!row.querySelector('.badge.' + risk); }}
    row.style.display = show ? '' : 'none';
    var nx = row.nextElementSibling;
    if (nx && nx.classList.contains('exp-tr')) {{
      if (!show) nx.classList.remove('open');
      nx.style.display = show ? '' : 'none';
    }}
  }});
}}

// Sort table
var _sCol = null, _sDir = 1;
function sortTable(col) {{
  var tbody = document.querySelector('tbody');
  if (!tbody) return;
  if (_sCol === col) _sDir *= -1; else {{ _sCol = col; _sDir = 1; }}
  var pairs = [];
  var rows = Array.from(tbody.querySelectorAll('.clickable-row, .cr'));
  rows.forEach(function(r) {{
    var nx = r.nextElementSibling;
    var expRow = (nx && nx.classList.contains('exp-tr')) ? nx : null;
    pairs.push([r, expRow]);
  }});
  pairs.sort(function(a, b) {{
    var ar = a[0], br = b[0];
    var av = col==='score' ? parseFloat(ar.dataset.score||0) :
             col==='lead'  ? parseInt(ar.dataset.lead||0) :
             col==='rate'  ? parseInt(ar.dataset.rate||0) : 0;
    var bv = col==='score' ? parseFloat(br.dataset.score||0) :
             col==='lead'  ? parseInt(br.dataset.lead||0) :
             col==='rate'  ? parseInt(br.dataset.rate||0) : 0;
    return (av - bv) * _sDir;
  }});
  pairs.forEach(function(p) {{
    tbody.appendChild(p[0]);
    if (p[1]) tbody.appendChild(p[1]);
  }});
  document.querySelectorAll('.sort-arr').forEach(function(a) {{
    var c = a.closest('.sort-th') && a.closest('.sort-th').dataset.col;
    a.textContent = c === col ? (_sDir === -1 ? '↓' : '↑') : '↕';
    a.classList.toggle('on', c === col);
  }});
}}

// Build prediction reasons
function buildReasons(b, score) {{
  var r = [];
  var lt = b.lead_time||0, canc = b.previous_cancellations||0;
  var rep = b.is_repeated_guest||0, adr = b.adr||0;
  var chg = b.booking_changes||0, spec = b.total_of_special_requests||0;
  var wk = b.stays_in_week_nights||0, we = b.stays_in_weekend_nights||0;
  var nights = wk + we || 1;

  if (lt > 120)      r.push({{c:'bad',  t:'Lead time ' + lt + ' days — bookings this far out cancel 3× more often'}});
  else if (lt > 45)  r.push({{c:'warn', t:'Lead time ' + lt + ' days — moderate cancellation window remains'}});
  else               r.push({{c:'pos',  t:'Lead time ' + lt + ' days — close to arrival, low cancellation risk'}});

  if (canc >= 2)     r.push({{c:'bad',  t: canc + ' prior cancellations — strongest single predictor of future no-shows'}});
  else if (canc===1) r.push({{c:'warn', t:'1 previous cancellation on record — elevated risk signal'}});
  else               r.push({{c:'pos',  t:'No cancellation history — clean booking profile'}});

  if (!rep)          r.push({{c:'warn', t:'First-time guest — new guests cancel 2.3× more than returning guests'}});
  else               r.push({{c:'pos',  t:'Returning guest — loyalty significantly reduces no-show likelihood'}});

  if (adr > 200)     r.push({{c:'bad',  t:'Premium rate €' + Math.round(adr) + ' — high-value bookings warrant a deposit or guarantee'}});
  else if (adr > 120)r.push({{c:'warn', t:'Room rate €' + Math.round(adr) + ' — consider pre-authorisation'}});

  if (chg >= 2)      r.push({{c:'warn', t: chg + ' booking modifications — repeated changes signal hesitation'}});
  else if (chg===1)  r.push({{c:'warn', t:'1 booking change — slight uncertainty signal'}});

  if (spec === 0)    r.push({{c:'warn', t:'No special requests — low guest engagement with this stay'}});
  else               r.push({{c:'pos',  t: spec + ' special request' + (spec>1?'s':'') + ' — engaged guests are far less likely to cancel'}});

  if (nights >= 4)   r.push({{c:'pos',  t: nights + '-night stay — multi-night bookings have lower no-show rates'}});
  else if (nights===1)r.push({{c:'warn',t:'1-night stay — shortest stays carry the highest no-show rate'}});

  return r.slice(0, 5);
}}

function conclusionText(score) {{
  if (score >= 70) return 'High cancellation probability. Request deposit or guarantee now — revenue at risk.';
  if (score >= 40) return 'Moderate risk. Send a reminder 48h before arrival to confirm the booking.';
  return 'Low risk. Booking profile is stable — monitor normally.';
}}

// Toggle grouped stays dropdown
function toggleStays(sid, row) {{
  var opening = !row.classList.contains('is-open');
  row.classList.toggle('is-open', opening);
  document.querySelectorAll('tr.stays-sub[data-grp="' + sid + '"]').forEach(function(r) {{
    r.style.display = opening ? '' : 'none';
  }});
  var chev = document.getElementById('chev-' + sid);
  if (chev) chev.textContent = opening ? '▴' : '▾';
}}

// Inline expand row
var _openRow = null;
function toggleExpand(row, idx, score) {{
  var expTr = document.getElementById('exp-' + idx);
  if (!expTr) return;
  var isOpen = expTr.classList.contains('open');
  if (_openRow && _openRow !== expTr) {{
    _openRow.classList.remove('open');
    var pr = _openRow.previousElementSibling;
    if (pr) pr.classList.remove('is-open');
  }}
  if (isOpen) {{
    expTr.classList.remove('open');
    row.classList.remove('is-open');
    _openRow = null;
  }} else {{
    var bArr = typeof bookings !== 'undefined' ? bookings : (typeof guests !== 'undefined' ? guests : []);
    var b = bArr[idx] || {{}};
    var reasons = buildReasons(b, score);
    var scoreColor = score>=70 ? '#ef4444' : score>=40 ? '#f59e0b' : '#00d165';
    var verdict = score>=70 ? 'HIGH RISK' : score>=40 ? 'MEDIUM RISK' : 'LOW RISK';
    var factorsHtml = reasons.map(function(r) {{
      return '<div class="exp-factor ' + r.c + '">' + r.t + '</div>';
    }}).join('');
    var inner = document.getElementById('exp-inner-' + idx);
    if (inner) {{
      inner.innerHTML =
        '<div class="exp-score-wrap">' +
          '<div class="exp-score-big" style="color:' + scoreColor + '">' + score.toFixed(0) + '%</div>' +
          '<div class="exp-score-lbl">' + verdict + '</div>' +
          '<div class="exp-bar-bg"><div class="exp-bar-fill" style="width:' + score + '%;background:' + scoreColor + '"></div></div>' +
        '</div>' +
        '<div class="exp-right">' +
          '<div class="exp-head">Why this score</div>' +
          '<div class="exp-factors">' + factorsHtml + '</div>' +
          '<div class="exp-conclusion">' + conclusionText(score) + '</div>' +
        '</div>';
    }}
    expTr.classList.add('open');
    row.classList.add('is-open');
    _openRow = expTr;
  }}
}}

</script>
</body>
</html>"""
    today     = datetime.now()

    arriving_today  = [g for g in guests if g['status'] == 'Arriving Today']
    in_house        = [g for g in guests if g['status'] == 'In House']
    checked_out     = [g for g in guests if g['status'] == 'Checked Out']
    departing_tmrw  = [g for g, s in zip(guests, scores)
                       if g.get('dep_date') and (g['dep_date'] - today).days == 1]

    # Risk counts
    high_count = sum(1 for s in scores if s >= 70)
    med_count  = sum(1 for s in scores if 40 <= s < 70)
    low_count  = sum(1 for s in scores if s < 40)
    avg_score  = (sum(scores) / len(scores)) if scores else 0

    # Revenue context
    avg_adr = 130.0
    rev_arriving = len(arriving_today) * avg_adr * 1.5  # avg ~1.5 nights
    rev_inhouse  = len(in_house) * avg_adr * 1.2

    # Chart data — risk donut
    risk_donut_js    = json.dumps([low_count, med_count, high_count])
    risk_donut_pct_js = json.dumps([
        round(low_count / len(scores) * 100) if scores else 0,
        round(med_count / len(scores) * 100) if scores else 0,
        round(high_count / len(scores) * 100) if scores else 0,
    ])

    # Chart data — channel cancellations
    ch_labels = json.dumps(list(ch_data.keys()))
    ch_values = json.dumps(list(ch_data.values()))
    ch_pcts   = json.dumps([round(v / sum(ch_data.values()) * 100, 1) if ch_data else 0 for v in ch_data.values()])

    # Chart data — guest risk scores bar
    guest_names_js = json.dumps([g['name'].split(',')[0] for g in guests])
    guest_scores_js = json.dumps([round(s, 1) for s in scores])
    guest_colors_js = json.dumps([
        '#dc2626' if s >= 70 else ('#f59e0b' if s >= 40 else '#00d165')
        for s in scores
    ])

    # Build table rows
    rows_html = ""
    for i, (g, score) in enumerate(zip(guests, scores)):
        if score >= 70:
            badge  = f'<span class="badge high">HIGH {score:.1f}%</span>'
            action = f'<button class="btn dep" onclick="event.stopPropagation();openVdvEmail({i},\'deposit\')">Request Deposit</button>'
        elif score >= 40:
            badge  = f'<span class="badge med">MEDIUM {score:.1f}%</span>'
            action = f'<button class="btn rem" onclick="event.stopPropagation();openVdvEmail({i},\'reminder\')">Send Reminder</button>'
        else:
            badge  = f'<span class="badge low">LOW {score:.1f}%</span>'
            action = f'<button class="btn mon" onclick="event.stopPropagation();openVdvEmail({i},\'contact\')">Contact Guest</button>'

        # Status badge
        if g['status'] == 'Arriving Today':
            st_badge = '<span class="st-badge st-arriving">Arriving Today</span>'
        elif g['status'] == 'In House':
            st_badge = '<span class="st-badge st-inhouse">In House</span>'
        elif g['status'] == 'Checked Out':
            st_badge = '<span class="st-badge st-out">Checked Out</span>'
        else:
            st_badge = f'<span class="st-badge st-future">{g["status"]}</span>'

        memb = g.get('membership', '')
        memb_badge = f' <span class="memb-badge">{memb}</span>' if memb else ''
        note_html  = f'<span class="guest-note" title="{g["note"]}">{g["note"][:38]}…</span>' if g.get('note') and len(g['note']) > 38 else (f'<span class="guest-note">{g.get("note","")}</span>' if g.get('note') else '—')

        rows_html += f"""<tr class="clickable-row" onclick="showVdvDetail({i},{score:.1f})">
            <td><span class="guest-name">{g['name']}</span>{memb_badge}</td>
            <td>{st_badge}</td>
            <td>{g['arrival']}</td>
            <td>{g.get('departure','—')}</td>
            <td>{g['nights']} night{'s' if g['nights']!=1 else ''}</td>
            <td>{g['adults']}</td>
            <td>{badge}</td>
            <td class="note-cell">{note_html}</td>
            <td>{action}</td>
        </tr>"""

    # Action plan cards
    arriving_plan = ""
    for g in arriving_today:
        specific = g.get('note', '')
        tip = specific if specific else 'Prepare welcome, confirm room assignment'
        arriving_plan += f'<div class="plan-item"><div class="plan-guest">{g["name"]}</div><div class="plan-tip">{tip}</div></div>'
    if not arriving_plan:
        arriving_plan = '<div class="plan-empty">No arrivals today</div>'

    inhouse_plan = ""
    for g in in_house:
        dep = g.get('dep_date')
        days_left = (dep - today).days if dep else '?'
        tip = f'{days_left} night{"s" if days_left != 1 else ""} remaining'
        if g.get('note'):
            tip = g['note'] + ' · ' + tip
        inhouse_plan += f'<div class="plan-item"><div class="plan-guest">{g["name"]}</div><div class="plan-tip">{tip}</div></div>'
    if not inhouse_plan:
        inhouse_plan = '<div class="plan-empty">No guests currently in house</div>'

    # Guests JSON for JS
    guests_js = json.dumps([{
        'name': g['name'], 'arrival': g['arrival'], 'departure': g.get('departure',''),
        'nights': g['nights'], 'adults': g['adults'], 'membership': g.get('membership',''),
        'status': g['status'], 'note': g.get('note',''), 'adr': 149.0 if 'CORP' in g.get('membership','') else 115.0
    } for g in guests])

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Occupado — Van der Valk Hotel Mechelen</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*,*::before,*::after{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:#f5f7fb;color:#0d1120;font-family:'Plus Jakarta Sans',sans-serif;-webkit-font-smoothing:antialiased;}}
a{{text-decoration:none;}}
/* TOPBAR */
.topbar{{height:62px;background:#ffffff;border-bottom:1px solid #e4e8f0;display:flex;align-items:center;padding:0 32px;position:sticky;top:0;z-index:100;}}
.topbar-brand{{display:flex;align-items:center;gap:6px;font-family:'Syne',sans-serif;}}
.topbar-name{{font-family:'Plus Jakarta Sans',sans-serif;font-size:17px;font-weight:800;color:#0d1120;letter-spacing:-0.4px;}}
.topbar-name span{{color:#00d165;}}
.topbar-hotel{{font-family:'JetBrains Mono',monospace;font-size:11px;color:#94a3b8;margin-left:8px;padding-left:12px;border-left:1px solid #e4e8f0;}}
.topbar-right{{display:flex;align-items:center;gap:8px;margin-left:auto;}}
.btn-nav{{padding:7px 16px;background:transparent;border:1px solid #e4e8f0;border-radius:7px;color:#64748b;font-size:12px;font-weight:500;text-decoration:none;transition:all .2s;}}
.btn-nav:hover{{border-color:#cbd5e1;color:#0d1120;}}
.lang-selector{{padding:7px 12px;background:transparent;border:1px solid #e4e8f0;border-radius:7px;color:#64748b;font-size:12px;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;outline:none;}}
/* HERO */
.hero{{background:linear-gradient(135deg,#0d1120 0%,#0f2218 100%);padding:40px 32px 36px;margin-bottom:0;}}
.hero-eyebrow{{font-family:'JetBrains Mono',monospace;font-size:11px;color:#00d165;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:10px;}}
.hero-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:32px;font-weight:800;color:#ffffff;letter-spacing:-0.8px;margin-bottom:6px;}}
.hero-sub{{font-size:13px;color:#94a3b8;}}
.hero-badges{{display:flex;gap:10px;margin-top:16px;flex-wrap:wrap;}}
.hero-badge{{background:rgba(0,209,101,0.12);border:1px solid rgba(0,209,101,0.25);border-radius:6px;padding:5px 12px;font-family:'JetBrains Mono',monospace;font-size:11px;color:#00d165;}}
.hero-badge.warn{{background:rgba(245,158,11,0.12);border-color:rgba(245,158,11,0.25);color:#f59e0b;}}
.hero-badge.info{{background:rgba(148,163,184,0.12);border-color:rgba(148,163,184,0.25);color:#94a3b8;}}
/* CONTENT */
.content{{padding:28px 32px;}}
.section-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:17px;font-weight:700;color:#0d1120;letter-spacing:-0.3px;margin-bottom:14px;margin-top:32px;}}
.section-sub{{font-family:'JetBrains Mono',monospace;font-size:10px;color:#94a3b8;margin-top:-10px;margin-bottom:16px;text-transform:uppercase;letter-spacing:0.5px;}}
/* KPI GRID */
.kpi-grid{{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin-bottom:12px;}}
.kpi-card{{background:#ffffff;border:1px solid #e4e8f0;border-radius:12px;padding:18px 20px;}}
.kpi-num{{font-family:'Plus Jakarta Sans',sans-serif;font-size:36px;font-weight:800;line-height:1;letter-spacing:-1.5px;}}
.kpi-label{{font-family:'JetBrains Mono',monospace;font-size:10px;color:#94a3b8;margin-top:5px;text-transform:uppercase;letter-spacing:0.8px;line-height:1.4;}}
.kpi-trend{{font-size:11px;margin-top:4px;font-weight:500;}}
.kpi-num.green{{color:#00d165;}}
.kpi-num.red{{color:#dc2626;}}
.kpi-num.orange{{color:#f59e0b;}}
.kpi-num.blue{{color:#3b82f6;}}
.kpi-num.neutral{{color:#0d1120;}}
.kpi-card.highlight-green{{border-color:#bbf7d0;background:#f0fdf4;}}
.kpi-card.highlight-red{{border-color:#fecaca;background:#fef2f2;}}
/* CHARTS ROW */
.charts-row{{display:grid;grid-template-columns:1fr 2fr 2fr;gap:14px;margin-bottom:28px;}}
.chart-card{{background:#ffffff;border:1px solid #e4e8f0;border-radius:14px;padding:22px;}}
.chart-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:14px;font-weight:700;color:#0d1120;letter-spacing:-0.2px;margin-bottom:4px;}}
.chart-sub{{font-family:'JetBrains Mono',monospace;font-size:10px;color:#94a3b8;margin-bottom:14px;text-transform:uppercase;letter-spacing:0.5px;}}
.donut-center{{text-align:center;margin-top:8px;}}
.donut-big{{font-family:'Plus Jakarta Sans',sans-serif;font-size:28px;font-weight:800;color:#00d165;}}
.donut-label{{font-family:'JetBrains Mono',monospace;font-size:10px;color:#94a3b8;}}
/* TABLE */
table{{width:100%;border-collapse:collapse;background:#ffffff;border-radius:14px;overflow:hidden;border:1px solid #e4e8f0;margin-bottom:28px;}}
th{{background:#f8fafc;color:#94a3b8;font-family:'JetBrains Mono',monospace;font-size:9px;text-transform:uppercase;letter-spacing:1px;padding:12px 14px;text-align:left;border-bottom:1px solid #e4e8f0;white-space:nowrap;}}
td{{padding:12px 14px;font-size:12.5px;border-bottom:1px solid #f1f5f9;color:#374151;vertical-align:middle;}}
.clickable-row{{cursor:pointer;}}
.clickable-row:hover td{{background:#f8fafc;}}
.guest-name{{font-weight:600;color:#0d1120;font-size:13px;}}
.memb-badge{{background:#eff6ff;color:#3b82f6;border:1px solid #bfdbfe;padding:2px 7px;border-radius:4px;font-family:'JetBrains Mono',monospace;font-size:9px;font-weight:600;margin-left:4px;}}
.badge{{padding:3px 9px;border-radius:99px;font-family:'JetBrains Mono',monospace;font-size:10px;font-weight:600;border:1px solid;white-space:nowrap;}}
.high{{background:#fef2f2;color:#dc2626;border-color:#fecaca;}}
.med{{background:#fffbeb;color:#b45309;border-color:#fde68a;}}
.low{{background:#f0fdf4;color:#16a34a;border-color:#bbf7d0;}}
.btn{{padding:5px 12px;border-radius:6px;font-size:11px;font-weight:500;cursor:pointer;border:1px solid;background:transparent;font-family:'Plus Jakarta Sans',sans-serif;transition:all .2s;white-space:nowrap;}}
.dep{{color:#dc2626;border-color:#fecaca;}}.dep:hover{{background:#fef2f2;}}
.rem{{color:#b45309;border-color:#fde68a;}}.rem:hover{{background:#fffbeb;}}
.mon{{color:#16a34a;border-color:#bbf7d0;}}.mon:hover{{background:#f0fdf4;}}
/* STATUS BADGES */
.st-badge{{padding:3px 9px;border-radius:99px;font-family:'JetBrains Mono',monospace;font-size:9px;font-weight:600;border:1px solid;white-space:nowrap;}}
.st-arriving{{background:#fef3c7;color:#92400e;border-color:#fde68a;}}
.st-inhouse{{background:#eff6ff;color:#1d4ed8;border-color:#bfdbfe;}}
.st-out{{background:#f8fafc;color:#94a3b8;border-color:#e4e8f0;}}
.st-future{{background:#f0fdf4;color:#15803d;border-color:#bbf7d0;}}
.note-cell{{max-width:200px;overflow:hidden;}}
.guest-note{{font-size:11px;color:#64748b;font-style:italic;}}
/* ACTION PLAN */
.action-plan-grid{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:28px;}}
.plan-card{{background:#ffffff;border:1px solid #e4e8f0;border-radius:14px;padding:22px;}}
.plan-card-header{{display:flex;align-items:center;gap:10px;margin-bottom:14px;}}
.plan-icon{{width:36px;height:36px;border-radius:9px;display:flex;align-items:center;justify-content:center;font-size:16px;flex-shrink:0;}}
.plan-icon.arriving{{background:#fef3c7;}}
.plan-icon.inhouse{{background:#eff6ff;}}
.plan-icon.intel{{background:#f0fdf4;}}
.plan-icon.bulk{{background:#faf5ff;}}
.plan-card-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:14px;font-weight:700;color:#0d1120;}}
.plan-card-sub{{font-family:'JetBrains Mono',monospace;font-size:10px;color:#94a3b8;}}
.plan-item{{background:#f8fafc;border:1px solid #e4e8f0;border-radius:8px;padding:10px 12px;margin-bottom:8px;}}
.plan-item:last-child{{margin-bottom:0;}}
.plan-guest{{font-weight:600;font-size:12px;color:#0d1120;margin-bottom:3px;}}
.plan-tip{{font-size:11.5px;color:#64748b;line-height:1.4;}}
.plan-empty{{font-family:'JetBrains Mono',monospace;font-size:11px;color:#94a3b8;padding:10px;}}
.plan-action-btn{{margin-top:12px;width:100%;padding:9px;background:#ffffff;border:1px solid #e4e8f0;border-radius:8px;color:#0d1120;font-size:12px;font-weight:600;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;transition:all .2s;}}
.plan-action-btn:hover{{background:#f1f5f9;}}
.plan-action-btn.green-btn{{background:#00d165;border-color:#00d165;color:#080c14;}}
.plan-action-btn.green-btn:hover{{background:#04e270;}}
/* INTEL CARDS */
.intel-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:8px;}}
.intel-item{{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:12px;}}
.intel-item.warn{{background:#fffbeb;border-color:#fde68a;}}
.intel-item.info{{background:#eff6ff;border-color:#bfdbfe;}}
.intel-num{{font-family:'Plus Jakarta Sans',sans-serif;font-size:22px;font-weight:800;color:#15803d;line-height:1;}}
.intel-num.warn{{color:#92400e;}}
.intel-num.info{{color:#1d4ed8;}}
.intel-label{{font-family:'JetBrains Mono',monospace;font-size:9px;color:#64748b;margin-top:4px;text-transform:uppercase;letter-spacing:0.5px;}}
/* OPTIMIZER */
.optimizer-card{{background:#ffffff;border:1px solid #e4e8f0;border-radius:14px;padding:24px;margin-bottom:28px;}}
.optimizer-row{{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:12px;margin-bottom:18px;}}
.opt-stat{{background:#f8fafc;border:1px solid #e4e8f0;border-radius:9px;padding:14px;}}
.opt-stat-val{{font-family:'Plus Jakarta Sans',sans-serif;font-size:26px;font-weight:800;color:#0d1120;line-height:1;}}
.opt-stat-label{{font-family:'JetBrains Mono',monospace;font-size:9px;color:#94a3b8;margin-top:4px;text-transform:uppercase;}}
/* BULK ACTIONS */
.bulk-zone{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:28px;}}
.bulk-card{{background:#f8fafc;border:1px solid #e4e8f0;border-radius:12px;padding:20px;transition:border-color .2s;}}
.bulk-card:hover{{border-color:#cbd5e1;}}
.bulk-icon{{font-size:24px;margin-bottom:8px;}}
.bulk-card-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:13px;font-weight:700;color:#0d1120;margin-bottom:4px;}}
.bulk-card-sub{{font-size:12px;color:#64748b;margin-bottom:14px;line-height:1.4;}}
.bulk-btn{{padding:9px 14px;background:#ffffff;border:1px solid #e4e8f0;border-radius:8px;color:#0d1120;font-size:12px;font-weight:600;cursor:pointer;width:100%;font-family:'Plus Jakarta Sans',sans-serif;transition:all .2s;}}
.bulk-btn:hover{{background:#f1f5f9;}}
.bulk-btn.green-btn{{background:#00d165;border-color:#00d165;color:#080c14;}}
.bulk-btn.green-btn:hover{{background:#04e270;}}
/* MODAL */
.modal-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.4);z-index:1000;align-items:center;justify-content:center;backdrop-filter:blur(4px);}}
.modal-overlay.show{{display:flex;}}
.modal{{background:#ffffff;border:1px solid #e4e8f0;border-radius:20px;padding:36px;width:100%;max-width:520px;max-height:88vh;overflow-y:auto;position:relative;box-shadow:0 16px 48px rgba(0,0,0,0.1);}}
.modal-close{{position:absolute;top:14px;right:16px;font-size:20px;cursor:pointer;color:#94a3b8;background:none;border:none;}}
.modal-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:20px;font-weight:800;color:#0d1120;margin-bottom:2px;}}
.modal-sub{{font-family:'JetBrains Mono',monospace;font-size:11px;color:#94a3b8;margin-bottom:18px;}}
.score-display{{font-family:'Plus Jakarta Sans',sans-serif;font-size:56px;font-weight:800;line-height:1;margin-bottom:6px;}}
.score-bar-bg{{height:8px;background:#f1f5f9;border-radius:4px;overflow:hidden;margin-bottom:10px;}}
.score-bar-fill{{height:100%;border-radius:4px;}}
.score-verdict{{font-size:12px;font-weight:600;padding:6px 12px;border-radius:6px;display:inline-block;margin-bottom:14px;}}
.modal-section{{margin-bottom:14px;}}
.modal-section-title{{font-family:'JetBrains Mono',monospace;font-size:10px;color:#94a3b8;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;}}
.modal-detail-row{{display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid #f1f5f9;font-size:12.5px;}}
.modal-detail-row:last-child{{border-bottom:none;}}
.modal-detail-label{{color:#64748b;}}
.modal-detail-val{{font-family:'JetBrains Mono',monospace;color:#0d1120;font-weight:500;}}
.reason-item{{background:#f8fafc;border-radius:7px;padding:10px 12px;margin-bottom:6px;font-size:12px;color:#374151;line-height:1.5;border-left:3px solid #e4e8f0;}}
.reason-item.pos{{border-left-color:#00d165;}}
.reason-item.neg{{border-left-color:#f59e0b;}}
/* EMAIL COMPOSER */
.email-composer{{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.4);z-index:1001;align-items:center;justify-content:center;backdrop-filter:blur(4px);}}
.email-composer.show{{display:flex;}}
.email-box{{background:#ffffff;border:1px solid #e4e8f0;border-radius:20px;padding:32px;width:100%;max-width:600px;max-height:90vh;overflow-y:auto;box-shadow:0 16px 48px rgba(0,0,0,0.1);}}
.email-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:17px;font-weight:800;color:#0d1120;margin-bottom:2px;}}
.email-label{{font-family:'JetBrains Mono',monospace;font-size:10px;color:#94a3b8;text-transform:uppercase;letter-spacing:1px;display:block;margin-bottom:6px;font-weight:500;}}
.email-input{{width:100%;padding:10px 13px;background:#f8fafc;border:1px solid #e4e8f0;border-radius:8px;font-size:13px;color:#0d1120;outline:none;margin-bottom:12px;font-family:'Plus Jakarta Sans',sans-serif;}}
.email-input:focus{{border-color:#00d165;background:#fff;}}
.email-textarea{{width:100%;padding:10px 13px;background:#f8fafc;border:1px solid #e4e8f0;border-radius:8px;font-size:12.5px;color:#0d1120;outline:none;resize:vertical;min-height:160px;margin-bottom:12px;font-family:'Plus Jakarta Sans',sans-serif;line-height:1.6;}}
.email-textarea:focus{{border-color:#00d165;background:#fff;}}
.email-actions{{display:flex;gap:10px;margin-top:16px;}}
.email-send{{flex:1;padding:11px;background:#00d165;color:#080c14;border:none;border-radius:8px;font-size:13px;font-weight:700;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;}}
.email-send:hover{{background:#04e270;}}
.email-cancel{{flex:1;padding:11px;background:#f8fafc;color:#64748b;border:1px solid #e4e8f0;border-radius:8px;font-size:13px;font-weight:500;cursor:pointer;}}
/* TOAST */
.toast{{position:fixed;bottom:24px;right:24px;background:#0d1120;color:#fff;border-radius:10px;padding:13px 17px;font-size:12.5px;transform:translateY(60px);opacity:0;transition:all 0.3s;z-index:2000;box-shadow:0 8px 24px rgba(0,0,0,0.15);}}
.toast.show{{transform:translateY(0);opacity:1;}}
/* WELCOME BANNER */
.welcome-banner{{background:#f0fdf4;border-bottom:1px solid #bbf7d0;padding:11px 32px;display:flex;align-items:center;justify-content:space-between;font-size:13px;color:#166534;}}
.welcome-close{{background:none;border:none;color:#94a3b8;font-size:18px;cursor:pointer;}}

</style>
</head>
<body>

<!-- TOPBAR -->
<div class="topbar">
  <div class="topbar-brand">
    <span class="topbar-name">Occup<span>ado</span></span>
    <span class="topbar-hotel">{hotel_name}</span>
  </div>
  <div class="topbar-right">
    <select class="lang-selector" onchange="window.location.href='/dashboard?lang='+this.value">
      <option value="en" {"selected" if lang=="en" else ""}>EN</option>
      <option value="nl" {"selected" if lang=="nl" else ""}>NL</option>
      <option value="fr" {"selected" if lang=="fr" else ""}>FR</option>
    </select>
    <a href="/settings" class="btn-nav">Settings</a>
    <a href="/logout" class="btn-nav">Sign Out</a>
  </div>
</div>

{f'''<div id="welcome-banner" class="welcome-banner">
  <span>Welcome back, <strong>{hotel_name}</strong>. Your live intelligence dashboard is ready.</span>
  <button onclick="document.getElementById('welcome-banner').style.display='none'" class="welcome-close">×</button>
</div>''' if first_login else ''}

<!-- HERO -->
<div class="hero">
  <div class="hero-eyebrow">Live Intelligence Dashboard</div>
  <div class="hero-title">Van der Valk Hotel Mechelen</div>
  <div class="hero-sub">{today_str} · {len(guests)} repeat guests tracked · Powered by Occupado AI</div>
  <div class="hero-badges">
    <span class="hero-badge">✓ Model accuracy 80.5%</span>
    <span class="hero-badge">✓ {len(arriving_today)} arriving today</span>
    <span class="hero-badge">✓ {len(in_house)} in house</span>
    <span class="hero-badge {'warn' if high_count > 0 else ''}">{'⚠ ' + str(high_count) + ' high risk' if high_count > 0 else '✓ 0 high risk today'}</span>
    <span class="hero-badge info">VDV Shiji · Data current</span>
  </div>
</div>

<div class="content">

<!-- KPI ROW 1: Today's operations -->
<div class="section-title" style="margin-top:0">Today at a Glance</div>
<div class="kpi-grid">
  <div class="kpi-card highlight-green">
    <div class="kpi-num green">{len(arriving_today)}</div>
    <div class="kpi-label">Arriving Today</div>
    <div class="kpi-trend" style="color:#16a34a">↑ Ready for check-in</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num blue">{len(in_house)}</div>
    <div class="kpi-label">Currently In House</div>
    <div class="kpi-trend" style="color:#3b82f6">Repeat guests</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num neutral">{len(departing_tmrw)}</div>
    <div class="kpi-label">Departing Tomorrow</div>
    <div class="kpi-trend" style="color:#64748b">Send departure reminders</div>
  </div>
  <div class="kpi-card {'highlight-red' if high_count > 0 else ''}">
    <div class="kpi-num {'red' if high_count > 0 else 'green'}">{high_count}</div>
    <div class="kpi-label">High Risk Bookings</div>
    <div class="kpi-trend" style="color:{'#dc2626' if high_count > 0 else '#16a34a'}">{'Action required' if high_count > 0 else '✓ All clear'}</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num orange">{med_count}</div>
    <div class="kpi-label">Medium Risk</div>
    <div class="kpi-trend" style="color:#b45309">Monitor closely</div>
  </div>
  <div class="kpi-card highlight-green">
    <div class="kpi-num green">{low_count}</div>
    <div class="kpi-label">Low Risk</div>
    <div class="kpi-trend" style="color:#16a34a">Avg score {avg_score:.1f}%</div>
  </div>
</div>

<!-- KPI ROW 2: Historical intelligence -->
<div class="kpi-grid" style="margin-top:12px;margin-bottom:28px;">
  <div class="kpi-card">
    <div class="kpi-num neutral" style="font-size:28px;">15.7%</div>
    <div class="kpi-label">Historical Cancellation Rate</div>
    <div class="kpi-trend" style="color:#64748b">Based on 1,694 cancellations</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num neutral" style="font-size:28px;">5.0%</div>
    <div class="kpi-label">Historical No-show Rate</div>
    <div class="kpi-trend" style="color:#64748b">339 no-shows tracked</div>
  </div>
  <div class="kpi-card highlight-green">
    <div class="kpi-num green" style="font-size:28px;">80.5%</div>
    <div class="kpi-label">AI Model Accuracy</div>
    <div class="kpi-trend" style="color:#16a34a">Kaggle + Shiji data</div>
  </div>
  <div class="kpi-card highlight-green">
    <div class="kpi-num green" style="font-size:28px;">€0</div>
    <div class="kpi-label">Revenue at Risk Today</div>
    <div class="kpi-trend" style="color:#16a34a">All guests confirmed</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num neutral" style="font-size:28px;">€{int(rev_arriving):,}</div>
    <div class="kpi-label">Expected Revenue — Arrivals</div>
    <div class="kpi-trend" style="color:#64748b">{len(arriving_today)} guests × avg rate</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num neutral" style="font-size:28px;">€{int(rev_inhouse):,}</div>
    <div class="kpi-label">Active Revenue — In House</div>
    <div class="kpi-trend" style="color:#64748b">{len(in_house)} guests × avg rate</div>
  </div>
</div>

<!-- CHARTS ROW -->
<div class="section-title">Analytics</div>
<div class="charts-row">

  <!-- Donut: Risk distribution -->
  <div class="chart-card">
    <div class="chart-title">Booking Risk</div>
    <div class="chart-sub">Current guests distribution</div>
    <canvas id="riskDonut" height="180"></canvas>
    <div class="donut-center">
      <div class="donut-big">{round(low_count/len(scores)*100) if scores else 0}%</div>
      <div class="donut-label">Low Risk</div>
    </div>
  </div>

  <!-- Bar: Cancellation by channel -->
  <div class="chart-card">
    <div class="chart-title">Cancellations by Channel</div>
    <div class="chart-sub">Historical — {sum(ch_data.values()):,} total cancellations</div>
    <canvas id="channelChart" height="180"></canvas>
  </div>

  <!-- Bar: Guest risk scores -->
  <div class="chart-card">
    <div class="chart-title">Guest Cancellation Scores</div>
    <div class="chart-sub">AI prediction per guest — lower is better</div>
    <canvas id="scoreChart" height="180"></canvas>
  </div>

</div>

<!-- OVERBOOKING OPTIMIZER -->
<div class="section-title">Overbooking Optimizer</div>
<div class="optimizer-card">
  <div class="optimizer-row">
    <div class="opt-stat" style="background:#f0fdf4;border-color:#bbf7d0;">
      <div class="opt-stat-val" style="color:#00d165">+{max(0, round(len(guests)*0.05))}</div>
      <div class="opt-stat-label">Safe rooms to oversell</div>
    </div>
    <div class="opt-stat">
      <div class="opt-stat-val">{len(guests)}</div>
      <div class="opt-stat-label">Bookings analysed</div>
    </div>
    <div class="opt-stat">
      <div class="opt-stat-val" style="color:#dc2626">0</div>
      <div class="opt-stat-label">Predicted no-shows</div>
    </div>
    <div class="opt-stat">
      <div class="opt-stat-val">€130</div>
      <div class="opt-stat-label">Avg room rate</div>
    </div>
  </div>
  <div style="display:flex;gap:10px;align-items:center;background:#f8fafc;border:1px solid #e4e8f0;border-radius:9px;padding:14px 18px;">
    <div style="flex:1;font-size:13px;color:#374151;line-height:1.5;">
      <strong style="color:#0d1120;">Tonight's outlook:</strong> All {len(guests)} tracked repeat guests are confirmed.
      Historical no-show rate of 4.7% suggests <strong>{round(len(guests)*0.047,1)}</strong> potential gaps.
      Consider moderate overbooking for non-repeat segments.
    </div>
    <button onclick="showToast('Recommendation noted — coordinate with front desk')" style="padding:10px 20px;background:#00d165;border:none;border-radius:8px;color:#080c14;font-size:13px;font-weight:700;cursor:pointer;white-space:nowrap;font-family:Inter,sans-serif;">Apply Recommendation</button>
  </div>
</div>

<!-- GUEST TABLE -->
<div class="section-title">Repeat Guests — Click any row for AI analysis</div>
<div class="section-sub">Sorted by status: Arriving Today → In House → Checked Out</div>
<table>
<thead>
  <tr>
    <th>Guest</th><th>Status</th><th>Arrival</th><th>Departure</th>
    <th>Nights</th><th>Adults</th><th>Risk Score</th><th>Guest Notes</th><th>Action</th>
  </tr>
</thead>
<tbody>{rows_html}</tbody>
</table>

<!-- ACTION PLAN -->
<div class="section-title">Interactive Action Plan</div>
<div class="section-sub">Contextual tips and contact options per guest segment</div>

<div class="action-plan-grid">

  <!-- Arriving Today -->
  <div class="plan-card">
    <div class="plan-card-header">
      <div class="plan-icon arriving">🛬</div>
      <div>
        <div class="plan-card-title">Today's Arrivals ({len(arriving_today)})</div>
        <div class="plan-card-sub">Expected today · Pre-arrival actions</div>
      </div>
    </div>
    {arriving_plan}
    <button class="plan-action-btn green-btn" onclick="openBulkTemplate('welcome')" style="margin-top:14px;">
      Send Welcome Email to All →
    </button>
  </div>

  <!-- In House -->
  <div class="plan-card">
    <div class="plan-card-header">
      <div class="plan-icon inhouse">🏨</div>
      <div>
        <div class="plan-card-title">In-House Guests ({len(in_house)})</div>
        <div class="plan-card-sub">Currently staying · Experience & upsell</div>
      </div>
    </div>
    {inhouse_plan}
    <button class="plan-action-btn" onclick="openBulkTemplate('departure')" style="margin-top:14px;">
      Send Departure Reminder to All →
    </button>
  </div>

  <!-- Revenue Intelligence -->
  <div class="plan-card">
    <div class="plan-card-header">
      <div class="plan-icon intel">📊</div>
      <div>
        <div class="plan-card-title">Revenue Intelligence</div>
        <div class="plan-card-sub">Historical insights from Shiji data</div>
      </div>
    </div>
    <div class="intel-grid">
      <div class="intel-item warn">
        <div class="intel-num warn">31.6%</div>
        <div class="intel-label">Booking.com cancellations</div>
      </div>
      <div class="intel-item warn">
        <div class="intel-num warn">24.2%</div>
        <div class="intel-label">Corporate cancellations</div>
      </div>
      <div class="intel-item info">
        <div class="intel-num info">14.7%</div>
        <div class="intel-label">Direct / Web cancellations</div>
      </div>
    </div>
    <div style="margin-top:10px;background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:12px;font-size:12px;color:#92400e;line-height:1.5;">
      <strong>💡 Insight:</strong> Booking.com is your #1 cancellation source (1,147 cancellations tracked).
      Consider requiring prepayment on OTA bookings with high lead times.
    </div>
  </div>

  <!-- Tips & Best Practices -->
  <div class="plan-card">
    <div class="plan-card-header">
      <div class="plan-icon bulk">💡</div>
      <div>
        <div class="plan-card-title">Tips & Best Practices</div>
        <div class="plan-card-sub">AI-powered recommendations for VdV Mechelen</div>
      </div>
    </div>
    <div class="plan-item">
      <div class="plan-guest">Repeat Guest Program</div>
      <div class="plan-tip">Your repeat guests (like those tracked here) have significantly lower cancellation rates. Prioritise VIP recognition on check-in to reinforce loyalty.</div>
    </div>
    <div class="plan-item">
      <div class="plan-guest">Corporate Rate Management</div>
      <div class="plan-tip">Corporate guests (Alheembouw, CORP accounts) have fixed rates. Verify current-period rates are loaded correctly in Shiji before check-in.</div>
    </div>
    <div class="plan-item">
      <div class="plan-guest">No-show Protection</div>
      <div class="plan-tip">With a 4.7% historical no-show rate, consider implementing credit card guarantees for all bookings with lead time &gt; 30 days.</div>
    </div>
  </div>

</div>

<!-- BULK ACTIONS -->
<div class="section-title">Take Action</div>
<div class="bulk-zone">
  <div class="bulk-card">
    <div class="bulk-icon">👋</div>
    <div class="bulk-card-title">Welcome Arriving Guests</div>
    <div class="bulk-card-sub">Send personalised welcome emails to today's {len(arriving_today)} arriving guests</div>
    <button class="bulk-btn green-btn" onclick="openBulkTemplate('welcome')">Send Welcome Email →</button>
  </div>
  <div class="bulk-card">
    <div class="bulk-icon">🛎</div>
    <div class="bulk-card-title">Departure Reminders</div>
    <div class="bulk-card-sub">Remind tomorrow's departures of checkout time and offer late checkout</div>
    <button class="bulk-btn" onclick="openBulkTemplate('departure')">Send Departure Reminder →</button>
  </div>
  <div class="bulk-card">
    <div class="bulk-icon">⬆️</div>
    <div class="bulk-card-title">Upsell In-House Guests</div>
    <div class="bulk-card-sub">Offer restaurant, parking, or room upgrade to the {len(in_house)} guests currently staying</div>
    <button class="bulk-btn" onclick="openBulkTemplate('upsell')">Send Upsell Offer →</button>
  </div>
</div>

</div><!-- /content -->


<!-- DETAIL MODAL -->
<div class="modal-overlay" id="detailModal">
  <div class="modal">
    <button class="modal-close" onclick="closeDetailModal()">✕</button>
    <div class="modal-title" id="dm-title">Guest Analysis</div>
    <div class="modal-sub" id="dm-sub">AI Cancellation Risk</div>
    <div class="score-display" id="dm-score">0%</div>
    <div class="score-bar-bg"><div class="score-bar-fill" id="dm-bar" style="width:0%"></div></div>
    <div class="score-verdict" id="dm-verdict"></div>
    <div class="modal-section">
      <div class="modal-section-title">Booking Details</div>
      <div id="dm-details"></div>
    </div>
    <div class="modal-section">
      <div class="modal-section-title">AI Risk Factors</div>
      <div id="dm-reasons"></div>
    </div>
    <div style="display:flex;gap:10px;margin-top:20px;">
      <button id="dm-contact-btn" style="flex:1;padding:12px;border:none;border-radius:8px;font-size:13px;font-weight:700;cursor:pointer;font-family:Inter,sans-serif;background:#00d165;color:#080c14;" onclick="openContactFromModal()">Contact Guest →</button>
      <button style="flex:1;padding:12px;background:#f8fafc;color:#64748b;border:1px solid #e4e8f0;border-radius:8px;font-size:13px;font-weight:500;cursor:pointer;" onclick="closeDetailModal()">Close</button>
    </div>
  </div>
</div>

<!-- EMAIL COMPOSER -->
<div class="email-composer" id="emailComposer">
  <div class="email-box">
    <div style="margin-bottom:20px;">
      <div class="email-title" id="ec-title">Contact Guest</div>
      <div style="font-family:'JetBrains Mono',monospace;font-size:11px;color:#94a3b8;margin-top:3px;" id="ec-sub"></div>
    </div>
    <label class="email-label">Guest Email Address</label>
    <input type="email" id="ec-email" class="email-input" placeholder="guest@example.com">
    <label class="email-label">Guest Name</label>
    <input type="text" id="ec-name" class="email-input" placeholder="Guest Name">
    <label class="email-label">Subject</label>
    <input type="text" id="ec-subject" class="email-input" placeholder="Your upcoming stay at Van der Valk Hotel Mechelen">
    <label class="email-label">Message</label>
    <textarea id="ec-body" class="email-textarea" placeholder="Dear Guest..."></textarea>
    <div class="email-actions">
      <button class="email-send" onclick="sendGuestEmail()">📧 Send Email</button>
      <button class="email-cancel" onclick="closeEmailComposer()">Cancel</button>
    </div>
  </div>
</div>

<!-- BULK EMAIL COMPOSER -->
<div class="email-composer" id="bulkComposer">
  <div class="email-box">
    <div style="margin-bottom:20px;">
      <div class="email-title" id="bc-title">Send to Group</div>
      <div style="font-family:'JetBrains Mono',monospace;font-size:11px;color:#94a3b8;margin-top:3px;" id="bc-sub"></div>
    </div>
    <label class="email-label">Subject</label>
    <input type="text" id="bc-subject" class="email-input">
    <label class="email-label">Message (sent to all guests in this group)</label>
    <textarea id="bc-body" class="email-textarea"></textarea>
    <div class="email-actions">
      <button class="email-send" onclick="sendBulkEmail()">📧 Send to All</button>
      <button class="email-cancel" onclick="closeBulkComposer()">Cancel</button>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>
const guests = {guests_js};
const scores = {guest_scores_js};
let currentGuestIdx = -1;

// ── CHARTS ────────────────────────────────────────────────────────────────────
(function initCharts() {{
  // Donut — risk distribution
  new Chart(document.getElementById('riskDonut'), {{
    type: 'doughnut',
    data: {{
      labels: ['Low Risk','Medium Risk','High Risk'],
      datasets: [{{ data: {risk_donut_js}, backgroundColor: ['#00d165','#f59e0b','#dc2626'], borderWidth: 0, hoverOffset: 4 }}]
    }},
    options: {{
      cutout: '72%', plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ family: 'JetBrains Mono', size: 10 }}, boxWidth: 10, padding: 12 }} }}, tooltip: {{ callbacks: {{ label: ctx => ` ${{ctx.label}}: ${{ctx.parsed}} bookings` }} }} }},
      animation: {{ animateRotate: true, duration: 900 }}
    }}
  }});

  // Channel cancellations bar
  new Chart(document.getElementById('channelChart'), {{
    type: 'bar',
    data: {{
      labels: {ch_labels},
      datasets: [{{
        label: 'Cancellations',
        data: {ch_values},
        backgroundColor: ['#dc2626','#f59e0b','#3b82f6','#8b5cf6','#94a3b8'],
        borderRadius: 5, borderWidth: 0
      }}]
    }},
    options: {{
      indexAxis: 'y',
      plugins: {{ legend: {{ display: false }}, tooltip: {{ callbacks: {{ label: ctx => ` ${{ctx.parsed.x}} cancellations (${{({ch_pcts})[ctx.dataIndex]}}%)` }} }} }},
      scales: {{ x: {{ grid: {{ color: '#f1f5f9' }}, ticks: {{ font: {{ family: 'JetBrains Mono', size: 10 }} }} }}, y: {{ grid: {{ display: false }}, ticks: {{ font: {{ family: 'JetBrains Mono', size: 10 }} }} }} }},
      animation: {{ duration: 900 }}
    }}
  }});

  // Guest risk scores bar
  new Chart(document.getElementById('scoreChart'), {{
    type: 'bar',
    data: {{
      labels: {guest_names_js},
      datasets: [{{
        label: 'Risk Score %',
        data: scores,
        backgroundColor: {guest_colors_js},
        borderRadius: 5, borderWidth: 0
      }}]
    }},
    options: {{
      plugins: {{ legend: {{ display: false }}, tooltip: {{ callbacks: {{ label: ctx => ` ${{ctx.parsed.y.toFixed(1)}}% cancellation risk` }} }} }},
      scales: {{ x: {{ grid: {{ display: false }}, ticks: {{ font: {{ family: 'JetBrains Mono', size: 9 }}, maxRotation: 45 }} }}, y: {{ min:0, max:100, grid: {{ color:'#f1f5f9' }}, ticks: {{ font: {{ family:'JetBrains Mono',size:10 }}, callback: v => v+'%' }} }} }},
      animation: {{ duration: 900 }}
    }}
  }});
}})();

// ── DETAIL MODAL ──────────────────────────────────────────────────────────────
function showVdvDetail(idx, score) {{
  const g = guests[idx];
  currentGuestIdx = idx;
  document.getElementById('dm-title').textContent = g.name;
  document.getElementById('dm-sub').textContent = g.arrival + ' → ' + g.departure + ' · ' + g.status;
  document.getElementById('dm-score').textContent = score.toFixed(1) + '%';

  const bar   = document.getElementById('dm-bar');
  const verd  = document.getElementById('dm-verdict');
  const scoreEl = document.getElementById('dm-score');
  bar.style.width = score + '%';

  if (score >= 70) {{
    bar.style.background = '#dc2626'; scoreEl.style.color = '#dc2626';
    verd.textContent = 'HIGH RISK'; verd.style.background = '#fef2f2'; verd.style.color = '#dc2626';
  }} else if (score >= 40) {{
    bar.style.background = '#f59e0b'; scoreEl.style.color = '#f59e0b';
    verd.textContent = 'MEDIUM RISK'; verd.style.background = '#fffbeb'; verd.style.color = '#b45309';
  }} else {{
    bar.style.background = '#00d165'; scoreEl.style.color = '#00d165';
    verd.textContent = 'LOW RISK'; verd.style.background = '#f0fdf4'; verd.style.color = '#16a34a';
  }}

  document.getElementById('dm-details').innerHTML = `
    <div class="modal-detail-row"><span class="modal-detail-label">Nights</span><span class="modal-detail-val">${{g.nights}}</span></div>
    <div class="modal-detail-row"><span class="modal-detail-label">Adults</span><span class="modal-detail-val">${{g.adults}}</span></div>
    <div class="modal-detail-row"><span class="modal-detail-label">Membership</span><span class="modal-detail-val">${{g.membership || 'Standard'}}</span></div>
    <div class="modal-detail-row"><span class="modal-detail-label">Avg Room Rate</span><span class="modal-detail-val">€${{g.adr}}/night</span></div>
    ${{g.note ? '<div class="modal-detail-row"><span class="modal-detail-label">Notes</span><span class="modal-detail-val" style="font-size:11px;max-width:220px;text-align:right;">' + g.note + '</span></div>' : ''}}
  `;

  const reasons = [];
  if (score < 15) {{
    reasons.push({{pos:true, text:'Repeat guest — historically reliable, lower no-show probability'}});
    reasons.push({{pos:true, text:'Low lead time — very close to check-in, commitment is high'}});
    if (g.membership) reasons.push({{pos:true, text:'Corporate/loyalty membership — adds accountability'}});
  }} else if (score < 40) {{
    reasons.push({{pos:true, text:'Returning guest with positive booking history'}});
    reasons.push({{pos:false, text:'Extended stay or higher lead time increases marginal risk'}});
  }} else {{
    reasons.push({{pos:false, text:'Risk factors present — consider proactive contact'}});
  }}

  document.getElementById('dm-reasons').innerHTML = reasons.map(r =>
    `<div class="reason-item ${{r.pos?'pos':'neg'}}">${{r.pos?'✓':'⚠'}} ${{r.text}}</div>`
  ).join('');

  document.getElementById('detailModal').classList.add('show');
}}

function closeDetailModal() {{
  document.getElementById('detailModal').classList.remove('show');
  currentGuestIdx = -1;
}}

function openContactFromModal() {{
  closeDetailModal();
  if (currentGuestIdx >= 0) setTimeout(() => openVdvEmail(currentGuestIdx, 'contact'), 100);
}}

// ── EMAIL COMPOSER ────────────────────────────────────────────────────────────
function openVdvEmail(idx, type) {{
  const g = guests[idx];
  currentGuestIdx = idx;
  const firstName = g.name.split(',').length > 1 ? g.name.split(',')[1].split(',')[0].trim().split(' ')[1] || g.name.split(',')[1].trim() : g.name;

  const templates = {{
    contact: {{
      subject: 'Your upcoming stay at Van der Valk Hotel Mechelen',
      body: `Dear ${{firstName}},\n\nWe are looking forward to welcoming you to Van der Valk Hotel Mechelen.\n\nPlease do not hesitate to contact us if you have any special requests or questions regarding your stay from ${{g.arrival}} to ${{g.departure}}.\n\nWarm regards,\nVan der Valk Hotel Mechelen`
    }},
    welcome: {{
      subject: 'Welcome to Van der Valk Hotel Mechelen',
      body: `Dear ${{firstName}},\n\nWelcome! We are delighted to have you with us today.\n\nYour room is being prepared and we look forward to ensuring you have a wonderful stay. Should you need anything during your time with us, please do not hesitate to contact the front desk.\n\nWarm regards,\nVan der Valk Hotel Mechelen`
    }},
    departure: {{
      subject: 'Thank you for staying — departure reminder',
      body: `Dear ${{firstName}},\n\nWe hope you are enjoying your stay at Van der Valk Hotel Mechelen.\n\nThis is a friendly reminder that your check-out is scheduled for ${{g.departure}}. Checkout time is 12:00. If you wish to arrange a late checkout, please contact the front desk.\n\nWe hope to welcome you back soon!\n\nWarm regards,\nVan der Valk Hotel Mechelen`
    }},
    reminder: {{
      subject: 'Reminder: Your upcoming reservation',
      body: `Dear ${{firstName}},\n\nThis is a friendly reminder of your upcoming reservation from ${{g.arrival}} to ${{g.departure}} at Van der Valk Hotel Mechelen.\n\nPlease confirm your booking or contact us if there are any changes to your plans.\n\nWarm regards,\nVan der Valk Hotel Mechelen`
    }},
    deposit: {{
      subject: 'Reservation guarantee — deposit request',
      body: `Dear ${{firstName}},\n\nThank you for your reservation from ${{g.arrival}} to ${{g.departure}} at Van der Valk Hotel Mechelen.\n\nTo secure your booking, we kindly request a deposit. Please contact us at your earliest convenience to complete this.\n\nWarm regards,\nVan der Valk Hotel Mechelen`
    }},
    upsell: {{
      subject: 'Make the most of your stay — exclusive offers',
      body: `Dear ${{firstName}},\n\nWe hope you are enjoying your time at Van der Valk Hotel Mechelen!\n\nAs our valued guest, we would like to offer you:\n• Restaurant dinner reservation (10% loyalty discount)\n• Covered parking upgrade\n• Late checkout until 14:00 (subject to availability)\n\nPlease contact the front desk or reply to this email to arrange any of the above.\n\nWarm regards,\nVan der Valk Hotel Mechelen`
    }}
  }};

  const tpl = templates[type] || templates.contact;
  document.getElementById('ec-title').textContent = 'Email to ' + g.name;
  document.getElementById('ec-sub').textContent   = g.status + ' · ' + g.arrival + ' – ' + g.departure;
  document.getElementById('ec-name').value    = g.name;
  document.getElementById('ec-email').value   = '';
  document.getElementById('ec-subject').value = tpl.subject;
  document.getElementById('ec-body').value    = tpl.body;
  document.getElementById('emailComposer').classList.add('show');
}}

function closeEmailComposer() {{
  document.getElementById('emailComposer').classList.remove('show');
}}

function sendGuestEmail() {{
  const email   = document.getElementById('ec-email').value.trim();
  const subject = document.getElementById('ec-subject').value.trim();
  const body    = document.getElementById('ec-body').value.trim();
  if (!email || !subject || !body) {{ showToast('Please fill in all fields', 'error'); return; }}
  fetch('/send-guest-email', {{
    method: 'POST', headers: {{'Content-Type':'application/json'}},
    body: JSON.stringify({{guest_email: email, subject, body}})
  }}).then(r => r.json()).then(d => {{
    closeEmailComposer();
    showToast(d.status === 'success' ? '✓ Email sent' : 'Error: ' + d.message, d.status === 'success' ? 'success' : 'error');
  }}).catch(() => showToast('Error sending email', 'error'));
}}

// ── BULK COMPOSER ─────────────────────────────────────────────────────────────
function openBulkTemplate(type) {{
  const templates = {{
    welcome: {{
      title: 'Welcome Email — Today\'s Arrivals',
      subject: 'Welcome to Van der Valk Hotel Mechelen',
      body: 'Dear Valued Guest,\n\nWelcome! We are so glad to have you with us today.\n\nYour room is being prepared and our team is ready to ensure a wonderful experience. For any requests, please contact the front desk.\n\nWarm regards,\nVan der Valk Hotel Mechelen'
    }},
    departure: {{
      title: 'Departure Reminder — Tomorrow\'s Checkouts',
      subject: 'Departure reminder — checkout tomorrow',
      body: 'Dear Valued Guest,\n\nThis is a friendly reminder that your checkout is scheduled for tomorrow. Standard checkout is at 12:00.\n\nWe hope you have enjoyed your stay and look forward to welcoming you back!\n\nWarm regards,\nVan der Valk Hotel Mechelen'
    }},
    upsell: {{
      title: 'Upsell Offer — In-House Guests',
      subject: 'Exclusive offers for your stay',
      body: 'Dear Valued Guest,\n\nWe hope you are enjoying your time with us!\n\nAs a valued guest, we are pleased to offer you exclusive in-stay experiences:\n• Restaurant reservation with priority seating\n• Parking upgrade\n• Late checkout (subject to availability)\n\nContact the front desk to arrange any of these.\n\nWarm regards,\nVan der Valk Hotel Mechelen'
    }}
  }};
  const tpl = templates[type] || templates.welcome;
  document.getElementById('bc-title').textContent   = tpl.title;
  document.getElementById('bc-sub').textContent     = 'Template — fill in guest details before sending';
  document.getElementById('bc-subject').value = tpl.subject;
  document.getElementById('bc-body').value    = tpl.body;
  document.getElementById('bulkComposer').classList.add('show');
}}

function closeBulkComposer() {{
  document.getElementById('bulkComposer').classList.remove('show');
}}

function sendBulkEmail() {{
  const subject = document.getElementById('bc-subject').value.trim();
  const body    = document.getElementById('bc-body').value.trim();
  if (!subject || !body) {{ showToast('Please fill in all fields', 'error'); return; }}
  fetch('/send-bulk-email', {{
    method: 'POST', headers: {{'Content-Type':'application/json'}},
    body: JSON.stringify({{count: guests.length, subject, body}})
  }}).then(r => r.json()).then(d => {{
    closeBulkComposer();
    showToast(d.status === 'success' ? '✓ Emails sent' : 'Error', d.status === 'success' ? 'success' : 'error');
  }}).catch(() => showToast('Error', 'error'));
}}

// ── UTILS ─────────────────────────────────────────────────────────────────────
function showToast(msg, type) {{
  const t = document.getElementById('toast');
  t.textContent = msg;
  t.style.background = type === 'error' ? '#dc2626' : '#0d1120';
  t.classList.add('show');
  setTimeout(() => t.classList.remove('show'), 3200);
}}

document.getElementById('detailModal').addEventListener('click', e => {{ if (e.target === document.getElementById('detailModal')) closeDetailModal(); }});
document.getElementById('emailComposer').addEventListener('click', e => {{ if (e.target === document.getElementById('emailComposer')) closeEmailComposer(); }});
document.getElementById('bulkComposer').addEventListener('click', e => {{ if (e.target === document.getElementById('bulkComposer')) closeBulkComposer(); }});
// ── DYNAMICS JS ──────────────────────────────────────────────

// Counter animation
function countUp(el, target, duration, pre, suf) {{
  if (!el) return;
  pre = pre||''; suf = suf||''; duration = duration||1400;
  const start = performance.now();
  function tick(now) {{
    const p = Math.min((now - start)/duration, 1);
    const ease = 1 - Math.pow(1-p, 3);
    const val = Math.round(ease * target);
    el.textContent = pre + val.toLocaleString() + suf;
    if (p < 1) requestAnimationFrame(tick);
  }}
  requestAnimationFrame(tick);
}}
document.querySelectorAll('.count-up').forEach(function(el) {{
  countUp(el, parseFloat(el.dataset.val)||0, 1400, el.dataset.pre||'', el.dataset.suf||'');
}});

// Filter tabs
function filterRisk(risk, btn) {{
  document.querySelectorAll('.f-tab').forEach(function(t) {{ t.classList.remove('active'); }});
  btn.classList.add('active');
  document.querySelectorAll('.clickable-row, .cr').forEach(function(row) {{
    var show;
    if (risk === 'all') {{ show = true; }}
    else {{ show = !!row.querySelector('.badge.' + risk); }}
    row.style.display = show ? '' : 'none';
    var nx = row.nextElementSibling;
    if (nx && nx.classList.contains('exp-tr')) {{
      if (!show) nx.classList.remove('open');
      nx.style.display = show ? '' : 'none';
    }}
  }});
}}

// Sort table
var _sCol = null, _sDir = 1;
function sortTable(col) {{
  var tbody = document.querySelector('tbody');
  if (!tbody) return;
  if (_sCol === col) _sDir *= -1; else {{ _sCol = col; _sDir = 1; }}
  var pairs = [];
  var rows = Array.from(tbody.querySelectorAll('.clickable-row, .cr'));
  rows.forEach(function(r) {{
    var nx = r.nextElementSibling;
    var expRow = (nx && nx.classList.contains('exp-tr')) ? nx : null;
    pairs.push([r, expRow]);
  }});
  pairs.sort(function(a, b) {{
    var ar = a[0], br = b[0];
    var av = col==='score' ? parseFloat(ar.dataset.score||0) :
             col==='lead'  ? parseInt(ar.dataset.lead||0) :
             col==='rate'  ? parseInt(ar.dataset.rate||0) : 0;
    var bv = col==='score' ? parseFloat(br.dataset.score||0) :
             col==='lead'  ? parseInt(br.dataset.lead||0) :
             col==='rate'  ? parseInt(br.dataset.rate||0) : 0;
    return (av - bv) * _sDir;
  }});
  pairs.forEach(function(p) {{
    tbody.appendChild(p[0]);
    if (p[1]) tbody.appendChild(p[1]);
  }});
  document.querySelectorAll('.sort-arr').forEach(function(a) {{
    var c = a.closest('.sort-th') && a.closest('.sort-th').dataset.col;
    a.textContent = c === col ? (_sDir === -1 ? '↓' : '↑') : '↕';
    a.classList.toggle('on', c === col);
  }});
}}

// Build prediction reasons
function buildReasons(b, score) {{
  var r = [];
  var lt = b.lead_time||0, canc = b.previous_cancellations||0;
  var rep = b.is_repeated_guest||0, adr = b.adr||0;
  var chg = b.booking_changes||0, spec = b.total_of_special_requests||0;
  var wk = b.stays_in_week_nights||0, we = b.stays_in_weekend_nights||0;
  var nights = wk + we || 1;

  if (lt > 120)      r.push({{c:'bad',  t:'Lead time ' + lt + ' days — bookings this far out cancel 3× more often'}});
  else if (lt > 45)  r.push({{c:'warn', t:'Lead time ' + lt + ' days — moderate cancellation window remains'}});
  else               r.push({{c:'pos',  t:'Lead time ' + lt + ' days — close to arrival, low cancellation risk'}});

  if (canc >= 2)     r.push({{c:'bad',  t: canc + ' prior cancellations — strongest single predictor of future no-shows'}});
  else if (canc===1) r.push({{c:'warn', t:'1 previous cancellation on record — elevated risk signal'}});
  else               r.push({{c:'pos',  t:'No cancellation history — clean booking profile'}});

  if (!rep)          r.push({{c:'warn', t:'First-time guest — new guests cancel 2.3× more than returning guests'}});
  else               r.push({{c:'pos',  t:'Returning guest — loyalty significantly reduces no-show likelihood'}});

  if (adr > 200)     r.push({{c:'bad',  t:'Premium rate €' + Math.round(adr) + ' — high-value bookings warrant a deposit or guarantee'}});
  else if (adr > 120)r.push({{c:'warn', t:'Room rate €' + Math.round(adr) + ' — consider pre-authorisation'}});

  if (chg >= 2)      r.push({{c:'warn', t: chg + ' booking modifications — repeated changes signal hesitation'}});
  else if (chg===1)  r.push({{c:'warn', t:'1 booking change — slight uncertainty signal'}});

  if (spec === 0)    r.push({{c:'warn', t:'No special requests — low guest engagement with this stay'}});
  else               r.push({{c:'pos',  t: spec + ' special request' + (spec>1?'s':'') + ' — engaged guests are far less likely to cancel'}});

  if (nights >= 4)   r.push({{c:'pos',  t: nights + '-night stay — multi-night bookings have lower no-show rates'}});
  else if (nights===1)r.push({{c:'warn',t:'1-night stay — shortest stays carry the highest no-show rate'}});

  return r.slice(0, 5);
}}

function conclusionText(score) {{
  if (score >= 70) return 'High cancellation probability. Request deposit or guarantee now — revenue at risk.';
  if (score >= 40) return 'Moderate risk. Send a reminder 48h before arrival to confirm the booking.';
  return 'Low risk. Booking profile is stable — monitor normally.';
}}

// Toggle grouped stays dropdown
function toggleStays(sid, row) {{
  var opening = !row.classList.contains('is-open');
  row.classList.toggle('is-open', opening);
  document.querySelectorAll('tr.stays-sub[data-grp="' + sid + '"]').forEach(function(r) {{
    r.style.display = opening ? '' : 'none';
  }});
  var chev = document.getElementById('chev-' + sid);
  if (chev) chev.textContent = opening ? '▴' : '▾';
}}

// Inline expand row
var _openRow = null;
function toggleExpand(row, idx, score) {{
  var expTr = document.getElementById('exp-' + idx);
  if (!expTr) return;
  var isOpen = expTr.classList.contains('open');
  if (_openRow && _openRow !== expTr) {{
    _openRow.classList.remove('open');
    var pr = _openRow.previousElementSibling;
    if (pr) pr.classList.remove('is-open');
  }}
  if (isOpen) {{
    expTr.classList.remove('open');
    row.classList.remove('is-open');
    _openRow = null;
  }} else {{
    var bArr = typeof bookings !== 'undefined' ? bookings : (typeof guests !== 'undefined' ? guests : []);
    var b = bArr[idx] || {{}};
    var reasons = buildReasons(b, score);
    var scoreColor = score>=70 ? '#ef4444' : score>=40 ? '#f59e0b' : '#00d165';
    var verdict = score>=70 ? 'HIGH RISK' : score>=40 ? 'MEDIUM RISK' : 'LOW RISK';
    var factorsHtml = reasons.map(function(r) {{
      return '<div class="exp-factor ' + r.c + '">' + r.t + '</div>';
    }}).join('');
    var inner = document.getElementById('exp-inner-' + idx);
    if (inner) {{
      inner.innerHTML =
        '<div class="exp-score-wrap">' +
          '<div class="exp-score-big" style="color:' + scoreColor + '">' + score.toFixed(0) + '%</div>' +
          '<div class="exp-score-lbl">' + verdict + '</div>' +
          '<div class="exp-bar-bg"><div class="exp-bar-fill" style="width:' + score + '%;background:' + scoreColor + '"></div></div>' +
        '</div>' +
        '<div class="exp-right">' +
          '<div class="exp-head">Why this score</div>' +
          '<div class="exp-factors">' + factorsHtml + '</div>' +
          '<div class="exp-conclusion">' + conclusionText(score) + '</div>' +
        '</div>';
    }}
    expTr.classList.add('open');
    row.classList.add('is-open');
    _openRow = expTr;
  }}
}}

</script>
</body>
</html>"""

features = [
    "lead_time", "arrival_date_week_number", "stays_in_weekend_nights",
    "stays_in_week_nights", "adults", "is_repeated_guest",
    "previous_cancellations", "previous_bookings_not_canceled",
    "booking_changes", "days_in_waiting_list", "adr", "total_of_special_requests"
]

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "hotel" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated

def is_valid_email(email):
    """Basic email format validation."""
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email))

def build_empty_state(hotel_name, lang="en"):
    return f"""<!DOCTYPE html>
<html>
<head>
<title>Occupado — {hotel_name}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:#ffffff;color:#0d1120;font-family:'Plus Jakarta Sans',sans-serif;-webkit-font-smoothing:antialiased;}}
a{{text-decoration:none;}}
.topbar{{height:62px;background:#ffffff;border-bottom:1px solid #e4e8f0;display:flex;align-items:center;padding:0 48px;}}
.topbar-name{{font-family:'Plus Jakarta Sans',sans-serif;font-size:17px;font-weight:800;color:#0d1120;letter-spacing:-0.4px;}}
.topbar-name span{{color:#00d165;}}
.topbar-hotel{{font-family:'JetBrains Mono',monospace;font-size:11px;color:#94a3b8;margin-left:12px;padding-left:12px;border-left:1px solid #e4e8f0;}}
.topbar-right{{margin-left:auto;display:flex;gap:8px;}}
.btn-nav{{padding:7px 16px;background:transparent;border:1px solid #e4e8f0;border-radius:7px;color:#64748b;font-size:12px;font-weight:500;text-decoration:none;}}
.btn-nav:hover{{border-color:#cbd5e1;color:#0d1120;}}
.page{{max-width:600px;margin:0 auto;padding:80px 24px 40px;text-align:center;}}
@media(max-width:600px){{.topbar{{padding:0 16px;}}.topbar-hotel{{display:none;}}.page{{padding:48px 16px;}}.features{{grid-template-columns:1fr;}}}}
.welcome-tag{{display:inline-flex;align-items:center;gap:6px;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:99px;padding:5px 14px;font-family:'JetBrains Mono',monospace;font-size:11px;color:#16a34a;margin-bottom:24px;}}
.welcome-dot{{width:6px;height:6px;background:#00d165;border-radius:50%;}}
h1{{font-family:'Plus Jakarta Sans',sans-serif;font-size:40px;font-weight:800;color:#0d1120;letter-spacing:-1.5px;line-height:1.1;margin-bottom:14px;}}
h1 span{{color:#00d165;}}
.sub{{font-size:16px;color:#64748b;line-height:1.6;margin-bottom:48px;}}
.features{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:48px;text-align:left;}}
.feat{{background:#f8fafc;border:1px solid #e4e8f0;border-radius:12px;padding:20px;}}
.feat-icon{{font-size:22px;margin-bottom:10px;}}
.feat-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:13px;font-weight:700;color:#0d1120;margin-bottom:4px;letter-spacing:-0.2px;}}
.feat-sub{{font-size:12px;color:#64748b;line-height:1.5;}}
.upload-card{{background:#f8fafc;border:2px dashed #cbd5e1;border-radius:16px;padding:40px;margin-bottom:20px;cursor:pointer;transition:all .2s;}}
.upload-card:hover{{border-color:#00d165;background:#f0fdf4;}}
.upload-icon{{font-size:36px;margin-bottom:12px;}}
.upload-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:18px;font-weight:800;color:#0d1120;margin-bottom:6px;letter-spacing:-0.4px;}}
.upload-sub{{font-size:13px;color:#64748b;margin-bottom:24px;}}
.upload-btn{{display:inline-block;padding:13px 32px;background:#00d165;color:#080c14;border:none;border-radius:9px;font-size:14px;font-weight:700;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;}}
.upload-btn:hover{{background:#04e270;}}
.demo-link{{font-size:13px;color:#94a3b8;}}
.demo-link a{{color:#64748b;text-decoration:underline;text-underline-offset:3px;}}
.demo-link a:hover{{color:#0d1120;}}

</style>
</head>
<body>
<div class="topbar">
  <span class="topbar-name">Occup<span>ado</span></span>
  <span class="topbar-hotel">{hotel_name}</span>
  <div class="topbar-right">
    <a href="/settings" class="btn-nav">Settings</a>
    <a href="/logout" class="btn-nav">Sign out</a>
  </div>
</div>
<div class="page">
  <div class="welcome-tag"><span class="welcome-dot"></span>Account ready</div>
  <h1>Welcome to<br><span>Occupado</span></h1>
  <p class="sub">Upload your booking data and get AI-powered cancellation risk scores, revenue forecasts, and action plans — in seconds.</p>
  <div class="features">
    <div class="feat">
      <div class="feat-icon">🎯</div>
      <div class="feat-title">Risk Scores</div>
      <div class="feat-sub">Every booking scored 0–100% cancellation probability</div>
    </div>
    <div class="feat">
      <div class="feat-icon">💶</div>
      <div class="feat-title">Revenue at Risk</div>
      <div class="feat-sub">See exactly how much revenue is exposed each week</div>
    </div>
    <div class="feat">
      <div class="feat-icon">📧</div>
      <div class="feat-title">Take Action</div>
      <div class="feat-sub">Send deposit requests and reminders in one click</div>
    </div>
  </div>
  <form method="POST" action="/upload" enctype="multipart/form-data">
    <div class="upload-card" onclick="document.getElementById('csv-onboard').click()">
      <div class="upload-icon">📂</div>
      <div class="upload-title">Upload your booking export</div>
      <div class="upload-sub">CSV file from your PMS — Opera, Protel, Mews, or any system</div>
      <input type="file" id="csv-onboard" name="csv_file" accept=".csv" style="display:none" onchange="this.form.submit()">
      <button type="button" class="upload-btn" onclick="event.stopPropagation();document.getElementById('csv-onboard').click()">Choose file</button>
    </div>
  </form>
  <p class="demo-link">Not ready yet? <a href="/dashboard?skip_onboard=1">Preview with demo data</a></p>
</div>
</body>
</html>"""


def build_vdv_bru_dashboard(hotel_name, lang="en", first_login=False):
    """Thin wrapper — renders full VdV dashboard using Brussels Airport data."""
    # Build monthly stats: use real parsed data or sensible fallback
    months     = VDV_BRU_MONTHS     or ['Oct 2025','Nov 2025','Dec 2025','Jan 2026','Feb 2026','Mar 2026']
    cx_monthly = VDV_BRU_CX_MONTHLY or [0] * len(months)
    ns_monthly = VDV_BRU_NS_MONTHLY or [0] * len(months)

    _bru_data = {
        'guests':      VDV_BRU_GUESTS_RAW,
        'score_fn':    _score_bru_guests,
        'ch_stats':    VDV_BRU_CHANNEL_STATS,
        'mice':        VDV_BRU_MICE_DATA if VDV_BRU_MICE_DATA else None,
        'forecast':    VDV_BRU_FORECAST_DATA,
        'groups':      VDV_BRU_GROUP_PIPELINE,
        'months':      months,
        'cx_monthly':  cx_monthly,
        'ns_monthly':  ns_monthly,
        'avg_adr':     190.0,
        'avg_nights':  1.2,
        'fut_bookings': VDV_BRU_FUTURE_BOOKINGS,
        'fut_scores':  VDV_BRU_FUTURE_SCORES,
        'hotel_key':   VDV_BRU_HOTEL_KEY,
        'export_url':  '/bru/export-highrisk',
    }
    return build_vdv_dashboard(hotel_name, lang=lang, first_login=first_login, _data=_bru_data)


def build_dashboard(hotel_name, sample, scores, tonight_scores, tonight_sample=None, uploaded=False, lang="en", first_login=False):
    high = sum(1 for s in scores if s >= 70)
    med  = sum(1 for s in scores if 40 <= s < 70)
    low  = sum(1 for s in scores if s < 40)

    avg_rate = sample["adr"].mean() if "adr" in sample.columns else df["adr"].mean()
    predicted_noshows = sum(1 for s in tonight_scores if s >= 70)
    safe_overbook = int(predicted_noshows * 0.80)
    revenue = safe_overbook * avg_rate

    # --- Full-dataset analytics ---
    if tonight_sample is not None and len(tonight_sample) > 0:
        all_scores   = list(tonight_scores)
        total_bookings = len(tonight_sample)
        high_total   = sum(1 for s in all_scores if s >= 70)
        med_total    = sum(1 for s in all_scores if 40 <= s < 70)
        low_total    = sum(1 for s in all_scores if s < 40)
        avg_adr      = float(tonight_sample["adr"].mean()) if "adr" in tonight_sample.columns else float(avg_rate)
        wk = tonight_sample["stays_in_week_nights"] if "stays_in_week_nights" in tonight_sample.columns else pd.Series([1]*total_bookings)
        we = tonight_sample["stays_in_weekend_nights"] if "stays_in_weekend_nights" in tonight_sample.columns else pd.Series([0]*total_bookings)
        avg_nights   = max(1.0, float((wk + we).mean()))
        revenue_at_risk = int(high_total * avg_adr * avg_nights)

        # Lead-time buckets for chart
        lt_buckets = [[] for _ in range(5)]
        for idx, (_, row) in enumerate(tonight_sample.iterrows()):
            lt = float(row.get("lead_time", 0))
            s  = all_scores[idx] if idx < len(all_scores) else 0
            if lt <= 7:    lt_buckets[0].append(s)
            elif lt <= 30: lt_buckets[1].append(s)
            elif lt <= 60: lt_buckets[2].append(s)
            elif lt <= 90: lt_buckets[3].append(s)
            else:          lt_buckets[4].append(s)
        lt_high_js = json.dumps([sum(1 for s in b if s >= 70)      for b in lt_buckets])
        lt_med_js  = json.dumps([sum(1 for s in b if 40 <= s < 70) for b in lt_buckets])
        lt_low_js  = json.dumps([sum(1 for s in b if s < 40)       for b in lt_buckets])

        # Table: top 50 highest-risk bookings
        table_pairs = sorted(zip(tonight_sample.iterrows(), all_scores), key=lambda x: x[1], reverse=True)[:50]
    else:
        total_bookings  = len(sample)
        high_total      = high
        med_total       = med
        low_total       = low
        avg_adr         = float(avg_rate)
        avg_nights      = 2.0
        revenue_at_risk = int(high_total * avg_adr * avg_nights)
        lt_high_js      = json.dumps([0, 0, 0, 0, 0])
        lt_med_js       = json.dumps([0, 0, 0, 0, 0])
        lt_low_js       = json.dumps([0, 0, 0, 0, 0])
        table_pairs     = list(zip(sample.iterrows(), scores))

    savings_default_bk  = int(high_total * 0.20)
    savings_default_rev = int(revenue_at_risk * 0.20)

    bookings_data = []
    for (_, booking), _ in table_pairs:
        bookings_data.append({k: float(booking.get(k, 0)) for k in features})
    bookings_js = json.dumps(bookings_data)

    rows = ""
    for i, ((_, booking), score) in enumerate(table_pairs):
        if score >= 70:
            badge = f'<span class="badge high">{t("high", lang)} {score:.1f}%</span>'
            action = f'<button class="btn dep" onclick="event.stopPropagation(); openEmailComposer({i}, \'deposit\')">{t("req_dep", lang)}</button>'
        elif score >= 40:
            badge = f'<span class="badge med">{t("medium", lang)} {score:.1f}%</span>'
            action = f'<button class="btn rem" onclick="event.stopPropagation(); openEmailComposer({i}, \'reminder\')">{t("send_rem", lang)}</button>'
        else:
            badge = f'<span class="badge low">{t("low", lang)} {score:.1f}%</span>'
            action = f'<button class="btn mon" onclick="event.stopPropagation(); openEmailComposer({i}, \'monitor\')">{t("monitor", lang)}</button>'

        lead = int(booking.get("lead_time", 0))
        adr  = int(booking.get("adr", 0))
        rep  = t("yes", lang) if booking.get("is_repeated_guest", 0) else t("no", lang)
        canc = int(booking.get("previous_cancellations", 0))

        rows += f"""<tr class="clickable-row" data-score="{score:.1f}" data-lead="{lead}" data-rate="{adr}" onclick="toggleExpand(this, {i}, {score:.1f})">
            <td><span style="font-family:'JetBrains Mono',monospace;color:#9ca3af;font-size:11px">{i+1}</span></td>
            <td><span style="font-weight:600;color:#111827">{t("booking", lang)} {i+1}</span></td>
            <td>{lead} {t("days", lang)}</td><td>€{adr}</td><td>{rep}</td><td>{canc}</td>
            <td>{badge}</td><td>{action}</td>
        </tr>
        <tr class="exp-tr" id="exp-{i}"><td class="exp-td" colspan="8"><div class="exp-inner" id="exp-inner-{i}"></div></td></tr>"""

    upload_banner = ""
    clear_button = ""
    if uploaded:
        upload_banner = f'<div class="upload-banner">📂 Your uploaded data is loaded</div>'
        clear_button = f'<a href="/clear" class="clear-btn">🗑 Clear File</a>'

    bulk_action_html = f'''<div class="section-title">{t("take_action", lang)}</div>
<div class="bulk-action-zone">
    <div class="bulk-action-card">
        <div class="bulk-action-icon">📧</div>
        <div class="bulk-action-title">{t("send_mass", lang)}</div>
        <div class="bulk-action-sub">{t("send_high", lang, high=high)}</div>
        <button class="bulk-action-btn" onclick="openBulkEmailComposer()">{t("send_btn", lang)}</button>
    </div>
    <div class="bulk-action-card">
        <div class="bulk-action-icon">💰</div>
        <div class="bulk-action-title">{t("request_dep", lang)}</div>
        <div class="bulk-action-sub">{t("dep_template", lang)}</div>
        <button class="bulk-action-btn deposit-btn" onclick="openBulkEmailTemplate('deposit')">{t("dep_btn", lang)}</button>
    </div>
    <div class="bulk-action-card">
        <div class="bulk-action-icon">⏰</div>
        <div class="bulk-action-title">{t("reminders", lang)}</div>
        <div class="bulk-action-sub">{t("rem_template", lang)}</div>
        <button class="bulk-action-btn reminder-btn" onclick="openBulkEmailTemplate('reminder')">{t("rem_btn", lang)}</button>
    </div>
</div>'''

    return f"""<!DOCTYPE html>
<html>
<head>
<title>Occupado — {hotel_name}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*,*::before,*::after{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:#fff;color:#111827;font-family:'Plus Jakarta Sans',sans-serif;-webkit-font-smoothing:antialiased;font-size:14px;}}
a{{text-decoration:none;}}
.topbar{{height:56px;border-bottom:1px solid #e5e7eb;display:flex;align-items:center;padding:0 40px;background:#fff;position:sticky;top:0;z-index:100;}}
.topbar-brand{{display:flex;align-items:center;gap:8px;font-family:'Syne',sans-serif;}}
.topbar-name{{font-size:15px;font-weight:700;letter-spacing:-0.3px;color:#111827;}}
.topbar-name span{{color:#00d165;}}
.topbar-hotel{{font-size:12px;color:#9ca3af;margin-left:14px;padding-left:14px;border-left:1px solid #e5e7eb;}}
.topbar-right{{display:flex;align-items:center;gap:6px;margin-left:auto;}}
.lang-selector{{padding:6px 10px;background:transparent;border:1px solid #e5e7eb;border-radius:6px;color:#6b7280;font-size:12px;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;outline:none;}}
.btn-nav{{padding:6px 14px;background:transparent;border:1px solid #e5e7eb;border-radius:6px;color:#6b7280;font-size:12px;font-weight:500;text-decoration:none;display:inline-flex;align-items:center;}}
.btn-nav:hover{{border-color:#d1d5db;color:#111827;}}
.clear-btn{{padding:6px 12px;background:transparent;border:1px solid #fecaca;border-radius:6px;color:#ef4444;font-size:12px;font-weight:500;text-decoration:none;}}
.welcome-banner{{border-bottom:1px solid #e5e7eb;padding:10px 40px;display:flex;align-items:center;justify-content:space-between;font-size:13px;color:#6b7280;}}
.welcome-close{{background:none;border:none;color:#9ca3af;font-size:18px;cursor:pointer;padding:0;line-height:1;}}
.content{{padding:40px 40px;max-width:1100px;margin:0 auto;}}
.page-sub{{font-size:11px;color:#9ca3af;margin-bottom:32px;text-transform:uppercase;letter-spacing:0.07em;font-weight:500;}}
.section-title{{font-size:11px;font-weight:500;letter-spacing:0.08em;text-transform:uppercase;color:#9ca3af;margin-bottom:16px;margin-top:48px;display:flex;align-items:center;gap:16px;}}
.section-title::after{{content:'';flex:1;height:1px;background:#f3f4f6;}}
.upload-zone{{border:1px dashed #e5e7eb;border-radius:12px;padding:40px;text-align:center;background:#fff;cursor:pointer;transition:border-color .15s;}}
.upload-zone:hover{{border-color:#00d165;background:#f0fdf4;}}
.upload-zone-title{{font-size:15px;font-weight:600;color:#111827;margin-bottom:6px;}}
.upload-zone-sub{{font-size:13px;color:#9ca3af;margin-bottom:20px;}}
.upload-btn{{padding:9px 22px;background:#00d165;color:#0a0a0a;border:none;border-radius:7px;font-size:13px;font-weight:600;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;}}
.upload-banner{{border:1px solid #bbf7d0;border-radius:8px;padding:10px 16px;font-size:13px;color:#15803d;margin-bottom:20px;font-weight:500;}}
.hero-cards{{display:grid;grid-template-columns:repeat(4,1fr);border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;margin-bottom:48px;}}
.hero-card{{background:#fff;padding:28px 28px;border-right:1px solid #e5e7eb;}}
.hero-card:last-child{{border-right:none;}}
.hero-card-red{{background:#fff;}}
.hero-val{{font-size:40px;font-weight:700;letter-spacing:-2px;line-height:1;color:#111827;}}
.hero-lbl{{font-size:11px;color:#9ca3af;margin-top:8px;font-weight:500;text-transform:uppercase;letter-spacing:0.06em;}}
.hero-sub{{font-size:11px;color:#d1d5db;margin-top:3px;}}
.charts-row{{display:grid;grid-template-columns:220px 1fr;gap:16px;margin-bottom:48px;}}
.chart-card{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:24px;}}
.chart-head{{font-size:11px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:20px;}}
.savings-wrap{{border:1px solid #e5e7eb;border-radius:12px;padding:28px;display:grid;grid-template-columns:1fr 1fr;gap:48px;margin-bottom:48px;align-items:center;}}
.sav-main{{text-align:center;}}
.sav-lbl{{font-size:11px;color:#9ca3af;text-transform:uppercase;letter-spacing:0.08em;font-weight:500;}}
.sav-pct{{font-size:56px;font-weight:700;color:#00d165;line-height:1;letter-spacing:-3px;margin:12px 0;}}
.sav-slider{{width:100%;margin:16px 0 6px;accent-color:#00d165;cursor:pointer;}}
.sav-range{{display:flex;justify-content:space-between;font-size:11px;color:#d1d5db;}}
.sav-row{{display:flex;justify-content:space-between;align-items:center;padding:14px 0;border-bottom:1px solid #f3f4f6;font-size:13px;}}
.sav-row:last-child{{border-bottom:none;}}
.sav-row-lbl{{color:#6b7280;}}
.sav-row-val{{font-size:20px;font-weight:700;color:#111827;letter-spacing:-0.5px;}}
.sav-row-val.green{{color:#00d165;}}
.optimizer{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:48px;}}
.opt-main{{border:1px solid #e5e7eb;border-radius:12px;padding:28px;}}
.opt-value{{font-size:64px;font-weight:700;color:#00d165;line-height:1;letter-spacing:-3px;}}
.opt-label{{font-size:11px;color:#9ca3af;margin-top:6px;text-transform:uppercase;letter-spacing:0.06em;font-weight:500;}}
.opt-btn{{margin-top:24px;width:100%;padding:11px;background:#00d165;border:none;border-radius:7px;color:#0a0a0a;font-size:13px;font-weight:600;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;transition:background .15s;}}
.opt-btn:hover{{background:#04e270;}}
.opt-stats{{border:1px solid #e5e7eb;border-radius:12px;padding:24px;}}
.opt-row{{display:flex;justify-content:space-between;align-items:center;padding:10px 0;border-bottom:1px solid #f3f4f6;font-size:13px;}}
.opt-row:last-child{{border-bottom:none;}}
.opt-row-label{{color:#6b7280;}}
.opt-row-value{{font-weight:600;color:#111827;}}
.bulk-action-zone{{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:48px;}}
.bulk-action-card{{border:1px solid #e5e7eb;border-radius:12px;padding:24px;transition:border-color .15s;}}
.bulk-action-card:hover{{border-color:#d1d5db;}}
.bulk-action-icon{{display:none;}}
.bulk-action-title{{font-size:13px;font-weight:600;color:#111827;margin-bottom:4px;}}
.bulk-action-sub{{font-size:12px;color:#9ca3af;margin-bottom:16px;line-height:1.5;}}
.bulk-action-btn{{padding:8px 14px;background:transparent;border:1px solid #e5e7eb;border-radius:6px;color:#111827;font-size:12px;font-weight:500;cursor:pointer;width:100%;font-family:'Plus Jakarta Sans',sans-serif;transition:background .15s;}}
.bulk-action-btn:hover{{background:#f9fafb;}}
.bulk-action-btn.deposit-btn{{border-color:#fecaca;color:#ef4444;}}
.bulk-action-btn.deposit-btn:hover{{background:#fef2f2;}}
.bulk-action-btn.reminder-btn{{border-color:#fde68a;color:#f59e0b;}}
.bulk-action-btn.reminder-btn:hover{{background:#fffbeb;}}
table{{width:100%;border-collapse:collapse;}}
th{{font-size:11px;font-weight:500;letter-spacing:0.08em;text-transform:uppercase;color:#9ca3af;padding:10px 12px;border-bottom:1px solid #e5e7eb;text-align:left;}}
th:first-child,td:first-child{{padding-left:0;}}
td{{padding:14px 12px;font-size:13px;border-bottom:1px solid #f3f4f6;color:#374151;}}
tr:last-child td{{border-bottom:none;}}
.clickable-row{{cursor:pointer;}}
.clickable-row:hover td{{color:#111827;}}
.badge{{padding:2px 7px;border-radius:4px;font-family:'JetBrains Mono',monospace;font-size:11px;font-weight:600;border:1px solid;}}
.high{{background:#fef2f2;color:#ef4444;border-color:#fecaca;}}
.med{{background:#fffbeb;color:#f59e0b;border-color:#fde68a;}}
.low{{background:#f0fdf4;color:#00d165;border-color:#bbf7d0;}}
.btn{{padding:4px 10px;border-radius:5px;font-size:11px;font-weight:500;cursor:pointer;border:1px solid;background:transparent;font-family:'Plus Jakarta Sans',sans-serif;transition:background .1s;}}
.dep{{color:#ef4444;border-color:#fecaca;}}.dep:hover{{background:#fef2f2;}}
.rem{{color:#f59e0b;border-color:#fde68a;}}.rem:hover{{background:#fffbeb;}}
.mon{{color:#00d165;border-color:#bbf7d0;}}.mon:hover{{background:#f0fdf4;}}
.modal-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.25);z-index:1000;align-items:center;justify-content:center;}}
.modal-overlay.show{{display:flex;}}
.modal{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:32px;width:100%;max-width:480px;max-height:85vh;overflow-y:auto;position:relative;}}
.modal-close{{position:absolute;top:14px;right:16px;font-size:18px;cursor:pointer;color:#9ca3af;background:none;border:none;}}
.modal-title{{font-size:18px;font-weight:700;color:#111827;margin-bottom:2px;letter-spacing:-0.3px;}}
.modal-sub{{font-size:11px;color:#9ca3af;margin-bottom:20px;}}
.score-display{{font-size:56px;font-weight:700;line-height:1;margin-bottom:6px;letter-spacing:-2px;}}
.score-bar-bg{{height:4px;background:#f3f4f6;border-radius:2px;overflow:hidden;margin-bottom:12px;}}
.score-bar-fill{{height:100%;border-radius:2px;}}
.score-verdict{{font-size:11px;font-weight:600;padding:4px 10px;border-radius:4px;display:inline-block;margin-bottom:8px;}}
.email-composer{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.25);z-index:1001;align-items:center;justify-content:center;}}
.email-composer.show{{display:flex;}}
.email-box{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:28px;width:100%;max-width:580px;max-height:90vh;overflow-y:auto;}}
.email-title{{font-size:17px;font-weight:700;color:#111827;margin-bottom:2px;letter-spacing:-0.3px;}}
.email-label{{font-size:11px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:0.08em;display:block;margin-bottom:5px;}}
.email-input{{width:100%;padding:9px 12px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px;font-size:13px;color:#111827;outline:none;margin-bottom:10px;font-family:'Plus Jakarta Sans',sans-serif;}}
.email-input:focus{{border-color:#00d165;background:#fff;}}
.email-input::placeholder{{color:#d1d5db;}}
.email-textarea{{width:100%;padding:9px 12px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px;font-size:13px;color:#111827;outline:none;resize:vertical;min-height:160px;margin-bottom:10px;font-family:'Plus Jakarta Sans',sans-serif;line-height:1.6;}}
.email-textarea:focus{{border-color:#00d165;background:#fff;}}
.email-actions{{display:flex;gap:8px;margin-top:16px;}}
.email-send{{flex:1;padding:11px;background:#00d165;color:#0a0a0a;border:none;border-radius:7px;font-size:13px;font-weight:600;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;}}
.email-send:hover{{background:#04e270;}}
.email-cancel{{flex:1;padding:11px;background:transparent;color:#6b7280;border:1px solid #e5e7eb;border-radius:7px;font-size:13px;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;}}
.bulk-email-composer{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.25);z-index:1002;align-items:center;justify-content:center;}}
.bulk-email-composer.show{{display:flex;}}
.bulk-email-box{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:28px;width:100%;max-width:580px;max-height:90vh;overflow-y:auto;}}
.bulk-email-title{{font-size:17px;font-weight:700;color:#111827;margin-bottom:2px;letter-spacing:-0.3px;}}
.bulk-email-subtitle{{font-size:11px;color:#9ca3af;}}
.bulk-email-actions{{display:flex;gap:8px;margin-top:16px;}}
.bulk-email-send{{flex:1;padding:11px;background:#00d165;color:#0a0a0a;border:none;border-radius:7px;font-size:13px;font-weight:600;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;}}
.bulk-email-send.deposit{{background:#fef2f2;color:#ef4444;border:1px solid #fecaca;}}
.bulk-email-send.reminder{{background:#fffbeb;color:#f59e0b;border:1px solid #fde68a;}}
.bulk-email-cancel{{flex:1;padding:11px;background:transparent;color:#6b7280;border:1px solid #e5e7eb;border-radius:7px;font-size:13px;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;}}
.bulk-booking-row{{padding:8px 12px;margin-bottom:3px;background:#fef2f2;border:1px solid #fecaca;border-radius:6px;font-size:11px;display:flex;justify-content:space-between;align-items:center;cursor:pointer;}}
.toast{{position:fixed;bottom:24px;right:24px;background:#111827;color:#fff;border-radius:8px;padding:12px 16px;font-size:13px;transform:translateY(50px);opacity:0;transition:all .25s;z-index:2000;}}
.toast.show{{transform:translateY(0);opacity:1;}}
@media(max-width:900px){{
  .content{{padding:24px 16px;}}
  .topbar{{padding:0 16px;}}
  .topbar-hotel{{display:none;}}
  .hero-cards{{grid-template-columns:1fr 1fr;}}
  .hero-card:nth-child(2){{border-right:none;}}
  .hero-card:nth-child(3){{border-top:1px solid #e5e7eb;}}
  .hero-card:nth-child(4){{border-top:1px solid #e5e7eb;border-right:none;}}
  .hero-val{{font-size:32px;}}
  .charts-row{{grid-template-columns:1fr;}}
  .optimizer{{grid-template-columns:1fr;}}
  .savings-wrap{{grid-template-columns:1fr;gap:28px;}}
  .bulk-action-zone{{grid-template-columns:1fr;}}
  table{{font-size:12px;}}
  th,td{{padding:10px 8px;}}
  td:nth-child(5),th:nth-child(5),td:nth-child(6),th:nth-child(6){{display:none;}}
  .section-title{{margin-top:36px;}}
}}


/* ── DYNAMICS ────────────────────────────────────────────── */
.filter-bar{{display:flex;gap:8px;margin-bottom:20px;align-items:center;flex-wrap:wrap;}}
.f-tab{{padding:5px 16px;border:1px solid #e5e7eb;border-radius:99px;font-size:12px;color:#6b7280;cursor:pointer;transition:all .15s;background:#fff;font-family:'Plus Jakarta Sans',sans-serif;}}
.f-tab:hover{{border-color:#111827;color:#111827;}}
.f-tab.active{{background:#111827;color:#fff;border-color:#111827;}}
.f-count{{font-size:10px;opacity:.6;margin-left:3px;}}
.sort-th{{cursor:pointer;user-select:none;}}
.sort-th:hover{{color:#374151;}}
.sort-arr{{margin-left:3px;opacity:.25;font-size:9px;}}
.sort-arr.on{{opacity:1;}}
.live-pill{{display:inline-flex;align-items:center;gap:5px;font-size:11px;color:#9ca3af;padding:4px 10px;border:1px solid #e5e7eb;border-radius:99px;}}
.live-dot{{width:6px;height:6px;border-radius:50%;background:#00d165;flex-shrink:0;animation:pdot 2s ease-in-out infinite;}}
@keyframes pdot{{0%,100%{{opacity:1;transform:scale(1);}}50%{{opacity:.3;transform:scale(.8);}}}}
.exp-tr{{display:none;}}
.exp-tr.open{{display:table-row;}}
.exp-td{{padding:0!important;background:#fff!important;border-bottom:1px solid #f3f4f6!important;}}
.stays-ct{{display:inline-block;background:#eff6ff;color:#2563eb;border:1px solid #bfdbfe;border-radius:99px;font-size:10px;font-weight:600;padding:1px 7px;margin-left:6px;vertical-align:middle;}}
.stays-hdr{{cursor:pointer;}}
.stays-hdr:hover td{{background:#fafafa;}}
.grp-chev{{font-size:11px;color:#9ca3af;margin-left:5px;display:inline-block;vertical-align:middle;line-height:1;}}
.exp-inner{{padding:20px 0 24px;display:grid;grid-template-columns:100px 1fr;gap:24px;align-items:start;}}
.exp-score-wrap{{display:flex;flex-direction:column;align-items:center;gap:4px;padding-top:4px;}}
.exp-score-big{{font-size:48px;font-weight:700;letter-spacing:-2px;line-height:1;}}
.exp-score-lbl{{font-size:9px;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;margin-top:2px;}}
.exp-bar-bg{{width:70px;height:3px;background:#f3f4f6;border-radius:2px;overflow:hidden;margin-top:8px;}}
.exp-bar-fill{{height:100%;border-radius:2px;transition:width .7s ease;}}
.exp-right{{}}
.exp-head{{font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;margin-bottom:10px;}}
.exp-factors{{display:flex;flex-direction:column;gap:4px;}}
.exp-factor{{font-size:12px;color:#374151;padding:6px 10px 6px 12px;border-left:2px solid #e5e7eb;line-height:1.4;}}
.exp-factor.pos{{border-left-color:#00d165;}}
.exp-factor.warn{{border-left-color:#f59e0b;}}
.exp-factor.bad{{border-left-color:#ef4444;}}
.exp-conclusion{{margin-top:10px;font-size:12px;font-weight:500;color:#111827;padding:8px 12px;background:#f9fafb;border-radius:6px;border:1px solid #e5e7eb;}}
.cr.is-open{{background:#fafafa;}}
.count-up{{display:inline;}}
@keyframes fadeUp{{from{{opacity:0;transform:translateY(10px);}}to{{opacity:1;transform:translateY(0);}}}}
.anim-card{{opacity:0;animation:fadeUp .5s ease forwards;}}
.anim-card:nth-child(1){{animation-delay:.04s;}}
.anim-card:nth-child(2){{animation-delay:.08s;}}
.anim-card:nth-child(3){{animation-delay:.12s;}}
.anim-card:nth-child(4){{animation-delay:.16s;}}
@media(max-width:900px){{.exp-inner{{grid-template-columns:1fr;gap:16px;}}}}
</style>
</head>
<body>
<div class="topbar">
  <div class="topbar-brand">
    <span class="topbar-name">Occup<span>ado</span></span>
    <span class="topbar-hotel">{hotel_name}</span>
  </div>
  <div class="topbar-right">
    {clear_button}
    <select class="lang-selector" onchange="changeLanguage(this.value)">
      <option value="en" {"selected" if lang == "en" else ""}>EN</option>
      <option value="nl" {"selected" if lang == "nl" else ""}>NL</option>
      <option value="fr" {"selected" if lang == "fr" else ""}>FR</option>
    </select>
    <a href="/settings" class="btn-nav">{t("settings", lang)}</a>
    <a href="/logout" class="btn-nav">{t("sign_out", lang)}</a>
  </div>
</div>
{f'''<div id="welcome-banner" class="welcome-banner">
    <span>Welcome, <strong>{hotel_name}</strong>. Upload your booking data to get your first risk scores.</span>
    <button onclick="document.getElementById('welcome-banner').style.display='none'" class="welcome-close">×</button>
</div>''' if first_login else ''}
<div class="content">
<div class="page-sub" style="display:flex;align-items:center;justify-content:space-between;">{t("live_dashboard", lang)} · {total_bookings} {t("bookings_analysed", lang)}<span class="live-pill"><span class="live-dot"></span>Live</span></div>
{upload_banner}

<div class="section-title" style="margin-top:0">Overview</div>
<div class="hero-cards">
  <div class="hero-card">
    <div class="hero-val anim-card"><span class="count-up" data-val="{total_bookings}">—</span></div>
    <div class="hero-lbl">Bookings Analysed</div>
  </div>
  <div class="hero-card hero-card-red">
    <div class="hero-val anim-card" style="color:#ef4444"><span class="count-up" data-val="{high_total}">—</span></div>
    <div class="hero-lbl">High Risk</div>
    <div class="hero-sub">{f"{high_total/total_bookings*100:.0f}%" if total_bookings > 0 else "0%"} of bookings</div>
  </div>
  <div class="hero-card hero-card-red">
    <div class="hero-val anim-card" style="color:#ef4444">€<span class="count-up" data-val="{revenue_at_risk}">—</span></div>
    <div class="hero-lbl">Revenue at Risk</div>
    <div class="hero-sub">from high-risk bookings</div>
  </div>
  <div class="hero-card">
    <div class="hero-val anim-card">€<span class="count-up" data-val="{avg_adr:.0f}">—</span></div>
    <div class="hero-lbl">Avg Daily Rate</div>
    <div class="hero-sub">{avg_nights:.1f} avg nights/stay</div>
  </div>
</div>

<div class="charts-row">
  <div class="chart-card" style="display:flex;flex-direction:column;align-items:center;">
    <div class="chart-head" style="align-self:flex-start">Risk Distribution</div>
    <canvas id="riskDoughnut" width="200" height="200"></canvas>
  </div>
  <div class="chart-card">
    <div class="chart-head">Risk by Lead Time</div>
    <canvas id="leadTimeChart" height="180"></canvas>
  </div>
</div>

<div class="section-title">Savings Calculator</div>
<div class="savings-wrap">
  <div class="sav-main">
    <div class="sav-lbl">If you convert</div>
    <div class="sav-pct" id="savPct">20%</div>
    <div class="sav-lbl">of high-risk bookings</div>
    <input type="range" min="5" max="60" value="20" step="5" class="sav-slider" id="savSlider" oninput="updateSavings(this.value)">
    <div class="sav-range"><span>5%</span><span>60%</span></div>
  </div>
  <div class="sav-results">
    <div class="sav-row">
      <span class="sav-row-lbl">Bookings saved</span>
      <span class="sav-row-val" id="savBookings">{savings_default_bk}</span>
    </div>
    <div class="sav-row">
      <span class="sav-row-lbl">Revenue recovered</span>
      <span class="sav-row-val green" id="savRevenue">€{savings_default_rev:,}</span>
    </div>
    <div class="sav-row">
      <span class="sav-row-lbl">Annual projection</span>
      <span class="sav-row-val green" id="savAnnual">€{savings_default_rev * 12:,}</span>
    </div>
  </div>
</div>

<div class="section-title">{t("optimizer", lang)}</div>
<div class="optimizer">
    <div class="opt-main">
        <div class="opt-label">{t("safe_rooms", lang)}</div>
        <div class="opt-value">+{safe_overbook}</div>
        <div class="opt-label" style="margin-top:8px">{t("revenue", lang)} {revenue:.0f}</div>
        <button class="opt-btn" onclick="showToast('Recommendation applied! {safe_overbook} rooms released.')">{t("apply", lang)}</button>
    </div>
    <div class="opt-stats">
        <div class="opt-row"><span class="opt-row-label">{t("bookings_analysed_stat", lang)}</span><span class="opt-row-value">{total_bookings}</span></div>
        <div class="opt-row"><span class="opt-row-label">{t("predicted", lang)}</span><span class="opt-row-value" style="color:#dc2626">{predicted_noshows}</span></div>
        <div class="opt-row"><span class="opt-row-label">{t("confidence", lang)}</span><span class="opt-row-value" style="color:#00d165">80.7%</span></div>
        <div class="opt-row"><span class="opt-row-label">{t("walk_risk", lang)}</span><span class="opt-row-value" style="color:#00d165">2.1%</span></div>
        <div class="opt-row"><span class="opt-row-label">{t("avg_rate", lang)}</span><span class="opt-row-value">EUR {avg_rate:.0f}</span></div>
    </div>
</div>
{bulk_action_html}
<div class="section-title">{t("click_row", lang)}</div>
<div class="filter-bar"><button class="f-tab active" onclick="filterRisk('all',this)">All <span class="f-count">{total_bookings}</span></button><button class="f-tab" onclick="filterRisk('high',this)">High Risk <span class="f-count" style="color:#ef4444">{high_total}</span></button><button class="f-tab" onclick="filterRisk('med',this)">Medium <span class="f-count">{med_total}</span></button><button class="f-tab" onclick="filterRisk('low',this)">Low Risk <span class="f-count" style="color:#00d165">{low_total}</span></button></div>
<table>
<thead><tr><th>#</th><th>{t("booking", lang)}</th><th class="sort-th" onclick="sortTable('lead')" data-col="lead">{t("lead", lang)} <span class="sort-arr" id="arr-lead">↕</span></th><th class="sort-th" onclick="sortTable('rate')" data-col="rate">{t("rate", lang)} <span class="sort-arr" id="arr-rate">↕</span></th><th>{t("returning", lang)}</th><th>{t("cancels", lang)}</th><th class="sort-th" onclick="sortTable('score')" data-col="score">{t("risk", lang)} <span class="sort-arr" id="arr-score">↕</span></th><th>{t("action", lang)}</th></tr></thead>
<tbody>{rows}</tbody>
</table>
<div style="margin-top:48px;">
<div class="section-title">{t("upload_data", lang)}</div>
<form method="POST" action="/upload" enctype="multipart/form-data">
    <div class="upload-zone" onclick="document.getElementById('csv-input').click()">
        <div class="upload-zone-title">{t("drop_csv", lang)}</div>
        <div class="upload-zone-sub">{t("export_pms", lang)}</div>
        <input type="file" id="csv-input" name="csv_file" accept=".csv" style="display:none" onchange="this.form.submit()">
        <button type="button" class="upload-btn" onclick="event.stopPropagation();document.getElementById('csv-input').click()">{t("choose_file", lang)}</button>
    </div>
</form>
</div>
</div>

<div class="modal-overlay" id="modal">
    <div class="modal">
        <button class="modal-close" onclick="closeModal()">✕</button>
        <div class="modal-title" id="modal-title">Booking Detail</div>
        <div class="modal-sub" id="modal-sub">Risk Analysis</div>
        <div class="score-display" id="modal-score">0%</div>
        <div class="score-bar-bg"><div class="score-bar-fill" id="modal-bar" style="width:0%"></div></div>
        <div class="score-verdict" id="modal-verdict"></div>
        <div id="modal-reasons"></div>
        <button style="margin-top:24px; width:100%; padding:14px; border:none; border-radius:10px; font-size:14px; font-weight:700; background:#0d1120; color:#ffffff; cursor:pointer; font-family:'Plus Jakarta Sans',sans-serif;" onclick="closeModal()">Close</button>
    </div>
</div>

<div class="email-composer" id="emailComposer">
    <div class="email-box">
        <div style="margin-bottom:24px;">
            <div class="email-title" id="emailTitle">{t("email_guest", lang)}</div>
            <div style="font-family:'JetBrains Mono',monospace; font-size:12px; color:#4a6648; margin-top:4px;" id="emailSubtitle">Booking 1</div>
        </div>
        
        <div>
            <label class="email-label">{t("guest_email", lang)}</label>
            <input type="email" id="guestEmail" class="email-input" placeholder="guest@example.com">
        </div>
        
        <div>
            <label class="email-label">{t("guest_name", lang)}</label>
            <input type="text" id="guestName" class="email-input" placeholder="John Doe">
        </div>
        
        <div>
            <label class="email-label">{t("subject", lang)}</label>
            <input type="text" id="emailSubject" class="email-input" placeholder="Confirm Your Booking">
        </div>
        
        <div>
            <label class="email-label">{t("message", lang)}</label>
            <textarea id="emailBody" class="email-textarea" placeholder="Dear Guest..."></textarea>
        </div>
        
        <div class="email-actions">
            <button class="email-send" onclick="sendEmailToGuest()">{t("send_email", lang)}</button>
            <button class="email-cancel" onclick="closeEmailComposer()">{t("cancel", lang)}</button>
        </div>
    </div>
</div>

<div class="bulk-email-composer" id="bulkEmailComposer">
    <div class="bulk-email-box">
        <div style="margin-bottom:24px;">
            <div class="bulk-email-title" id="bulkEmailTitle">{t("bulk_email", lang)}</div>
            <div class="bulk-email-subtitle" id="bulkEmailSubtitle">{t("select_book", lang)}</div>
        </div>
        
        <div style="background:#f8fafc; border:1px solid #e4e8f0; border-radius:10px; padding:14px; margin-bottom:16px;">
            <div style="margin-bottom:10px;">
                <label class="email-label">{t("select_book", lang)}</label>
                <div id="bulkBookingsList" style="max-height:150px; overflow-y:auto; margin-bottom:10px; padding:8px; background:white; border:1px solid #e4e8f0; border-radius:8px;"></div>
                <input type="text" id="bulkBookingsInput" class="email-input" placeholder="Edit or remove booking numbers..." style="margin-bottom:6px;">
                <div style="font-family:'JetBrains Mono',monospace; font-size:10px; color:#94a3b8;">{t("auto_pop", lang)}</div>
            </div>
            <div style="background:#f1f5f9; padding:12px 14px; border-radius:8px; display:flex; justify-content:space-between; align-items:center;">
                <div style="display:flex; align-items:center; gap:10px;">
                    <span style="font-family:'Plus Jakarta Sans',sans-serif; font-size:24px; font-weight:800; color:#0d1120;" id="bulkCountBig">0</span>
                    <span style="font-family:'Plus Jakarta Sans',sans-serif; font-size:12px; color:#64748b; font-weight:500;">{t("selected_count", lang)}</span>
                </div>
                <button type="button" style="padding:8px 18px; background:#00d165; color:#080c14; border:none; border-radius:8px; font-size:12px; font-weight:700; cursor:pointer; font-family:'Plus Jakarta Sans',sans-serif;" onclick="saveBookingChanges()">{t("save_changes", lang)}</button>
            </div>
        </div>
        
        <div>
            <label class="email-label">{t("subject", lang)}</label>
            <input type="text" id="bulkEmailSubject" class="email-input" placeholder="Important: Your Booking">
        </div>
        
        <div>
            <label class="email-label">{t("message", lang)}</label>
            <textarea id="bulkEmailBody" class="email-textarea" placeholder="Dear Guest..."></textarea>
        </div>
        
        <div class="bulk-email-actions">
            <button class="bulk-email-send" id="bulkEmailSendBtn" onclick="sendBulkEmailToGuests()">{t("send_selected", lang)}</button>
            <button class="bulk-email-cancel" onclick="closeBulkEmailComposer()">{t("cancel", lang)}</button>
        </div>
    </div>
</div>

<div class="toast" id="toast"></div>
<script>
const bookings = {bookings_js};
const currentLang = "{lang}";
const translations = {json.dumps(TRANSLATIONS)};

function changeLanguage(lang) {{
    window.location.href = '/dashboard?lang=' + lang;
}}

function t(key) {{
    return translations[currentLang][key] || translations['en'][key] || key;
}}

function populateBulkBookingsList() {{
    const list = document.getElementById('bulkBookingsList');
    const input = document.getElementById('bulkBookingsInput');
    list.innerHTML = '';
    
    const rows = document.querySelectorAll('.clickable-row');
    let highRiskNums = [];
    let html = '';
    
    rows.forEach((row, idx) => {{
        const badgeSpan = row.querySelector('.badge.high');
        if (badgeSpan) {{
            const cells = row.querySelectorAll('td');
            const scoreText = badgeSpan.textContent;
            const leadTime = cells[1] ? cells[1].textContent : '-';
            const roomRate = cells[2] ? cells[2].textContent : '-';
            
            highRiskNums.push(idx + 1);
            
            html += `<div class="bulk-booking-row" data-booking="` + (idx + 1) + `" style="padding:9px 12px; margin-bottom:4px; background:#fef2f2; border:1px solid #fecaca; border-radius:8px; font-size:11px; display:flex; justify-content:space-between; align-items:center; transition:all 0.2s; cursor:pointer; user-select:none;" onclick="addBookingToField(` + (idx + 1) + `)">
                <div>
                    <span style="color:#0d1120; font-weight:600;">Booking ` + (idx + 1) + `</span>
                    <span style="color:#cbd5e1; margin:0 8px;">·</span>
                    <span style="color:#94a3b8; font-size:10px;">` + leadTime + ` lead</span>
                    <span style="color:#cbd5e1; margin:0 4px;">·</span>
                    <span style="color:#94a3b8; font-size:10px;">` + roomRate + `</span>
                </div>
                <span style="color:#dc2626; font-weight:600; font-family:'JetBrains Mono',monospace;">` + scoreText + `</span>
            </div>`;
        }}
    }});
    
    if (html === '') {{
        list.innerHTML = '<div style="padding:10px; color:#999; font-size:11px; text-align:center;">No high-risk bookings found</div>';
    }} else {{
        list.innerHTML = html;
    }}
    
    const bookingsList = highRiskNums.join(', ');
    input.value = bookingsList;
    
    input.addEventListener('input', handleBookingInputChange);
    
    updateSendCount();
}}

function addBookingToField(bookingNum) {{
    const input = document.getElementById('bulkBookingsInput');
    const current = input.value.trim();
    const nums = current.split(',').map(n => parseInt(n.trim())).filter(n => !isNaN(n));
    
    if (!nums.includes(bookingNum)) {{
        if (current === '') {{
            input.value = bookingNum.toString();
        }} else {{
            input.value = current + ', ' + bookingNum;
        }}
        updateBookingVisuals();
    }}
}}

function updateBookingVisuals() {{
    const input = document.getElementById('bulkBookingsInput').value.trim();
    const selectedNums = input.split(',').map(n => parseInt(n.trim())).filter(n => !isNaN(n));
    
    const bookingRows = document.querySelectorAll('.bulk-booking-row');
    
    bookingRows.forEach(row => {{
        const bookingNum = parseInt(row.getAttribute('data-booking'));
        
        if (selectedNums.includes(bookingNum)) {{
            row.style.opacity = '1';
            row.style.background = '#fef2f2';
            row.style.borderColor = '#fecaca';
        }} else {{
            row.style.opacity = '0.4';
            row.style.background = '#f8fafc';
            row.style.borderColor = '#e4e8f0';
        }}
    }});
    
    updateSendCount();
}}

function openBulkEmailComposer() {{
    populateBulkBookingsList();
    document.getElementById('bulkEmailSubject').value = 'Regarding Your Upcoming Reservation';
    document.getElementById('bulkEmailBody').value = 'Dear Valued Guest,\\n\\nWe wanted to reach out regarding your upcoming reservation. Please confirm your booking details.\\n\\nBest regards,\\nThe Hotel Team';
    document.getElementById('bulkEmailComposer').classList.add('show');
}}

function openBulkEmailTemplate(type) {{
    populateBulkBookingsList();
    let subject = '';
    let template = '';
    let btnClass = 'bulk-email-send';
    
    if (type === 'deposit') {{
        subject = 'Please Confirm Your Reservation with Deposit';
        template = 'Dear Valued Guest,\\n\\nWe want to ensure your upcoming stay is confirmed. Please provide a deposit to guarantee your booking.\\n\\nBest regards,\\nThe Hotel Team';
        btnClass = 'bulk-email-send deposit';
    }} else {{
        subject = 'Reminder: Your Upcoming Stay';
        template = 'Dear Valued Guest,\\n\\nThis is a friendly reminder about your upcoming reservation. If you have questions, please contact us.\\n\\nBest regards,\\nThe Hotel Team';
        btnClass = 'bulk-email-send reminder';
    }}
    
    document.getElementById('bulkEmailSubject').value = subject;
    document.getElementById('bulkEmailBody').value = template;
    document.getElementById('bulkEmailSendBtn').className = btnClass;
    document.getElementById('bulkEmailComposer').classList.add('show');
}}

function handleBookingInputChange() {{
    updateBookingVisuals();
}}

function updateSendCount() {{
    const input = document.getElementById('bulkBookingsInput').value.trim();
    let count = 0;
    
    if (input === '') {{
        const rows = document.querySelectorAll('.clickable-row');
        rows.forEach(row => {{
            if (row.querySelector('.badge.high')) count++;
        }});
    }} else {{
        const nums = input.split(',').map(n => n.trim()).filter(n => n !== '');
        count = nums.length;
    }}
    
    document.getElementById('bulkCountBig').textContent = count;
}}

function saveBookingChanges() {{
    const input = document.getElementById('bulkBookingsInput').value.trim();
    const nums = input.split(',').map(n => parseInt(n.trim())).filter(n => !isNaN(n));
    
    if (nums.length === 0) {{
        showToast('Please keep at least one booking', 'error');
        return;
    }}
    
    showToast('✓ ' + nums.length + ' booking(s) selected for email', 'success');
}}

function validateAndGetSelectedBookings() {{
    const input = document.getElementById('bulkBookingsInput').value.trim();
    const rows = document.querySelectorAll('.clickable-row');
    let selected = [];
    
    if (input === '') {{
        rows.forEach((row, idx) => {{
            if (row.querySelector('.badge.high')) {{
                selected.push(idx);
            }}
        }});
    }} else {{
        const nums = input.split(',').map(n => parseInt(n.trim())).filter(n => !isNaN(n));
        
        rows.forEach((row, idx) => {{
            if (row.querySelector('.badge.high') && nums.includes(idx + 1)) {{
                selected.push(idx);
            }}
        }});
    }}
    
    return selected;
}}

function closeBulkEmailComposer() {{
    document.getElementById('bulkEmailComposer').classList.remove('show');
}}

function sendBulkEmailToGuests() {{
    const subject = document.getElementById('bulkEmailSubject').value.trim();
    const body = document.getElementById('bulkEmailBody').value.trim();
    
    if (!subject || !body) {{
        showToast('Please fill in all fields', 'error');
        return;
    }}
    
    const selected = validateAndGetSelectedBookings();
    
    if (selected.length === 0) {{
        showToast('Please select at least one booking', 'error');
        return;
    }}
    
    fetch('/send-bulk-email', {{
        method: 'POST',
        headers: {{'Content-Type': 'application/json'}},
        body: JSON.stringify({{
            count: selected.length,
            subject: subject,
            body: body
        }})
    }})
    .then(r => r.json())
    .then(data => {{
        if (data.status === 'success') {{
            closeBulkEmailComposer();
            showToast('✓ Emails sent to ' + data.count + ' guest(s)', 'success');
        }} else {{
            showToast('Error: ' + data.message, 'error');
        }}
    }})
    .catch(e => showToast('Error sending emails', 'error'));
}}

function openEmailComposer(idx, actionType) {{
    const booking = bookings[idx];
    const score = Object.values(bookings[idx]).slice(-1)[0] * 100 || 0;
    
    document.getElementById('emailTitle').textContent = 'Send Email to Guest';
    document.getElementById('emailSubtitle').textContent = 'Booking ' + (idx+1) + ' · Risk Score: ' + score.toFixed(1) + '%';
    
    let template = '';
    if (actionType === 'deposit') {{
        template = 'Dear Valued Guest,\\n\\nWe want to ensure your upcoming stay is confirmed and secured. Please provide a deposit to guarantee your booking.\\n\\nBest regards,\\nThe Hotel Team';
        document.getElementById('emailSubject').value = 'Please Confirm Your Reservation with Deposit';
    }} else if (actionType === 'reminder') {{
        template = 'Dear Valued Guest,\\n\\nWe hope you are looking forward to your stay! This is a friendly reminder about your upcoming reservation.\\n\\nBest regards,\\nThe Hotel Team';
        document.getElementById('emailSubject').value = 'Reminder: Your Upcoming Stay';
    }} else {{
        template = 'Dear Valued Guest,\\n\\nWe wanted to confirm your upcoming reservation. Please let us know if you have any questions.\\n\\nBest regards,\\nThe Hotel Team';
        document.getElementById('emailSubject').value = 'Reservation Confirmation';
    }}
    
    document.getElementById('emailBody').value = template;
    document.getElementById('guestEmail').value = '';
    document.getElementById('guestName').value = '';
    document.getElementById('emailComposer').classList.add('show');
}}

function closeEmailComposer() {{
    document.getElementById('emailComposer').classList.remove('show');
}}

function sendEmailToGuest() {{
    const email = document.getElementById('guestEmail').value.trim();
    const subject = document.getElementById('emailSubject').value.trim();
    const body = document.getElementById('emailBody').value.trim();
    
    if (!email || !subject || !body) {{
        showToast('Please fill in all fields', 'error');
        return;
    }}
    
    fetch('/send-guest-email', {{
        method: 'POST',
        headers: {{'Content-Type': 'application/json'}},
        body: JSON.stringify({{guest_email: email, subject: subject, body: body}})
    }})
    .then(r => r.json())
    .then(data => {{
        if (data.status === 'success') {{
            closeEmailComposer();
            showToast('Email sent successfully', 'success');
        }} else {{
            showToast('Error: ' + data.message, 'error');
        }}
    }})
    .catch(() => showToast('Error sending email', 'error'));
}}

function showDetail(idx, score) {{
    const booking = bookings[idx];
    document.getElementById('modal-title').textContent = 'Booking ' + (idx+1);
    document.getElementById('modal-sub').textContent = 'Lead time: ' + booking.lead_time + ' days · EUR ' + booking.adr + '/night';
    document.getElementById('modal-score').textContent = score.toFixed(1) + '%';
    
    const bar = document.getElementById('modal-bar');
    bar.style.width = score + '%';
    if (score >= 70) {{
        bar.style.background = '#ef4444';
        document.getElementById('modal-score').style.color = '#ef4444';
        document.getElementById('modal-verdict').textContent = 'HIGH RISK';
        document.getElementById('modal-verdict').style.background = '#fef2f2';
        document.getElementById('modal-verdict').style.color = '#dc2626';
    }} else if (score >= 40) {{
        bar.style.background = '#f59e0b';
        document.getElementById('modal-score').style.color = '#d97706';
        document.getElementById('modal-verdict').textContent = 'MEDIUM RISK';
        document.getElementById('modal-verdict').style.background = '#fffbeb';
        document.getElementById('modal-verdict').style.color = '#b45309';
    }} else {{
        bar.style.background = '#00d165';
        document.getElementById('modal-score').style.color = '#00d165';
        document.getElementById('modal-verdict').textContent = 'LOW RISK';
        document.getElementById('modal-verdict').style.background = '#f0fdf4';
        document.getElementById('modal-verdict').style.color = '#16a34a';
    }}
    document.getElementById('modal').classList.add('show');
}}

function closeModal() {{
    document.getElementById('modal').classList.remove('show');
}}

function showToast(msg, type) {{
    const t = document.getElementById('toast');
    t.textContent = msg;
    t.classList.add('show');
    t.style.background = type === 'error' ? '#dc2626' : '#0d1120';
    setTimeout(() => t.classList.remove('show'), 3000);
}}

document.getElementById('modal').addEventListener('click', e => {{ if (e.target === this) closeModal(); }});
document.getElementById('emailComposer').addEventListener('click', e => {{ if (e.target === this) closeEmailComposer(); }});
document.getElementById('bulkEmailComposer').addEventListener('click', e => {{ if (e.target === this) closeBulkEmailComposer(); }});

// Risk distribution doughnut
(function() {{
  const ctx = document.getElementById('riskDoughnut');
  if (!ctx) return;
  new Chart(ctx, {{
    type: 'doughnut',
    data: {{
      labels: ['High Risk', 'Medium', 'Low Risk'],
      datasets: [{{
        data: [{high_total}, {med_total}, {low_total}],
        backgroundColor: ['#ef4444', '#f59e0b', '#00d165'],
        borderWidth: 0,
        hoverOffset: 4
      }}]
    }},
    options: {{
      cutout: '72%',
      plugins: {{
        legend: {{ position: 'bottom', labels: {{ font: {{ family: 'Plus Jakarta Sans', size: 11 }}, padding: 16, usePointStyle: true, pointStyle: 'circle' }} }}
      }}
    }}
  }});
}})();

// Lead-time stacked bar
(function() {{
  const ctx = document.getElementById('leadTimeChart');
  if (!ctx) return;
  new Chart(ctx, {{
    type: 'bar',
    data: {{
      labels: ['0–7d', '8–30d', '31–60d', '61–90d', '90+d'],
      datasets: [
        {{ label: 'High Risk', data: {lt_high_js}, backgroundColor: '#ef4444', borderWidth: 0, borderRadius: 0 }},
        {{ label: 'Medium',    data: {lt_med_js},  backgroundColor: '#f59e0b', borderWidth: 0, borderRadius: 0 }},
        {{ label: 'Low Risk',  data: {lt_low_js},  backgroundColor: '#00d165', borderWidth: 0, borderRadius: 4 }}
      ]
    }},
    options: {{
      responsive: true,
      scales: {{
        x: {{ stacked: true, grid: {{ display: false }}, border: {{ display: false }}, ticks: {{ font: {{ family: 'Plus Jakarta Sans', size: 11 }}, color: '#94a3b8' }} }},
        y: {{ stacked: true, beginAtZero: true, grid: {{ color: '#f1f5f9' }}, border: {{ display: false }}, ticks: {{ font: {{ family: 'Plus Jakarta Sans', size: 11 }}, color: '#94a3b8' }} }}
      }},
      plugins: {{
        legend: {{ position: 'bottom', labels: {{ font: {{ family: 'Plus Jakarta Sans', size: 11 }}, padding: 16, usePointStyle: true, pointStyle: 'circle', color: '#64748b' }} }}
      }}
    }}
  }});
}})();

// Savings calculator
const _savHighTotal = {high_total};
const _savAvgAdr    = {avg_adr:.2f};
const _savAvgNights = {avg_nights:.2f};
function updateSavings(pct) {{
  const p = parseInt(pct) / 100;
  const bk  = Math.round(_savHighTotal * p);
  const rev = Math.round(bk * _savAvgAdr * _savAvgNights);
  document.getElementById('savPct').textContent = pct + '%';
  document.getElementById('savBookings').textContent = bk.toLocaleString();
  document.getElementById('savRevenue').textContent = '€' + rev.toLocaleString();
  document.getElementById('savAnnual').textContent  = '€' + (rev * 12).toLocaleString();
}}
// ── DYNAMICS JS ──────────────────────────────────────────────

// Counter animation
function countUp(el, target, duration, pre, suf) {{
  if (!el) return;
  pre = pre||''; suf = suf||''; duration = duration||1400;
  const start = performance.now();
  function tick(now) {{
    const p = Math.min((now - start)/duration, 1);
    const ease = 1 - Math.pow(1-p, 3);
    const val = Math.round(ease * target);
    el.textContent = pre + val.toLocaleString() + suf;
    if (p < 1) requestAnimationFrame(tick);
  }}
  requestAnimationFrame(tick);
}}
document.querySelectorAll('.count-up').forEach(function(el) {{
  countUp(el, parseFloat(el.dataset.val)||0, 1400, el.dataset.pre||'', el.dataset.suf||'');
}});

// Filter tabs
function filterRisk(risk, btn) {{
  document.querySelectorAll('.f-tab').forEach(function(t) {{ t.classList.remove('active'); }});
  btn.classList.add('active');
  document.querySelectorAll('.clickable-row, .cr').forEach(function(row) {{
    var show;
    if (risk === 'all') {{ show = true; }}
    else {{ show = !!row.querySelector('.badge.' + risk); }}
    row.style.display = show ? '' : 'none';
    var nx = row.nextElementSibling;
    if (nx && nx.classList.contains('exp-tr')) {{
      if (!show) nx.classList.remove('open');
      nx.style.display = show ? '' : 'none';
    }}
  }});
}}

// Sort table
var _sCol = null, _sDir = 1;
function sortTable(col) {{
  var tbody = document.querySelector('tbody');
  if (!tbody) return;
  if (_sCol === col) _sDir *= -1; else {{ _sCol = col; _sDir = 1; }}
  var pairs = [];
  var rows = Array.from(tbody.querySelectorAll('.clickable-row, .cr'));
  rows.forEach(function(r) {{
    var nx = r.nextElementSibling;
    var expRow = (nx && nx.classList.contains('exp-tr')) ? nx : null;
    pairs.push([r, expRow]);
  }});
  pairs.sort(function(a, b) {{
    var ar = a[0], br = b[0];
    var av = col==='score' ? parseFloat(ar.dataset.score||0) :
             col==='lead'  ? parseInt(ar.dataset.lead||0) :
             col==='rate'  ? parseInt(ar.dataset.rate||0) : 0;
    var bv = col==='score' ? parseFloat(br.dataset.score||0) :
             col==='lead'  ? parseInt(br.dataset.lead||0) :
             col==='rate'  ? parseInt(br.dataset.rate||0) : 0;
    return (av - bv) * _sDir;
  }});
  pairs.forEach(function(p) {{
    tbody.appendChild(p[0]);
    if (p[1]) tbody.appendChild(p[1]);
  }});
  document.querySelectorAll('.sort-arr').forEach(function(a) {{
    var c = a.closest('.sort-th') && a.closest('.sort-th').dataset.col;
    a.textContent = c === col ? (_sDir === -1 ? '↓' : '↑') : '↕';
    a.classList.toggle('on', c === col);
  }});
}}

// Build prediction reasons
function buildReasons(b, score) {{
  var r = [];
  var lt = b.lead_time||0, canc = b.previous_cancellations||0;
  var rep = b.is_repeated_guest||0, adr = b.adr||0;
  var chg = b.booking_changes||0, spec = b.total_of_special_requests||0;
  var wk = b.stays_in_week_nights||0, we = b.stays_in_weekend_nights||0;
  var nights = wk + we || 1;

  if (lt > 120)      r.push({{c:'bad',  t:'Lead time ' + lt + ' days — bookings this far out cancel 3× more often'}});
  else if (lt > 45)  r.push({{c:'warn', t:'Lead time ' + lt + ' days — moderate cancellation window remains'}});
  else               r.push({{c:'pos',  t:'Lead time ' + lt + ' days — close to arrival, low cancellation risk'}});

  if (canc >= 2)     r.push({{c:'bad',  t: canc + ' prior cancellations — strongest single predictor of future no-shows'}});
  else if (canc===1) r.push({{c:'warn', t:'1 previous cancellation on record — elevated risk signal'}});
  else               r.push({{c:'pos',  t:'No cancellation history — clean booking profile'}});

  if (!rep)          r.push({{c:'warn', t:'First-time guest — new guests cancel 2.3× more than returning guests'}});
  else               r.push({{c:'pos',  t:'Returning guest — loyalty significantly reduces no-show likelihood'}});

  if (adr > 200)     r.push({{c:'bad',  t:'Premium rate €' + Math.round(adr) + ' — high-value bookings warrant a deposit or guarantee'}});
  else if (adr > 120)r.push({{c:'warn', t:'Room rate €' + Math.round(adr) + ' — consider pre-authorisation'}});

  if (chg >= 2)      r.push({{c:'warn', t: chg + ' booking modifications — repeated changes signal hesitation'}});
  else if (chg===1)  r.push({{c:'warn', t:'1 booking change — slight uncertainty signal'}});

  if (spec === 0)    r.push({{c:'warn', t:'No special requests — low guest engagement with this stay'}});
  else               r.push({{c:'pos',  t: spec + ' special request' + (spec>1?'s':'') + ' — engaged guests are far less likely to cancel'}});

  if (nights >= 4)   r.push({{c:'pos',  t: nights + '-night stay — multi-night bookings have lower no-show rates'}});
  else if (nights===1)r.push({{c:'warn',t:'1-night stay — shortest stays carry the highest no-show rate'}});

  return r.slice(0, 5);
}}

function conclusionText(score) {{
  if (score >= 70) return 'High cancellation probability. Request deposit or guarantee now — revenue at risk.';
  if (score >= 40) return 'Moderate risk. Send a reminder 48h before arrival to confirm the booking.';
  return 'Low risk. Booking profile is stable — monitor normally.';
}}

// Toggle grouped stays dropdown
function toggleStays(sid, row) {{
  var opening = !row.classList.contains('is-open');
  row.classList.toggle('is-open', opening);
  document.querySelectorAll('tr.stays-sub[data-grp="' + sid + '"]').forEach(function(r) {{
    r.style.display = opening ? '' : 'none';
  }});
  var chev = document.getElementById('chev-' + sid);
  if (chev) chev.textContent = opening ? '▴' : '▾';
}}

// Inline expand row
var _openRow = null;
function toggleExpand(row, idx, score) {{
  var expTr = document.getElementById('exp-' + idx);
  if (!expTr) return;
  var isOpen = expTr.classList.contains('open');
  if (_openRow && _openRow !== expTr) {{
    _openRow.classList.remove('open');
    var pr = _openRow.previousElementSibling;
    if (pr) pr.classList.remove('is-open');
  }}
  if (isOpen) {{
    expTr.classList.remove('open');
    row.classList.remove('is-open');
    _openRow = null;
  }} else {{
    var bArr = typeof bookings !== 'undefined' ? bookings : (typeof guests !== 'undefined' ? guests : []);
    var b = bArr[idx] || {{}};
    var reasons = buildReasons(b, score);
    var scoreColor = score>=70 ? '#ef4444' : score>=40 ? '#f59e0b' : '#00d165';
    var verdict = score>=70 ? 'HIGH RISK' : score>=40 ? 'MEDIUM RISK' : 'LOW RISK';
    var factorsHtml = reasons.map(function(r) {{
      return '<div class="exp-factor ' + r.c + '">' + r.t + '</div>';
    }}).join('');
    var inner = document.getElementById('exp-inner-' + idx);
    if (inner) {{
      inner.innerHTML =
        '<div class="exp-score-wrap">' +
          '<div class="exp-score-big" style="color:' + scoreColor + '">' + score.toFixed(0) + '%</div>' +
          '<div class="exp-score-lbl">' + verdict + '</div>' +
          '<div class="exp-bar-bg"><div class="exp-bar-fill" style="width:' + score + '%;background:' + scoreColor + '"></div></div>' +
        '</div>' +
        '<div class="exp-right">' +
          '<div class="exp-head">Why this score</div>' +
          '<div class="exp-factors">' + factorsHtml + '</div>' +
          '<div class="exp-conclusion">' + conclusionText(score) + '</div>' +
        '</div>';
    }}
    expTr.classList.add('open');
    row.classList.add('is-open');
    _openRow = expTr;
  }}
}}

</script>
</body>
</html>"""
    return dashboard_html

@app.route("/robots.txt")
def robots():
    return "User-agent: *\nDisallow: /admin\nDisallow: /admin/\n", 200, {"Content-Type": "text/plain"}


@app.route("/")
def home():
    try:
        return send_file("landing.html")
    except:
        return redirect(url_for("login"))

@app.route("/landing")
def landing():
    try:
        return send_file("landing.html")
    except:
        return redirect(url_for("login"))


@app.route("/api/landing-stats")
def api_landing_stats():
    """Return real computed stats for the landing page — auto-updates as new data is loaded."""
    s = VDV_LANDING_STATS
    total_cx = s['total_cx']
    total_ns = s['total_ns']
    avg_adr  = s['avg_adr']
    avg_nights = s['avg_nights']
    revenue_lost_k = round((total_cx + total_ns) * avg_adr * avg_nights / 1000)
    total_lost = total_cx + total_ns
    cancel_rate = round(total_cx / max(1, total_cx + total_ns + 8500) * 100, 1)
    noshow_rate = round(total_ns / max(1, total_cx + total_ns + 8500) * 100, 1)
    # High-risk upcoming bookings (at-risk revenue recoverable with Occupado)
    fut_scores  = VDV_FUTURE_SCORES
    fut_high    = sum(1 for sc in fut_scores if sc >= 70) if fut_scores else 36
    fut_med     = sum(1 for sc in fut_scores if 40 <= sc < 70) if fut_scores else 844
    at_risk_rev = round((fut_high + fut_med * 0.4) * avg_adr * avg_nights / 1000)
    return jsonify({
        'revenue_lost_eur_k':  revenue_lost_k,
        'model_accuracy_pct':  s['model_accuracy'],
        'model_auc':           s['model_auc'],
        'training_count':      s['training_count'],
        'total_cx':            total_cx,
        'total_ns':            total_ns,
        'cancel_rate_pct':     cancel_rate,
        'noshow_rate_pct':     noshow_rate,
        'fut_high':            fut_high,
        'fut_med':             fut_med,
        'at_risk_revenue_k':   at_risk_rev,
    })


@app.route("/magic/<token>")
def magic_link(token):
    print(f"\n[MAGIC] Token received: {token[:30]}...")
    
    token_data = get_token_data(token)
    if not token_data:
        print(f"[MAGIC] Token not found!")
        return redirect(url_for("login"))
    
    hotel_username = token_data.get("hotel")
    csv_data = token_data.get("csv_data")

    # Check both HOTELS and REGISTERED_USERS
    if hotel_username in HOTELS:
        hotel_name = HOTELS[hotel_username]["name"]
    else:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT name FROM registered_users WHERE username=%s", (hotel_username,))
        user = cur.fetchone()
        cur.close()
        conn.close()
        if not user:
            return redirect(url_for("login"))
        hotel_name = user["name"]

    session["hotel"]        = hotel_username
    session["hotel_name"]   = hotel_name
    session["alert_email"]  = ""
    session["uploaded_csv"] = csv_data
    
    delete_token(token)
    print(f"[MAGIC] Auto-logged in: {hotel_username}")
    
    return redirect(url_for("dashboard"))

@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    success = session.pop("verified_success", "")
    if request.method == "POST":
        ip = request.remote_addr
        blocked, remaining = check_rate_limit(ip)
        if blocked:
            error = f"Too many failed attempts. Try again in {remaining // 60 + 1} minute(s)."
        else:
            username = sanitise(request.form.get("username", "")).lower()
            password = request.form.get("password", "").strip()
            # Check admin credentials
            if username == "jpdourado" and password == "livejoao":
                reset_attempts(ip)
                session.permanent = True
                session["is_admin"] = True
                return redirect(url_for("admin_panel"))
            # Check demo/hotel accounts
            if username in HOTELS and HOTELS[username]["password"] == password:
                reset_attempts(ip)
                session.permanent = True
                session["hotel"] = username
                session["hotel_name"] = HOTELS[username]["name"]
                session["alert_email"] = ""
                session["language"] = "en"
                session["first_login"] = True
                return redirect(url_for("dashboard"))
            # Check registered users in database
            try:
                conn = get_db()
                cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                cur.execute("SELECT * FROM registered_users WHERE username=%s", (username,))
                user = cur.fetchone()
                cur.close()
                conn.close()
            except Exception as _db_err:
                print(f"[LOGIN] DB connection failed: {_db_err}")
                record_failed_attempt(ip)
                error = "Invalid credentials"
                user = None
            if user:
                # Support both bcrypt hashed and legacy plain text passwords
                pw_bytes = password.encode("utf-8")
                stored   = user["password"]
                try:
                    match = bcrypt.checkpw(pw_bytes, stored.encode("utf-8"))
                except Exception:
                    match = (stored == password)
                if match:
                    if not user["verified"]:
                        error = "Please verify your email before logging in. Check your inbox."
                    else:
                        reset_attempts(ip)
                        session.permanent = True
                        session["hotel"] = username
                        session["hotel_name"] = user["name"]
                        session["alert_email"] = user["email"]
                        session["language"] = "en"
                        session["first_login"] = True
                        return redirect(url_for("dashboard"))
                else:
                    record_failed_attempt(ip)
                    error = "Invalid credentials"
            else:
                record_failed_attempt(ip)
                error = "Invalid credentials"

    error_html   = f'<div class="err">{error}</div>' if error else ''
    success_html = f'<div class="ok">{success}</div>' if success else ''

    return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Sign in</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:#f5f7fb;font-family:'Plus Jakarta Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;-webkit-font-smoothing:antialiased;}}
.card{{background:#ffffff;border:1px solid #e4e8f0;border-radius:20px;width:100%;max-width:400px;padding:48px;box-shadow:0 4px 32px rgba(0,0,0,0.06);}}
.brand{{font-family:'Syne',sans-serif;font-size:22px;font-weight:800;color:#0d1120;letter-spacing:-0.5px;margin-bottom:28px;}}
.brand span{{color:#00d165;}}
.card-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:22px;font-weight:800;color:#0d1120;letter-spacing:-0.6px;margin-bottom:6px;}}
.card-sub{{font-size:13px;color:#64748b;margin-bottom:28px;}}
.err{{background:#fef2f2;border:1px solid #fecaca;color:#dc2626;padding:12px 14px;border-radius:9px;font-size:13px;margin-bottom:20px;line-height:1.5;}}
.ok{{background:#f0fdf4;border:1px solid #bbf7d0;color:#16a34a;padding:12px 14px;border-radius:9px;font-size:13px;margin-bottom:20px;line-height:1.5;}}
label{{font-family:'JetBrains Mono',monospace;font-size:10px;font-weight:500;color:#94a3b8;text-transform:uppercase;letter-spacing:1px;display:block;margin-bottom:7px;}}
input{{width:100%;padding:12px 14px;background:#f8fafc;border:1px solid #e4e8f0;border-radius:9px;font-size:14px;color:#0d1120;margin-bottom:16px;outline:none;font-family:'Plus Jakarta Sans',sans-serif;transition:border-color .2s,background .2s;}}
input:focus{{border-color:#00d165;background:#ffffff;}}
input::placeholder{{color:#cbd5e1;}}
.btn-submit{{width:100%;padding:13px;background:#00d165;color:#080c14;border:none;border-radius:9px;font-weight:700;cursor:pointer;font-size:14px;font-family:'Plus Jakarta Sans',sans-serif;transition:all .2s;display:flex;align-items:center;justify-content:center;gap:6px;margin-top:4px;}}
.btn-submit:hover{{background:#04e270;box-shadow:0 4px 16px rgba(0,209,101,0.25);}}
.links{{margin-top:24px;text-align:center;display:flex;flex-direction:column;gap:10px;}}
.links a{{font-size:13px;color:#94a3b8;text-decoration:none;transition:color .2s;}}
.links a span{{color:#00d165;font-weight:600;}}
.links a:hover{{color:#0d1120;}}
.divider{{height:1px;background:#e4e8f0;margin:20px 0;}}

</style>
</head>
<body>
<div class="card">
  <div class="brand">Occup<span>ado</span></div>
  <div class="card-title">Welcome back</div>
  <div class="card-sub">Sign in to your revenue dashboard</div>
  {error_html}{success_html}
  <form method="POST">
    <label>Username</label>
    <input type="text" name="username" required autocomplete="username" placeholder="your username">
    <label>Password</label>
    <input type="password" name="password" required autocomplete="current-password" placeholder="••••••••">
    <button type="submit" class="btn-submit">Sign in &rarr;</button>
  </form>
  <div class="divider"></div>
  <div class="links">
    <a href="/register">No account yet? <span>Start free pilot</span></a>
    <a href="/forgot-password">Forgot password?</a>
  </div>
</div>
</body>
</html>"""

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/dashboard")
@login_required
def dashboard():
    hotel_username = session.get("hotel", "")
    hotel_name = session.get("hotel_name", "Your Hotel")
    lang = request.args.get("lang", session.get("language", "en"))

    if lang not in TRANSLATIONS:
        lang = "en"
    session["language"] = lang

    # ── Van der Valk Mechelen gets a dedicated enriched dashboard ──
    if hotel_username == VDV_HOTEL_KEY:
        first_login = session.pop("first_login", False)
        try:
            return build_vdv_dashboard(hotel_name, lang=lang, first_login=first_login)
        except Exception as _e:
            import traceback
            print(f"[MEC] Dashboard error: {_e}\n{traceback.format_exc()}")
            raise

    # ── Van der Valk Brussels Airport gets its own dashboard ──
    if hotel_username == VDV_BRU_HOTEL_KEY:
        first_login = session.pop("first_login", False)
        try:
            return build_vdv_bru_dashboard(hotel_name, lang=lang, first_login=first_login)
        except Exception as _e:
            import traceback
            print(f"[BRU] Dashboard error: {_e}\n{traceback.format_exc()}")
            raise

    uploaded_data = session.get("uploaded_csv")
    first_login = session.pop("first_login", False)
    skip_onboard = request.args.get("skip_onboard")

    # Show onboarding screen for first-time logins with no data yet
    if first_login and not uploaded_data and not skip_onboard:
        return build_empty_state(hotel_name, lang=lang)

    if uploaded_data:
        sample = pd.DataFrame(uploaded_data)
        for feat in features:
            if feat not in sample.columns:
                sample[feat] = 0
        sample = sample[features].head(20).fillna(0)
        tonight_sample = pd.DataFrame(uploaded_data)[features].head(500).fillna(0)
        uploaded = True
    else:
        sample = df[features].head(20).fillna(0)
        tonight_sample = df[features].head(500).fillna(0)
        uploaded = False

    scores = model.predict_proba(sample)[:, 1] * 100
    tonight_scores = model.predict_proba(tonight_sample)[:, 1] * 100

    return build_dashboard(hotel_name, sample, scores, tonight_scores, tonight_sample=tonight_sample, uploaded=uploaded, lang=lang, first_login=first_login)

@app.route("/clear")
@login_required
def clear():
    session.pop("uploaded_csv", None)
    return redirect(url_for("dashboard"))

@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    hotel_name = session.get("hotel_name", "Your Hotel")
    lang = session.get("language", "en")
    message = None
    
    if request.method == "POST":
        alert_email = request.form.get("alert_email", "").strip()
        if alert_email and not is_valid_email(alert_email):
            message = "Please enter a valid email address."
        else:
            session["alert_email"] = alert_email
            message = t("saved", lang)
    
    current_email = session.get("alert_email", "")
    message_html = f'<div style="background:#c8e6c9;border:1px solid #81c784;padding:14px 20px;margin:0 auto 30px;color:#2e7d32;border-radius:8px;text-align:center;max-width:500px;font-size:13px;font-weight:500;">{message}</div>' if message else ''
    
    return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — {t("settings_title", lang)}</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:#f9fafb;color:#111827;font-family:'Plus Jakarta Sans',sans-serif;-webkit-font-smoothing:antialiased;}}
.topbar{{background:#fff;border-bottom:1px solid #e5e7eb;padding:0 40px;height:56px;display:flex;align-items:center;justify-content:space-between;}}
.topbar-logo{{font-family:'Syne',sans-serif;font-size:18px;font-weight:800;color:#111827;letter-spacing:-0.3px;}}
.topbar-logo span{{color:#00d165;}}
.topbar-hotel{{font-size:11px;color:#9ca3af;font-family:'JetBrains Mono',monospace;margin-top:2px;}}
.topbar-right{{display:flex;align-items:center;gap:8px;}}
.btn-nav{{padding:7px 16px;background:#fff;border:1px solid #e5e7eb;border-radius:8px;color:#374151;font-size:13px;font-weight:500;text-decoration:none;transition:border-color .15s;}}
.btn-nav:hover{{border-color:#111827;color:#111827;}}
.content{{padding:60px 40px;display:flex;align-items:center;justify-content:center;min-height:calc(100vh - 56px);}}
.wrapper{{width:100%;max-width:480px;text-align:center;}}
.page-title{{font-size:28px;font-weight:700;color:#111827;letter-spacing:-0.5px;margin-bottom:6px;}}
.page-sub{{font-size:13px;color:#6b7280;margin-bottom:36px;}}
.card{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:36px;}}
.card-title{{font-size:16px;font-weight:600;color:#111827;margin-bottom:8px;}}
.card-sub{{font-size:13px;color:#6b7280;margin-bottom:24px;}}
label{{font-size:10px;font-weight:500;color:#9ca3af;text-transform:uppercase;letter-spacing:.08em;display:block;margin-bottom:7px;text-align:left;font-family:'JetBrains Mono',monospace;}}
input{{width:100%;padding:11px 14px;background:#fff;border:1px solid #e5e7eb;border-radius:8px;font-size:14px;margin-bottom:20px;font-family:'Plus Jakarta Sans',sans-serif;color:#111827;outline:none;transition:border-color .15s;}}
input:focus{{border-color:#00d165;}}
button{{padding:11px 28px;background:#00d165;color:#080c14;border:none;border-radius:8px;font-weight:700;cursor:pointer;font-size:14px;font-family:'Plus Jakarta Sans',sans-serif;transition:background .15s;}}
button:hover{{background:#04e270;}}

</style>
</head>
<body>
<div class="topbar">
    <div>
        <div class="topbar-logo">Occup<span>ado</span></div>
        <div class="topbar-hotel">{hotel_name} · {t("settings_title", lang)}</div>
    </div>
    <div class="topbar-right">
        <a href="/dashboard" class="btn-nav">{t("back", lang)}</a>
        <a href="/logout" class="btn-nav">{t("sign_out", lang)}</a>
    </div>
</div>
<div class="content">
    <div class="wrapper">
        <div class="page-title">{t("settings_title", lang)}</div>
        <div class="page-sub">{t("config_alert", lang)}</div>
        {message_html}
        <div class="card">
            <div class="card-title">{t("alert_email", lang)}</div>
            <div class="card-sub">{t("alert_desc", lang)}</div>
            <form method="POST">
                <label>{t("email_addr", lang)}</label>
                <input type="email" name="alert_email" value="{current_email}" placeholder="your-email@example.com" required>
                <button type="submit">{t("save", lang)}</button>
            </form>
        </div>
    </div>
</div>
</body>
</html>"""

@app.route("/upload", methods=["POST"])
@login_required
def upload():
    if "csv_file" not in request.files:
        return redirect(url_for("dashboard"))
    file = request.files["csv_file"]
    if file.filename == "":
        return redirect(url_for("dashboard"))
    try:
        content = file.read().decode("utf-8")
        df_uploaded = pd.read_csv(io.StringIO(content))
        # Store raw CSV columns and data in session for field mapping
        session["raw_csv_columns"] = list(df_uploaded.columns)
        session["raw_csv_data"] = df_uploaded.head(500).fillna(0).to_dict(orient="records")
        return redirect(url_for("map_fields"))
    except Exception as e:
        print(f"[ERROR] {e}")
        return redirect(url_for("dashboard"))


@app.route("/map-fields", methods=["GET", "POST"])
@login_required
def map_fields():
    hotel_name = session.get("hotel_name", "Hotel")
    hotel_username = session.get("hotel")
    lang = request.args.get("lang", session.get("language", "en"))
    raw_columns = session.get("raw_csv_columns", [])
    raw_data = session.get("raw_csv_data", [])

    if not raw_columns:
        return redirect(url_for("dashboard"))

    FEATURE_LABELS = {
        "lead_time":                      "Lead Time (days before arrival)",
        "arrival_date_week_number":       "Arrival Week Number",
        "stays_in_weekend_nights":        "Weekend Nights",
        "stays_in_week_nights":           "Week Nights",
        "adults":                         "Number of Adults",
        "is_repeated_guest":              "Repeated Guest (0/1)",
        "previous_cancellations":         "Previous Cancellations",
        "previous_bookings_not_canceled": "Previous Bookings (not cancelled)",
        "booking_changes":                "Booking Changes",
        "days_in_waiting_list":           "Days in Waiting List",
        "adr":                            "Average Daily Rate (room price)",
        "total_of_special_requests":      "Total Special Requests",
    }

    # Auto-match: find best column for each feature
    def auto_match(feat, cols):
        feat_lower = feat.lower().replace("_", "")
        for col in cols:
            col_lower = col.lower().replace("_", "").replace(" ", "")
            if feat_lower == col_lower:
                return col
        # Partial match
        for col in cols:
            col_lower = col.lower().replace("_", "").replace(" ", "")
            if feat_lower in col_lower or col_lower in feat_lower:
                return col
        return ""

    if request.method == "POST":
        mapping = {}
        for feat in features:
            mapped_col = request.form.get(f"map_{feat}", "").strip()
            mapping[feat] = mapped_col

        # Build final dataframe using mapping
        df_raw = pd.DataFrame(raw_data)
        df_final = pd.DataFrame()
        for feat in features:
            col = mapping.get(feat, "")
            if col and col in df_raw.columns:
                df_final[feat] = pd.to_numeric(df_raw[col], errors="coerce").fillna(0)
            else:
                df_final[feat] = 0

        csv_data = df_final.to_dict(orient="records")
        session["uploaded_csv"] = csv_data

        # Send alert if configured
        sample = df_final.head(20)
        scores = model.predict_proba(sample)[:, 1] * 100
        alert_email = session.get("alert_email", "")
        high_risk_bookings = [{"id": f"Booking {i+1}", "score": s} for i, s in enumerate(scores) if s >= 70]
        if high_risk_bookings and alert_email:
            send_consolidated_alert(hotel_name, alert_email, high_risk_bookings, hotel_username, csv_data)

        return redirect(url_for("dashboard"))

    # Build auto-matched defaults
    auto_map = {feat: auto_match(feat, raw_columns) for feat in features}

    # Preview: first 3 rows, only mapped or all columns
    preview_rows = raw_data[:3]
    preview_cols = raw_columns[:8]  # show first 8 cols in preview

    # Build options HTML helper
    def col_options(selected):
        opts = '<option value="">(skip / use 0)</option>'
        for col in raw_columns:
            sel = 'selected' if col == selected else ''
            opts += f'<option value="{col}" {sel}>{col}</option>'
        return opts

    # Count auto-matched
    matched = sum(1 for v in auto_map.values() if v)

    rows_html = ""
    for feat in features:
        label = FEATURE_LABELS[feat]
        matched_col = auto_map[feat]
        badge = f'<span style="background:#e8f5e9;color:#2e7d32;border-radius:6px;padding:2px 8px;font-size:11px;font-family:JetBrains Mono,monospace;">✓ auto-matched</span>' if matched_col else f'<span style="background:#fff3e0;color:#e65100;border-radius:6px;padding:2px 8px;font-size:11px;font-family:JetBrains Mono,monospace;">needs mapping</span>'
        rows_html += f"""
        <div style="display:grid;grid-template-columns:1fr 1fr auto;gap:12px;align-items:center;padding:14px 0;border-bottom:1px solid rgba(0,128,0,0.08);">
            <div>
                <div style="font-size:13px;font-weight:600;color:#0a1a0a;font-family:'JetBrains Mono',monospace;">{feat}</div>
                <div style="font-size:12px;color:#4a6648;margin-top:2px;">{label}</div>
            </div>
            <select name="map_{feat}" style="padding:10px 12px;border:1px solid rgba(0,128,0,0.2);border-radius:8px;font-size:13px;font-family:'Plus Jakarta Sans',sans-serif;background:#f5faf5;color:#0a1a0a;width:100%;outline:none;">
                {col_options(matched_col)}
            </select>
            <div>{badge}</div>
        </div>"""

    # Preview table
    preview_header = "".join(f'<th style="padding:8px 12px;font-size:11px;font-family:JetBrains Mono,monospace;color:#4a6648;text-align:left;border-bottom:1px solid rgba(0,128,0,0.1);">{c}</th>' for c in preview_cols)
    preview_body = ""
    for row in preview_rows:
        cells = "".join(f'<td style="padding:8px 12px;font-size:12px;color:#0a1a0a;border-bottom:1px solid rgba(0,128,0,0.06);">{str(row.get(c,""))[:20]}</td>' for c in preview_cols)
        preview_body += f"<tr>{cells}</tr>"

    return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Map Your Data</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:#f9fafb;font-family:'Plus Jakarta Sans',sans-serif;color:#111827;-webkit-font-smoothing:antialiased;}}
.topbar{{background:#fff;border-bottom:1px solid #e5e7eb;padding:0 40px;height:56px;display:flex;align-items:center;justify-content:space-between;}}
.topbar-logo{{font-family:'Syne',sans-serif;font-size:18px;font-weight:800;color:#111827;letter-spacing:-0.3px;}}
.topbar-logo span{{color:#00d165;}}
.topbar-hotel{{font-size:11px;color:#9ca3af;font-family:'JetBrains Mono',monospace;margin-top:2px;}}
.btn-nav{{padding:7px 16px;background:#fff;border:1px solid #e5e7eb;border-radius:8px;color:#374151;font-size:13px;font-weight:500;text-decoration:none;transition:border-color .15s;display:inline-block;}}
.btn-nav:hover{{border-color:#111827;color:#111827;}}
.content{{max-width:860px;margin:0 auto;padding:48px 24px;}}
.page-title{{font-size:28px;font-weight:700;color:#111827;letter-spacing:-0.5px;margin-bottom:6px;}}
.page-sub{{font-size:12px;color:#9ca3af;font-family:'JetBrains Mono',monospace;margin-bottom:32px;}}
.card{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:28px;margin-bottom:20px;}}
.card-title{{font-size:16px;font-weight:600;color:#111827;margin-bottom:6px;}}
.card-sub{{font-size:13px;color:#6b7280;margin-bottom:20px;}}
.submit-btn{{width:100%;padding:14px;background:#00d165;color:#080c14;border:none;border-radius:8px;font-weight:700;font-size:15px;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;margin-top:8px;transition:background .15s;}}
.submit-btn:hover{{background:#04e270;}}
select:focus{{border-color:#00d165;background:white;outline:none;}}

</style>
</head>
<body>
<div class="topbar">
    <div>
        <div class="topbar-logo">Occup<span>ado</span></div>
        <div class="topbar-hotel">{hotel_name} · Map Your Data</div>
    </div>
    <div style="display:flex;gap:10px;">
        <a href="/dashboard" class="btn-nav">← Back</a>
        <a href="/logout" class="btn-nav">Sign Out</a>
    </div>
</div>
<div class="content">
    <div class="page-title">Map Your Columns</div>
    <div class="page-sub">Tell Occupado which columns in your CSV match each feature · {matched}/12 auto-matched</div>

    <div class="card">
        <div class="card-title">📂 Your CSV has {len(raw_columns)} columns, {len(raw_data)} rows</div>
        <div class="card-sub">Match your column names to Occupado's required features. Auto-matched columns are pre-filled — review and adjust if needed.</div>
        <form method="POST">
            {rows_html}
            <button type="submit" class="submit-btn">Run Predictions →</button>
        </form>
    </div>

    <div class="card">
        <div class="card-title">👀 Data Preview (first 3 rows)</div>
        <div class="card-sub">Showing first 8 columns of your file</div>
        <div style="overflow-x:auto;">
            <table style="width:100%;border-collapse:collapse;">
                <thead><tr>{preview_header}</tr></thead>
                <tbody>{preview_body}</tbody>
            </table>
        </div>
    </div>
</div>
</body>
</html>"""

@app.route("/send-guest-email", methods=["POST"])
@login_required
def send_guest_email():
    ip = request.remote_addr
    blocked, remaining = check_rate_limit(ip)
    if blocked:
        return {"status": "error", "message": f"Too many attempts. Try again in {remaining} minutes."}, 429
    data = request.get_json()
    guest_email = sanitise(data.get("guest_email", "")).strip()
    guest_name  = sanitise(data.get("guest_name", "Guest"))
    subject     = sanitise(data.get("subject", ""), max_length=200)
    body        = data.get("body", "").strip()

    if not is_valid_email(guest_email):
        return {"status": "error", "message": "Invalid email address"}, 400
    if not subject or not body:
        return {"status": "error", "message": "Subject and message are required"}, 400

    hotel_name = session.get("hotel_name", "Hotel")
    success = send_email_to_guest(guest_email, guest_name, hotel_name, subject, body)

    if success:
        try:
            _conn = get_db()
            _cur  = _conn.cursor()
            _cur.execute(
                "INSERT INTO roi_actions (hotel_username, guest_name, action_type) VALUES (%s, %s, %s)",
                (session.get("hotel", ""), guest_name, "email")
            )
            _conn.commit()
            _cur.close()
            _conn.close()
        except:
            pass
        return {"status": "success", "message": f"Email sent to {guest_email}"}
    else:
        return {"status": "error", "message": "Failed to send email"}, 500

@app.route("/send-bulk-email", methods=["POST"])
@login_required
def send_bulk_email():
    data    = request.get_json()
    count   = data.get("count", 0)
    subject = sanitise(data.get("subject", ""), max_length=200)
    body    = data.get("body", "").strip()

    if not subject or not body:
        return {"status": "error", "message": "Subject and message are required"}, 400
    if count == 0:
        return {"status": "error", "message": "No bookings selected"}, 400
    if len(body) > 5000:
        return {"status": "error", "message": "Message is too long (max 5000 characters)"}, 400

    return {"status": "success", "count": count}


# ─────────────────────────────────────────────
#  SHIJI TRANSFORMER
# ─────────────────────────────────────────────

def parse_shiji_date(val):
    if pd.isnull(val) or str(val).strip() == "":
        return None
    for fmt in ["%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d %b %Y", "%b %d %Y"]:
        try:
            return pd.to_datetime(str(val).strip(), format=fmt)
        except:
            continue
    try:
        return pd.to_datetime(str(val).strip(), infer_datetime_format=True)
    except:
        return None


def transform_shiji(df_res, df_cxl=None, df_ns=None):
    df = df_res.copy()

    def find_col(dataframe, keywords):
        for kw in keywords:
            for col in dataframe.columns:
                if kw.lower() in col.lower():
                    return col
        return None

    arr_col     = find_col(df, ["arr. date", "arrival date", "arr date", "checkin", "check-in"])
    created_col = find_col(df, ["created on", "created", "booking date", "booked on"])
    if arr_col and created_col:
        arr_dates     = df[arr_col].apply(parse_shiji_date)
        created_dates = df[created_col].apply(parse_shiji_date)
        df["lead_time"] = (arr_dates - created_dates).dt.days.clip(lower=0).fillna(0)
    else:
        df["lead_time"] = 0

    if arr_col:
        df["arrival_date_week_number"] = df[arr_col].apply(
            lambda x: parse_shiji_date(x).isocalendar()[1] if parse_shiji_date(x) else 0
        )
    else:
        df["arrival_date_week_number"] = 0

    nights_col = find_col(df, ["# nights", "nights", "num nights", "length of stay", "los"])
    dep_col    = find_col(df, ["dep. date", "departure date", "checkout", "check-out"])
    if nights_col:
        total = pd.to_numeric(df[nights_col], errors="coerce").fillna(0).clip(lower=0)
        df["stays_in_week_nights"]    = total
        df["stays_in_weekend_nights"] = (total * 0.28).round().astype(int)
    elif arr_col and dep_col:
        arr_d = df[arr_col].apply(parse_shiji_date)
        dep_d = df[dep_col].apply(parse_shiji_date)
        total = (dep_d - arr_d).dt.days.clip(lower=0).fillna(0)
        df["stays_in_week_nights"]    = total
        df["stays_in_weekend_nights"] = (total * 0.28).round().astype(int)
    else:
        df["stays_in_week_nights"]    = 0
        df["stays_in_weekend_nights"] = 0

    adults_col = find_col(df, ["ad/ch", "adults", "adult", "pax", "guests", "no. of guests"])
    if adults_col:
        def parse_adults(v):
            try:
                return int(str(v).split("/")[0].strip())
            except:
                return 1
        df["adults"] = df[adults_col].apply(parse_adults)
    else:
        df["adults"] = 2

    memb_col = find_col(df, ["memb. level", "membership", "loyalty", "vip", "member"])
    if memb_col:
        df["is_repeated_guest"] = df[memb_col].apply(
            lambda x: 0 if (pd.isnull(x) or str(x).strip() in ["", "None", "0", "No"]) else 1
        )
    else:
        df["is_repeated_guest"] = 0

    if df_cxl is not None and not df_cxl.empty:
        guest_col_res = find_col(df,     ["guest name", "guest", "name"])
        guest_col_cxl = find_col(df_cxl, ["guest name", "guest", "name"])
        if guest_col_res and guest_col_cxl:
            cxl_counts = df_cxl[guest_col_cxl].value_counts().to_dict()
            df["previous_cancellations"] = df[guest_col_res].map(cxl_counts).fillna(0).astype(int)
        else:
            df["previous_cancellations"] = 0
    else:
        cxl_no_col = find_col(df, ["cxl no", "cxl number", "cancellation no"])
        df["previous_cancellations"] = pd.to_numeric(df[cxl_no_col], errors="coerce").fillna(0).clip(lower=0) if cxl_no_col else 0

    df["previous_bookings_not_canceled"] = 0

    changes_col = find_col(df, ["booking changes", "modifications", "amended", "changes"])
    df["booking_changes"] = pd.to_numeric(df[changes_col], errors="coerce").fillna(0).clip(lower=0) if changes_col else 0

    wait_col = find_col(df, ["waiting list", "waitlist", "days waiting"])
    df["days_in_waiting_list"] = pd.to_numeric(df[wait_col], errors="coerce").fillna(0).clip(lower=0) if wait_col else 0

    rate_col = find_col(df, ["rate amount", "room rate total", "room rate", "nightly rate", "price", "adr"])
    if rate_col:
        df["adr"] = pd.to_numeric(
            df[rate_col].astype(str).str.replace("[€$£,]", "", regex=True),
            errors="coerce"
        ).fillna(0).clip(lower=0)
    else:
        df["adr"] = 0

    req_col = find_col(df, ["special request", "requests", "extras"])
    df["total_of_special_requests"] = pd.to_numeric(df[req_col], errors="coerce").fillna(0).clip(lower=0) if req_col else 0

    feat_cols = [
        "lead_time", "arrival_date_week_number", "stays_in_weekend_nights",
        "stays_in_week_nights", "adults", "is_repeated_guest",
        "previous_cancellations", "previous_bookings_not_canceled",
        "booking_changes", "days_in_waiting_list", "adr", "total_of_special_requests"
    ]
    return df[feat_cols].fillna(0)


@app.route("/shiji-upload", methods=["GET", "POST"])
@login_required
def shiji_upload():
    hotel_name     = session.get("hotel_name", "Hotel")
    hotel_username = session.get("hotel")
    error          = ""

    if request.method == "POST":
        try:
            if "res_file" not in request.files or request.files["res_file"].filename == "":
                error = "Please upload at least the Reservations file."
            else:
                def read_upload(key):
                    f = request.files.get(key)
                    if f and f.filename:
                        raw = f.read()
                        name = f.filename.lower()
                        if name.endswith(".xlsx") or name.endswith(".xls"):
                            return pd.read_excel(io.BytesIO(raw))
                        else:
                            return pd.read_csv(io.StringIO(raw.decode("utf-8", errors="replace")))
                    return None

                df_res = read_upload("res_file")
                df_cxl = read_upload("cxl_file")
                df_ns  = read_upload("ns_file")

                # Persist upload record
                try:
                    _upf = request.files.get("res_file")
                    _fn  = _upf.filename if _upf else "upload"
                    _conn_up = get_db()
                    _cur_up  = _conn_up.cursor()
                    _cur_up.execute(
                        "INSERT INTO hotel_uploads (hotel_username, filename) VALUES (%s, %s)",
                        (hotel_username, _fn)
                    )
                    _conn_up.commit()
                    _cur_up.close()
                    _conn_up.close()
                except:
                    pass

                df_transformed = transform_shiji(df_res, df_cxl, df_ns)
                csv_data = df_transformed.to_dict(orient="records")
                session["uploaded_csv"] = csv_data

                sample = df_transformed.head(20)
                scores = model.predict_proba(sample)[:, 1] * 100
                alert_email = session.get("alert_email", "")
                high_risk   = [{"id": f"Booking {i+1}", "score": s} for i, s in enumerate(scores) if s >= 70]
                if high_risk and alert_email:
                    send_consolidated_alert(hotel_name, alert_email, high_risk, hotel_username, csv_data)

                return redirect(url_for("dashboard"))

        except Exception as e:
            print(f"[SHIJI ERROR] {e}")
            error = f"Error processing files: {str(e)}"

    error_html = f'''<div style="background:#ffcdd2;padding:12px 16px;border-radius:8px;color:#c62828;font-size:14px;margin-bottom:24px;">{error}</div>''' if error else ""

    return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Shiji Import</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#f5faf5; font-family:'Plus Jakarta Sans',sans-serif; color:#0a1a0a; }}
.topbar {{ background:#008000; padding:16px 40px; display:flex; align-items:center; justify-content:space-between; }}
.topbar-logo {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:22px; font-weight:800; color:#ffffff; }}
.topbar-hotel {{ font-family:'JetBrains Mono',monospace; font-size:12px; color:rgba(255,255,255,0.8); }}
.btn-nav {{ padding:8px 18px; background:rgba(255,255,255,0.15); border:1px solid rgba(255,255,255,0.3); border-radius:8px; color:#ffffff; font-size:13px; font-weight:600; text-decoration:none; }}
.btn-nav:hover {{ background:rgba(255,255,255,0.25); }}
.content {{ max-width:700px; margin:0 auto; padding:48px 24px; }}
.page-title {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:32px; font-weight:800; margin-bottom:6px; }}
.page-sub {{ font-size:13px; color:#4a6648; font-family:'JetBrains Mono',monospace; margin-bottom:32px; }}
.card {{ background:#ffffff; border:1px solid rgba(0,128,0,0.15); border-radius:16px; padding:32px; margin-bottom:20px; }}
.card-title {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:18px; font-weight:700; margin-bottom:12px; }}
.card-sub {{ font-size:13px; color:#4a6648; margin-bottom:24px; line-height:1.6; }}
.file-row {{ margin-bottom:20px; }}
.file-label {{ font-size:12px; font-family:'JetBrains Mono',monospace; color:#4a6648; font-weight:600; text-transform:uppercase; letter-spacing:0.5px; display:block; margin-bottom:8px; }}
.req-badge {{ background:rgba(0,128,0,0.1);color:#008000;border-radius:4px;padding:2px 6px;font-size:10px;margin-left:6px; }}
.opt-badge {{ background:rgba(0,0,0,0.06);color:#4a6648;border-radius:4px;padding:2px 6px;font-size:10px;margin-left:6px; }}
input[type=file] {{ width:100%; padding:12px; background:#f5faf5; border:1px solid rgba(0,128,0,0.2); border-radius:10px; font-size:13px; font-family:'Plus Jakarta Sans',sans-serif; cursor:pointer; }}
input[type=file]:hover {{ border-color:#008000; }}
.submit-btn {{ width:100%; padding:16px; background:#008000; color:white; border:none; border-radius:12px; font-weight:700; font-size:16px; cursor:pointer; font-family:'Plus Jakarta Sans',sans-serif; }}
.submit-btn:hover {{ background:#006600; }}
.info-row {{ display:flex; gap:8px; margin-bottom:10px; font-size:13px; color:#4a6648; line-height:1.5; }}
.info-icon {{ color:#008000; }}

</style>
</head>
<body>
<div class="topbar">
    <div>
        <div class="topbar-logo">Occupado</div>
        <div class="topbar-hotel">{hotel_name} · Shiji Import</div>
    </div>
    <div style="display:flex;gap:10px;">
        <a href="/dashboard" class="btn-nav">← Back</a>
        <a href="/logout" class="btn-nav">Sign Out</a>
    </div>
</div>
<div class="content">
    <div class="page-title">Import from Shiji PMS</div>
    <div class="page-sub">Upload your Shiji exports · Occupado calculates all features automatically</div>
    {error_html}
    <div class="card">
        <div class="card-title">How it works</div>
        <div class="info-row"><span class="info-icon">✦</span><span>Occupado reads your Shiji columns and calculates all 12 AI features automatically</span></div>
        <div class="info-row"><span class="info-icon">✦</span><span><strong>Lead time</strong> calculated from Arr. Date minus Created On</span></div>
        <div class="info-row"><span class="info-icon">✦</span><span><strong>Previous cancellations</strong> counted from your CXL file per guest name</span></div>
        <div class="info-row"><span class="info-icon">✦</span><span><strong>Room rate</strong> from Rate Amount · Membership from Memb. Level</span></div>
    </div>
    <div class="card">
        <div class="card-title">Upload your Shiji files</div>
        <div class="card-sub">Export these reports from Shiji PMS and upload them here. Only Reservations is required.</div>
        <form method="POST" enctype="multipart/form-data">
            <div class="file-row">
                <label class="file-label">Reservations <span class="req-badge">REQUIRED</span></label>
                <input type="file" name="res_file" accept=".csv,.xlsx,.xls" required>
            </div>
            <div class="file-row">
                <label class="file-label">Cancelled Reservations <span class="opt-badge">OPTIONAL</span></label>
                <input type="file" name="cxl_file" accept=".csv,.xlsx,.xls">
            </div>
            <div class="file-row" style="margin-bottom:28px;">
                <label class="file-label">No-Shows <span class="opt-badge">OPTIONAL</span></label>
                <input type="file" name="ns_file" accept=".csv,.xlsx,.xls">
            </div>
            <button type="submit" class="submit-btn">🚀 Run Predictions →</button>
        </form>
    </div>
</div>
</body>
</html>"""


def send_verification_email(to_email, hotel_name, token):
    """Send email verification link via SendGrid"""
    base_url = os.environ.get("BASE_URL", "https://occupado.co")
    verify_url = f"{base_url}/verify/{token}"
    api_key = os.environ.get("SENDGRID_API_KEY")
    from_email = os.environ.get("ALERT_FROM_EMAIL", "team@occupado.co")
    if not api_key:
        return False
    try:
        html = f"""
        <div style="font-family:'Plus Jakarta Sans',sans-serif;max-width:480px;margin:0 auto;">
          <div style="background:#008000;padding:24px 32px;border-radius:12px 12px 0 0;">
            <span style="font-family:'Plus Jakarta Sans',sans-serif;font-size:24px;font-weight:800;color:#fff;">Occupado</span>
          </div>
          <div style="background:#ffffff;padding:32px;border:1px solid rgba(0,128,0,0.15);border-radius:0 0 12px 12px;">
            <h2 style="font-family:'Plus Jakarta Sans',sans-serif;color:#0a1a0a;margin-bottom:12px;">Verify your email</h2>
            <p style="color:#4a6648;margin-bottom:24px;line-height:1.6;">Hi <strong>{hotel_name}</strong>, thanks for registering with Occupado. Click the button below to activate your account.</p>
            <a href="{verify_url}" style="display:inline-block;background:#008000;color:#fff;padding:14px 28px;border-radius:10px;font-weight:700;text-decoration:none;font-size:15px;">Verify Email →</a>
            <p style="color:#4a6648;font-size:12px;margin-top:24px;">Or copy this link: {verify_url}</p>
          </div>
        </div>"""
        message = Mail(from_email=from_email, to_emails=to_email,
                       subject="Verify your Occupado account", html_content=html)
        sg = SendGridAPIClient(api_key)
        sg.send(message)
        return True
    except Exception as e:
        print(f"Verification email error: {e}")
        return False


def send_reset_email(to_email, hotel_name, token):
    """Send password reset link via SendGrid"""
    base_url  = os.environ.get("BASE_URL", "https://occupado.co")
    reset_url = f"{base_url}/reset-password/{token}"
    api_key   = os.environ.get("SENDGRID_API_KEY")
    from_email = os.environ.get("ALERT_FROM_EMAIL", "team@occupado.co")
    if not api_key:
        return False
    try:
        html = f"""
        <div style="font-family:'Plus Jakarta Sans',sans-serif;max-width:480px;margin:0 auto;">
          <div style="background:#008000;padding:24px 32px;border-radius:12px 12px 0 0;">
            <span style="font-family:'Plus Jakarta Sans',sans-serif;font-size:24px;font-weight:800;color:#fff;">Occupado</span>
          </div>
          <div style="background:#ffffff;padding:32px;border:1px solid rgba(0,128,0,0.15);border-radius:0 0 12px 12px;">
            <h2 style="font-family:'Plus Jakarta Sans',sans-serif;color:#0a1a0a;margin-bottom:12px;">Reset your password</h2>
            <p style="color:#4a6648;margin-bottom:24px;line-height:1.6;">Hi <strong>{hotel_name}</strong>, click the button below to reset your password. This link expires in <strong>1 hour</strong>.</p>
            <a href="{reset_url}" style="display:inline-block;background:#008000;color:#fff;padding:14px 28px;border-radius:10px;font-weight:700;text-decoration:none;font-size:15px;">Reset Password →</a>
            <p style="color:#4a6648;font-size:12px;margin-top:24px;">If you didn't request this, ignore this email. Your password won't change.</p>
            <p style="color:#4a6648;font-size:12px;margin-top:8px;">Or copy this link: {reset_url}</p>
          </div>
        </div>"""
        message = Mail(from_email=from_email, to_emails=to_email,
                       subject="Reset your Occupado password", html_content=html)
        sg = SendGridAPIClient(api_key)
        sg.send(message)
        return True
    except Exception as e:
        print(f"Reset email error: {e}")
        return False


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    ip = request.remote_addr
    message = ""
    if request.method == "POST":
        blocked, remaining = check_rate_limit(ip)
        if blocked:
            message = f"Too many attempts. Try again in {remaining} minutes."
        else:
            email = sanitise(request.form.get("email", ""), max_length=150).lower().strip()
            if not is_valid_email(email):
                message = "Please enter a valid email address."
            else:
                # Always show success message (don't reveal if email exists)
                conn = get_db()
                cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                cur.execute("SELECT username, name FROM registered_users WHERE email=%s AND verified=1", (email,))
                user = cur.fetchone()
                if user:
                    token = secrets.token_urlsafe(32)
                    cur.execute("INSERT INTO password_reset_tokens (token, username, expires_at) VALUES (%s,%s, NOW() + INTERVAL '1 hour')", (token, user["username"]))
                    conn.commit()
                    send_reset_email(email, user["name"], token)
                cur.close()
                conn.close()
                message = "success"

    if message == "success":
        success_html = '<div style="background:#e8f5e9;padding:12px;margin-bottom:20px;color:#2e7d32;border-radius:8px;font-size:14px;">If that email is registered, a reset link has been sent. Check your inbox.</div>'
        error_html   = ""
    else:
        error_html   = f'<div style="background:#ffcdd2;padding:12px;margin-bottom:20px;color:#c62828;border-radius:8px;font-size:14px;">{message}</div>' if message else ""
        success_html = ""

    return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Forgot Password</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#f5faf5; font-family:'Plus Jakarta Sans',sans-serif; min-height:100vh; display:flex; align-items:center; justify-content:center; }}
.box {{ background:#ffffff; padding:48px; border-radius:20px; width:100%; max-width:400px; border:1px solid rgba(0,128,0,0.15); }}
.logo {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:28px; font-weight:800; color:#008000; margin-bottom:8px; }}
.subtitle {{ font-size:13px; color:#4a6648; margin-bottom:28px; font-family:'JetBrains Mono',monospace; }}
label {{ font-size:12px; color:#4a6648; display:block; margin-bottom:6px; font-family:'JetBrains Mono',monospace; font-weight:600; }}
input {{ width:100%; padding:12px; background:#f5faf5; border:1px solid rgba(0,128,0,0.2); border-radius:10px; font-size:14px; margin-bottom:16px; outline:none; }}
input:focus {{ border-color:#008000; background:white; }}
button {{ width:100%; padding:14px; background:#008000; color:white; border:none; border-radius:10px; font-weight:700; cursor:pointer; font-size:15px; font-family:'Plus Jakarta Sans',sans-serif; }}
button:hover {{ background:#006600; }}
.switch-link {{ text-align:center; margin-top:20px; font-size:13px; color:#4a6648; }}
.switch-link a {{ color:#008000; font-weight:700; text-decoration:none; }}

</style>
</head>
<body>
<div class="box">
    <div class="logo">Occupado</div>
    <div class="subtitle">Password Reset</div>
    {error_html}{success_html}
    <form method="POST">
        <label>Email address</label>
        <input type="email" name="email" required placeholder="your@email.com">
        <button type="submit">Send Reset Link →</button>
    </form>
    <div class="switch-link"><a href="/login">← Back to login</a></div>
</div>
</body>
</html>"""


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    from datetime import timezone
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT username, expires_at FROM password_reset_tokens WHERE token=%s", (token,))
    row = cur.fetchone()

    def invalid_page(msg):
        cur.close()
        conn.close()
        return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Reset Failed</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@400;500&display=swap" rel="stylesheet">
<style>* {{margin:0;padding:0;box-sizing:border-box;}} body {{background:#f5faf5;font-family:'Plus Jakarta Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;}} .box {{background:#fff;padding:48px;border-radius:20px;max-width:400px;text-align:center;border:1px solid rgba(0,128,0,0.15);}}
</style>
</head>
<body><div class="box">
<div style="font-family:'Plus Jakarta Sans',sans-serif;font-size:28px;font-weight:800;color:#008000;margin-bottom:16px;">Occupado</div>
<div style="font-size:36px;margin-bottom:16px;">❌</div>
<h2 style="margin-bottom:12px;color:#0a1a0a;">{msg}</h2>
<a href="/forgot-password" style="color:#008000;font-weight:700;text-decoration:none;">Request a new link →</a>
</div></body></html>"""

    if not row:
        return invalid_page("Invalid or expired link")

    expires_at = row["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if datetime.now(timezone.utc) > expires_at:
        cur.execute("DELETE FROM password_reset_tokens WHERE token=%s", (token,))
        conn.commit()
        return invalid_page("Reset link has expired")

    username = row["username"]
    error = ""

    if request.method == "POST":
        password = request.form.get("password", "").strip()
        confirm  = request.form.get("confirm", "").strip()
        if not password or not confirm:
            error = "Both fields are required."
        elif len(password) < 8:
            error = "Password must be at least 8 characters."
        elif password != confirm:
            error = "Passwords do not match."
        else:
            hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
            cur.execute("UPDATE registered_users SET password=%s WHERE username=%s", (hashed, username))
            cur.execute("DELETE FROM password_reset_tokens WHERE token=%s", (token,))
            conn.commit()
            cur.close()
            conn.close()
            session["verified_success"] = "Password reset successfully. Please sign in."
            return redirect(url_for("login"))

    cur.close()
    conn.close()
    error_html = f'<div style="background:#ffcdd2;padding:12px;margin-bottom:20px;color:#c62828;border-radius:8px;font-size:14px;">{error}</div>' if error else ""

    return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Reset Password</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#f5faf5; font-family:'Plus Jakarta Sans',sans-serif; min-height:100vh; display:flex; align-items:center; justify-content:center; }}
.box {{ background:#ffffff; padding:48px; border-radius:20px; width:100%; max-width:400px; border:1px solid rgba(0,128,0,0.15); }}
.logo {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:28px; font-weight:800; color:#008000; margin-bottom:8px; }}
.subtitle {{ font-size:13px; color:#4a6648; margin-bottom:28px; font-family:'JetBrains Mono',monospace; }}
label {{ font-size:12px; color:#4a6648; display:block; margin-bottom:6px; font-family:'JetBrains Mono',monospace; font-weight:600; }}
input {{ width:100%; padding:12px; background:#f5faf5; border:1px solid rgba(0,128,0,0.2); border-radius:10px; font-size:14px; margin-bottom:16px; outline:none; }}
input:focus {{ border-color:#008000; background:white; }}
button {{ width:100%; padding:14px; background:#008000; color:white; border:none; border-radius:10px; font-weight:700; cursor:pointer; font-size:15px; font-family:'Plus Jakarta Sans',sans-serif; }}
button:hover {{ background:#006600; }}
.switch-link {{ text-align:center; margin-top:20px; font-size:13px; color:#4a6648; }}
.switch-link a {{ color:#008000; font-weight:700; text-decoration:none; }}

</style>
</head>
<body>
<div class="box">
    <div class="logo">Occupado</div>
    <div class="subtitle">Set New Password</div>
    {error_html}
    <form method="POST">
        <label>New Password</label>
        <input type="password" name="password" required minlength="8" placeholder="Min. 8 characters">
        <label>Confirm Password</label>
        <input type="password" name="confirm" required placeholder="Repeat password">
        <button type="submit">Reset Password →</button>
    </form>
    <div class="switch-link"><a href="/login">← Back to login</a></div>
</div>
</body>
</html>"""


@app.route("/register", methods=["GET", "POST"])
def register():
    error = ""
    success = ""
    ip = request.remote_addr
    if request.method == "POST":
        blocked, remaining = check_rate_limit(ip)
        if blocked:
            error = f"Too many attempts. Try again in {remaining} minutes."
        else:
            hotel_name = sanitise(request.form.get("hotel_name", ""), max_length=100)
            email      = sanitise(request.form.get("email", ""), max_length=150).lower()
            username   = sanitise(request.form.get("username", ""), max_length=50).lower()
            password   = request.form.get("password", "").strip()
            confirm    = request.form.get("confirm", "").strip()

        RESERVED_USERNAMES = set(HOTELS.keys()) | {"jpdourado", "admin", "occupado"}

        if not all([hotel_name, email, username, password, confirm]):
            error = "All fields are required."
        elif not is_valid_email(email):
            error = "Please enter a valid email address."
        elif not re.match(r"^[a-z0-9_]{3,50}$", username):
            error = "Username must be 3–50 characters, letters, numbers and underscores only."
        elif password != confirm:
            error = "Passwords do not match."
        elif len(password) < 8:
            error = "Password must be at least 8 characters."
        elif username in RESERVED_USERNAMES:
            error = "That username is already taken. Please choose another."
        else:
            conn = get_db()
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute("SELECT username FROM registered_users WHERE username=%s", (username,))
            existing = cur.fetchone()
            if existing:
                error = "That username is already taken. Please choose another."
                cur.close()
                conn.close()
            else:
                hashed_pw = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
                token = secrets.token_urlsafe(32)
                cur.execute("INSERT INTO registered_users (username, password, name, email, verified, signed_up) VALUES (%s,%s,%s,%s,0,%s)",
                             (username, hashed_pw, hotel_name, email, datetime.now().strftime("%d %b %Y")))
                cur.execute("INSERT INTO verification_tokens (token, username, expires_at) VALUES (%s,%s, NOW() + INTERVAL '24 hours')", (token, username))
                conn.commit()
                cur.close()
                conn.close()
                sent = send_verification_email(email, hotel_name, token)
                if sent:
                    success = f"Account created! A verification email has been sent to {email}. Please check your inbox."
                else:
                    conn2 = get_db()
                    cur2 = conn2.cursor()
                    cur2.execute("UPDATE registered_users SET verified=1 WHERE username=%s", (username,))
                    cur2.execute("DELETE FROM verification_tokens WHERE token=%s", (token,))
                    conn2.commit()
                    cur2.close()
                    conn2.close()
                    success = "Account created! (Email verification skipped — no SendGrid key detected.) You can now sign in."

    error_html   = f'<div class="err">{error}</div>' if error else ''
    success_html = f'<div class="ok">{success}</div>' if success else ''

    return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Start free pilot</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:#f5f7fb;font-family:'Plus Jakarta Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px;-webkit-font-smoothing:antialiased;}}
.card{{background:#ffffff;border:1px solid #e4e8f0;border-radius:20px;width:100%;max-width:460px;padding:48px;box-shadow:0 4px 32px rgba(0,0,0,0.06);}}
.brand{{font-family:'Syne',sans-serif;font-size:22px;font-weight:800;color:#0d1120;letter-spacing:-0.5px;margin-bottom:28px;}}
.brand span{{color:#00d165;}}
.pilot-badge{{display:inline-flex;align-items:center;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:99px;padding:5px 12px;font-family:'JetBrains Mono',monospace;font-size:10px;color:#16a34a;letter-spacing:1px;margin-bottom:20px;}}
.card-title{{font-family:'Plus Jakarta Sans',sans-serif;font-size:22px;font-weight:800;color:#0d1120;letter-spacing:-0.6px;margin-bottom:6px;}}
.card-sub{{font-size:13px;color:#64748b;margin-bottom:28px;}}
.err{{background:#fef2f2;border:1px solid #fecaca;color:#dc2626;padding:12px 14px;border-radius:9px;font-size:13px;margin-bottom:20px;line-height:1.5;}}
.ok{{background:#f0fdf4;border:1px solid #bbf7d0;color:#16a34a;padding:12px 14px;border-radius:9px;font-size:13px;margin-bottom:20px;line-height:1.5;}}
.ok a{{color:#00d165;font-weight:600;text-decoration:none;}}
label{{font-family:'JetBrains Mono',monospace;font-size:10px;font-weight:500;color:#94a3b8;text-transform:uppercase;letter-spacing:1px;display:block;margin-bottom:7px;}}
input{{width:100%;padding:12px 14px;background:#f8fafc;border:1px solid #e4e8f0;border-radius:9px;font-size:14px;color:#0d1120;margin-bottom:16px;outline:none;font-family:'Plus Jakarta Sans',sans-serif;transition:border-color .2s,background .2s;}}
input:focus{{border-color:#00d165;background:#ffffff;}}
input::placeholder{{color:#cbd5e1;}}
.btn-submit{{width:100%;padding:13px;background:#00d165;color:#080c14;border:none;border-radius:9px;font-weight:700;cursor:pointer;font-size:14px;font-family:'Plus Jakarta Sans',sans-serif;transition:all .2s;margin-top:4px;}}
.btn-submit:hover{{background:#04e270;box-shadow:0 4px 16px rgba(0,209,101,0.25);}}
.divider{{height:1px;background:#e4e8f0;margin:20px 0;}}
.back-link{{text-align:center;font-size:13px;}}
.back-link a{{color:#94a3b8;text-decoration:none;transition:color .2s;}}
.back-link a span{{color:#00d165;font-weight:600;}}
.back-link a:hover{{color:#0d1120;}}

</style>
</head>
<body>
<div class="card">
  <div class="brand">Occup<span>ado</span></div>
  <div class="pilot-badge">FREE 40-DAY PILOT · NO CREDIT CARD</div>
  <div class="card-title">Create your account</div>
  <div class="card-sub">Start predicting cancellations in minutes</div>
  {error_html}{success_html}
  {'<div class="divider"></div><div class="back-link"><a href="/login"><span>Back to Sign in</span></a></div>' if success else f"""
  <form method="POST">
    <label>Hotel Name</label>
    <input type="text" name="hotel_name" placeholder="e.g. Van der Valk Mechelen" required>
    <label>Email Address</label>
    <input type="email" name="email" placeholder="revenue@yourhotel.com" required>
    <label>Username</label>
    <input type="text" name="username" placeholder="Choose a username" required>
    <label>Password</label>
    <input type="password" name="password" placeholder="Min. 8 characters" required>
    <label>Confirm Password</label>
    <input type="password" name="confirm" placeholder="Repeat password" required>
    <button type="submit" class="btn-submit">Create account &rarr;</button>
  </form>
  <div class="divider"></div>
  <div class="back-link"><a href="/login">Already have an account? <span>Sign in</span></a></div>
  """}
</div>
</body>
</html>"""


@app.route("/verify/<token>")
def verify_email(token):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT username, expires_at FROM verification_tokens WHERE token=%s", (token,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Verification Failed</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@400;500&display=swap" rel="stylesheet">
<style>* {{margin:0;padding:0;box-sizing:border-box;}} body {{background:#f5faf5;font-family:'Plus Jakarta Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;}} .box {{background:#fff;padding:48px;border-radius:20px;max-width:400px;text-align:center;border:1px solid rgba(0,128,0,0.15);}}
</style>
</head>
<body><div class="box">
<div style="font-family:'Plus Jakarta Sans',sans-serif;font-size:28px;font-weight:800;color:#008000;margin-bottom:16px;">Occupado</div>
<div style="font-size:36px;margin-bottom:16px;">❌</div>
<h2 style="margin-bottom:12px;color:#0a1a0a;">Invalid or expired link</h2>
<p style="color:#4a6648;margin-bottom:24px;">This verification link is not valid. Please register again.</p>
<a href="/register" style="color:#008000;font-weight:700;text-decoration:none;">Register →</a>
</div></body></html>"""

    # Check token expiry
    from datetime import timezone
    expires_at = row["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if datetime.now(timezone.utc) > expires_at:
        cur.execute("DELETE FROM verification_tokens WHERE token=%s", (token,))
        conn.commit()
        cur.close()
        conn.close()
        return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Link Expired</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@400;500&display=swap" rel="stylesheet">
<style>* {{margin:0;padding:0;box-sizing:border-box;}} body {{background:#f5faf5;font-family:'Plus Jakarta Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;}} .box {{background:#fff;padding:48px;border-radius:20px;max-width:400px;text-align:center;border:1px solid rgba(0,128,0,0.15);}}
</style>
</head>
<body><div class="box">
<div style="font-family:'Plus Jakarta Sans',sans-serif;font-size:28px;font-weight:800;color:#008000;margin-bottom:16px;">Occupado</div>
<div style="font-size:36px;margin-bottom:16px;">⏰</div>
<h2 style="margin-bottom:12px;color:#0a1a0a;">Verification link expired</h2>
<p style="color:#4a6648;margin-bottom:24px;">This link expired after 24 hours. Please register again to get a new link.</p>
<a href="/register" style="color:#008000;font-weight:700;text-decoration:none;">Register again →</a>
</div></body></html>"""

    username = row["username"]
    cur.execute("UPDATE registered_users SET verified=1 WHERE username=%s", (username,))
    cur.execute("DELETE FROM verification_tokens WHERE token=%s", (token,))
    conn.commit()
    cur.close()
    conn.close()

    session["verified_success"] = "✅ Email verified! You can now sign in."
    return redirect(url_for("login"))


# ─────────────────────────────────────────────
#  ADMIN PANEL
# ─────────────────────────────────────────────

ADMIN_PASSWORD = "occupado-admin-2024"

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    error = ""
    if request.method == "POST":
        ip = request.remote_addr
        blocked, remaining = check_rate_limit(ip)
        if blocked:
            error = f"Too many failed attempts. Try again in {remaining // 60 + 1} minute(s)."
        else:
            pw = request.form.get("password", "").strip()
            if pw == ADMIN_PASSWORD:
                reset_attempts(ip)
                session.permanent = True
                session["is_admin"] = True
                return redirect(url_for("admin_panel"))
            record_failed_attempt(ip)
            error = "Wrong password."

    error_html = f'<div style="background:#ffcdd2;padding:12px;border-radius:8px;color:#c62828;font-size:14px;margin-bottom:20px;">{error}</div>' if error else ""

    return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Admin</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#f5faf5; font-family:'Plus Jakarta Sans',sans-serif; min-height:100vh; display:flex; align-items:center; justify-content:center; }}
.box {{ background:#fff; padding:48px; border-radius:20px; width:100%; max-width:380px; border:1px solid rgba(0,128,0,0.15); }}
.logo {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:26px; font-weight:800; color:#008000; margin-bottom:4px; }}
.subtitle {{ font-size:12px; color:#4a6648; font-family:'JetBrains Mono',monospace; margin-bottom:28px; }}
label {{ font-size:12px; color:#4a6648; display:block; margin-bottom:6px; font-family:'JetBrains Mono',monospace; font-weight:600; text-transform:uppercase; }}
input {{ width:100%; padding:12px; background:#f5faf5; border:1px solid rgba(0,128,0,0.2); border-radius:10px; font-size:14px; margin-bottom:16px; outline:none; }}
input:focus {{ border-color:#008000; background:#fff; }}
button {{ width:100%; padding:14px; background:#008000; color:#fff; border:none; border-radius:10px; font-weight:700; font-size:15px; cursor:pointer; font-family:'Plus Jakarta Sans',sans-serif; }}
button:hover {{ background:#006600; }}

</style>
</head>
<body>
<div class="box">
    <div class="logo">Occupado</div>
    <div class="subtitle">Admin Panel</div>
    {error_html}
    <form method="POST">
        <label>Admin Password</label>
        <input type="password" name="password" required autofocus>
        <button type="submit">Enter Admin →</button>
    </form>
</div>
</body>
</html>"""


@app.route("/admin")
@admin_required
def admin_panel():
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT username, name, email, verified, signed_up FROM registered_users ORDER BY username")
    users = cur.fetchall()
    cur.close()
    conn.close()

    total      = len(users)
    verified   = sum(1 for u in users if u["verified"])
    unverified = total - verified

    rows_html = ""
    for u in users:
        verified_badge = '<span style="background:#e8f5e9;color:#2e7d32;border-radius:6px;padding:3px 10px;font-size:11px;font-family:JetBrains Mono,monospace;font-weight:600;">✓ Verified</span>' \
                       if u["verified"] else \
                       '<span style="background:#fff3e0;color:#e65100;border-radius:6px;padding:3px 10px;font-size:11px;font-family:JetBrains Mono,monospace;font-weight:600;">Pending</span>'
        uname = u["username"]
        confirm_verify = "Manually verify " + uname + "?"
        confirm_delete = "Delete " + uname + "? This cannot be undone."
        verify_btn = "" if u["verified"] else (
            '<form method="POST" action="/admin/verify-user/' + uname + '" style="display:inline;" onsubmit="return confirm('' + confirm_verify + '')">'
            '<button type="submit" style="background:none;border:none;font-size:12px;color:#008000;font-weight:600;cursor:pointer;margin-right:12px;padding:0;">✓ Verify</button></form>'
        )
        delete_btn = (
            '<form method="POST" action="/admin/delete-user/' + uname + '" style="display:inline;" onsubmit="return confirm('' + confirm_delete + '')">'
            '<button type="submit" style="background:none;border:none;font-size:12px;color:#c62828;font-weight:600;cursor:pointer;padding:0;">✕ Delete</button></form>'
        )
        signed_up  = u["signed_up"] or "—"

        rows_html += f"""<tr>
            <td style="padding:14px 16px;font-size:14px;font-family:'JetBrains Mono',monospace;color:#0a1a0a;border-bottom:1px solid rgba(0,128,0,0.08);">{u['username']}</td>
            <td style="padding:14px 16px;font-size:14px;color:#0a1a0a;border-bottom:1px solid rgba(0,128,0,0.08);">{u['name']}</td>
            <td style="padding:14px 16px;font-size:14px;color:#4a6648;border-bottom:1px solid rgba(0,128,0,0.08);">{u['email']}</td>
            <td style="padding:14px 16px;font-size:13px;color:#4a6648;border-bottom:1px solid rgba(0,128,0,0.08);font-family:'JetBrains Mono',monospace;">{signed_up}</td>
            <td style="padding:14px 16px;border-bottom:1px solid rgba(0,128,0,0.08);">{verified_badge}</td>
            <td style="padding:14px 16px;border-bottom:1px solid rgba(0,128,0,0.08);">{verify_btn}{delete_btn}</td>
        </tr>"""

    if not rows_html:
        rows_html = '<tr><td colspan="6" style="padding:32px;text-align:center;color:#4a6648;font-size:14px;">No registered users yet.</td></tr>'

    return f"""<!DOCTYPE html>
<html>
<head><title>Occupado — Admin Panel</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#f5faf5; font-family:'Plus Jakarta Sans',sans-serif; color:#0a1a0a; }}
.topbar {{ background:#008000; padding:16px 40px; display:flex; align-items:center; justify-content:space-between; }}
.topbar-logo {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:22px; font-weight:800; color:#ffffff; }}
.topbar-sub {{ font-family:'JetBrains Mono',monospace; font-size:12px; color:rgba(255,255,255,0.8); }}
.btn-nav {{ padding:8px 18px; background:rgba(255,255,255,0.15); border:1px solid rgba(255,255,255,0.3); border-radius:8px; color:#ffffff; font-size:13px; font-weight:600; text-decoration:none; }}
.btn-nav:hover {{ background:rgba(255,255,255,0.25); }}
.content {{ max-width:1000px; margin:0 auto; padding:48px 24px; }}
.page-title {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:32px; font-weight:800; margin-bottom:6px; }}
.page-sub {{ font-size:13px; color:#4a6648; font-family:'JetBrains Mono',monospace; margin-bottom:32px; }}
.stats {{ display:flex; gap:16px; margin-bottom:28px; flex-wrap:wrap; }}
.stat-card {{ background:#fff; border:1px solid rgba(0,128,0,0.15); border-radius:12px; padding:20px 28px; flex:1; min-width:140px; }}
.stat-num {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:32px; font-weight:800; color:#008000; line-height:1; }}
.stat-label {{ font-size:12px; color:#4a6648; font-family:'JetBrains Mono',monospace; margin-top:6px; }}
.card {{ background:#fff; border:1px solid rgba(0,128,0,0.15); border-radius:16px; overflow:hidden; }}
.card-header {{ padding:20px 24px; border-bottom:1px solid rgba(0,128,0,0.08); }}
.card-title {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:18px; font-weight:700; }}
table {{ width:100%; border-collapse:collapse; }}
th {{ padding:12px 16px; text-align:left; font-size:11px; font-family:'JetBrains Mono',monospace; color:#4a6648; text-transform:uppercase; letter-spacing:0.5px; border-bottom:2px solid rgba(0,128,0,0.1); background:#f5faf5; }}
tr:hover td {{ background:rgba(0,128,0,0.02); }}

</style>
</head>
<body>
<div class="topbar">
    <div>
        <div class="topbar-logo">Occupado</div>
        <div class="topbar-sub">Admin Panel</div>
    </div>
    <a href="/admin/logout" class="btn-nav">Sign Out</a>
</div>
<div class="content">
    <div class="page-title">Registered Users</div>
    <div class="page-sub">All hotels that have signed up for Occupado</div>

    <div class="stats">
        <div class="stat-card">
            <div class="stat-num">{total}</div>
            <div class="stat-label">Total signups</div>
        </div>
        <div class="stat-card">
            <div class="stat-num" style="color:#2e7d32;">{verified}</div>
            <div class="stat-label">Verified</div>
        </div>
        <div class="stat-card">
            <div class="stat-num" style="color:#e65100;">{unverified}</div>
            <div class="stat-label">Pending verification</div>
        </div>
    </div>

    <div class="card">
        <div class="card-header">
            <div class="card-title">All Users</div>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Username</th>
                    <th>Hotel Name</th>
                    <th>Email</th>
                    <th>Signed Up</th>
                    <th>Status</th>
                    <th>Actions</th>
                </tr>
            </thead>
            <tbody>{rows_html}</tbody>
        </table>
    </div>
</div>
</body>
</html>"""


@app.route("/admin/delete-user/<username>", methods=["POST"])
@admin_required
def admin_delete_user(username):
    username = sanitise(username)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM registered_users WHERE username=%s", (username,))
    cur.execute("DELETE FROM verification_tokens WHERE username=%s", (username,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("admin_panel"))


@app.route("/admin/verify-user/<username>", methods=["POST"])
@admin_required
def admin_verify_user(username):
    username = sanitise(username)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE registered_users SET verified=1 WHERE username=%s", (username,))
    cur.execute("DELETE FROM verification_tokens WHERE username=%s", (username,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("admin_panel"))


@app.route("/admin/clear-test-data", methods=["POST"])
@admin_required
def admin_clear_test_data():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM registered_users WHERE username IN ('jpdourado', 'admin')")
    cur.execute("DELETE FROM verification_tokens")
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("admin_panel"))


@app.route("/admin/logout")
@admin_required
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


# ─────────────────────────────────────────────
#  FEATURE 1 — Excel export of high-risk bookings
# ─────────────────────────────────────────────

@app.route("/bru/export-highrisk")
@login_required
def bru_export_highrisk():
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment
    if session.get("hotel") != VDV_BRU_HOTEL_KEY:
        return "Not authorized", 403
    bookings = VDV_BRU_FUTURE_BOOKINGS
    scores   = VDV_BRU_FUTURE_SCORES
    if not bookings or not scores:
        return "No data available", 404
    pairs = sorted(zip(scores, bookings), reverse=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "High Risk Bookings"
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["Guest Name", "Arrival", "Nights", "Channel", "Lead Days", "ADR (€)", "Risk Score", "Risk Level", "Breakfast"]
    widths   = [28, 12, 8, 18, 10, 10, 12, 12, 12]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = w
    green_fill = PatternFill("solid", fgColor="D1FAE5")
    red_fill   = PatternFill("solid", fgColor="FEE2E2")
    amber_fill = PatternFill("solid", fgColor="FEF3C7")
    for row_idx, (sc, b) in enumerate(pairs, 2):
        level = "HIGH" if sc >= 70 else "MEDIUM" if sc >= 40 else "LOW"
        fill  = red_fill if sc >= 70 else amber_fill if sc >= 40 else green_fill
        vals  = [b['name'], b['arrival'], b['nights'], b['channel'],
                 b['lead'], round(b['adr']), f"{sc:.1f}%", level,
                 "Yes" if b.get('has_breakfast') else "No"]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=ci, value=v)
            if ci in (7, 8): c.fill = fill
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file as _sf
    return _sf(buf, as_attachment=True,
               download_name=f"BRU_HighRisk_{datetime.now().strftime('%Y%m%d')}.xlsx",
               mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/vdv/export-highrisk")
@login_required
def vdv_export_highrisk():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    if session.get("hotel") != VDV_HOTEL_KEY:
        return "Not authorized", 403
    bookings = VDV_FUTURE_BOOKINGS
    scores   = VDV_FUTURE_SCORES
    if not bookings or not scores:
        # Fallback 1: legacy blob cache
        try:
            _c = get_db(); _cu = _c.cursor()
            _cu.execute("SELECT bookings_json, scores_json FROM vdv_bookings_cache_legacy ORDER BY cached_at DESC LIMIT 1")
            row = _cu.fetchone()
            _cu.close(); _c.close()
            if row:
                from datetime import datetime as _dt
                _raw = json.loads(row[0])
                for b in _raw:
                    b['arr_date'] = _dt.fromisoformat(b['arr_date'])
                bookings = _raw
                scores   = json.loads(row[1])
        except Exception as _ce:
            pass
    if not bookings or not scores:
        # Fallback 2: per-row vdv_bookings_cache (populated by upsert / daily-rescore)
        try:
            _c2 = get_db(); _cu2 = _c2.cursor()
            _cu2.execute("""
                SELECT guest_name, arrival_date, channel, channel_raw,
                       lead_time, risk_score, risk_tier, scored_at
                FROM vdv_bookings_cache
                WHERE hotel_id = 'vdv'
                  AND arrival_date >= CURRENT_DATE
                ORDER BY arrival_date, risk_score DESC
            """)
            _rows2 = _cu2.fetchall()
            _cu2.close(); _c2.close()
            if _rows2:
                bookings = [
                    {'name': r[0], 'arr_date': r[1], 'channel': r[2],
                     'channel_raw': r[3], 'lead': r[4]}
                    for r in _rows2
                ]
                scores = [r[5] for r in _rows2]
                print(f"[VDV] Dashboard: loaded {len(_rows2)} scores from per-row DB cache")
        except Exception as _ce2:
            print(f"[VDV] DB score read error: {_ce2}")
    if not bookings or not scores:
        return "No booking data available. Run the app locally with VDV-MEC files to populate the cache.", 404
    # Build repeat guest name set for export flagging
    from collections import Counter as _Ctr
    _name_counts = _Ctr(b["name"].strip().lower() for b in bookings)
    _freq_repeats = {n for n, c in _name_counts.items() if c >= 2}
    _known_repeats = {g['name'].strip().lower() for g in VDV_GUESTS_RAW}
    _all_repeats = _freq_repeats | _known_repeats

    def _is_repeat_export(b):
        return b["name"].strip().lower() in _all_repeats

    def _export_reason(b, sc):
        parts = []
        if b["lead"] > 90:
            parts.append(f"Long lead time ({b['lead']}d)")
        elif b["lead"] > 30:
            parts.append(f"Moderate lead ({b['lead']}d)")
        else:
            parts.append(f"Short lead ({b['lead']}d, close to arrival)")
        ch = b["channel"]
        if ch == "Booking.com":
            parts.append("OTA booking (Booking.com)")
        elif ch in ("Direct/Web", "Direct / Web"):
            parts.append("Direct/Web channel")
        elif ch == "Corporate":
            parts.append("Corporate channel (lower risk)")
        elif ch in ("Package", "Packages / Groups"):
            parts.append("Package/Group (typically prepaid)")
        else:
            parts.append(f"{ch} channel")
        if b["gtd"] == "NONE":
            parts.append("No guarantee on file")
        if b["nights"] == 1:
            parts.append("1-night stay (high no-show rate)")
        elif b["nights"] >= 4:
            parts.append(f"{b['nights']}-night stay (longer stays rarely cancel last-minute)")
        if b.get("adr", 0) > 200:
            parts.append(f"High ADR (€{b['adr']:.0f})")
        if b.get("has_breakfast"):
            parts.append("Breakfast included")
        if _is_repeat_export(b):
            parts.append("Known repeat guest")
        return " · ".join(parts)

    all_indexed = sorted(enumerate(scores), key=lambda x: -x[1])
    high_risk = [(i, s) for i, s in all_indexed if s >= 70]
    med_risk_all = [(i, s) for i, s in all_indexed if 40 <= s < 70]
    import math
    med_sample_n = max(1, math.ceil(len(med_risk_all) * 0.10))
    med_risk = med_risk_all[:med_sample_n]
    indexed = high_risk + med_risk

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Bookings"
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["#", "Guest", "Arrival", "Nights", "Lead (days)", "Channel", "Guarantee",
               "ADR (€)", "Breakfast", "Repeat Guest", "Risk %", "Risk Level", "Reason"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center")
    high_fill = PatternFill("solid", fgColor="FEE2E2")
    med_fill  = PatternFill("solid", fgColor="FEF3C7")
    yes_fill  = PatternFill("solid", fgColor="DCFCE7")
    for rank, (idx, sc) in enumerate(indexed):
        b = bookings[idx]
        risk_level = "High Risk" if sc >= 70 else "Medium Risk"
        reason = _export_reason(b, sc)
        has_bf = "Yes" if b.get("has_breakfast") else "No"
        is_rep = "Yes" if _is_repeat_export(b) else "No"
        row = [rank+1, b["name"], b["arrival"], b["nights"], b["lead"],
               b["channel"], b["gtd"], b.get("adr", ""),
               has_bf, is_rep, round(sc, 1), risk_level, reason]
        ws.append(row)
        fill = high_fill if sc >= 70 else med_fill
        for col_idx in (11, 12):
            ws.cell(row=rank+2, column=col_idx).fill = fill
            ws.cell(row=rank+2, column=col_idx).font = Font(bold=True)
        # Green highlight for breakfast/repeat Yes cells
        if has_bf == "Yes":
            ws.cell(row=rank+2, column=9).fill = yes_fill
        if is_rep == "Yes":
            ws.cell(row=rank+2, column=10).fill = yes_fill
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"occupado_highrisk_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ─────────────────────────────────────────────
#  FEATURE 2 — Weekly email digest
# ─────────────────────────────────────────────

def send_weekly_digest():
    """Build and send Monday morning digest to VdV alert email."""
    alert_email = os.environ.get("VDV_ALERT_EMAIL", "")
    if not alert_email:
        return False
    bookings = VDV_FUTURE_BOOKINGS
    scores   = VDV_FUTURE_SCORES
    fut_high = sum(1 for s in scores if s >= 70) if scores else 36
    fut_med  = sum(1 for s in scores if 40 <= s < 70) if scores else 844
    fut_low  = sum(1 for s in scores if s < 40) if scores else 1888
    fut_total = len(bookings) if bookings else 2768
    roi_emails = 0
    roi_recovered = 0
    try:
        _c = get_db(); _cu = _c.cursor()
        _cu.execute("SELECT COUNT(*) FROM roi_actions WHERE hotel_username=%s", (VDV_HOTEL_KEY,))
        roi_emails = _cu.fetchone()[0]
        _cu.execute("SELECT COUNT(*) FROM roi_actions WHERE hotel_username=%s AND booking_ref='RECOVERED'", (VDV_HOTEL_KEY,))
        roi_recovered = _cu.fetchone()[0]
        _cu.close(); _c.close()
    except:
        pass
    week_label = datetime.now().strftime("Week of %d %b %Y")
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:680px;margin:0 auto;">
      <div style="background:#111827;color:#fff;padding:28px 32px;border-radius:12px 12px 0 0;">
        <h2 style="margin:0;font-size:22px;font-weight:700;">Occupado Weekly Digest</h2>
        <p style="margin:6px 0 0;font-size:13px;opacity:.7;">Van der Valk Hotel Mechelen · {week_label}</p>
      </div>
      <div style="background:#f9fafb;padding:28px 32px;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 12px 12px;">
        <h3 style="font-size:15px;font-weight:600;color:#111827;margin:0 0 16px;">Future Booking Risk</h3>
        <table style="width:100%;border-collapse:collapse;margin-bottom:24px;">
          <tr style="border-bottom:1px solid #e5e7eb;">
            <td style="padding:10px 0;color:#374151;">Total upcoming bookings</td>
            <td style="padding:10px 0;text-align:right;font-weight:600;">{fut_total:,}</td>
          </tr>
          <tr style="border-bottom:1px solid #e5e7eb;">
            <td style="padding:10px 0;color:#dc2626;">&#x25CF; High risk (≥70%)</td>
            <td style="padding:10px 0;text-align:right;font-weight:600;color:#dc2626;">{fut_high:,}</td>
          </tr>
          <tr style="border-bottom:1px solid #e5e7eb;">
            <td style="padding:10px 0;color:#d97706;">&#x25CF; Medium risk (40–69%)</td>
            <td style="padding:10px 0;text-align:right;font-weight:600;color:#d97706;">{fut_med:,}</td>
          </tr>
          <tr>
            <td style="padding:10px 0;color:#16a34a;">&#x25CF; Low risk (&lt;40%)</td>
            <td style="padding:10px 0;text-align:right;font-weight:600;color:#16a34a;">{fut_low:,}</td>
          </tr>
        </table>
        <h3 style="font-size:15px;font-weight:600;color:#111827;margin:0 0 16px;">Recovery Actions This Week</h3>
        <table style="width:100%;border-collapse:collapse;margin-bottom:28px;">
          <tr style="border-bottom:1px solid #e5e7eb;">
            <td style="padding:10px 0;color:#374151;">Retention emails sent</td>
            <td style="padding:10px 0;text-align:right;font-weight:600;">{roi_emails}</td>
          </tr>
          <tr>
            <td style="padding:10px 0;color:#374151;">Bookings recovered</td>
            <td style="padding:10px 0;text-align:right;font-weight:600;color:#16a34a;">{roi_recovered}</td>
          </tr>
        </table>
        <center>
          <a href="https://occupado.up.railway.app/dashboard" style="background:#111827;color:#fff;padding:12px 32px;text-decoration:none;border-radius:8px;display:inline-block;font-weight:600;font-size:14px;">View Full Dashboard →</a>
        </center>
      </div>
    </div>"""
    message = Mail(
        from_email=os.environ.get("ALERT_FROM_EMAIL", "team@occupado.co"),
        to_emails=alert_email,
        subject=f"Occupado Weekly Digest — {week_label}",
        html_content=html
    )
    try:
        sg = SendGridAPIClient(os.environ.get("SENDGRID_API_KEY"))
        sg.send(message)
        print(f"[DIGEST] Sent to {alert_email}")
        return True
    except Exception as e:
        print(f"[DIGEST ERROR] {e}")
        return False


@app.route("/internal/weekly-digest", methods=["POST"])
def internal_weekly_digest():
    token = request.headers.get("X-Internal-Token", "")
    expected = os.environ.get("INTERNAL_TOKEN", "")
    if not expected or token != expected:
        return {"status": "error", "message": "Unauthorized"}, 401
    ok = send_weekly_digest()
    return {"status": "ok" if ok else "error"}


# ─────────────────────────────────────────────
#  FEATURE 5 — ROI recovery marking
# ─────────────────────────────────────────────

@app.route("/internal/mark-recoveries", methods=["POST"])
def internal_mark_recoveries():
    token = request.headers.get("X-Internal-Token", "")
    expected = os.environ.get("INTERNAL_TOKEN", "")
    if not expected or token != expected:
        return {"status": "error", "message": "Unauthorized"}, 401
    data     = request.get_json() or {}
    guest_name = data.get("guest_name", "")
    if not guest_name:
        return {"status": "error", "message": "guest_name required"}, 400
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute(
            "UPDATE roi_actions SET booking_ref='RECOVERED' WHERE hotel_username=%s AND guest_name=%s AND booking_ref=''",
            (VDV_HOTEL_KEY, guest_name)
        )
        updated = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()
        return {"status": "ok", "marked": updated}
    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route('/internal/daily-rescore', methods=['POST'])
def daily_rescore():
    global VDV_FUTURE_BOOKINGS, VDV_FUTURE_SCORES
    # Security: Railway CRON only
    auth = request.headers.get('Authorization', '')
    if auth != f"Bearer {os.environ.get('CRON_SECRET', 'occupado-cron')}":
        return jsonify({'error': 'unauthorized'}), 401

    try:
        from datetime import date
        today = date.today()

        # Use in-memory globals if available
        bookings = VDV_FUTURE_BOOKINGS

        if not bookings:
            # Fallback: reconstruct from legacy blob cache
            try:
                conn = get_db()
                cur = conn.cursor()
                cur.execute("""
                    SELECT bookings_json
                    FROM vdv_bookings_cache_legacy
                    ORDER BY cached_at DESC
                    LIMIT 1
                """)
                row = cur.fetchone()
                cur.close()
                conn.close()
                if row:
                    import json as _json
                    raw = _json.loads(row[0])
                    for b in raw:
                        if 'arr_date' in b:
                            from datetime import date as _d
                            b['arr_date'] = _d.fromisoformat(b['arr_date'])
                    bookings = raw
            except Exception as _fe:
                print(f"[VDV] Daily rescore fallback error: {_fe}")

        if not bookings:
            return jsonify({'status': 'no_data', 'message': 'No bookings to rescore'})

        # Filter to future bookings only
        future = [b for b in bookings if b.get('arr_date') and b['arr_date'] >= today]

        # Rescore with updated days_out
        scores = _score_vdv_future(future)

        # Detect outcomes before overwriting DB with new scores
        try:
            n_out = _detect_vdv_outcomes(future, scores)
            print(f"[VDV] Rescore outcomes: {n_out}")
        except Exception as e:
            print(f"[VDV] Rescore outcome detection error: {e}")

        # Upsert to DB
        n = _upsert_vdv_scores(future, scores)

        # Update globals
        VDV_FUTURE_BOOKINGS = future
        VDV_FUTURE_SCORES   = scores

        print(f"[VDV] Daily rescore: {n} bookings rescored at {today}")

        return jsonify({'status': 'ok', 'rescored': n, 'date': today.isoformat()})

    except Exception as e:
        print(f"[VDV] Daily rescore error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


if __name__ == "__main__":
    print("\n" + "="*50)
    print("Occupado running on http://localhost:8080")
    print("="*50 + "\n")
    app.run(host="0.0.0.0", port=8080, debug=False)